import pytest
import os
import json
from openpyxl import load_workbook


class TestWorkbookBuilderInit:
    """Tests for WorkbookBuilder initialization."""

    def test_init(self, temp_workbook):
        """Test basic initialization."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)

        assert builder.workbook_path == temp_workbook


class TestWorkbookBuilderCreateTemplate:
    """Tests for template workbook creation."""

    def test_create_template_workbook(self, temp_workbook):
        """Test creating a template workbook."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)
        result = builder.create_template_workbook()

        assert result == temp_workbook
        assert os.path.exists(temp_workbook)

    def test_template_has_config_sheet(self, temp_workbook):
        """Test that template has config sheet."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)
        builder.create_template_workbook()

        wb = load_workbook(temp_workbook)
        assert "config" in wb.sheetnames

    def test_template_has_prompts_sheet(self, temp_workbook):
        """Test that template has prompts sheet."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)
        builder.create_template_workbook()

        wb = load_workbook(temp_workbook)
        assert "prompts" in wb.sheetnames

    def test_template_config_has_required_fields(self, temp_workbook):
        """Test that config sheet has required fields."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)
        builder.create_template_workbook()

        wb = load_workbook(temp_workbook)
        ws = wb["config"]

        fields = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}

        assert "model" in fields
        assert "api_key_env" in fields
        assert "max_retries" in fields
        assert "temperature" in fields
        assert "max_tokens" in fields
        assert "system_instructions" in fields

    def test_template_prompts_has_required_headers(self, temp_workbook):
        """Test that prompts sheet has required headers."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)
        builder.create_template_workbook()

        wb = load_workbook(temp_workbook)
        ws = wb["prompts"]

        headers = [cell.value for cell in ws[1]]

        assert "sequence" in headers
        assert "prompt_name" in headers
        assert "prompt" in headers
        assert "history" in headers


class TestWorkbookBuilderValidate:
    """Tests for workbook validation."""

    def test_validate_workbook_success(self, temp_workbook_with_data):
        """Test validating a valid workbook."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)
        result = builder.validate_workbook()

        assert result is True

    def test_validate_workbook_not_found(self, temp_workbook):
        """Test validating non-existent workbook."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)

        with pytest.raises(FileNotFoundError):
            builder.validate_workbook()

    def test_validate_workbook_missing_config(self, temp_workbook):
        """Test validating workbook missing config sheet."""
        from src.orchestrator.workbook_builder import WorkbookBuilder
        from openpyxl import Workbook

        wb = Workbook()
        wb.create_sheet("prompts")
        del wb["Sheet"]
        wb.save(temp_workbook)

        builder = WorkbookBuilder(temp_workbook)

        with pytest.raises(ValueError, match="Missing required sheet: config"):
            builder.validate_workbook()

    def test_validate_workbook_missing_prompts(self, temp_workbook):
        """Test validating workbook missing prompts sheet."""
        from src.orchestrator.workbook_builder import WorkbookBuilder
        from openpyxl import Workbook

        wb = Workbook()
        wb.create_sheet("config")
        del wb["Sheet"]
        wb.save(temp_workbook)

        builder = WorkbookBuilder(temp_workbook)

        with pytest.raises(ValueError, match="Missing required sheet: prompts"):
            builder.validate_workbook()

    def test_validate_workbook_missing_columns(self, temp_workbook):
        """Test validating workbook missing required columns."""
        from src.orchestrator.workbook_builder import WorkbookBuilder
        from openpyxl import Workbook

        wb = Workbook()
        wb.create_sheet("config")
        wb.create_sheet("prompts")
        ws = wb["prompts"]
        ws["A1"] = "sequence"
        del wb["Sheet"]
        wb.save(temp_workbook)

        builder = WorkbookBuilder(temp_workbook)

        with pytest.raises(ValueError, match="Missing required columns"):
            builder.validate_workbook()


class TestWorkbookBuilderLoadConfig:
    """Tests for loading configuration."""

    def test_load_config(self, temp_workbook_with_data, sample_config):
        """Test loading configuration from workbook."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)
        config = builder.load_config()

        assert config["model"] == sample_config["model"]
        assert config["api_key_env"] == sample_config["api_key_env"]
        assert config["max_retries"] == sample_config["max_retries"]
        assert config["temperature"] == sample_config["temperature"]

    def test_load_config_type_conversion(self, temp_workbook):
        """Test that config values are converted to correct types."""
        from src.orchestrator.workbook_builder import WorkbookBuilder
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "max_retries"
        ws["B2"] = "5"
        ws["A3"] = "temperature"
        ws["B3"] = "0.5"
        ws["A4"] = "max_tokens"
        ws["B4"] = "8192"
        wb.save(temp_workbook)

        builder = WorkbookBuilder(temp_workbook)
        config = builder.load_config()

        assert isinstance(config["max_retries"], int)
        assert isinstance(config["temperature"], float)
        assert isinstance(config["max_tokens"], int)


class TestWorkbookBuilderLoadPrompts:
    """Tests for loading prompts."""

    def test_load_prompts(self, temp_workbook_with_data, sample_prompts):
        """Test loading prompts from workbook."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)
        prompts = builder.load_prompts()

        assert len(prompts) == 3
        assert prompts[0]["prompt_name"] == "greeting"
        assert prompts[1]["prompt_name"] == "math"
        assert prompts[2]["prompt_name"] == "followup"

    def test_load_prompts_sorted_by_sequence(self, temp_workbook_with_data):
        """Test that prompts are sorted by sequence number."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)
        prompts = builder.load_prompts()

        sequences = [p["sequence"] for p in prompts]
        assert sequences == sorted(sequences)

    def test_load_prompts_parses_history(self, temp_workbook_with_data):
        """Test that history column is parsed correctly."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)
        prompts = builder.load_prompts()

        followup = next(p for p in prompts if p["prompt_name"] == "followup")
        assert followup["history"] == ["math", "greeting"]


class TestWorkbookBuilderParseHistoryString:
    """Tests for parsing history strings."""

    def test_parse_history_string_json_array(self, temp_workbook):
        """Test parsing JSON array format."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)

        result = builder.parse_history_string('["a", "b", "c"]')

        assert result == ["a", "b", "c"]

    def test_parse_history_string_quoted(self, temp_workbook):
        """Test parsing quoted format."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)

        result = builder.parse_history_string('["math", "greeting"]')

        assert result == ["math", "greeting"]

    def test_parse_history_string_none(self, temp_workbook):
        """Test parsing None."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)

        result = builder.parse_history_string(None)

        assert result is None

    def test_parse_history_string_empty(self, temp_workbook):
        """Test parsing empty string."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)

        result = builder.parse_history_string("")

        assert result is None

    def test_parse_history_string_already_list(self, temp_workbook):
        """Test parsing already-list input."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)

        result = builder.parse_history_string(["a", "b"])

        assert result == ["a", "b"]

    def test_parse_history_string_comma_separated(self, temp_workbook):
        """Test parsing comma-separated format."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook)

        result = builder.parse_history_string("[a, b, c]")

        assert result == ["a", "b", "c"]


class TestWorkbookBuilderWriteResults:
    """Tests for writing results."""

    def test_write_results(self, temp_workbook_with_data):
        """Test writing results to workbook."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)

        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello?",
                "history": None,
                "response": "Hi!",
                "status": "success",
                "attempts": 1,
                "error": None,
            }
        ]

        sheet_name = builder.write_results(results, "results_test")

        assert sheet_name == "results_test"

        wb = load_workbook(temp_workbook_with_data)
        assert "results_test" in wb.sheetnames

    def test_write_results_creates_new_sheet_name(self, temp_workbook_with_data):
        """Test that duplicate sheet names get timestamp suffix."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)

        builder.write_results([], "results_test")
        sheet_name = builder.write_results([], "results_test")

        assert sheet_name != "results_test"

    def test_write_results_has_all_columns(self, temp_workbook_with_data):
        """Test that results sheet has all required columns."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)

        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello?",
                "history": ["a"],
                "response": "Hi!",
                "status": "success",
                "attempts": 1,
                "error": None,
            }
        ]

        builder.write_results(results, "results_test")

        wb = load_workbook(temp_workbook_with_data)
        ws = wb["results_test"]

        headers = [cell.value for cell in ws[1]]

        assert "sequence" in headers
        assert "prompt_name" in headers
        assert "prompt" in headers
        assert "history" in headers
        assert "response" in headers
        assert "status" in headers
        assert "attempts" in headers
        assert "error" in headers

    def test_write_results_converts_history_to_json(self, temp_workbook_with_data):
        """Test that history is converted to JSON string."""
        from src.orchestrator.workbook_builder import WorkbookBuilder

        builder = WorkbookBuilder(temp_workbook_with_data)

        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello?",
                "history": ["a", "b"],
                "response": "Hi!",
                "status": "success",
                "attempts": 1,
                "error": None,
            }
        ]

        builder.write_results(results, "results_test")

        wb = load_workbook(temp_workbook_with_data)
        ws = wb["results_test"]

        history_cell = ws.cell(row=2, column=4).value

        assert history_cell == '["a", "b"]'
