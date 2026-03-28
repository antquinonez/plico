# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

import os
from unittest.mock import MagicMock

import pytest


class TestExcelOrchestratorInit:
    """Tests for ExcelOrchestrator initialization."""

    def test_init(self, temp_workbook, mock_ffmistralsmall):
        """Test basic initialization."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        assert orchestrator.workbook_path == temp_workbook
        assert orchestrator.client == mock_ffmistralsmall
        assert orchestrator.results == []

    def test_init_with_config_overrides(self, temp_workbook, mock_ffmistralsmall):
        """Test initialization with config overrides."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            temp_workbook, mock_ffmistralsmall, config_overrides={"temperature": 0.5}
        )

        assert orchestrator.config_overrides == {"temperature": 0.5}


class TestExcelOrchestratorWorkbookInit:
    """Tests for workbook initialization."""

    def test_init_workbook_creates_new(self, temp_workbook, mock_ffmistralsmall):
        """Test that missing workbook is created."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()

        assert os.path.exists(temp_workbook)

    def test_init_workbook_validates_existing(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that existing workbook is validated."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)

        assert orchestrator._init_workbook() is None


class TestExcelOrchestratorConfig:
    """Tests for configuration loading."""

    def test_load_config(self, temp_workbook_with_data, mock_ffmistralsmall, sample_config):
        """Test loading configuration."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._load_config()

        assert orchestrator.config["model"] == sample_config["model"]
        assert orchestrator.config["max_retries"] == sample_config["max_retries"]

    def test_load_config_with_overrides(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that config overrides are applied."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_data,
            mock_ffmistralsmall,
            config_overrides={"temperature": 0.1, "max_retries": 10},
        )
        orchestrator._load_config()

        assert orchestrator.config["temperature"] == 0.1
        assert orchestrator.config["max_retries"] == 10


class TestExcelOrchestratorDependencyValidation:
    """Tests for dependency validation."""

    def test_validate_dependencies_success(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test successful dependency validation."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator.config = {}

        orchestrator._validate()

    def test_validate_dependencies_missing_reference(self, temp_workbook, mock_ffmistralsmall):
        """Test validation fails for missing dependency reference."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["A2"] = 1
        ws["B2"] = "test"
        ws["C2"] = "Hello"
        ws["D2"] = '["nonexistent"]'
        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator.config = {}

        with pytest.raises(ValueError, match="Validation failed"):
            orchestrator._validate()

    def test_validate_dependencies_wrong_order(self, temp_workbook, mock_ffmistralsmall):
        """Test validation fails when dependency defined after use."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = '["second"]'
        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "World"
        ws["D3"] = ""
        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator.config = {}

        with pytest.raises(ValueError, match="must come before"):
            orchestrator._validate()


class TestExcelOrchestratorExecute:
    """Tests for prompt execution."""

    def test_execute_single_prompt(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test executing a single prompt."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()[:1]
        orchestrator._init_client()

        results = orchestrator.execute()

        assert len(results) == 1
        assert results[0]["status"] == "success"

    def test_execute_all_prompts(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test executing all prompts."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute()

        assert len(results) == 3
        assert all(r["status"] == "success" for r in results)

    def test_execute_with_history(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test executing prompt with history dependencies."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute()

        followup = next(r for r in results if r["prompt_name"] == "followup")
        assert followup["history"] == ["math", "greeting"]

    def test_execute_handles_failure(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that execution handles failures gracefully."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        mock_ffmistralsmall.generate_response = MagicMock(side_effect=Exception("API Error"))
        mock_ffmistralsmall.clone = MagicMock(return_value=mock_ffmistralsmall)

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()[:1]
        orchestrator._init_client()

        results = orchestrator.execute()

        assert results[0]["status"] == "failed"
        assert "API Error" in results[0]["error"]

    def test_execute_retries_on_failure(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that execution retries on failure."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        call_count = [0]

        def flaky_generate(prompt, **kwargs):
            call_count[0] += 1
            if call_count[0] < 3:
                raise Exception("Temporary error")
            return "Success!"

        mock_ffmistralsmall.generate_response = flaky_generate
        mock_ffmistralsmall.clone = MagicMock(return_value=mock_ffmistralsmall)

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()[:1]
        orchestrator._init_client()

        results = orchestrator.execute()

        assert results[0]["status"] == "success"
        assert results[0]["attempts"] == 3


class TestExcelOrchestratorRun:
    """Tests for the main run method."""

    def test_run_creates_results_sheet(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that run creates a results sheet."""
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        results_sheet = orchestrator.run()

        assert results_sheet.startswith("results_")

        wb = load_workbook(temp_workbook_with_data)
        assert results_sheet in wb.sheetnames

    def test_run_returns_sheet_name(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that run returns the results sheet name."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        results_sheet = orchestrator.run()

        assert isinstance(results_sheet, str)
        assert results_sheet.startswith("results_")


class TestExcelOrchestratorSummary:
    """Tests for execution summary."""

    def test_get_summary_before_run(self, temp_workbook, mock_ffmistralsmall):
        """Test getting summary before execution."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        summary = orchestrator.get_summary()

        assert summary["status"] == "not_run"

    def test_get_summary_after_run(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test getting summary after execution."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["total_prompts"] == 3
        assert summary["successful"] == 3
        assert summary["failed"] == 0
        assert "workbook" in summary

    def test_get_summary_with_failures(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test getting summary with failures."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        mock_ffmistralsmall.generate_response = MagicMock(side_effect=Exception("Error"))
        mock_ffmistralsmall.clone = MagicMock(return_value=mock_ffmistralsmall)

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 0
        assert summary["failed"] == 3


class TestExcelOrchestratorParallelExecution:
    """Tests for parallel execution."""

    def test_build_execution_graph(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test building execution graph with dependency levels."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator.prompts = orchestrator.builder.load_prompts()

        nodes = orchestrator._build_execution_graph()

        assert len(nodes) == 3

    def test_parallel_execution_basic(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test parallel execution with mock client."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_data, mock_ffmistralsmall, concurrency=2
        )
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute_parallel()

        assert len(results) == 3
        assert all(r["status"] == "success" for r in results)

    def test_parallel_execution_respects_dependencies(self, temp_workbook, mock_ffmistralsmall):
        """Test that parallel execution respects dependencies via execution graph."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Task A"

        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "Task B"

        ws["A4"] = 3
        ws["B4"] = "third"
        ws["C4"] = "Task C depends on first"
        ws["D4"] = '["first"]'

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall, concurrency=2)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()

        nodes = orchestrator._build_execution_graph()

        assert nodes[1].level == 0
        assert nodes[2].level == 0
        assert nodes[3].level == 1

        assert nodes[1].dependencies == set()
        assert nodes[2].dependencies == set()
        assert nodes[3].dependencies == {1}

    def test_concurrency_parameter_clamping(self, temp_workbook, mock_ffmistralsmall):
        """Test that concurrency is clamped to valid range."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall, concurrency=0)
        assert orchestrator.concurrency == 1

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall, concurrency=15)
        assert orchestrator.concurrency == 10

    def test_sequential_when_concurrency_is_one(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that concurrency=1 uses sequential execution."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_data, mock_ffmistralsmall, concurrency=1
        )
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute()

        assert len(results) == 3
        assert all(r["status"] == "success" for r in results)


class TestExcelOrchestratorBatchVariableResolution:
    """Tests for variable resolution in batch mode."""

    def test_resolve_variables_basic(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test basic variable resolution."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)

        data_row = {"region": "north", "product": "widget_a", "price": 10}

        result = orchestrator._resolve_variables(
            "Analyze {{region}} region, product {{product}}.", data_row
        )

        assert result == "Analyze north region, product widget_a."

    def test_resolve_variables_multiple_occurrences(
        self, temp_workbook_with_batch_data, mock_ffmistralsmall
    ):
        """Test variable resolution with multiple occurrences."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)

        data_row = {"region": "east", "price": 20}

        result = orchestrator._resolve_variables(
            "{{region}} has price {{price}} and tax on {{price}}.", data_row
        )

        assert result == "east has price 20 and tax on 20."

    def test_resolve_variables_missing_keeps_placeholder(
        self, temp_workbook_with_batch_data, mock_ffmistralsmall
    ):
        """Test that missing variables keep placeholder."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)

        data_row = {"region": "north"}

        result = orchestrator._resolve_variables(
            "Region: {{region}}, Missing: {{unknown}}", data_row
        )

        assert result == "Region: north, Missing: {{unknown}}"

    def test_resolve_prompt_variables(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test resolving variables in prompt dict."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)

        prompt = {
            "sequence": 1,
            "prompt_name": "test",
            "prompt": "Price is {{price}}, quantity is {{quantity}}.",
        }
        data_row = {"price": 10, "quantity": 100}

        result = orchestrator._resolve_prompt_variables(prompt, data_row)

        assert result["prompt"] == "Price is 10, quantity is 100."
        assert result["sequence"] == 1

    def test_resolve_batch_name_with_template(
        self, temp_workbook_with_batch_data, mock_ffmistralsmall
    ):
        """Test batch name resolution with template."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)

        data_row = {
            "region": "north",
            "product": "widget_a",
            "batch_name": "{{region}}_{{product}}",
        }

        result = orchestrator._resolve_batch_name(data_row, 1)

        assert result == "north_widget_a"

    def test_resolve_batch_name_default(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test batch name resolution with default."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)

        data_row = {"region": "north", "product": "widget_a"}

        result = orchestrator._resolve_batch_name(data_row, 5)

        assert result == "batch_5"


class TestExcelOrchestratorBatchExecution:
    """Tests for batch execution."""

    def test_batch_mode_detection(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test that batch mode is detected when data sheet exists."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.batch_data = orchestrator.builder.load_data()

        assert orchestrator.is_batch_mode is False
        orchestrator.is_batch_mode = len(orchestrator.batch_data) > 0
        assert orchestrator.is_batch_mode is True

    def test_execute_batch_single_batch(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test executing a single batch."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()
        orchestrator.batch_data = orchestrator.builder.load_data()[:1]

        results = orchestrator.execute_batch()

        assert len(results) == 3
        assert all(r["batch_id"] == 1 for r in results)
        assert all(r["batch_name"] == "north_widget_a" for r in results)

    def test_execute_batch_all_batches(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test executing all batches."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()
        orchestrator.batch_data = orchestrator.builder.load_data()

        results = orchestrator.execute_batch()

        assert len(results) == 9
        batch_ids = set(r["batch_id"] for r in results)
        assert batch_ids == {1, 2, 3}

    def test_execute_batch_parallel(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test parallel batch execution."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_batch_data, mock_ffmistralsmall, concurrency=2
        )
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()
        orchestrator.batch_data = orchestrator.builder.load_data()

        results = orchestrator.execute_batch_parallel()

        assert len(results) == 9
        assert all(r["status"] == "success" for r in results)

    def test_execute_batch_with_error_continue(
        self, temp_workbook_with_batch_data, mock_ffmistralsmall
    ):
        """Test batch execution continues on error."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        call_count = [0]

        def failing_generate(prompt, **kwargs):
            call_count[0] += 1
            if call_count[0] <= 3:
                raise Exception("API Error")
            return "Success"

        mock_ffmistralsmall.generate_response = failing_generate
        mock_ffmistralsmall.clone = MagicMock(return_value=mock_ffmistralsmall)

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_batch_data,
            mock_ffmistralsmall,
            config_overrides={"on_batch_error": "continue"},
        )
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()
        orchestrator.batch_data = orchestrator.builder.load_data()

        results = orchestrator.execute_batch()

        assert len(results) == 9
        assert any(r["status"] == "failed" for r in results)


class TestExcelOrchestratorBatchRun:
    """Tests for the run method with batch mode."""

    def test_run_batch_mode_creates_combined_results(
        self, temp_workbook_with_batch_data, mock_ffmistralsmall
    ):
        """Test that batch run creates combined results sheet."""
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)
        results_sheet = orchestrator.run()

        assert results_sheet.startswith("results_")

        wb = load_workbook(temp_workbook_with_batch_data)
        assert results_sheet in wb.sheetnames

        ws = wb[results_sheet]
        headers = [cell.value for cell in ws[1]]
        assert "batch_id" in headers
        assert "batch_name" in headers

    def test_run_batch_mode_summary(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test that batch run summary includes batch info."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)
        orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["batch_mode"] is True
        assert summary["total_batches"] == 3
        assert summary["total_prompts"] == 9

    def test_run_non_batch_mode(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that non-batch mode still works correctly."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator.run()
        summary = orchestrator.get_summary()

        assert "batch_mode" not in summary or summary.get("batch_mode") is False
        assert summary["total_prompts"] == 3


class TestExcelOrchestratorBoolParsing:
    """Tests for boolean override parsing."""

    def test_parse_bool_from_string_true(self, temp_workbook, mock_ffmistralsmall):
        """Test parsing 'true', 'yes', '1' as True."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        assert orchestrator._parse_bool_override("true") is True
        assert orchestrator._parse_bool_override("True") is True
        assert orchestrator._parse_bool_override("TRUE") is True
        assert orchestrator._parse_bool_override("yes") is True
        assert orchestrator._parse_bool_override("YES") is True
        assert orchestrator._parse_bool_override("1") is True

    def test_parse_bool_from_string_false(self, temp_workbook, mock_ffmistralsmall):
        """Test parsing 'false', 'no', '0' as False."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        assert orchestrator._parse_bool_override("false") is False
        assert orchestrator._parse_bool_override("False") is False
        assert orchestrator._parse_bool_override("FALSE") is False
        assert orchestrator._parse_bool_override("no") is False
        assert orchestrator._parse_bool_override("NO") is False
        assert orchestrator._parse_bool_override("0") is False

    def test_parse_bool_from_bool(self, temp_workbook, mock_ffmistralsmall):
        """Test passthrough of boolean values."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        assert orchestrator._parse_bool_override(True) is True
        assert orchestrator._parse_bool_override(False) is False

    def test_parse_bool_none_returns_none(self, temp_workbook, mock_ffmistralsmall):
        """Test that None returns None."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        assert orchestrator._parse_bool_override(None) is None

    def test_parse_bool_invalid_returns_none(self, temp_workbook, mock_ffmistralsmall):
        """Test invalid values return None."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        assert orchestrator._parse_bool_override("invalid") is None
        assert orchestrator._parse_bool_override("maybe") is None
        assert orchestrator._parse_bool_override(123) is None


class TestExcelOrchestratorConditions:
    """Tests for conditional execution."""

    def test_execute_skips_when_condition_false(self, temp_workbook, mock_ffmistralsmall):
        """Test that prompts are skipped when condition is false."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["E1"] = "condition"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = ""
        ws["E2"] = ""

        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "World"
        ws["D3"] = ""
        ws["E3"] = '{{first.response}} == "nonexistent"'

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute()

        assert results[0]["status"] == "success"
        assert results[1]["status"] == "skipped"
        assert results[1]["condition_result"] is False

    def test_condition_dependencies_in_graph(self, temp_workbook, mock_ffmistralsmall):
        """Test that condition references are added to dependency graph."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["E1"] = "condition"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = ""
        ws["E2"] = ""

        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "World"
        ws["D3"] = ""
        ws["E3"] = '{{first.status}} == "success"'

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator.prompts = orchestrator.builder.load_prompts()

        nodes = orchestrator._build_execution_graph()

        assert 1 in nodes[2].dependencies

    def test_condition_error_captured(self, temp_workbook, mock_ffmistralsmall):
        """Test that condition evaluation errors are captured."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["E1"] = "condition"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = ""
        ws["E2"] = ""

        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "World"
        ws["D3"] = ""
        ws["E3"] = "{{nonexistent.status}}.invalid_method()"

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute()

        assert results[1]["condition_error"] is not None

    def test_summary_includes_condition_count(self, temp_workbook, mock_ffmistralsmall):
        """Test that summary includes prompts_with_conditions."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["E1"] = "condition"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = ""
        ws["E2"] = ""

        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "World"
        ws["D3"] = ""
        ws["E3"] = '{{first.status}} == "success"'

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator.run()
        summary = orchestrator.get_summary()

        assert "prompts_with_conditions" in summary
        assert summary["prompts_with_conditions"] == 1


class TestExcelOrchestratorBatchEdgeCases:
    """Tests for batch execution edge cases."""

    def test_batch_error_stop_mode(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test that on_batch_error: stop halts execution."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        mock_ffmistralsmall.generate_response = MagicMock(side_effect=Exception("API Error"))
        mock_ffmistralsmall.clone = MagicMock(return_value=mock_ffmistralsmall)

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_batch_data,
            mock_ffmistralsmall,
            config_overrides={"on_batch_error": "stop"},
        )
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()
        orchestrator.batch_data = orchestrator.builder.load_data()

        results = orchestrator.execute_batch()

        assert len(results) < 9

    def test_batch_output_separate_sheets(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test batch_output: separate_sheets creates multiple sheets."""
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_batch_data,
            mock_ffmistralsmall,
            config_overrides={"batch_output": "separate_sheets"},
        )
        sheet_names = orchestrator.run()

        assert "," in sheet_names

        wb = load_workbook(temp_workbook_with_batch_data)
        sheets = [s.strip() for s in sheet_names.split(",")]
        for sheet in sheets:
            assert sheet in wb.sheetnames

    def test_batch_summary_with_failures(self, temp_workbook_with_batch_data, mock_ffmistralsmall):
        """Test that summary includes batches_with_failures."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        mock_ffmistralsmall.generate_response = MagicMock(side_effect=Exception("API Error"))
        mock_ffmistralsmall.clone = MagicMock(return_value=mock_ffmistralsmall)

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_batch_data,
            mock_ffmistralsmall,
            config_overrides={"max_retries": 1},
        )
        orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["failed"] > 0
        assert "batches_with_failures" in summary


class TestExcelOrchestratorParallelEdgeCases:
    """Tests for parallel execution edge cases."""

    def test_parallel_skipped_count(self, temp_workbook, mock_ffmistralsmall):
        """Test that skipped prompts are counted correctly in parallel execution."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["E1"] = "condition"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = ""
        ws["E2"] = ""

        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "World"
        ws["D3"] = ""
        ws["E3"] = '{{first.response}} == "nonexistent"'

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall, concurrency=2)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute_parallel()

        skipped = [r for r in results if r["status"] == "skipped"]
        assert len(skipped) == 1

    def test_parallel_exception_in_thread(self, temp_workbook, mock_ffmistralsmall):
        """Test that exceptions in threads are captured properly."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = ""

        wb.save(temp_workbook)

        mock_ffmistralsmall.generate_response = MagicMock(side_effect=RuntimeError("Thread error"))
        mock_ffmistralsmall.clone = MagicMock(return_value=mock_ffmistralsmall)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall, concurrency=2)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute_parallel()

        assert len(results) == 1
        assert results[0]["status"] == "failed"
        assert "Thread error" in results[0]["error"]


class TestExcelOrchestratorMultiClient:
    """Tests for multi-client registry."""

    def test_init_client_registry_loads_clients(self, temp_workbook, mock_ffmistralsmall):
        """Test that client registry is initialized from workbook."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["A2"] = 1
        ws["B2"] = "test"
        ws["C2"] = "Hello"
        ws["D2"] = ""

        ws = wb.create_sheet("clients")
        ws["A1"] = "name"
        ws["B1"] = "client_type"
        ws["C1"] = "model"
        ws["A2"] = "fast"
        ws["B2"] = "mistral-small"
        ws["C2"] = "mistral-small-2503"

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator._init_client_registry()

        assert orchestrator.has_multi_client is True
        assert orchestrator.client_registry is not None

    @pytest.mark.integration
    def test_execute_with_named_client(self, temp_workbook, mock_ffmistralsmall):
        """Test executing prompt with specific client from registry."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["E1"] = "client"
        ws["A2"] = 1
        ws["B2"] = "test"
        ws["C2"] = "Hello"
        ws["D2"] = ""
        ws["E2"] = "fast"

        ws = wb.create_sheet("clients")
        ws["A1"] = "name"
        ws["B1"] = "client_type"
        ws["C1"] = "model"
        ws["A2"] = "fast"
        ws["B2"] = "mistral-small"
        ws["C2"] = "mistral-small-2503"

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator._init_client_registry()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        results = orchestrator.execute()

        assert results[0]["status"] == "success"
        assert results[0]["client"] == "fast"

    def test_client_registry_empty_when_no_clients_sheet(self, temp_workbook, mock_ffmistralsmall):
        """Test that client registry is not initialized when no clients sheet."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["A2"] = 1
        ws["B2"] = "test"
        ws["C2"] = "Hello"
        ws["D2"] = ""

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator._init_client_registry()

        assert orchestrator.has_multi_client is False
        assert orchestrator.client_registry is None


class TestExcelOrchestratorDocuments:
    """Tests for document/reference injection."""

    def test_inject_references_returns_original_when_no_documents(
        self, temp_workbook, mock_ffmistralsmall
    ):
        """Test that original prompt is returned when no documents configured."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        prompt = {"prompt": "Hello world", "references": ["doc1"]}
        result = orchestrator._inject_references(prompt)

        assert result == "Hello world"

    def test_inject_references_returns_original_when_no_references_field(
        self, temp_workbook, mock_ffmistralsmall
    ):
        """Test that original prompt is returned when no references field."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        prompt = {"prompt": "Hello world"}
        result = orchestrator._inject_references(prompt)

        assert result == "Hello world"

    def test_inject_references_with_empty_references_list(self, temp_workbook, mock_ffmistralsmall):
        """Test that original prompt is returned when references list is empty."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)

        prompt = {"prompt": "Hello world", "references": []}
        result = orchestrator._inject_references(prompt)

        assert result == "Hello world"

    def test_init_documents_no_documents_sheet(self, temp_workbook, mock_ffmistralsmall):
        """Test that document registry is not created when no documents sheet."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["A2"] = 1
        ws["B2"] = "test"
        ws["C2"] = "Hello"
        ws["D2"] = ""

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator._init_documents()

        assert orchestrator.has_documents is False
        assert orchestrator.document_registry is None


class TestExcelOrchestratorIsolatedFFAI:
    """Tests for isolated FFAI creation."""

    def test_get_isolated_ffai_without_registry(self, temp_workbook, mock_ffmistralsmall):
        """Test that FFAI is created with cloned client when no registry."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        ffai = orchestrator._get_isolated_ffai()

        assert ffai is not None
        assert ffai.client is not mock_ffmistralsmall

    @pytest.mark.integration
    def test_get_isolated_ffai_with_registry(self, temp_workbook, mock_ffmistralsmall):
        """Test that FFAI is created with cloned client from registry."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["A2"] = 1
        ws["B2"] = "test"
        ws["C2"] = "Hello"
        ws["D2"] = ""

        ws = wb.create_sheet("clients")
        ws["A1"] = "name"
        ws["B1"] = "client_type"
        ws["C1"] = "model"
        ws["A2"] = "fast"
        ws["B2"] = "mistral-small"
        ws["C2"] = "mistral-small-2503"

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator._init_client_registry()

        ffai = orchestrator._get_isolated_ffai("fast")

        assert ffai is not None

    def test_get_isolated_ffai_creates_distinct_instances(self, temp_workbook, mock_ffmistralsmall):
        """Test that each call to _get_isolated_ffai returns a distinct instance."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        ffai1 = orchestrator._get_isolated_ffai()
        ffai2 = orchestrator._get_isolated_ffai()

        assert ffai1 is not ffai2
        assert ffai1.client is not ffai2.client


class TestExcelOrchestratorSummaryEdgeCases:
    """Tests for summary edge cases."""

    def test_summary_with_skipped_prompts(self, temp_workbook, mock_ffmistralsmall):
        """Test that summary includes skipped count."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"
        ws["E1"] = "condition"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = ""
        ws["E2"] = ""

        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "World"
        ws["D3"] = ""
        ws["E3"] = '{{first.response}} == "nonexistent"'

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["skipped"] == 1

    def test_summary_total_attempts(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that summary includes total attempts."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator.run()
        summary = orchestrator.get_summary()

        assert "total_attempts" in summary
        assert summary["total_attempts"] >= 3
