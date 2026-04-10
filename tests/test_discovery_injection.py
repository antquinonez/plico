# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Tests for auto-discovery runtime injection in ExcelOrchestrator."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pytest

from src.orchestrator.discovery import discover_documents
from src.orchestrator.excel_orchestrator import ExcelOrchestrator


class TestDiscoverDocumentsAbsolutePaths:
    """Tests for discover_documents with absolute_paths=True."""

    def test_absolute_paths_mode(self, tmp_path):
        folder = tmp_path / "resumes"
        folder.mkdir()
        (folder / "alice.pdf").write_text("resume content")

        result = discover_documents(folder, absolute_paths=True)
        assert len(result) == 1
        assert Path(result[0]["file_path"]).is_absolute()
        assert result[0]["file_path"].endswith("alice.pdf")

    def test_relative_paths_mode_default(self, tmp_path):
        folder = tmp_path / "resumes"
        folder.mkdir()
        (folder / "alice.pdf").write_text("resume content")

        result = discover_documents(folder, absolute_paths=False)
        assert len(result) == 1
        assert result[0]["file_path"] == "alice.pdf"
        assert not Path(result[0]["file_path"]).is_absolute()

    def test_absolute_paths_with_tags(self, tmp_path):
        folder = tmp_path / "resumes"
        folder.mkdir()
        (folder / "bob.docx").write_text("resume")

        result = discover_documents(folder, absolute_paths=True, tags=["resume"])
        assert result[0]["tags"] == "resume"
        assert Path(result[0]["file_path"]).is_absolute()


class TestResolveJdDocument:
    """Tests for ExcelOrchestrator._resolve_jd_document."""

    def test_valid_jd_file(self, tmp_path):
        jd = tmp_path / "job_description.md"
        jd.write_text("Senior Engineer role...")

        result = ExcelOrchestrator._resolve_jd_document(str(jd))
        assert result["reference_name"] == "job_description"
        assert result["common_name"] == "Job Description"
        assert result["tags"] == "jd"
        assert Path(result["file_path"]).is_absolute()
        assert "job_description.md" in result["file_path"]

    def test_jd_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError, match="Job description file not found"):
            ExcelOrchestrator._resolve_jd_document(str(tmp_path / "nonexistent.md"))


class TestExcelOrchestratorInjection:
    """Tests for ExcelOrchestrator with resumes_path and jd_path."""

    def _create_minimal_workbook(self, path: Path) -> None:
        """Create a minimal workbook with prompts sheet only."""
        from openpyxl import Workbook

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["C1"] = "notes"
        ws_config["A2"] = "name"
        ws_config["B2"] = "test"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "notes"
        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "test_prompt"
        ws_prompts["C2"] = "Say hello"

        wb.save(str(path))

    def test_inject_jd_only(self, tmp_path):
        workbook = tmp_path / "test.xlsx"
        self._create_minimal_workbook(workbook)

        jd = tmp_path / "jd.md"
        jd.write_text("Senior Engineer role...")

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
            jd_path=str(jd),
        )
        orchestrator._load_source()

        assert orchestrator.has_documents
        assert "job_description" in orchestrator.document_registry.documents

    def test_inject_resumes_only(self, tmp_path):
        workbook = tmp_path / "test.xlsx"
        self._create_minimal_workbook(workbook)

        resumes = tmp_path / "resumes"
        resumes.mkdir()
        (resumes / "alice.pdf").write_text("Alice resume")
        (resumes / "bob.docx").write_text("Bob resume")

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
            resumes_path=str(resumes),
        )
        orchestrator._load_source()

        assert orchestrator.has_documents
        assert orchestrator.is_batch_mode
        assert len(orchestrator.batch_data) == 2
        assert "alice" in orchestrator.document_registry.documents
        assert "bob" in orchestrator.document_registry.documents

    def test_inject_both_jd_and_resumes(self, tmp_path):
        workbook = tmp_path / "test.xlsx"
        self._create_minimal_workbook(workbook)

        jd = tmp_path / "jd.md"
        jd.write_text("Senior Engineer role...")

        resumes = tmp_path / "resumes"
        resumes.mkdir()
        (resumes / "alice.pdf").write_text("Alice resume")
        (resumes / "bob.docx").write_text("Bob resume")

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
            resumes_path=str(resumes),
            jd_path=str(jd),
        )
        orchestrator._load_source()

        assert orchestrator.has_documents
        assert orchestrator.is_batch_mode
        assert len(orchestrator.batch_data) == 2
        assert "job_description" in orchestrator.document_registry.documents
        assert "alice" in orchestrator.document_registry.documents
        assert "bob" in orchestrator.document_registry.documents

    def test_batch_data_has_documents_binding(self, tmp_path):
        workbook = tmp_path / "test.xlsx"
        self._create_minimal_workbook(workbook)

        resumes = tmp_path / "resumes"
        resumes.mkdir()
        (resumes / "alice.pdf").write_text("Alice resume")

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
            resumes_path=str(resumes),
        )
        orchestrator._load_source()

        row = orchestrator.batch_data[0]
        assert row["candidate_name"] == "alice"
        assert row["_documents"] == '["alice"]'

    def test_no_injection_when_no_paths(self, tmp_path):
        workbook = tmp_path / "test.xlsx"
        self._create_minimal_workbook(workbook)

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
        )
        orchestrator._load_source()

        assert not orchestrator.has_documents
        assert not orchestrator.is_batch_mode

    def test_inject_merges_with_workbook_documents(self, tmp_path):
        """When workbook has existing documents, discovery merges them."""
        from openpyxl import Workbook

        workbook = tmp_path / "test.xlsx"
        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["C1"] = "notes"
        ws_config["A2"] = "name"
        ws_config["B2"] = "test"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "notes"
        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "test_prompt"
        ws_prompts["C2"] = "Say hello"

        ws_docs = wb.create_sheet(title="documents")
        ws_docs["A1"] = "reference_name"
        ws_docs["B1"] = "common_name"
        ws_docs["C1"] = "file_path"
        ws_docs["D1"] = "tags"
        ws_docs["E1"] = "notes"
        ws_docs["A2"] = "existing_doc"
        ws_docs["B2"] = "Existing Doc"
        ws_docs["C2"] = str(tmp_path / "existing.txt")
        ws_docs["D2"] = "test"
        ws_docs["E2"] = ""

        (tmp_path / "existing.txt").write_text("existing content")
        wb.save(str(workbook))

        resumes = tmp_path / "resumes"
        resumes.mkdir()
        (resumes / "alice.pdf").write_text("Alice resume")

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
            resumes_path=str(resumes),
        )
        orchestrator._load_source()

        assert "existing_doc" in orchestrator.document_registry.documents
        assert "alice" in orchestrator.document_registry.documents
        assert len(orchestrator.batch_data) == 1

    def test_inject_merges_with_workbook_batch_data(self, tmp_path):
        """When workbook has existing batch data, discovery appends rows."""
        from openpyxl import Workbook

        workbook = tmp_path / "test.xlsx"
        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["C1"] = "notes"
        ws_config["A2"] = "name"
        ws_config["B2"] = "test"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "notes"
        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "test_prompt"
        ws_prompts["C2"] = "Say hello"

        ws_data = wb.create_sheet(title="data")
        ws_data["A1"] = "id"
        ws_data["B1"] = "batch_name"
        ws_data["C1"] = "candidate_name"
        ws_data["D1"] = "_documents"
        ws_data["A2"] = 1
        ws_data["B2"] = "existing_batch"
        ws_data["C2"] = "Existing Candidate"
        ws_data["D2"] = '["existing_doc"]'

        wb.save(str(workbook))

        resumes = tmp_path / "resumes"
        resumes.mkdir()
        (resumes / "alice.pdf").write_text("Alice resume")

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
            resumes_path=str(resumes),
        )
        orchestrator._load_source()

        assert len(orchestrator.batch_data) == 2
        assert orchestrator.batch_data[0]["batch_name"] == "existing_batch"
        assert orchestrator.batch_data[1]["batch_name"] == "alice"

    def test_empty_resumes_path_warns(self, tmp_path):
        workbook = tmp_path / "test.xlsx"
        self._create_minimal_workbook(workbook)

        empty_folder = tmp_path / "empty"
        empty_folder.mkdir()

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
            resumes_path=str(empty_folder),
        )
        orchestrator._load_source()

        assert not orchestrator.is_batch_mode
        assert len(orchestrator.batch_data) == 0

    def test_jd_paths_are_absolute(self, tmp_path):
        workbook = tmp_path / "test.xlsx"
        self._create_minimal_workbook(workbook)

        jd = tmp_path / "jd.md"
        jd.write_text("Senior Engineer role...")

        resumes = tmp_path / "resumes"
        resumes.mkdir()
        (resumes / "alice.pdf").write_text("Alice resume")

        client = MagicMock()
        orchestrator = ExcelOrchestrator(
            workbook_path=str(workbook),
            client=client,
            resumes_path=str(resumes),
            jd_path=str(jd),
        )
        orchestrator._load_source()

        for doc in orchestrator.document_registry.documents.values():
            assert Path(doc["file_path"]).is_absolute()
