# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

from unittest.mock import MagicMock

from src.FFAIClientBase import FFAIClientBase
from src.orchestrator.condition_evaluator import ConditionEvaluator
from src.orchestrator.results.builder import ResultBuilder
from src.orchestrator.results.result import PromptResult


class ScriptedClient(FFAIClientBase):
    """Client that returns scripted responses in order across all clones."""

    model = "test-model"
    system_instructions = "test"
    _shared_responses: list[str]
    _shared_lock: object

    def __init__(self, responses=None, _shared=None, _lock=None):
        if _shared is not None:
            self._shared_responses = _shared
            self._shared_lock = _lock
        else:
            self._shared_responses = list(responses or [])
            import threading

            self._shared_lock = threading.Lock()
        self._last_usage = None
        self._last_cost_usd = 0.0
        self._conversation_history = []

    def generate_response(self, prompt, **kwargs):
        with self._shared_lock:
            if not self._shared_responses:
                return "default response"
            resp = self._shared_responses.pop(0)
        self._conversation_history.append({"role": "user", "content": prompt})
        self._conversation_history.append({"role": "assistant", "content": resp})
        return resp

    def clone(self):
        return ScriptedClient(_shared=self._shared_responses, _lock=self._shared_lock)

    def clear_conversation(self):
        self._conversation_history = []

    def get_conversation_history(self):
        return list(self._conversation_history)

    def set_conversation_history(self, history):
        self._conversation_history = list(history)


class TestResultBuilderAborted:
    def test_as_aborted_sets_status_and_response(self):
        prompt = {"sequence": 1, "prompt_name": "test", "prompt": "hi"}
        result = ResultBuilder(prompt).as_aborted("some trace").build_dict()
        assert result["status"] == "aborted"
        assert result["response"] == "-1"
        assert result["abort_trace"] == "some trace"

    def test_as_aborted_without_trace(self):
        prompt = {"sequence": 1, "prompt_name": "test", "prompt": "hi"}
        result = ResultBuilder(prompt).as_aborted().build_dict()
        assert result["status"] == "aborted"
        assert result["response"] == "-1"
        assert result["abort_trace"] is None

    def test_with_abort_triggered(self):
        prompt = {"sequence": 1, "prompt_name": "test", "prompt": "hi"}
        result = (
            ResultBuilder(prompt)
            .with_response("ok")
            .with_abort_triggered("trace here")
            .build_dict()
        )
        assert result["status"] == "success"
        assert result["response"] == "ok"
        assert result["abort_trace"] == "trace here"

    def test_with_batch_and_aborted(self):
        prompt = {"sequence": 1, "prompt_name": "test", "prompt": "hi"}
        result = ResultBuilder(prompt).with_batch(1, "batch_a").as_aborted("abort!").build_dict()
        assert result["batch_id"] == 1
        assert result["batch_name"] == "batch_a"
        assert result["status"] == "aborted"
        assert result["response"] == "-1"


class TestPromptResultAbortedStatus:
    def test_valid_statuses_includes_aborted(self):
        assert "aborted" in PromptResult.VALID_STATUSES

    def test_abort_trace_field(self):
        result = PromptResult(sequence=1, abort_trace="resolved trace")
        assert result.abort_trace == "resolved trace"

    def test_from_dict_with_abort_trace(self):
        data = {"sequence": 1, "abort_trace": "test trace", "status": "aborted"}
        result = PromptResult.from_dict(data)
        assert result.abort_trace == "test trace"
        assert result.status == "aborted"


class TestEvaluateAbortCondition:
    def test_no_abort_condition_returns_false(self):
        prompt = {"sequence": 1, "prompt_name": "p1", "prompt": "hi"}
        results = {}
        triggered, trace = _eval_abort(prompt, results)
        assert triggered is False
        assert trace is None

    def test_abort_condition_true(self):
        prompt = {
            "sequence": 1,
            "prompt_name": "check",
            "prompt": "check stuff",
            "abort_condition": '{{check.status}} == "success"',
        }
        results = {"check": {"status": "success", "response": "ok"}}
        triggered, trace = _eval_abort(prompt, results)
        assert triggered is True
        assert trace is not None

    def test_abort_condition_false(self):
        prompt = {
            "sequence": 1,
            "prompt_name": "check",
            "prompt": "check stuff",
            "abort_condition": '{{check.status}} == "failed"',
        }
        results = {"check": {"status": "success", "response": "ok"}}
        triggered, trace = _eval_abort(prompt, results)
        assert triggered is False

    def test_abort_condition_with_json_get(self):
        prompt = {
            "sequence": 1,
            "prompt_name": "check",
            "prompt": "check stuff",
            "abort_condition": 'json_get({{check.response}}, "score") < 5',
        }
        results = {"check": {"status": "success", "response": '{"score": 2}'}}
        triggered, trace = _eval_abort(prompt, results)
        assert triggered is True


def _eval_abort(prompt, results_by_name):
    abort_condition = prompt.get("abort_condition")
    if not abort_condition or not str(abort_condition).strip():
        return False, None

    from src.orchestrator.graph import evaluate_condition_with_trace

    should_abort, _, _, abort_trace = evaluate_condition_with_trace(
        prompt, results_by_name, condition_field="abort_condition"
    )
    return should_abort, abort_trace


class TestAbortConditionValidation:
    def test_valid_abort_condition_passes(self):
        is_valid, error = ConditionEvaluator.validate_syntax('{{check.status}} == "success"')
        assert is_valid is True
        assert error is None

    def test_invalid_abort_condition_fails(self):
        is_valid, error = ConditionEvaluator.validate_syntax("{{check.}} ==")
        assert is_valid is False
        assert error is not None


class TestAbortInSequentialExecution:
    def test_sequential_abort_skips_remaining(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "max_retries"
        ws_config["B2"] = 1

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "abort_condition"

        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "check"
        ws_prompts["C2"] = "Check something"
        ws_prompts["E2"] = '{{check.response}} == "abort"'

        ws_prompts["A3"] = 20
        ws_prompts["B3"] = "after_check"
        ws_prompts["C3"] = "Should be aborted"

        ws_prompts["A4"] = 30
        ws_prompts["B4"] = "also_aborted"
        ws_prompts["C4"] = "Also aborted"

        path = str(tmp_path / "abort_seq.xlsx")
        wb.save(path)

        client = ScriptedClient(responses=["abort"])
        orchestrator = ExcelOrchestrator(path, client=client, concurrency=1)
        orchestrator.run()

        results = orchestrator.results
        assert len(results) == 3

        assert results[0]["prompt_name"] == "check"
        assert results[0]["status"] == "success"
        assert results[0]["response"] == "abort"
        assert results[0]["abort_trace"] is not None

        assert results[1]["status"] == "aborted"
        assert results[1]["response"] == "-1"
        assert results[1]["prompt_name"] == "after_check"

        assert results[2]["status"] == "aborted"
        assert results[2]["response"] == "-1"
        assert results[2]["prompt_name"] == "also_aborted"

    def test_sequential_no_abort_runs_all(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "max_retries"
        ws_config["B2"] = 1

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "abort_condition"

        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "check"
        ws_prompts["C2"] = "Check something"
        ws_prompts["E2"] = '{{check.response}} == "abort"'

        ws_prompts["A3"] = 20
        ws_prompts["B3"] = "after_check"
        ws_prompts["C3"] = "Should run normally"

        path = str(tmp_path / "no_abort.xlsx")
        wb.save(path)

        client = ScriptedClient(responses=["continue"])
        orchestrator = ExcelOrchestrator(path, client=client, concurrency=1)
        orchestrator.run()

        results = orchestrator.results
        assert len(results) == 2
        assert results[0]["status"] == "success"
        assert results[0]["abort_trace"] is None
        assert results[1]["status"] == "success"


class TestAbortInBatchExecution:
    def test_batch_abort_per_row(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "max_retries"
        ws_config["B2"] = 1

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "abort_condition"

        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "check"
        ws_prompts["C2"] = "Check {{candidate}}"
        ws_prompts["E2"] = '{{check.response}} == "reject"'

        ws_prompts["A3"] = 20
        ws_prompts["B3"] = "evaluate"
        ws_prompts["C3"] = "Evaluate {{candidate}}"

        ws_prompts["A4"] = 30
        ws_prompts["B4"] = "summarize"
        ws_prompts["C4"] = "Summarize {{candidate}}"

        ws_data = wb.create_sheet(title="data")
        ws_data["A1"] = "id"
        ws_data["B1"] = "batch_name"
        ws_data["C1"] = "candidate"

        ws_data["A2"] = 1
        ws_data["B2"] = "alice"
        ws_data["C2"] = "Alice"

        ws_data["A3"] = 2
        ws_data["B3"] = "bob"
        ws_data["C3"] = "Bob"

        path = str(tmp_path / "abort_batch.xlsx")
        wb.save(path)

        client = ScriptedClient(responses=["reject", "accept"])
        orchestrator = ExcelOrchestrator(path, client=client, concurrency=1)
        orchestrator.run()

        results = orchestrator.results
        assert len(results) == 6

        alice_check = results[0]
        assert alice_check["batch_name"] == "alice"
        assert alice_check["status"] == "success"
        assert alice_check["response"] == "reject"
        assert alice_check["abort_trace"] is not None

        alice_eval = results[1]
        assert alice_eval["batch_name"] == "alice"
        assert alice_eval["status"] == "aborted"
        assert alice_eval["response"] == "-1"

        alice_summary = results[2]
        assert alice_summary["batch_name"] == "alice"
        assert alice_summary["status"] == "aborted"

        bob_check = results[3]
        assert bob_check["batch_name"] == "bob"
        assert bob_check["status"] == "success"
        assert bob_check["response"] == "accept"
        assert bob_check["abort_trace"] is None

        bob_eval = results[4]
        assert bob_eval["batch_name"] == "bob"
        assert bob_eval["status"] == "success"

        bob_summary = results[5]
        assert bob_summary["batch_name"] == "bob"
        assert bob_summary["status"] == "success"

    def test_batch_abort_parallel(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "max_retries"
        ws_config["B2"] = 1

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "abort_condition"

        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "check"
        ws_prompts["C2"] = "Check {{candidate}}"
        ws_prompts["E2"] = '{{check.response}} == "reject"'

        ws_prompts["A3"] = 20
        ws_prompts["B3"] = "evaluate"
        ws_prompts["C3"] = "Evaluate {{candidate}}"

        ws_data = wb.create_sheet(title="data")
        ws_data["A1"] = "id"
        ws_data["B1"] = "batch_name"
        ws_data["C1"] = "candidate"

        ws_data["A2"] = 1
        ws_data["B2"] = "alice"
        ws_data["C2"] = "Alice"

        ws_data["A3"] = 2
        ws_data["B3"] = "bob"
        ws_data["C3"] = "Bob"

        path = str(tmp_path / "abort_batch_parallel.xlsx")
        wb.save(path)

        client = ScriptedClient(responses=["reject", "accept"])
        orchestrator = ExcelOrchestrator(path, client=client, concurrency=2)
        orchestrator.run()

        results = orchestrator.results
        results.sort(key=lambda r: (r["batch_id"], r["sequence"]))

        assert len(results) == 4

        alice_results = [r for r in results if r["batch_name"] == "alice"]
        assert alice_results[0]["status"] == "success"
        assert alice_results[0]["abort_trace"] is not None
        assert alice_results[1]["status"] == "aborted"

        bob_results = [r for r in results if r["batch_name"] == "bob"]
        assert bob_results[0]["status"] == "success"
        assert bob_results[0].get("abort_trace") is None
        assert bob_results[1]["status"] == "success"


class TestAbortConditionParseFromWorkbook:
    def test_abort_condition_parsed_from_prompts_sheet(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "max_retries"
        ws_config["B2"] = 3

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "abort_condition"

        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "check"
        ws_prompts["C2"] = "Check stuff"
        ws_prompts["E2"] = '{{check.status}} == "failed"'

        ws_prompts["A3"] = 20
        ws_prompts["B3"] = "no_abort"
        ws_prompts["C3"] = "No abort condition here"

        path = str(tmp_path / "parse_abort.xlsx")
        wb.save(path)

        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(path)
        prompts = parser.load_prompts()

        assert len(prompts) == 2
        assert prompts[0]["abort_condition"] == '{{check.status}} == "failed"'
        assert prompts[1]["abort_condition"] is None

    def test_abort_condition_in_pivot_headers(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser.__new__(WorkbookParser)
        parser._config = MagicMock()
        parser._config.workbook.sheet_names.config = "config"
        parser._config.workbook.sheet_names.prompts = "prompts"
        parser._config.workbook.sheet_names.data = "data"
        parser._config.workbook.sheet_names.clients = "clients"
        parser._config.workbook.sheet_names.documents = "documents"
        parser._config.workbook.sheet_names.tools = "tools"
        parser._config.workbook.sheet_names.scoring = "scoring"
        parser._config.workbook.sheet_names.synthesis = "synthesis"
        parser._config.workbook.defaults = MagicMock()
        parser._config.workbook.batch = MagicMock()
        parser._config.workbook.formatting = MagicMock()

        assert "abort_condition" in parser.PROMPTS_HEADERS


class TestAbortSummaryCounts:
    def test_summary_includes_aborted_count(self):
        from src.orchestrator.results.frame import ResultsFrame

        results = [
            {
                "sequence": 1,
                "prompt_name": "a",
                "prompt": "x",
                "status": "success",
                "response": "ok",
                "attempts": 1,
                "abort_trace": None,
            },
            {
                "sequence": 2,
                "prompt_name": "b",
                "prompt": "x",
                "status": "aborted",
                "response": "-1",
                "attempts": 0,
                "abort_trace": "trace",
            },
            {
                "sequence": 3,
                "prompt_name": "c",
                "prompt": "x",
                "status": "aborted",
                "response": "-1",
                "attempts": 0,
                "abort_trace": None,
            },
        ]
        frame = ResultsFrame(results)
        summary = frame.summary()
        assert summary["successful"] == 1
        assert summary["aborted"] == 2


class TestAbortConditionWithGraphConditionField:
    def test_evaluate_condition_field_parameter(self):
        from src.orchestrator.graph import evaluate_condition_with_trace

        prompt = {
            "sequence": 1,
            "prompt_name": "p",
            "prompt": "hi",
            "condition": "ignored",
            "abort_condition": '{{p.status}} == "success"',
        }

        should_run, _, _, _ = evaluate_condition_with_trace(
            prompt, {"p": {"status": "success"}}, condition_field="abort_condition"
        )
        assert should_run is True

        should_run2, _, _, _ = evaluate_condition_with_trace(
            prompt, {"p": {"status": "failed"}}, condition_field="abort_condition"
        )
        assert should_run2 is False


class TestAbortInNonBatchParallel:
    def test_parallel_abort_skips_remaining(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "max_retries"
        ws_config["B2"] = 1

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "abort_condition"

        ws_prompts["A2"] = 10
        ws_prompts["B2"] = "check"
        ws_prompts["C2"] = "Check something"
        ws_prompts["E2"] = '{{check.response}} == "abort"'

        ws_prompts["A3"] = 20
        ws_prompts["B3"] = "after_check"
        ws_prompts["C3"] = "Depends on check"
        ws_prompts["D3"] = '["check"]'

        ws_prompts["A4"] = 30
        ws_prompts["B4"] = "also_aborted"
        ws_prompts["C4"] = "Depends on after_check"
        ws_prompts["D4"] = '["after_check"]'

        path = str(tmp_path / "abort_parallel.xlsx")
        wb.save(path)

        client = ScriptedClient(responses=["abort", "should not run", "should not run"])
        orchestrator = ExcelOrchestrator(path, client=client, concurrency=2)
        orchestrator.run()

        results = orchestrator.results
        assert len(results) == 3

        results.sort(key=lambda r: r["sequence"])

        assert results[0]["prompt_name"] == "check"
        assert results[0]["status"] == "success"
        assert results[0]["abort_trace"] is not None

        assert results[1]["prompt_name"] == "after_check"
        assert results[1]["status"] == "aborted"
        assert results[1]["response"] == "-1"

        assert results[2]["prompt_name"] == "also_aborted"
        assert results[2]["status"] == "aborted"
        assert results[2]["response"] == "-1"


class TestAbortGraphDependencies:
    def test_abort_condition_creates_dependency_edge(self):
        from src.orchestrator.graph import build_execution_graph_with_edges

        prompts = [
            {
                "sequence": 10,
                "prompt_name": "gate",
                "prompt": "Gate check",
            },
            {
                "sequence": 20,
                "prompt_name": "eval",
                "prompt": "Evaluate",
                "abort_condition": '{{gate.status}} == "success"',
            },
        ]
        graph = build_execution_graph_with_edges(prompts)

        assert 10 in graph.nodes[20].dependencies

        abort_edges = [e for e in graph.edges if e.source == "abort_condition"]
        assert len(abort_edges) == 1
        assert abort_edges[0].from_seq == 10
        assert abort_edges[0].to_seq == 20

    def test_abort_condition_does_not_duplicate_edges(self):
        from src.orchestrator.graph import build_execution_graph_with_edges

        prompts = [
            {
                "sequence": 10,
                "prompt_name": "gate",
                "prompt": "Gate check",
            },
            {
                "sequence": 20,
                "prompt_name": "eval",
                "prompt": "Evaluate",
                "condition": '{{gate.status}} == "success"',
                "abort_condition": '{{gate.response}} contains "pass"',
            },
        ]
        graph = build_execution_graph_with_edges(prompts)

        assert 10 in graph.nodes[20].dependencies

        edges_from_10 = [e for e in graph.edges if e.from_seq == 10 and e.to_seq == 20]
        assert len(edges_from_10) == 2
        sources = {e.source for e in edges_from_10}
        assert sources == {"condition", "abort_condition"}
