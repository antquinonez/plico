# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Tests for src/orchestrator/discovery.py."""

from __future__ import annotations

from pathlib import Path

import pytest

from src.orchestrator.discovery import (
    DEFAULT_EXTENSIONS,
    _derive_names,
    _sanitize_for_reference,
    create_data_rows_from_documents,
    create_evaluation_workbook,
    discover_documents,
)
from src.orchestrator.workbook_parser import WorkbookParser


class TestSanitizeForReference:
    """Tests for _sanitize_for_reference naming convention."""

    def test_simple_name(self):
        assert _sanitize_for_reference("alice_chen") == "alice_chen"

    def test_spaces_to_underscores(self):
        assert _sanitize_for_reference("Alice Chen") == "alice_chen"

    def test_hyphens_and_dots(self):
        assert _sanitize_for_reference("resume-v2-final") == "resume_v2_final"

    def test_mixed_case_and_special(self):
        assert _sanitize_for_reference("Bob Martinez CV") == "bob_martinez_cv"

    def test_leading_trailing_underscores(self):
        assert _sanitize_for_reference("__test__") == "test"

    def test_consecutive_special_chars(self):
        assert _sanitize_for_reference("resume  (draft)") == "resume_draft"

    def test_with_extension_in_path(self):
        assert _sanitize_for_reference("alice_chen.pdf") == "alice_chen"

    def test_numbers_preserved(self):
        assert _sanitize_for_reference("Candidate_2024_Final") == "candidate_2024_final"


class TestDeriveNames:
    """Tests for _derive_names file naming conventions."""

    def test_standard_filename(self):
        ref, common, batch = _derive_names(Path("alice_chen.pdf"))
        assert ref == "alice_chen"
        assert common == "alice_chen"
        assert batch == "alice_chen"

    def test_spaced_filename(self):
        ref, common, batch = _derive_names(Path("Alice Chen CV.pdf"))
        assert ref == "alice_chen_cv"
        assert common == "Alice Chen CV"
        assert batch == "Alice Chen CV"

    def test_extension_preserved_in_none(self):
        ref, common, batch = _derive_names(Path("resume_v2.docx"))
        assert ref == "resume_v2"
        assert common == "resume_v2"
        assert batch == "resume_v2"


class TestDiscoverDocuments:
    """Tests for discover_documents."""

    def _create_temp_files(self, tmp_path: Path, files: list[str]) -> Path:
        folder = tmp_path / "docs"
        folder.mkdir()
        for name in files:
            (folder / name).write_text(f"content of {name}")
        return folder

    def test_discovers_supported_files(self, tmp_path):
        folder = self._create_temp_files(
            tmp_path, ["alice.pdf", "bob.docx", "carol.txt", "david.md"]
        )
        result = discover_documents(folder)
        assert len(result) == 4
        assert result[0]["reference_name"] == "alice"
        assert result[1]["reference_name"] == "bob"

    def test_sorted_alphabetically(self, tmp_path):
        folder = self._create_temp_files(tmp_path, ["charlie.pdf", "alice.pdf", "bob.pdf"])
        result = discover_documents(folder)
        names = [d["reference_name"] for d in result]
        assert names == ["alice", "bob", "charlie"]

    def test_skips_unsupported_extensions(self, tmp_path):
        folder = self._create_temp_files(
            tmp_path, ["alice.pdf", "image.png", "data.csv", "bob.docx"]
        )
        result = discover_documents(folder)
        names = [d["reference_name"] for d in result]
        assert names == ["alice", "bob"]

    def test_custom_extensions(self, tmp_path):
        folder = self._create_temp_files(
            tmp_path, ["alice.pdf", "bob.docx", "data.csv", "notes.txt"]
        )
        result = discover_documents(folder, extensions={".csv", ".txt"})
        names = [d["reference_name"] for d in result]
        assert names == ["data", "notes"]

    def test_empty_folder(self, tmp_path):
        folder = self._create_temp_files(tmp_path, [])
        result = discover_documents(folder)
        assert result == []

    def test_folder_not_exists(self, tmp_path):
        with pytest.raises(ValueError, match="Folder does not exist"):
            discover_documents(tmp_path / "nonexistent")

    def test_tags_applied(self, tmp_path):
        folder = self._create_temp_files(tmp_path, ["alice.pdf"])
        result = discover_documents(folder, tags=["resume", "engineering"])
        assert result[0]["tags"] == "resume, engineering"

    def test_no_tags(self, tmp_path):
        folder = self._create_temp_files(tmp_path, ["alice.pdf"])
        result = discover_documents(folder)
        assert result[0]["tags"] == ""

    def test_subfolders_ignored(self, tmp_path):
        folder = self._create_temp_files(tmp_path, ["alice.pdf"])
        (folder / "subfolder").mkdir()
        (folder / "subfolder" / "hidden.pdf").write_text("hidden")
        result = discover_documents(folder)
        assert len(result) == 1

    def test_output_matches_documents_headers(self, tmp_path):
        folder = self._create_temp_files(tmp_path, ["alice.pdf"])
        result = discover_documents(folder)
        doc = result[0]
        for header in (
            "reference_name",
            "common_name",
            "file_path",
            "tags",
            "chunking_strategy",
            "notes",
        ):
            assert header in doc

    def test_file_path_relative_to_folder(self, tmp_path):
        folder = self._create_temp_files(tmp_path, ["alice.pdf"])
        result = discover_documents(folder)
        assert result[0]["file_path"] == "alice.pdf"
        assert not Path(result[0]["file_path"]).is_absolute()

    def test_hidden_files_included(self, tmp_path):
        folder = self._create_temp_files(tmp_path, [".hidden.pdf", "visible.pdf"])
        result = discover_documents(folder)
        names = [d["reference_name"] for d in result]
        assert "hidden" in names
        assert "visible" in names

    def test_doc_extension_supported(self, tmp_path):
        folder = self._create_temp_files(tmp_path, ["legacy.doc"])
        result = discover_documents(folder)
        assert len(result) == 1
        assert result[0]["reference_name"] == "legacy"

    def test_notes_field_empty(self, tmp_path):
        folder = self._create_temp_files(tmp_path, ["alice.pdf"])
        result = discover_documents(folder)
        assert result[0]["notes"] == ""


class TestCreateDataRowsFromDocuments:
    """Tests for create_data_rows_from_documents."""

    def test_basic_row_generation(self):
        docs = [
            {"reference_name": "resume_alice", "common_name": "Alice Chen"},
            {"reference_name": "resume_bob", "common_name": "Bob Martinez"},
        ]
        rows = create_data_rows_from_documents(docs)
        assert len(rows) == 2
        assert rows[0]["id"] == 1
        assert rows[0]["batch_name"] == "resume_alice"
        assert rows[0]["candidate_name"] == "Alice Chen"
        assert rows[1]["id"] == 2

    def test_documents_column_value(self):
        docs = [{"reference_name": "resume_alice", "common_name": "Alice"}]
        rows = create_data_rows_from_documents(docs)
        assert rows[0]["_documents"] == '["resume_alice"]'

    def test_custom_documents_column(self):
        docs = [{"reference_name": "ref_a", "common_name": "A"}]
        rows = create_data_rows_from_documents(docs, documents_column="_docs")
        assert "_docs" in rows[0]
        assert rows[0]["_docs"] == '["ref_a"]'

    def test_empty_list(self):
        rows = create_data_rows_from_documents([])
        assert rows == []

    def test_id_sequential(self):
        docs = [
            {"reference_name": "a", "common_name": "A"},
            {"reference_name": "b", "common_name": "B"},
            {"reference_name": "c", "common_name": "C"},
        ]
        rows = create_data_rows_from_documents(docs)
        assert [r["id"] for r in rows] == [1, 2, 3]

    def test_fallback_common_name(self):
        docs = [{"reference_name": "resume_alice"}]
        rows = create_data_rows_from_documents(docs)
        assert rows[0]["candidate_name"] == "resume_alice"

    def test_batch_name_from_reference_name(self):
        docs = [{"reference_name": "resume_bob", "common_name": "Bob"}]
        rows = create_data_rows_from_documents(docs)
        assert rows[0]["batch_name"] == "resume_bob"


class TestCreateEvaluationWorkbook:
    """Tests for create_evaluation_workbook."""

    def _create_doc_files(self, tmp_path: Path) -> Path:
        folder = tmp_path / "resumes"
        folder.mkdir()
        for name in ["alice.pdf", "bob.docx", "carol.txt"]:
            (folder / name).write_text(f"content of {name}")
        return folder

    def test_creates_workbook_file(self, tmp_path):
        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "output" / "test.xlsx"
        result = create_evaluation_workbook(output, folder)
        assert Path(result).exists()

    def test_returns_absolute_path(self, tmp_path):
        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        result = create_evaluation_workbook(output, folder)
        assert Path(result).is_absolute()

    def test_all_sheets_present(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder)

        wb = load_workbook(output)
        expected_sheets = {
            "config",
            "prompts",
            "data",
            "documents",
            "scoring",
            "synthesis",
        }
        assert set(wb.sheetnames) == expected_sheets

    def test_config_sheet_has_evaluation_strategy(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder, evaluation_strategy="potential")

        wb = load_workbook(output)
        ws = wb["config"]
        config_dict = {}
        for row in range(2, ws.max_row + 1):
            field = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            if field:
                config_dict[field] = value

        assert config_dict.get("evaluation_strategy") == "potential"

    def test_documents_sheet_populated(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder)

        wb = load_workbook(output)
        ws = wb["documents"]
        assert ws.max_row >= 4  # header + 3 documents

    def test_shared_documents_prepended(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        shared = [
            {
                "reference_name": "job_desc",
                "common_name": "Job Description",
                "file_path": "library/jd.md",
                "tags": "jd",
                "notes": "",
            }
        ]
        create_evaluation_workbook(output, folder, shared_documents=shared)

        wb = load_workbook(output)
        ws = wb["documents"]
        first_ref = ws.cell(row=2, column=1).value
        assert first_ref == "job_desc"

    def test_data_sheet_has_documents_column(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder)

        wb = load_workbook(output)
        ws = wb["data"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "_documents" in headers

    def test_scoring_sheet_headers_only(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder)

        wb = load_workbook(output)
        ws = wb["scoring"]
        assert ws.max_row == 1
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers == WorkbookParser.SCORING_HEADERS

    def test_synthesis_sheet_headers_only(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder)

        wb = load_workbook(output)
        ws = wb["synthesis"]
        assert ws.max_row == 1
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers == WorkbookParser.SYNTHESIS_HEADERS

    def test_prompts_sheet_headers_only(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder)

        wb = load_workbook(output)
        ws = wb["prompts"]
        assert ws.max_row == 1

    def test_empty_folder_raises(self, tmp_path):
        folder = tmp_path / "empty"
        folder.mkdir()
        output = tmp_path / "test.xlsx"
        with pytest.raises(ValueError, match="No documents found"):
            create_evaluation_workbook(output, folder)

    def test_creates_parent_directories(self, tmp_path):
        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "nested" / "deep" / "test.xlsx"
        create_evaluation_workbook(output, folder)
        assert output.exists()

    def test_custom_extensions(self, tmp_path):
        folder = tmp_path / "data"
        folder.mkdir()
        (folder / "report.csv").write_text("data")
        output = tmp_path / "test.xlsx"
        result = create_evaluation_workbook(output, folder, extensions={".csv"})
        assert Path(result).exists()

    def test_tags_passed_through(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder, tags=["resume"])

        wb = load_workbook(output)
        ws = wb["documents"]
        for row in range(2, ws.max_row + 1):
            tags = ws.cell(row=row, column=4).value
            assert tags == "resume"

    def test_config_sheet_has_batch_fields(self, tmp_path):
        from openpyxl import load_workbook

        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder)

        wb = load_workbook(output)
        ws = wb["config"]
        fields = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "batch_mode" in fields
        assert "on_batch_error" in fields

    def test_workbook_parseable_by_parser(self, tmp_path):
        folder = self._create_doc_files(tmp_path)
        output = tmp_path / "test.xlsx"
        create_evaluation_workbook(output, folder)

        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(output)
        docs = parser.load_documents()
        assert len(docs) >= 3

        data = parser.load_data()
        assert len(data) >= 3

        config = parser.load_config()
        assert config.get("evaluation_strategy") == "balanced"


class TestDefaultExtensions:
    """Tests for DEFAULT_EXTENSIONS constant."""

    def test_includes_expected(self):
        assert ".pdf" in DEFAULT_EXTENSIONS
        assert ".docx" in DEFAULT_EXTENSIONS
        assert ".doc" in DEFAULT_EXTENSIONS
        assert ".txt" in DEFAULT_EXTENSIONS
        assert ".md" in DEFAULT_EXTENSIONS

    def test_is_set(self):
        assert isinstance(DEFAULT_EXTENSIONS, set)
