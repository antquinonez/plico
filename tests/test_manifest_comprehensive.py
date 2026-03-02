# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Comprehensive tests for manifest-based orchestration."""

from __future__ import annotations

import json
import os
from unittest.mock import MagicMock

import pytest
import yaml


class TestWorkbookManifestExporterExtended:
    """Extended tests for WorkbookManifestExporter class."""

    def test_export_with_clients_yaml(self, temp_workbook_with_clients, tmp_path):
        """Test export with client configurations creates clients.yaml."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_clients)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "manifest.yaml"), encoding="utf-8") as f:
            manifest = yaml.safe_load(f)

        assert manifest["has_clients"] is True
        assert os.path.exists(os.path.join(manifest_path, "clients.yaml"))

        with open(os.path.join(manifest_path, "clients.yaml"), encoding="utf-8") as f:
            clients = yaml.safe_load(f)

        assert "clients" in clients
        assert len(clients["clients"]) == 2
        assert clients["clients"][0]["name"] == "writer"
        assert clients["clients"][1]["name"] == "analyzer"

    def test_export_with_documents_yaml(self, temp_workbook_with_documents, tmp_path):
        """Test export with document references creates documents.yaml."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_documents)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "manifest.yaml"), encoding="utf-8") as f:
            manifest = yaml.safe_load(f)

        assert manifest["has_documents"] is True
        assert os.path.exists(os.path.join(manifest_path, "documents.yaml"))

        with open(os.path.join(manifest_path, "documents.yaml"), encoding="utf-8") as f:
            docs = yaml.safe_load(f)

        assert "documents" in docs
        assert len(docs["documents"]) == 2

    def test_export_with_all_features(self, temp_workbook_full, tmp_path):
        """Test export with data, clients, and documents."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_full)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "manifest.yaml"), encoding="utf-8") as f:
            manifest = yaml.safe_load(f)

        assert manifest["has_data"] is True
        assert manifest["has_clients"] is True
        assert manifest["has_documents"] is True

        assert os.path.exists(os.path.join(manifest_path, "data.yaml"))
        assert os.path.exists(os.path.join(manifest_path, "clients.yaml"))
        assert os.path.exists(os.path.join(manifest_path, "documents.yaml"))

    def test_export_with_custom_manifest_dir(self, temp_workbook_with_data, tmp_path):
        """Test export to a custom manifest directory."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        custom_dir = str(tmp_path / "custom_manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=custom_dir)

        assert custom_dir in manifest_path
        assert os.path.exists(os.path.join(manifest_path, "manifest.yaml"))


class TestManifestOrchestratorClientRegistry:
    """Tests for ManifestOrchestrator client registry functionality."""

    def test_init_client_registry_no_clients(self, tmp_path, mock_ffmistralsmall):
        """Test _init_client_registry when no clients.yaml exists."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.manifest_meta = {"has_clients": False}
        orchestrator._init_client_registry()

        assert orchestrator.client_registry is None
        assert orchestrator.has_multi_client is False

    def test_init_client_registry_with_clients(self, tmp_path, mock_ffmistralsmall):
        """Test _init_client_registry loads and registers clients."""
        from src.orchestrator.manifest import ManifestOrchestrator

        clients_yaml = {
            "clients": [
                {
                    "name": "writer",
                    "client_type": "mistral-small",
                    "temperature": 0.9,
                },
                {
                    "name": "analyzer",
                    "client_type": "mistral-small",
                    "temperature": 0.3,
                },
            ]
        }

        with open(tmp_path / "clients.yaml", "w", encoding="utf-8") as f:
            yaml.dump(clients_yaml, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.manifest_meta = {"has_clients": True}
        orchestrator._init_client_registry()

        assert orchestrator.client_registry is not None
        assert orchestrator.has_multi_client is True
        assert orchestrator.client_registry.has_client("writer")
        assert orchestrator.client_registry.has_client("analyzer")

    def test_get_isolated_ffai_default(self, tmp_path, mock_ffmistralsmall):
        """Test _get_isolated_ffai clones default client."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator._init_client()

        ffai = orchestrator._get_isolated_ffai()

        assert ffai is not None
        assert ffai.client is not mock_ffmistralsmall

    def test_get_isolated_ffai_named(self, tmp_path, mock_ffmistralsmall):
        """Test _get_isolated_ffai gets named client from registry."""
        from src.orchestrator.manifest import ManifestOrchestrator

        clients_yaml = {
            "clients": [{"name": "writer", "client_type": "mistral-small", "temperature": 0.9}]
        }

        with open(tmp_path / "clients.yaml", "w", encoding="utf-8") as f:
            yaml.dump(clients_yaml, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.manifest_meta = {"has_clients": True}
        orchestrator._init_client_registry()

        ffai = orchestrator._get_isolated_ffai("writer")

        assert ffai is not None


class TestManifestOrchestratorDocuments:
    """Tests for ManifestOrchestrator document handling."""

    def test_init_documents_no_documents(self, tmp_path, mock_ffmistralsmall):
        """Test _init_documents when no documents.yaml exists."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.manifest_meta = {"has_documents": False}
        orchestrator._init_documents()

        assert orchestrator.document_registry is None
        assert orchestrator.has_documents is False

    def test_inject_references_no_refs(self, tmp_path, mock_ffmistralsmall):
        """Test _inject_references with prompt without references."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.has_documents = False

        prompt = {"prompt": "Hello world", "references": None}
        result = orchestrator._inject_references(prompt)

        assert result == "Hello world"

    def test_inject_references_empty_refs_list(self, tmp_path, mock_ffmistralsmall):
        """Test _inject_references with empty references list."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.has_documents = True
        orchestrator.document_registry = MagicMock()
        orchestrator.document_registry.get_reference_names.return_value = []

        prompt = {"prompt": "Hello world", "references": []}
        result = orchestrator._inject_references(prompt)

        assert result == "Hello world"

    def test_inject_references_no_document_registry(self, tmp_path, mock_ffmistralsmall):
        """Test _inject_references raises when registry not initialized."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.has_documents = True
        orchestrator.document_registry = None

        prompt = {"prompt": "Hello world", "references": ["doc1"]}

        with pytest.raises(ValueError, match="Document registry not initialized"):
            orchestrator._inject_references(prompt)

    def test_inject_references_missing_document(self, tmp_path, mock_ffmistralsmall):
        """Test _inject_references raises for missing document reference."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.has_documents = True
        orchestrator.document_registry = MagicMock()
        orchestrator.document_registry.get_reference_names.return_value = ["doc2"]

        prompt = {"prompt": "Hello world", "references": ["doc1", "missing_doc"]}

        with pytest.raises(ValueError, match="Referenced documents not found"):
            orchestrator._inject_references(prompt)

    def test_inject_references_success(self, tmp_path, mock_ffmistralsmall):
        """Test _inject_references successfully injects content."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.has_documents = True
        orchestrator.document_registry = MagicMock()
        orchestrator.document_registry.get_reference_names.return_value = ["doc1"]
        orchestrator.document_registry.inject_references_into_prompt.return_value = (
            "Hello world\n\n---\nDocument content"
        )

        prompt = {"prompt": "Hello world", "references": ["doc1"]}
        result = orchestrator._inject_references(prompt)

        assert "Document content" in result

    def test_inject_semantic_query_disabled(self, tmp_path, mock_ffmistralsmall):
        """Test _inject_references with semantic_query when RAG is disabled."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.has_documents = False

        prompt = {"prompt": "Hello", "semantic_query": "search term"}
        result = orchestrator._inject_references(prompt)

        assert result == "Hello"

    def test_parse_bool_override_none(self, tmp_path, mock_ffmistralsmall):
        """Test _parse_bool_override with None value."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        assert orchestrator._parse_bool_override(None) is None

    def test_parse_bool_override_bool(self, tmp_path, mock_ffmistralsmall):
        """Test _parse_bool_override with boolean value."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        assert orchestrator._parse_bool_override(True) is True
        assert orchestrator._parse_bool_override(False) is False

    def test_parse_bool_override_string(self, tmp_path, mock_ffmistralsmall):
        """Test _parse_bool_override with string values."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        assert orchestrator._parse_bool_override("true") is True
        assert orchestrator._parse_bool_override("yes") is True
        assert orchestrator._parse_bool_override("1") is True
        assert orchestrator._parse_bool_override("false") is False
        assert orchestrator._parse_bool_override("no") is False
        assert orchestrator._parse_bool_override("0") is False
        assert orchestrator._parse_bool_override("  TRUE  ") is True
        assert orchestrator._parse_bool_override("invalid") is None


class TestManifestOrchestratorVariables:
    """Tests for ManifestOrchestrator variable resolution."""

    def test_resolve_variables_basic(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_variables replaces placeholders."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        text = "Region: {{region}}, Product: {{product}}"
        data_row = {"region": "North", "product": "Widget"}

        result = orchestrator._resolve_variables(text, data_row)

        assert result == "Region: North, Product: Widget"

    def test_resolve_variables_missing(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_variables keeps placeholder for missing variable."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        text = "Region: {{region}}, Missing: {{unknown}}"
        data_row = {"region": "North"}

        result = orchestrator._resolve_variables(text, data_row)

        assert result == "Region: North, Missing: {{unknown}}"

    def test_resolve_variables_empty_text(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_variables with empty text."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        result = orchestrator._resolve_variables("", {"region": "North"})

        assert result == ""

    def test_resolve_variables_none_value(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_variables with None value in data."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        text = "Value: {{value}}"
        data_row = {"value": None}

        result = orchestrator._resolve_variables(text, data_row)

        assert result == "Value: {{value}}"

    def test_resolve_prompt_variables(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_prompt_variables resolves all fields."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        prompt = {
            "prompt": "Hello {{name}}",
            "prompt_name": "greeting_{{region}}",
            "sequence": 1,
        }
        data_row = {"name": "World", "region": "North"}

        result = orchestrator._resolve_prompt_variables(prompt, data_row)

        assert result["prompt"] == "Hello World"
        assert result["prompt_name"] == "greeting_North"
        assert result["sequence"] == 1

    def test_resolve_batch_name_custom(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_batch_name with custom template."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        data_row = {"batch_name": "{{region}}_{{product}}", "region": "north", "product": "widget"}
        result = orchestrator._resolve_batch_name(data_row, 1)

        assert result == "north_widget"

    def test_resolve_batch_name_sanitizes_special_chars(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_batch_name sanitizes special characters."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        data_row = {"batch_name": "test @#$ batch!!!", "region": "north"}
        result = orchestrator._resolve_batch_name(data_row, 1)

        assert "@" not in result
        assert "#" not in result
        assert "!" not in result

    def test_resolve_batch_name_default(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_batch_name uses default when no batch_name field."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        data_row = {"region": "north"}
        result = orchestrator._resolve_batch_name(data_row, 5)

        assert result == "batch_5"

    def test_resolve_batch_name_truncates_long_names(self, tmp_path, mock_ffmistralsmall):
        """Test _resolve_batch_name truncates names longer than 50 chars."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        long_name = "a" * 100
        data_row = {"batch_name": long_name}
        result = orchestrator._resolve_batch_name(data_row, 1)

        assert len(result) == 50


class TestManifestOrchestratorConditionEvaluation:
    """Tests for ManifestOrchestrator condition evaluation."""

    def test_evaluate_condition_empty(self, tmp_path, mock_ffmistralsmall):
        """Test _evaluate_condition with empty condition."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        prompt = {"condition": ""}
        results_by_name = {}

        should_exec, cond_result, cond_error = orchestrator._evaluate_condition(
            prompt, results_by_name
        )

        assert should_exec is True
        assert cond_result is None
        assert cond_error is None

    def test_evaluate_condition_whitespace_only(self, tmp_path, mock_ffmistralsmall):
        """Test _evaluate_condition with whitespace-only condition."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        prompt = {"condition": "   "}
        results_by_name = {}

        should_exec, cond_result, cond_error = orchestrator._evaluate_condition(
            prompt, results_by_name
        )

        assert should_exec is True

    def test_evaluate_condition_true(self, tmp_path, mock_ffmistralsmall):
        """Test _evaluate_condition when condition passes."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        prompt = {"condition": "1 == 1"}
        results_by_name = {}

        should_exec, cond_result, cond_error = orchestrator._evaluate_condition(
            prompt, results_by_name
        )

        assert should_exec is True
        assert cond_result == "True"
        assert cond_error is None

    def test_evaluate_condition_false(self, tmp_path, mock_ffmistralsmall):
        """Test _evaluate_condition when condition fails."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        prompt = {"condition": "1 == 2"}
        results_by_name = {}

        should_exec, cond_result, cond_error = orchestrator._evaluate_condition(
            prompt, results_by_name
        )

        assert should_exec is False
        assert cond_result == "False"
        assert cond_error is None

    def test_evaluate_condition_with_status_reference(self, tmp_path, mock_ffmistralsmall):
        """Test _evaluate_condition references another prompt's status."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        prompt = {"condition": "{{a.status}} == 'success'"}
        results_by_name = {"a": {"status": "success"}}

        should_exec, cond_result, cond_error = orchestrator._evaluate_condition(
            prompt, results_by_name
        )

        assert should_exec is True


class TestManifestOrchestratorExecute:
    """Tests for ManifestOrchestrator execute methods."""

    def test_execute_prompt_success(self, tmp_path, mock_ffmistralsmall):
        """Test _execute_prompt succeeds."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3}

        prompt = {
            "sequence": 1,
            "prompt_name": "test",
            "prompt": "Hello",
            "history": None,
        }

        result = orchestrator._execute_prompt(prompt, {})

        assert result["status"] == "success"
        assert result["response"] == "This is a test response."
        assert result["attempts"] == 1

    def test_execute_prompt_with_retry_success(self, tmp_path, mock_ffmistralsmall):
        """Test _execute_prompt retries and succeeds."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3}

        call_count = [0]

        original_clone = mock_ffmistralsmall.clone

        def mock_clone():
            cloned = original_clone()
            original_generate = cloned.generate_response

            def flaky_generate(*args, **kwargs):
                call_count[0] += 1
                if call_count[0] < 3:
                    raise Exception("Temporary failure")
                return original_generate(*args, **kwargs)

            cloned.generate_response = flaky_generate
            return cloned

        mock_ffmistralsmall.clone = mock_clone

        prompt = {
            "sequence": 1,
            "prompt_name": "test",
            "prompt": "Hello",
            "history": None,
        }

        result = orchestrator._execute_prompt(prompt, {})

        assert result["status"] == "success"
        assert result["attempts"] == 3

    def test_execute_prompt_max_retries_exceeded(self, tmp_path, mock_ffmistralsmall):
        """Test _execute_prompt fails after max retries."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 2}

        original_clone = mock_ffmistralsmall.clone

        def mock_clone():
            cloned = original_clone()

            def always_fail(*args, **kwargs):
                raise Exception("Always fails")

            cloned.generate_response = always_fail
            return cloned

        mock_ffmistralsmall.clone = mock_clone

        prompt = {
            "sequence": 1,
            "prompt_name": "test",
            "prompt": "Hello",
            "history": None,
        }

        result = orchestrator._execute_prompt(prompt, {})

        assert result["status"] == "failed"
        assert result["attempts"] == 2
        assert "Always fails" in result["error"]

    def test_execute_prompt_condition_skip(self, tmp_path, mock_ffmistralsmall):
        """Test _execute_prompt skips when condition is false."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3}

        prompt = {
            "sequence": 1,
            "prompt_name": "test",
            "prompt": "Hello",
            "history": None,
            "condition": "1 == 2",
        }

        result = orchestrator._execute_prompt(prompt, {})

        assert result["status"] == "skipped"
        assert result["attempts"] == 0

    def test_execute_with_client_name(self, tmp_path, mock_ffmistralsmall):
        """Test _execute_prompt uses named client."""
        from src.orchestrator.manifest import ManifestOrchestrator

        clients_yaml = {
            "clients": [{"name": "writer", "client_type": "mistral-small", "temperature": 0.9}]
        }

        with open(tmp_path / "clients.yaml", "w", encoding="utf-8") as f:
            yaml.dump(clients_yaml, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.manifest_meta = {"has_clients": True}
        orchestrator._init_client_registry()
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3}

        prompt = {
            "sequence": 1,
            "prompt_name": "test",
            "prompt": "Hello",
            "history": None,
            "client": "writer",
        }

        result = orchestrator._execute_prompt(prompt, {})

        assert result["status"] == "success"

    def test_execute_with_progress_callback(self, tmp_path, mock_ffmistralsmall):
        """Test execute calls progress callback."""
        from src.orchestrator.manifest import ManifestOrchestrator

        progress_calls = []

        def progress_callback(*args, **kwargs):
            progress_calls.append((args, kwargs))

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            progress_callback=progress_callback,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "a", "prompt": "Hello", "history": None},
            {"sequence": 2, "prompt_name": "b", "prompt": "World", "history": None},
        ]

        orchestrator.execute()

        assert len(progress_calls) >= 2

    def test_create_result_dict(self, tmp_path, mock_ffmistralsmall):
        """Test _create_result_dict creates expected structure."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        prompt = {
            "sequence": 5,
            "prompt_name": "test_prompt",
            "prompt": "Test prompt text",
            "history": ["prev"],
            "client": "writer",
            "condition": "x > 5",
            "references": ["doc1"],
            "semantic_query": "search term",
            "semantic_filter": '{"type": "pdf"}',
            "query_expansion": "true",
            "rerank": "false",
        }

        result = orchestrator._create_result_dict(prompt)

        assert result["sequence"] == 5
        assert result["prompt_name"] == "test_prompt"
        assert result["prompt"] == "Test prompt text"
        assert result["history"] == ["prev"]
        assert result["client"] == "writer"
        assert result["condition"] == "x > 5"
        assert result["status"] == "pending"
        assert result["attempts"] == 0
        assert result["references"] == ["doc1"]
        assert result["semantic_query"] == "search term"


class TestManifestOrchestratorExecuteParallel:
    """Tests for ManifestOrchestrator parallel execution."""

    def test_execute_parallel_basic(self, tmp_path, mock_ffmistralsmall):
        """Test execute_parallel runs prompts in parallel."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=2,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "a", "prompt": "Hello A", "history": None},
            {"sequence": 2, "prompt_name": "b", "prompt": "Hello B", "history": None},
            {"sequence": 3, "prompt_name": "c", "prompt": "Hello C", "history": None},
        ]

        results = orchestrator.execute_parallel()

        assert len(results) == 3
        assert all(r["status"] == "success" for r in results)

    def test_execute_parallel_with_dependencies(self, tmp_path, mock_ffmistralsmall):
        """Test execute_parallel respects dependencies - independent prompts run in parallel."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=3,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "first", "prompt": "A", "history": None},
            {"sequence": 2, "prompt_name": "second", "prompt": "B", "history": ["first"]},
            {"sequence": 3, "prompt_name": "third", "prompt": "C", "history": ["first"]},
            {"sequence": 4, "prompt_name": "fourth", "prompt": "D", "history": ["second", "third"]},
        ]

        results = orchestrator.execute_parallel()

        assert len(results) == 4
        assert all(r["status"] == "success" for r in results)

        results_by_name = {r["prompt_name"]: r for r in results}
        assert results_by_name["first"]["status"] == "success"
        assert results_by_name["second"]["status"] == "success"
        assert results_by_name["third"]["status"] == "success"
        assert results_by_name["fourth"]["status"] == "success"

    def test_execute_parallel_with_failure(self, tmp_path, mock_ffmistralsmall):
        """Test execute_parallel handles failures gracefully."""
        from src.orchestrator.manifest import ManifestOrchestrator

        original_clone = mock_ffmistralsmall.clone
        call_count = [0]

        def mock_clone():
            cloned = original_clone()
            original_generate = cloned.generate_response

            def fail_on_second(prompt, **kwargs):
                call_count[0] += 1
                if call_count[0] == 2:
                    raise Exception("Forced failure")
                return original_generate(prompt, **kwargs)

            cloned.generate_response = fail_on_second
            return cloned

        mock_ffmistralsmall.clone = mock_clone

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=2,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 1}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "a", "prompt": "Hello", "history": None},
            {"sequence": 2, "prompt_name": "b", "prompt": "World", "history": None},
        ]

        results = orchestrator.execute_parallel()

        statuses = [r["status"] for r in results]
        assert "failed" in statuses


class TestManifestOrchestratorBatchExecution:
    """Tests for ManifestOrchestrator batch execution."""

    def test_execute_batch_basic(self, tmp_path, mock_ffmistralsmall):
        """Test execute_batch runs all batches sequentially."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=1,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3, "on_batch_error": "continue"}

        orchestrator.prompts = [
            {
                "sequence": 1,
                "prompt_name": "intro",
                "prompt": "Region: {{region}}",
                "history": None,
            },
        ]
        orchestrator.batch_data = [
            {"region": "north", "product": "widget_a"},
            {"region": "south", "product": "widget_b"},
        ]
        orchestrator.is_batch_mode = True

        results = orchestrator.execute_batch()

        assert len(results) == 2
        assert all(r["status"] == "success" for r in results)
        assert results[0]["batch_id"] == 1
        assert results[1]["batch_id"] == 2

    def test_execute_batch_resolves_variables(self, tmp_path, mock_ffmistralsmall):
        """Test execute_batch resolves variables in prompts."""
        from src.orchestrator.manifest import ManifestOrchestrator

        prompts_sent = []
        original_clone = mock_ffmistralsmall.clone

        def mock_clone():
            cloned = original_clone()
            original_generate = cloned.generate_response

            def capture_prompt(prompt, **kwargs):
                prompts_sent.append(prompt)
                return original_generate(prompt, **kwargs)

            cloned.generate_response = capture_prompt
            return cloned

        mock_ffmistralsmall.clone = mock_clone

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=1,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3, "on_batch_error": "continue"}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "test", "prompt": "Region: {{region}}", "history": None},
        ]
        orchestrator.batch_data = [
            {"region": "north"},
            {"region": "south"},
        ]
        orchestrator.is_batch_mode = True

        orchestrator.execute_batch()

        assert "Region: north" in prompts_sent
        assert "Region: south" in prompts_sent

    def test_execute_batch_with_condition(self, tmp_path, mock_ffmistralsmall):
        """Test execute_batch handles conditional execution per batch."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=1,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3, "on_batch_error": "continue"}

        orchestrator.prompts = [
            {
                "sequence": 1,
                "prompt_name": "first",
                "prompt": "Hello",
                "history": None,
            },
            {
                "sequence": 2,
                "prompt_name": "test",
                "prompt": "World",
                "history": None,
                "condition": "{{first.status}} == 'success'",
            },
        ]
        orchestrator.batch_data = [
            {"id": 1},
            {"id": 2},
        ]
        orchestrator.is_batch_mode = True

        results = orchestrator.execute_batch()

        assert results[0]["status"] == "success"
        assert results[1]["status"] == "success"

    def test_execute_batch_on_error_continue(self, tmp_path, mock_ffmistralsmall):
        """Test execute_batch continues on error when configured."""
        from src.orchestrator.manifest import ManifestOrchestrator

        original_clone = mock_ffmistralsmall.clone
        call_count = [0]

        def mock_clone():
            cloned = original_clone()
            original_generate = cloned.generate_response

            def fail_on_second(prompt, **kwargs):
                call_count[0] += 1
                if call_count[0] == 2:
                    raise Exception("Batch failure")
                return original_generate(prompt, **kwargs)

            cloned.generate_response = fail_on_second
            return cloned

        mock_ffmistralsmall.clone = mock_clone

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=1,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 1, "on_batch_error": "continue"}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "test", "prompt": "Hello", "history": None},
        ]
        orchestrator.batch_data = [
            {"id": 1},
            {"id": 2},
            {"id": 3},
        ]
        orchestrator.is_batch_mode = True

        results = orchestrator.execute_batch()

        assert len(results) == 3
        statuses = [r["status"] for r in results]
        assert "failed" in statuses
        assert "success" in statuses

    def test_execute_batch_on_error_stop(self, tmp_path, mock_ffmistralsmall):
        """Test execute_batch stops processing remaining prompts in batch on error."""
        from src.orchestrator.manifest import ManifestOrchestrator

        original_clone = mock_ffmistralsmall.clone
        call_count = [0]

        def mock_clone():
            cloned = original_clone()
            original_generate = cloned.generate_response

            def fail_on_second_prompt(prompt, **kwargs):
                call_count[0] += 1
                if call_count[0] == 2:
                    raise Exception("Batch failure")
                return original_generate(prompt, **kwargs)

            cloned.generate_response = fail_on_second_prompt
            return cloned

        mock_ffmistralsmall.clone = mock_clone

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=1,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 1, "on_batch_error": "stop"}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "prompt1", "prompt": "First", "history": None},
            {"sequence": 2, "prompt_name": "prompt2", "prompt": "Second", "history": None},
            {"sequence": 3, "prompt_name": "prompt3", "prompt": "Third", "history": None},
        ]
        orchestrator.batch_data = [
            {"id": 1},
        ]
        orchestrator.is_batch_mode = True

        results = orchestrator.execute_batch()

        assert len(results) == 2
        assert results[0]["status"] == "success"
        assert results[1]["status"] == "failed"


class TestManifestOrchestratorBatchParallel:
    """Tests for ManifestOrchestrator parallel batch execution."""

    def test_execute_batch_parallel_basic(self, tmp_path, mock_ffmistralsmall):
        """Test execute_batch_parallel runs batches in parallel."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=3,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3, "on_batch_error": "continue"}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "test", "prompt": "Region: {{region}}", "history": None},
        ]
        orchestrator.batch_data = [
            {"region": "north"},
            {"region": "south"},
            {"region": "east"},
        ]
        orchestrator.is_batch_mode = True

        results = orchestrator.execute_batch_parallel()

        assert len(results) == 3
        assert all(r["status"] == "success" for r in results)

    def test_execute_batch_parallel_with_failure(self, tmp_path, mock_ffmistralsmall):
        """Test execute_batch_parallel handles failures."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=2,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 1, "on_batch_error": "continue"}

        call_count = [0]

        def fail_some(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] % 2 == 0:
                raise Exception("Batch failure")
            return "Success"

        orchestrator.ffai.generate_response = fail_some

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "test", "prompt": "Hello", "history": None},
        ]
        orchestrator.batch_data = [
            {"id": 1},
            {"id": 2},
            {"id": 3},
        ]
        orchestrator.is_batch_mode = True

        results = orchestrator.execute_batch_parallel()

        assert len(results) == 3

    def test_execute_batch_parallel_with_progress_callback(self, tmp_path, mock_ffmistralsmall):
        """Test execute_batch_parallel calls progress callback."""
        from src.orchestrator.manifest import ManifestOrchestrator

        progress_calls = []

        def progress_callback(*args, **kwargs):
            progress_calls.append((args, kwargs))

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=2,
            progress_callback=progress_callback,
        )
        orchestrator._init_client()
        orchestrator.config = {"max_retries": 3, "on_batch_error": "continue"}

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "test", "prompt": "Hello", "history": None},
        ]
        orchestrator.batch_data = [
            {"id": 1},
            {"id": 2},
        ]
        orchestrator.is_batch_mode = True

        orchestrator.execute_batch_parallel()

        assert len(progress_calls) >= 1


class TestManifestOrchestratorRun:
    """Tests for ManifestOrchestrator run method."""

    def test_run_non_batch_sequential(self, temp_workbook_with_data, tmp_path, mock_ffmistralsmall):
        """Test run() with sequential execution."""
        from src.orchestrator.manifest import (
            ManifestOrchestrator,
            WorkbookManifestExporter,
        )

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_path,
            client=mock_ffmistralsmall,
            concurrency=1,
        )

        parquet_path = orchestrator.run()

        assert os.path.exists(parquet_path)
        assert len(orchestrator.results) == 3

    def test_run_non_batch_parallel(self, temp_workbook_with_data, tmp_path, mock_ffmistralsmall):
        """Test run() with parallel execution."""
        from src.orchestrator.manifest import (
            ManifestOrchestrator,
            WorkbookManifestExporter,
        )

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_path,
            client=mock_ffmistralsmall,
            concurrency=2,
        )

        parquet_path = orchestrator.run()

        assert os.path.exists(parquet_path)
        assert len(orchestrator.results) == 3

    def test_run_batch_sequential(
        self, temp_workbook_with_batch_data, tmp_path, mock_ffmistralsmall
    ):
        """Test run() with batch mode and sequential execution."""
        from src.orchestrator.manifest import (
            ManifestOrchestrator,
            WorkbookManifestExporter,
        )

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_batch_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_path,
            client=mock_ffmistralsmall,
            concurrency=1,
        )

        parquet_path = orchestrator.run()

        assert os.path.exists(parquet_path)
        assert len(orchestrator.results) == 9  # 3 batches * 3 prompts

    def test_run_batch_parallel(self, temp_workbook_with_batch_data, tmp_path, mock_ffmistralsmall):
        """Test run() with batch mode and parallel execution."""
        from src.orchestrator.manifest import (
            ManifestOrchestrator,
            WorkbookManifestExporter,
        )

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_batch_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_path,
            client=mock_ffmistralsmall,
            concurrency=2,
        )

        parquet_path = orchestrator.run()

        assert os.path.exists(parquet_path)
        assert len(orchestrator.results) == 9

    def test_run_creates_output_directory(self, tmp_path, mock_ffmistralsmall):
        """Test run() creates output directory if needed."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest"
        manifest_dir.mkdir()

        manifest_data = {
            "version": "1.0",
            "source_workbook": "test.xlsx",
            "has_data": False,
            "has_clients": False,
            "has_documents": False,
            "prompt_count": 1,
        }

        with open(manifest_dir / "manifest.yaml", "w", encoding="utf-8") as f:
            yaml.dump(manifest_data, f)

        config_data = {"model": "mistral-small-2503", "max_retries": 3}
        with open(manifest_dir / "config.yaml", "w", encoding="utf-8") as f:
            yaml.dump(config_data, f)

        prompts_data = {
            "prompts": [{"sequence": 1, "prompt_name": "test", "prompt": "Hello", "history": []}]
        }
        with open(manifest_dir / "prompts.yaml", "w", encoding="utf-8") as f:
            yaml.dump(prompts_data, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
            concurrency=1,
        )

        parquet_path = orchestrator.run()

        assert os.path.exists(os.path.dirname(parquet_path))


class TestManifestOrchestratorSummaryExtended:
    """Extended tests for ManifestOrchestrator summary."""

    def test_get_summary_not_run(self, tmp_path, mock_ffmistralsmall):
        """Test get_summary when not run yet."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        summary = orchestrator.get_summary()

        assert summary["status"] == "not_run"

    def test_get_summary_with_conditions(self, tmp_path, mock_ffmistralsmall):
        """Test get_summary includes prompts with conditions."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.results = [
            {"status": "success", "attempts": 1, "condition": "x > 5"},
            {"status": "skipped", "attempts": 0, "condition": "y == 10"},
            {"status": "success", "attempts": 1},
        ]

        summary = orchestrator.get_summary()

        assert summary["prompts_with_conditions"] == 2

    def test_get_summary_batch_mode(self, tmp_path, mock_ffmistralsmall):
        """Test get_summary in batch mode."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.is_batch_mode = True
        orchestrator.results = [
            {"status": "success", "attempts": 1, "batch_id": 1},
            {"status": "success", "attempts": 1, "batch_id": 1},
            {"status": "success", "attempts": 1, "batch_id": 2},
            {"status": "failed", "attempts": 3, "batch_id": 2},
        ]

        summary = orchestrator.get_summary()

        assert summary["batch_mode"] is True
        assert summary["total_batches"] == 2
        assert "batches_with_failures" in summary
        assert 2 in summary["batches_with_failures"]

    def test_get_summary_no_batch_failures(self, tmp_path, mock_ffmistralsmall):
        """Test get_summary in batch mode without failures."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.is_batch_mode = True
        orchestrator.results = [
            {"status": "success", "attempts": 1, "batch_id": 1},
            {"status": "success", "attempts": 1, "batch_id": 2},
        ]

        summary = orchestrator.get_summary()

        assert "batches_with_failures" not in summary

    def test_get_summary_basic_counts(self, tmp_path, mock_ffmistralsmall):
        """Test get_summary basic counts."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.results = [
            {"status": "success", "attempts": 1},
            {"status": "success", "attempts": 2},
            {"status": "failed", "attempts": 3},
            {"status": "skipped", "attempts": 0},
        ]

        summary = orchestrator.get_summary()

        assert summary["total_prompts"] == 4
        assert summary["successful"] == 2
        assert summary["failed"] == 1
        assert summary["skipped"] == 1
        assert summary["total_attempts"] == 6


class TestManifestOrchestratorValidateDependencies:
    """Tests for ManifestOrchestrator dependency validation."""

    def test_validate_dependencies_success(self, tmp_path, mock_ffmistralsmall):
        """Test _validate_dependencies with valid dependencies."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "first", "prompt": "A", "history": None},
            {"sequence": 2, "prompt_name": "second", "prompt": "B", "history": ["first"]},
            {"sequence": 3, "prompt_name": "third", "prompt": "C", "history": ["first", "second"]},
        ]

        orchestrator._validate_dependencies()

    def test_validate_dependencies_missing_reference(self, tmp_path, mock_ffmistralsmall):
        """Test _validate_dependencies raises for missing reference."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "first", "prompt": "A", "history": None},
            {"sequence": 2, "prompt_name": "second", "prompt": "B", "history": ["nonexistent"]},
        ]

        with pytest.raises(ValueError, match="Dependency validation failed"):
            orchestrator._validate_dependencies()

    def test_validate_dependencies_wrong_order(self, tmp_path, mock_ffmistralsmall):
        """Test _validate_dependencies raises when dependency defined later."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "first", "prompt": "A", "history": ["second"]},
            {"sequence": 2, "prompt_name": "second", "prompt": "B", "history": None},
        ]

        with pytest.raises(ValueError, match="must be defined before"):
            orchestrator._validate_dependencies()

    def test_validate_dependencies_no_history(self, tmp_path, mock_ffmistralsmall):
        """Test _validate_dependencies with prompts without history."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        orchestrator.prompts = [
            {"sequence": 1, "prompt_name": "first", "prompt": "A", "history": None},
            {"sequence": 2, "prompt_name": None, "prompt": "B", "history": None},
        ]

        orchestrator._validate_dependencies()


class TestManifestOrchestratorLoadYaml:
    """Tests for ManifestOrchestrator YAML loading."""

    def test_load_yaml_file_missing(self, tmp_path, mock_ffmistralsmall):
        """Test _load_yaml_file returns empty dict for missing file."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        result = orchestrator._load_yaml_file("nonexistent.yaml")

        assert result == {}

    def test_load_yaml_file_empty(self, tmp_path, mock_ffmistralsmall):
        """Test _load_yaml_file handles empty file."""
        from src.orchestrator.manifest import ManifestOrchestrator

        with open(tmp_path / "empty.yaml", "w", encoding="utf-8") as f:
            f.write("")

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        result = orchestrator._load_yaml_file("empty.yaml")

        assert result == {}

    def test_load_yaml_file_valid(self, tmp_path, mock_ffmistralsmall):
        """Test _load_yaml_file loads valid YAML."""
        from src.orchestrator.manifest import ManifestOrchestrator

        data = {"key": "value", "nested": {"item": 123}}
        with open(tmp_path / "test.yaml", "w", encoding="utf-8") as f:
            yaml.dump(data, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        result = orchestrator._load_yaml_file("test.yaml")

        assert result == data

    def test_load_manifest_with_config_overrides(self, tmp_path, mock_ffmistralsmall):
        """Test _load_manifest applies config overrides."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_data = {
            "version": "1.0",
            "source_workbook": "test.xlsx",
            "has_data": False,
            "has_clients": False,
            "has_documents": False,
            "prompt_count": 1,
        }
        with open(tmp_path / "manifest.yaml", "w", encoding="utf-8") as f:
            yaml.dump(manifest_data, f)

        config_data = {"model": "original-model", "max_retries": 3}
        with open(tmp_path / "config.yaml", "w", encoding="utf-8") as f:
            yaml.dump(config_data, f)

        prompts_data = {
            "prompts": [{"sequence": 1, "prompt_name": "test", "prompt": "Hello", "history": []}]
        }
        with open(tmp_path / "prompts.yaml", "w", encoding="utf-8") as f:
            yaml.dump(prompts_data, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            config_overrides={"model": "overridden-model", "temperature": 0.5},
        )

        orchestrator._load_manifest()

        assert orchestrator.config["model"] == "overridden-model"
        assert orchestrator.config["temperature"] == 0.5
        assert orchestrator.config["max_retries"] == 3


# Additional fixtures needed for the tests
@pytest.fixture
def temp_workbook_with_clients(temp_workbook, sample_prompts, sample_config):
    """Create a workbook with client configurations."""
    from openpyxl import Workbook

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", sample_config["model"]),
        ("max_retries", sample_config["max_retries"]),
        ("temperature", sample_config["temperature"]),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    for idx, p in enumerate(sample_prompts, start=2):
        ws_prompts[f"A{idx}"] = p["sequence"]
        ws_prompts[f"B{idx}"] = p["prompt_name"]
        ws_prompts[f"C{idx}"] = p["prompt"]

    ws_clients = wb.create_sheet(title="clients")
    ws_clients["A1"] = "name"
    ws_clients["B1"] = "client_type"
    ws_clients["C1"] = "temperature"

    clients_data = [
        ("writer", "mistral-small", 0.9),
        ("analyzer", "mistral-small", 0.3),
    ]

    for idx, (name, client_type, temp) in enumerate(clients_data, start=2):
        ws_clients[f"A{idx}"] = name
        ws_clients[f"B{idx}"] = client_type
        ws_clients[f"C{idx}"] = temp

    wb.save(temp_workbook)
    return temp_workbook


@pytest.fixture
def temp_workbook_with_documents(temp_workbook, sample_prompts, sample_config, tmp_path):
    """Create a workbook with document references."""
    from openpyxl import Workbook

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", sample_config["model"]),
        ("max_retries", sample_config["max_retries"]),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"
    ws_prompts["E1"] = "references"

    for idx, p in enumerate(sample_prompts[:2], start=2):
        ws_prompts[f"A{idx}"] = p["sequence"]
        ws_prompts[f"B{idx}"] = p["prompt_name"]
        ws_prompts[f"C{idx}"] = p["prompt"]
        ws_prompts[f"D{idx}"] = ""
        ws_prompts[f"E{idx}"] = json.dumps(["doc1"])

    ws_docs = wb.create_sheet(title="documents")
    ws_docs["A1"] = "reference_name"
    ws_docs["B1"] = "common_name"
    ws_docs["C1"] = "file_path"
    ws_docs["D1"] = "notes"

    doc_file = tmp_path / "sample_doc.txt"
    doc_file.write_text("Sample document content")

    docs_data = [
        ("doc1", "Sample Doc", str(doc_file), "A sample document"),
        ("doc2", "Data Doc", str(doc_file), "Data document"),
    ]

    for idx, (ref, common, path, notes) in enumerate(docs_data, start=2):
        ws_docs[f"A{idx}"] = ref
        ws_docs[f"B{idx}"] = common
        ws_docs[f"C{idx}"] = path
        ws_docs[f"D{idx}"] = notes

    wb.save(temp_workbook)
    return temp_workbook


@pytest.fixture
def temp_workbook_full(temp_workbook, sample_config, sample_batch_data, tmp_path):
    """Create a workbook with all features: data, clients, documents."""
    from openpyxl import Workbook

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", sample_config["model"]),
        ("max_retries", sample_config["max_retries"]),
        ("batch_mode", "per_row"),
        ("batch_output", "combined"),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    batch_prompts = [
        {"sequence": 1, "prompt_name": "intro", "prompt": "Analyze {{region}}."},
        {"sequence": 2, "prompt_name": "calc", "prompt": "Price: {{price}}."},
    ]

    for idx, p in enumerate(batch_prompts, start=2):
        ws_prompts[f"A{idx}"] = p["sequence"]
        ws_prompts[f"B{idx}"] = p["prompt_name"]
        ws_prompts[f"C{idx}"] = p["prompt"]
        ws_prompts[f"D{idx}"] = ""

    ws_data = wb.create_sheet(title="data")
    data_headers = ["id", "region", "product", "price"]
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(sample_batch_data[:2], start=2):
        for col_idx, header in enumerate(data_headers, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=row_data.get(header))

    ws_clients = wb.create_sheet(title="clients")
    ws_clients["A1"] = "name"
    ws_clients["B1"] = "client_type"
    ws_clients["A2"] = "writer"
    ws_clients["B2"] = "mistral-small"

    ws_docs = wb.create_sheet(title="documents")
    ws_docs["A1"] = "reference_name"
    ws_docs["B1"] = "common_name"
    ws_docs["C1"] = "file_path"

    doc_file = tmp_path / "sample_doc.txt"
    doc_file.write_text("Sample document content")
    ws_docs["A2"] = "doc1"
    ws_docs["B2"] = "Sample"
    ws_docs["C2"] = str(doc_file)

    wb.save(temp_workbook)
    return temp_workbook
