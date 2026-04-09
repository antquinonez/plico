# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

import pytest


class TestSynthesisExecutorSourceScope:
    """Tests for source_scope resolution."""

    def test_resolve_all(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [{"batch_id": 1}, {"batch_id": 2}, {"batch_id": 3}]
        result = executor.resolve_source_scope("all", entries)
        assert len(result) == 3

    def test_resolve_top_n(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [{"batch_id": i} for i in range(10)]
        result = executor.resolve_source_scope("top:3", entries)
        assert len(result) == 3

    def test_resolve_top_n_larger_than_available(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [{"batch_id": 1}, {"batch_id": 2}]
        result = executor.resolve_source_scope("top:10", entries)
        assert len(result) == 2

    def test_resolve_top_n_one(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [{"batch_id": i} for i in range(5)]
        result = executor.resolve_source_scope("top:1", entries)
        assert len(result) == 1

    def test_invalid_scope_raises(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [{"batch_id": 1}]
        with pytest.raises(ValueError, match="Invalid source_scope"):
            executor.resolve_source_scope("bottom:5", entries)

    def test_top_zero_raises(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [{"batch_id": 1}]
        with pytest.raises(ValueError, match="N > 0"):
            executor.resolve_source_scope("top:0", entries)

    def test_empty_entries_all(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        result = executor.resolve_source_scope("all", [])
        assert result == []

    def test_scope_with_spaces(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [{"batch_id": i} for i in range(5)]
        result = executor.resolve_source_scope(" top:2 ", entries)
        assert len(result) == 2


class TestSynthesisExecutorSortEntries:
    """Tests for entry sorting with tiebreaking."""

    def _make_batch_results(self, entries):
        results = []
        for entry in entries:
            results.append([entry])
        return results

    def test_sort_by_composite_score(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [
            {"batch_id": 1, "composite_score": 5.0, "batch_name": "c", "scores": {}},
            {"batch_id": 2, "composite_score": 8.0, "batch_name": "a", "scores": {}},
            {"batch_id": 3, "composite_score": 6.0, "batch_name": "b", "scores": {}},
        ]
        sorted_entries = executor.sort_entries(self._make_batch_results(entries), has_scoring=True)
        scores = [e["composite_score"] for e in sorted_entries]
        assert scores == [8.0, 6.0, 5.0]

    def test_tiebreak_by_criteria_order(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        criteria = [{"criteria_name": "skills"}, {"criteria_name": "education"}]
        entries = [
            {
                "batch_id": 1,
                "composite_score": 7.0,
                "batch_name": "a",
                "scores": {"skills": 7, "education": 7},
            },
            {
                "batch_id": 2,
                "composite_score": 7.0,
                "batch_name": "b",
                "scores": {"skills": 8, "education": 6},
            },
        ]
        sorted_entries = executor.sort_entries(
            self._make_batch_results(entries),
            scoring_criteria=criteria,
            has_scoring=True,
        )
        assert sorted_entries[0]["batch_id"] == 2
        assert sorted_entries[1]["batch_id"] == 1

    def test_tiebreak_by_batch_id(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [
            {
                "batch_id": 3,
                "composite_score": 7.0,
                "batch_name": "c",
                "scores": {"skills": 7},
            },
            {
                "batch_id": 1,
                "composite_score": 7.0,
                "batch_name": "a",
                "scores": {"skills": 7},
            },
            {
                "batch_id": 2,
                "composite_score": 7.0,
                "batch_name": "b",
                "scores": {"skills": 7},
            },
        ]
        sorted_entries = executor.sort_entries(self._make_batch_results(entries), has_scoring=True)
        ids = [e["batch_id"] for e in sorted_entries]
        assert ids == [1, 2, 3]

    def test_sort_without_scoring_alphabetical(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entries = [
            {"batch_id": 1, "batch_name": "charlie"},
            {"batch_id": 2, "batch_name": "alice"},
            {"batch_id": 3, "batch_name": "bob"},
        ]
        sorted_entries = executor.sort_entries(self._make_batch_results(entries), has_scoring=False)
        names = [e["batch_name"] for e in sorted_entries]
        assert names == ["alice", "bob", "charlie"]

    def test_sort_empty_entries(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        sorted_entries = executor.sort_entries([], has_scoring=True)
        assert sorted_entries == []


class TestSynthesisExecutorContextFormatting:
    """Tests for XML context formatting."""

    def test_basic_xml_structure(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor(max_context_chars=50000)
        entries = [
            {
                "batch_id": 1,
                "composite_score": 8.0,
                "batch_name": "alice",
                "scores": {"skills": 8},
                "_all_results": {
                    "extract_profile": {"response": "Alice is a developer"},
                },
            },
        ]
        context = executor.format_entry_context(
            entries, ["extract_profile"], include_scores=True, strategy="balanced"
        )
        assert "<ENTRIES" in context
        assert 'strategy="balanced"' in context
        assert 'total_count="1"' in context
        assert 'included_count="1"' in context
        assert "<ENTRY" in context
        assert 'rank="1"' in context
        assert "<SCORES>" in context
        assert "skills:" in context
        assert "<EXTRACT_PROFILE>" in context
        assert "Alice is a developer" in context
        assert "</ENTRIES>" in context

    def test_multiple_entries(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor(max_context_chars=50000)
        entries = [
            {
                "batch_id": 1,
                "composite_score": 9.0,
                "batch_name": "alice",
                "scores": {},
                "_all_results": {},
            },
            {
                "batch_id": 2,
                "composite_score": 7.0,
                "batch_name": "bob",
                "scores": {},
                "_all_results": {},
            },
        ]
        context = executor.format_entry_context(entries, [], include_scores=False, strategy="")
        assert 'rank="1"' in context
        assert 'rank="2"' in context

    def test_no_scores_when_include_false(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor(max_context_chars=50000)
        entries = [
            {
                "batch_id": 1,
                "composite_score": 8.0,
                "batch_name": "alice",
                "scores": {"skills": 8},
                "_all_results": {},
            },
        ]
        context = executor.format_entry_context(entries, [], include_scores=False)
        assert "<SCORES>" not in context

    def test_empty_entries_returns_empty(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        context = executor.format_entry_context([], [], True)
        assert context == ""

    def test_context_truncation(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor(max_context_chars=500)
        long_response = "x" * 1000
        entries = [
            {
                "batch_id": 1,
                "composite_score": 8.0,
                "batch_name": "alice",
                "scores": {"skills": 8},
                "_all_results": {
                    "profile": {"response": long_response},
                },
            },
        ]
        context = executor.format_entry_context(entries, ["profile"], include_scores=True)
        assert "[...truncated" in context

    def test_emergency_reduction(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor(max_context_chars=800)
        long_response = "y" * 500
        entries = [
            {
                "batch_id": i,
                "composite_score": float(10 - i),
                "batch_name": f"entry_{i}",
                "scores": {},
                "_all_results": {
                    "profile": {"response": long_response},
                },
            }
            for i in range(5)
        ]
        context = executor.format_entry_context(entries, ["profile"], include_scores=False)
        assert "included_count=" in context
        count = int(context.split('included_count="')[1].split('"')[0])
        assert count < 5

    def test_single_entry_cant_fit_raises(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor(max_context_chars=10)
        entries = [
            {
                "batch_id": 1,
                "composite_score": 8.0,
                "batch_name": "alice",
                "scores": {"skills": 8, "education": 7},
                "_all_results": {
                    "profile": {"response": "some text"},
                },
            },
        ]
        with pytest.raises(ValueError, match="Cannot fit any entries"):
            executor.format_entry_context(entries, ["profile"], include_scores=True)

    def test_missing_prompt_response(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor(max_context_chars=50000)
        entries = [
            {
                "batch_id": 1,
                "composite_score": 8.0,
                "batch_name": "alice",
                "scores": {},
                "_all_results": {
                    "known_prompt": {"response": "has response"},
                },
            },
        ]
        context = executor.format_entry_context(
            entries, ["known_prompt", "unknown_prompt"], include_scores=False
        )
        assert "has response" in context


class TestSynthesisExecutorGetEntryName:
    """Tests for entry name extraction."""

    def test_name_from_data_column(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entry = {
            "batch_id": 1,
            "batch_name": "alice_chen",
            "candidate_name": "Alice Chen",
            "scores": {},
        }
        assert executor.get_entry_name(entry) == "Alice Chen"

    def test_fallback_to_batch_name(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entry = {
            "batch_id": 1,
            "batch_name": "alice_chen",
            "scores": {},
        }
        assert executor.get_entry_name(entry) == "alice_chen"

    def test_fallback_to_unknown(self):
        from src.orchestrator.synthesis import SynthesisExecutor

        executor = SynthesisExecutor()
        entry = {"batch_id": 1, "scores": {}}
        assert executor.get_entry_name(entry) == "unknown"


class TestSynthesisPrompt:
    """Tests for SynthesisPrompt dataclass."""

    def test_defaults(self):
        from src.orchestrator.synthesis import SynthesisPrompt

        sp = SynthesisPrompt(sequence=1, prompt_name="test", prompt="Hello")
        assert sp.source_scope == "all"
        assert sp.source_prompts == []
        assert sp.include_scores is True
        assert sp.history is None
        assert sp.condition is None
        assert sp.client is None

    def test_all_fields(self):
        from src.orchestrator.synthesis import SynthesisPrompt

        sp = SynthesisPrompt(
            sequence=1,
            prompt_name="rank",
            prompt="Rank them",
            source_scope="top:5",
            source_prompts=["profile", "assessment"],
            include_scores=False,
            history=["other"],
            condition="{{other.response}} % 'PASS'",
            client="anthropic",
        )
        assert sp.source_scope == "top:5"
        assert sp.source_prompts == ["profile", "assessment"]
        assert sp.include_scores is False


class TestBuildEntryResultsMap:
    """Tests for build_entry_results_map helper."""

    def test_basic_mapping(self):
        from src.orchestrator.synthesis import build_entry_results_map

        batch_results = [
            [
                {"batch_id": 0, "prompt_name": "intro", "response": "hello"},
                {"batch_id": 0, "prompt_name": "eval", "response": "good"},
            ],
            [
                {"batch_id": 1, "prompt_name": "intro", "response": "world"},
                {"batch_id": 1, "prompt_name": "eval", "response": "great"},
            ],
        ]
        mapping = build_entry_results_map(batch_results)
        assert len(mapping) == 2
        assert mapping[0]["intro"]["response"] == "hello"
        assert mapping[1]["intro"]["response"] == "world"

    def test_empty_results(self):
        from src.orchestrator.synthesis import build_entry_results_map

        mapping = build_entry_results_map([])
        assert mapping == {}

    def test_empty_batch_entry(self):
        from src.orchestrator.synthesis import build_entry_results_map

        mapping = build_entry_results_map([[]])
        assert mapping == {}


class TestSynthesisValidation:
    """Tests for synthesis validation in OrchestratorValidator."""

    def test_valid_synthesis_prompts(self):
        from src.orchestrator.validation import OrchestratorValidator

        prompts = [
            {"sequence": 1, "prompt_name": "profile", "prompt": "Extract profile"},
            {"sequence": 2, "prompt_name": "eval", "prompt": "Evaluate"},
        ]
        synthesis = [
            {
                "sequence": 100,
                "prompt_name": "rank",
                "prompt": "Rank them",
                "source_scope": "top:3",
                "source_prompts": ["profile"],
                "include_scores": True,
            },
        ]
        scoring = [{"criteria_name": "skills", "source_prompt": "eval"}]

        validator = OrchestratorValidator(
            prompts=prompts,
            config={},
            scoring_criteria=scoring,
            synthesis_prompts=synthesis,
        )
        result = validator.validate()
        error_codes = [e.code for e in result.errors if e.severity == "error"]
        assert not error_codes

    def test_invalid_source_scope(self):
        from src.orchestrator.validation import OrchestratorValidator

        prompts = [
            {"sequence": 1, "prompt_name": "profile", "prompt": "Extract profile"},
        ]
        synthesis = [
            {
                "sequence": 100,
                "prompt_name": "rank",
                "prompt": "Rank",
                "source_scope": "bottom:5",
                "source_prompts": ["profile"],
            },
        ]

        validator = OrchestratorValidator(
            prompts=prompts,
            config={},
            synthesis_prompts=synthesis,
        )
        result = validator.validate()
        error_codes = [e.code for e in result.errors if e.severity == "error"]
        assert "INVALID_SOURCE_SCOPE" in error_codes

    def test_top_zero_invalid(self):
        from src.orchestrator.validation import OrchestratorValidator

        synthesis = [
            {
                "sequence": 100,
                "prompt_name": "rank",
                "prompt": "Rank",
                "source_scope": "top:0",
                "source_prompts": [],
            },
        ]

        validator = OrchestratorValidator(
            prompts=[{"sequence": 1, "prompt_name": "p", "prompt": "hi"}],
            config={},
            synthesis_prompts=synthesis,
        )
        result = validator.validate()
        error_codes = [e.code for e in result.errors if e.severity == "error"]
        assert "INVALID_SOURCE_SCOPE" in error_codes

    def test_unknown_source_prompt(self):
        from src.orchestrator.validation import OrchestratorValidator

        prompts = [
            {"sequence": 1, "prompt_name": "profile", "prompt": "Extract"},
        ]
        synthesis = [
            {
                "sequence": 100,
                "prompt_name": "rank",
                "prompt": "Rank",
                "source_scope": "top:3",
                "source_prompts": ["nonexistent_prompt"],
            },
        ]

        validator = OrchestratorValidator(
            prompts=prompts,
            config={},
            synthesis_prompts=synthesis,
        )
        result = validator.validate()
        error_codes = [e.code for e in result.errors if e.severity == "error"]
        assert "INVALID_SYNTHESIS_SOURCE_PROMPT" in error_codes

    def test_invalid_synthesis_history_reference(self):
        from src.orchestrator.validation import OrchestratorValidator

        prompts = [
            {"sequence": 1, "prompt_name": "profile", "prompt": "Extract"},
        ]
        synthesis = [
            {
                "sequence": 100,
                "prompt_name": "rank",
                "prompt": "Rank",
                "source_scope": "top:3",
                "source_prompts": ["profile"],
            },
            {
                "sequence": 110,
                "prompt_name": "compare",
                "prompt": "Compare",
                "source_scope": "top:3",
                "source_prompts": ["profile"],
                "history": ["nonexistent_synthesis"],
            },
        ]

        validator = OrchestratorValidator(
            prompts=prompts,
            config={},
            synthesis_prompts=synthesis,
        )
        result = validator.validate()
        error_codes = [e.code for e in result.errors if e.severity == "error"]
        assert "INVALID_SYNTHESIS_HISTORY" in error_codes

    def test_valid_synthesis_history(self):
        from src.orchestrator.validation import OrchestratorValidator

        prompts = [
            {"sequence": 1, "prompt_name": "profile", "prompt": "Extract"},
        ]
        synthesis = [
            {
                "sequence": 100,
                "prompt_name": "rank",
                "prompt": "Rank",
                "source_scope": "top:3",
                "source_prompts": ["profile"],
            },
            {
                "sequence": 110,
                "prompt_name": "recommendation",
                "prompt": "Recommend",
                "source_scope": "top:1",
                "source_prompts": ["profile"],
                "history": ["rank"],
            },
        ]

        validator = OrchestratorValidator(
            prompts=prompts,
            config={},
            synthesis_prompts=synthesis,
        )
        result = validator.validate()
        error_codes = [e.code for e in result.errors if e.severity == "error"]
        assert "INVALID_SYNTHESIS_HISTORY" not in error_codes

    def test_synthesis_without_scoring_warns(self):
        from src.orchestrator.validation import OrchestratorValidator

        prompts = [
            {"sequence": 1, "prompt_name": "profile", "prompt": "Extract"},
        ]
        synthesis = [
            {
                "sequence": 100,
                "prompt_name": "rank",
                "prompt": "Rank",
                "source_scope": "all",
                "source_prompts": ["profile"],
            },
        ]

        validator = OrchestratorValidator(
            prompts=prompts,
            config={},
            synthesis_prompts=synthesis,
        )
        result = validator.validate()
        warning_codes = [e.code for e in result.errors if e.severity == "warning"]
        assert "SYNTHESIS_WITHOUT_SCORING" in warning_codes

    def test_duplicate_synthesis_name(self):
        from src.orchestrator.validation import OrchestratorValidator

        prompts = [
            {"sequence": 1, "prompt_name": "profile", "prompt": "Extract"},
        ]
        synthesis = [
            {
                "sequence": 100,
                "prompt_name": "rank",
                "prompt": "Rank",
                "source_scope": "all",
                "source_prompts": ["profile"],
            },
            {
                "sequence": 110,
                "prompt_name": "rank",
                "prompt": "Rank again",
                "source_scope": "all",
                "source_prompts": ["profile"],
            },
        ]

        validator = OrchestratorValidator(
            prompts=prompts,
            config={},
            synthesis_prompts=synthesis,
        )
        result = validator.validate()
        error_codes = [e.code for e in result.errors if e.severity == "error"]
        assert "DUPLICATE_SYNTHESIS_NAME" in error_codes

    def test_no_synthesis_no_errors(self):
        from src.orchestrator.validation import OrchestratorValidator

        prompts = [
            {"sequence": 1, "prompt_name": "hi", "prompt": "Hello"},
        ]

        validator = OrchestratorValidator(
            prompts=prompts,
            config={},
        )
        result = validator.validate()
        error_codes = [e.code for e in result.errors if e.severity == "error"]
        assert not error_codes


class TestSynthesisParser:
    """Tests for workbook_parser synthesis loading."""

    def test_load_synthesis_sheet(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "name"
        ws_config["B2"] = "Test"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["A2"] = 1
        ws_prompts["B2"] = "profile"
        ws_prompts["C2"] = "Extract profile"

        ws_data = wb.create_sheet(title="data")
        ws_data["A1"] = "id"
        ws_data["B1"] = "batch_name"

        ws_synthesis = wb.create_sheet(title="synthesis")
        headers = [
            "sequence",
            "prompt_name",
            "prompt",
            "source_scope",
            "source_prompts",
            "include_scores",
            "history",
            "condition",
        ]
        for col, h in enumerate(headers, start=1):
            ws_synthesis.cell(row=1, column=col, value=h)

        ws_synthesis.cell(row=2, column=1, value=100)
        ws_synthesis.cell(row=2, column=2, value="rank_summary")
        ws_synthesis.cell(row=2, column=3, value="Rank these entries")
        ws_synthesis.cell(row=2, column=4, value="top:5")
        ws_synthesis.cell(row=2, column=5, value='["profile"]')
        ws_synthesis.cell(row=2, column=6, value="true")

        ws_synthesis.cell(row=3, column=1, value=110)
        ws_synthesis.cell(row=3, column=2, value="recommendation")
        ws_synthesis.cell(row=3, column=3, value="Recommend")
        ws_synthesis.cell(row=3, column=4, value="top:1")
        ws_synthesis.cell(row=3, column=5, value='["profile"]')
        ws_synthesis.cell(row=3, column=6, value="false")
        ws_synthesis.cell(row=3, column=7, value='["rank_summary"]')

        path = str(tmp_path / "test_synthesis.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        assert parser.has_synthesis_sheet() is True

        synthesis = parser.load_synthesis()
        assert len(synthesis) == 2
        assert synthesis[0]["prompt_name"] == "rank_summary"
        assert synthesis[0]["source_scope"] == "top:5"
        assert synthesis[0]["source_prompts"] == ["profile"]
        assert synthesis[0]["include_scores"] is True
        assert synthesis[0]["history"] is None

        assert synthesis[1]["prompt_name"] == "recommendation"
        assert synthesis[1]["source_scope"] == "top:1"
        assert synthesis[1]["include_scores"] is False
        assert synthesis[1]["history"] == ["rank_summary"]

    def test_load_synthesis_empty_when_no_sheet(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "name"
        ws_config["B2"] = "Test"

        path = str(tmp_path / "no_synthesis.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        assert parser.has_synthesis_sheet() is False
        assert parser.load_synthesis() == []

    def test_synthesis_headers_defined(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        expected = [
            "sequence",
            "prompt_name",
            "prompt",
            "source_scope",
            "source_prompts",
            "include_scores",
            "history",
            "condition",
        ]
        assert WorkbookParser.SYNTHESIS_HEADERS == expected

    def test_create_template_with_synthesis_sheet(self, tmp_path):
        from openpyxl import load_workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        path = str(tmp_path / "template_synthesis.xlsx")
        parser = WorkbookParser(path)
        parser.create_template_workbook(with_synthesis_sheet=True)

        wb = load_workbook(path)
        assert "synthesis" in wb.sheetnames
        ws = wb["synthesis"]
        assert ws.cell(row=1, column=1).value == "sequence"
        assert ws.cell(row=1, column=4).value == "source_scope"

    def test_create_template_without_synthesis_sheet(self, tmp_path):
        from openpyxl import load_workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        path = str(tmp_path / "template_no_synthesis.xlsx")
        parser = WorkbookParser(path)
        parser.create_template_workbook()

        wb = load_workbook(path)
        assert "synthesis" not in wb.sheetnames
