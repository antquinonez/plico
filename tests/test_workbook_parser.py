# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

import json

import pytest
from openpyxl import load_workbook


class TestDeserializeJsonColumns:
    """Tests for _deserialize_json_columns helper."""

    def test_deserializes_scores_string_to_dict(self):
        from src.orchestrator.results.frame import _deserialize_json_columns

        rows = [{"scores": '{"skills": 8, "education": 7}', "batch_name": "alice"}]
        result = _deserialize_json_columns(rows)
        assert isinstance(result[0]["scores"], dict)
        assert result[0]["scores"]["skills"] == 8
        assert result[0]["scores"]["education"] == 7

    def test_deserializes_history_list(self):
        from src.orchestrator.results.frame import _deserialize_json_columns

        rows = [{"history": '["a", "b"]', "prompt_name": "test"}]
        result = _deserialize_json_columns(rows)
        assert isinstance(result[0]["history"], list)
        assert result[0]["history"] == ["a", "b"]

    def test_leaves_non_json_strings_unchanged(self):
        from src.orchestrator.results.frame import _deserialize_json_columns

        rows = [{"response": "plain text", "status": "success"}]
        result = _deserialize_json_columns(rows)
        assert result[0]["response"] == "plain text"
        assert result[0]["status"] == "success"

    def test_leaves_none_unchanged(self):
        from src.orchestrator.results.frame import _deserialize_json_columns

        rows = [{"scores": None, "batch_name": "alice"}]
        result = _deserialize_json_columns(rows)
        assert result[0]["scores"] is None

    def test_leaves_already_native_types_unchanged(self):
        from src.orchestrator.results.frame import _deserialize_json_columns

        rows = [{"scores": {"skills": 8}, "history": ["a"]}]
        result = _deserialize_json_columns(rows)
        assert result[0]["scores"] == {"skills": 8}
        assert result[0]["history"] == ["a"]


class TestWorkbookParserInit:
    """Tests for WorkbookParser initialization."""

    def test_init_sets_workbook_path(self, temp_workbook):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook)
        assert parser.workbook_path == temp_workbook

    def test_has_data_sheet_false_when_no_file(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "nonexistent.xlsx"))
        assert parser.has_data_sheet() is False

    def test_has_clients_sheet_false_when_no_file(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "nonexistent.xlsx"))
        assert parser.has_clients_sheet() is False


class TestWorkbookParserTemplate:
    """Tests for template workbook creation."""

    def test_create_basic_template(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        path = str(tmp_path / "template.xlsx")
        parser = WorkbookParser(path)
        result = parser.create_template_workbook()
        assert result == path

        wb = load_workbook(path)
        assert "config" in wb.sheetnames
        assert "prompts" in wb.sheetnames
        assert "data" not in wb.sheetnames

    def test_create_template_with_data_sheet(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        path = str(tmp_path / "template_data.xlsx")
        parser = WorkbookParser(path)
        parser.create_template_workbook(with_data_sheet=True)

        wb = load_workbook(path)
        assert "data" in wb.sheetnames

    def test_create_template_with_all_sheets(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        path = str(tmp_path / "template_all.xlsx")
        parser = WorkbookParser(path)
        parser.create_template_workbook(
            with_data_sheet=True,
            with_clients_sheet=True,
            with_documents_sheet=True,
            with_tools_sheet=True,
            with_scoring_sheet=True,
            with_synthesis_sheet=True,
        )

        wb = load_workbook(path)
        for sheet in [
            "config",
            "prompts",
            "data",
            "clients",
            "documents",
            "tools",
            "scoring",
            "synthesis",
        ]:
            assert sheet in wb.sheetnames

    def test_template_has_created_at(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        path = str(tmp_path / "template_ts.xlsx")
        parser = WorkbookParser(path)
        parser.create_template_workbook()

        wb = load_workbook(path)
        ws = wb["config"]
        fields = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "created_at" in fields


class TestWorkbookParserValidation:
    """Tests for workbook validation."""

    def test_validate_valid_workbook(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        assert parser.validate_workbook() is True

    def test_validate_missing_workbook(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "missing.xlsx"))
        with pytest.raises(FileNotFoundError):
            parser.validate_workbook()

    def test_validate_missing_config_sheet(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "prompts"
        wb["prompts"]["A1"] = "sequence"
        wb["prompts"]["B1"] = "prompt_name"
        wb["prompts"]["C1"] = "prompt"
        wb["prompts"]["D1"] = "history"
        path = str(tmp_path / "no_config.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        with pytest.raises(ValueError, match="Missing required sheet"):
            parser.validate_workbook()

    def test_validate_missing_required_columns(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        path = str(tmp_path / "missing_cols.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        with pytest.raises(ValueError, match="Missing required columns"):
            parser.validate_workbook()


class TestWorkbookParserLoadConfig:
    """Tests for config loading."""

    def test_load_config_returns_dict(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        config = parser.load_config()
        assert isinstance(config, dict)
        assert "model" in config
        assert config["model"] == "mistral-small-2503"

    def test_load_config_type_coercion(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        config = parser.load_config()
        assert config.get("max_retries") == 3
        assert config.get("temperature") == pytest.approx(0.8)


class TestWorkbookParserLoadPrompts:
    """Tests for prompt loading."""

    def test_load_prompts_returns_list(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        prompts = parser.load_prompts()
        assert isinstance(prompts, list)
        assert len(prompts) == 3

    def test_load_prompts_sorted_by_sequence(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        prompts = parser.load_prompts()
        sequences = [p["sequence"] for p in prompts]
        assert sequences == sorted(sequences)

    def test_load_prompts_fields(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        prompts = parser.load_prompts()
        p = prompts[0]
        assert p["sequence"] == 1
        assert p["prompt_name"] == "greeting"
        assert p["prompt"] == "Hello, how are you?"


class TestWorkbookParserLoadData:
    """Tests for batch data loading."""

    def test_load_data_no_sheet_returns_empty(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        assert parser.load_data() == []

    def test_load_data_with_batch_sheet(self, temp_workbook_with_batch_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_batch_data)
        data = parser.load_data()
        assert len(data) == 3
        assert data[0]["region"] == "north"

    def test_get_data_columns(self, temp_workbook_with_batch_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_batch_data)
        cols = parser.get_data_columns()
        assert "id" in cols
        assert "batch_name" in cols
        assert "region" in cols


class TestWorkbookParserWriteResults:
    """Tests for write_results and write_batch_results."""

    def test_write_results_creates_sheet(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "hello",
                "response": "world",
                "status": "success",
                "attempts": 1,
                "scores": {"skill": 8.0},
                "history": None,
                "batch_id": 0,
                "batch_name": "",
            }
        ]
        sheet_name = parser.write_results(results, "results")
        assert sheet_name == "results"

        wb = load_workbook(temp_workbook_with_data)
        assert "results" in wb.sheetnames
        ws = wb["results"]
        assert ws.cell(row=1, column=1).value == "batch_id"
        assert ws.cell(row=2, column=1).value == 0

    def test_write_results_scores_as_json_string(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        results = [
            {
                "sequence": 1,
                "prompt_name": "eval",
                "prompt": "test",
                "response": "ok",
                "status": "success",
                "attempts": 1,
                "scores": {"skills": 8.0, "education": 7.0},
                "history": None,
                "batch_id": 0,
                "batch_name": "alice",
            }
        ]
        parser.write_results(results, "results")

        wb = load_workbook(temp_workbook_with_data)
        ws = wb["results"]
        scores_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "scores":
                scores_col = col
                break
        assert scores_col is not None
        cell_val = ws.cell(row=2, column=scores_col).value
        parsed = json.loads(cell_val)
        assert parsed == {"skills": 8.0, "education": 7.0}

    def test_write_batch_results_creates_sheet(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "hello",
                "response": "world",
                "status": "success",
                "attempts": 1,
                "history": None,
                "batch_id": 1,
                "batch_name": "alice",
            }
        ]
        sheet_name = parser.write_batch_results(results, "alice")
        assert "alice" in sheet_name

    def test_write_results_dedup_sheet_name(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        results = [
            {"sequence": 1, "prompt_name": "t", "prompt": "", "status": "success", "attempts": 1}
        ]
        first = parser.write_results(results, "results")
        second = parser.write_results(results, "results")
        assert first != second


class TestWorkbookParserHistoryParsing:
    """Tests for history string parsing."""

    def test_parse_valid_json_list(self):
        from src.orchestrator.workbook_parser import parse_history_string

        result = parse_history_string('["a", "b", "c"]')
        assert result == ["a", "b", "c"]

    def test_parse_single_value(self):
        from src.orchestrator.workbook_parser import parse_history_string

        result = parse_history_string("my_prompt")
        assert result == ["my_prompt"]

    def test_parse_none_returns_none(self):
        from src.orchestrator.workbook_parser import parse_history_string

        assert parse_history_string(None) is None

    def test_parse_empty_returns_none(self):
        from src.orchestrator.workbook_parser import parse_history_string

        assert parse_history_string("") is None

    def test_parse_list_returns_list(self):
        from src.orchestrator.workbook_parser import parse_history_string

        result = parse_history_string(["a", "b"])
        assert result == ["a", "b"]

    def test_parse_smart_quotes(self):
        from src.orchestrator.workbook_parser import parse_history_string

        result = parse_history_string("[\u201ca\u201d, \u201cb\u201d]")
        assert isinstance(result, list)
        assert result == ["a", "b"]

    def test_comma_separated_no_brackets_returns_single(self):
        from src.orchestrator.workbook_parser import parse_history_string

        result = parse_history_string("alpha, beta, gamma")
        assert result == ["alpha, beta, gamma"]


class TestWorkbookParserClientsSheet:
    """Tests for clients sheet loading."""

    def test_load_clients_empty_when_no_sheet(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(temp_workbook_with_data)
        assert parser.load_clients() == []

    def test_load_clients_from_workbook(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"

        ws_clients = wb.create_sheet(title="clients")
        headers = ["name", "client_type", "api_key_env", "model"]
        for col, h in enumerate(headers, start=1):
            ws_clients.cell(row=1, column=col, value=h)
        ws_clients.cell(row=2, column=1, value="fast")
        ws_clients.cell(row=2, column=2, value="mistral-small")
        ws_clients.cell(row=2, column=3, value="MISTRALSMALL_KEY")
        ws_clients.cell(row=2, column=4, value="mistral-small-2503")

        path = str(tmp_path / "clients.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        clients = parser.load_clients()
        assert len(clients) == 1
        assert clients[0]["name"] == "fast"
        assert clients[0]["client_type"] == "mistral-small"


class TestSynthesisRunnerScoresDeserialization:
    """Integration test verifying scores survive the ResultsFrame round-trip."""

    def test_scores_survive_results_frame_round_trip(self):
        from src.orchestrator.results import ResultsFrame
        from src.orchestrator.results.frame import _deserialize_json_columns

        results = [
            {
                "sequence": 1,
                "prompt_name": "eval",
                "prompt": "test",
                "status": "success",
                "attempts": 1,
                "batch_id": 0,
                "batch_name": "alice",
                "scores": {"skills": 8.0, "education": 7.0},
                "composite_score": 7.5,
            },
            {
                "sequence": 2,
                "prompt_name": "eval",
                "prompt": "test",
                "status": "success",
                "attempts": 1,
                "batch_id": 1,
                "batch_name": "bob",
                "scores": {"skills": 5.0, "education": 6.0},
                "composite_score": 5.5,
            },
        ]

        frame = ResultsFrame(results)
        batch_groups = frame.by_batch()
        batch_results_list = [
            _deserialize_json_columns(batch_groups[bid].df.to_dicts())
            for bid in sorted(batch_groups.keys())
        ]

        assert isinstance(batch_results_list[0][0]["scores"], dict)
        assert batch_results_list[0][0]["scores"]["skills"] == 8.0
        assert isinstance(batch_results_list[1][0]["scores"], dict)
        assert batch_results_list[1][0]["scores"]["skills"] == 5.0

    def test_scores_used_for_tiebreaking_after_deserialization(self):
        from src.orchestrator.results import ResultsFrame
        from src.orchestrator.results.frame import _deserialize_json_columns
        from src.orchestrator.synthesis import SynthesisExecutor

        results = [
            {
                "sequence": 1,
                "prompt_name": "eval",
                "prompt": "test",
                "status": "success",
                "attempts": 1,
                "batch_id": 0,
                "batch_name": "alice",
                "scores": {"skills": 8.0, "education": 7.0},
                "composite_score": 7.5,
            },
            {
                "sequence": 1,
                "prompt_name": "eval",
                "prompt": "test",
                "status": "success",
                "attempts": 1,
                "batch_id": 1,
                "batch_name": "bob",
                "scores": {"skills": 6.0, "education": 9.0},
                "composite_score": 7.5,
            },
        ]

        frame = ResultsFrame(results)
        batch_groups = frame.by_batch()
        batch_results_list = [
            _deserialize_json_columns(batch_groups[bid].df.to_dicts())
            for bid in sorted(batch_groups.keys())
        ]

        executor = SynthesisExecutor()
        criteria = [{"criteria_name": "skills"}, {"criteria_name": "education"}]
        sorted_entries = executor.sort_entries(
            batch_results_list, scoring_criteria=criteria, has_scoring=True
        )
        assert sorted_entries[0]["batch_name"] == "alice"
        assert sorted_entries[0]["scores"]["skills"] == 8.0

    def test_scores_appear_in_context_format_after_deserialization(self):
        from src.orchestrator.results import ResultsFrame
        from src.orchestrator.results.frame import _deserialize_json_columns
        from src.orchestrator.synthesis import SynthesisExecutor

        results = [
            {
                "sequence": 1,
                "prompt_name": "eval",
                "prompt": "test",
                "status": "success",
                "attempts": 1,
                "batch_id": 0,
                "batch_name": "alice",
                "scores": {"skills": 8.0},
                "composite_score": 8.0,
                "response": "Good candidate",
            },
        ]

        frame = ResultsFrame(results)
        batch_groups = frame.by_batch()
        batch_results_list = [
            _deserialize_json_columns(batch_groups[bid].df.to_dicts())
            for bid in sorted(batch_groups.keys())
        ]

        executor = SynthesisExecutor(max_context_chars=50000)
        context = executor.format_entry_context(batch_results_list[0], [], include_scores=True)
        assert "skills: 8.0/10" in context


class TestWorkbookParserWriteScoresPivot:
    """Tests for write_scores_pivot method."""

    def test_pivot_created_with_normalized_score_criteria(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        builder = WorkbookParser(temp_workbook_with_data)

        results = [
            {
                "batch_id": 1,
                "batch_name": "alice_chen",
                "sequence": 1,
                "prompt_name": "eval",
                "prompt": "Evaluate",
                "response": '{"skills_match": 8}',
                "status": "success",
                "attempts": 1,
                "scores": {"skills_match": 8.0, "education": 7.0},
                "composite_score": 7.5,
                "scoring_status": "ok",
            },
            {
                "batch_id": 2,
                "batch_name": "bob_martinez",
                "sequence": 1,
                "prompt_name": "eval",
                "prompt": "Evaluate",
                "response": '{"skills_match": 5}',
                "status": "success",
                "attempts": 1,
                "scores": {"skills_match": 5.0, "education": 6.0},
                "composite_score": 5.5,
                "scoring_status": "ok",
            },
        ]

        scoring_criteria = [
            {
                "criteria_name": "skills_match",
                "description": "Skills alignment",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 1.0,
                "source_prompt": "eval",
                "score_type": "normalized_score",
            },
            {
                "criteria_name": "education",
                "description": "Education quality",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 0.8,
                "source_prompt": "eval",
                "score_type": "normalized_score",
            },
        ]

        sheet_name = builder.write_scores_pivot(results, scoring_criteria)

        assert sheet_name is not None
        wb = load_workbook(temp_workbook_with_data)
        assert sheet_name in wb.sheetnames

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "batch_name",
            "criteria_name",
            "label_1",
            "label_2",
            "label_3",
            "normalized_score",
            "weight",
            "weight_tier",
            "weighted_score",
            "rank",
            "percentile",
            "percent_rank",
            "scale_min",
            "scale_max",
            "description",
        ]

        assert ws.max_row == 7

        rows_data = []
        for row in range(2, ws.max_row + 1):
            rows_data.append(
                {
                    "batch_name": ws.cell(row=row, column=1).value,
                    "criteria_name": ws.cell(row=row, column=2).value,
                    "normalized_score": ws.cell(row=row, column=6).value,
                    "weight": ws.cell(row=row, column=7).value,
                    "weighted_score": ws.cell(row=row, column=9).value,
                    "rank": ws.cell(row=row, column=10).value,
                    "percentile": ws.cell(row=row, column=11).value,
                    "percent_rank": ws.cell(row=row, column=12).value,
                }
            )

        assert rows_data[0]["batch_name"] == "alice_chen"
        assert rows_data[0]["criteria_name"] == "skills_match"
        assert rows_data[0]["normalized_score"] == 8.0
        assert rows_data[0]["weight"] == 1.0
        assert rows_data[0]["weighted_score"] == 8.0
        assert rows_data[0]["rank"] == 1
        assert rows_data[0]["percentile"] == 100
        assert rows_data[0]["percent_rank"] == 100
        assert rows_data[1]["batch_name"] == "alice_chen"
        assert rows_data[1]["criteria_name"] == "education"
        assert rows_data[1]["normalized_score"] == 7.0
        assert rows_data[1]["weight"] == 0.8
        assert rows_data[1]["weighted_score"] == pytest.approx(5.6)
        assert rows_data[1]["rank"] == 1
        assert rows_data[1]["percentile"] == 100
        assert rows_data[1]["percent_rank"] == 100
        assert rows_data[2]["batch_name"] == "bob_martinez"
        assert rows_data[2]["criteria_name"] == "skills_match"
        assert rows_data[2]["normalized_score"] == 5.0
        assert rows_data[2]["weight"] == 1.0
        assert rows_data[2]["weighted_score"] == 5.0
        assert rows_data[2]["rank"] == 2
        assert rows_data[2]["percentile"] == 0
        assert rows_data[2]["percent_rank"] == 0
        assert rows_data[3]["batch_name"] == "bob_martinez"
        assert rows_data[3]["criteria_name"] == "education"
        assert rows_data[3]["rank"] == 2
        assert rows_data[3]["percentile"] == 0
        assert rows_data[3]["percent_rank"] == 0
        assert rows_data[4]["batch_name"] == "alice_chen"
        assert rows_data[4]["criteria_name"] == "_composite"
        assert rows_data[4]["weighted_score"] == 7.5
        assert rows_data[4]["rank"] == 1
        assert rows_data[4]["percentile"] == 100
        assert rows_data[4]["percent_rank"] == 100
        assert rows_data[5]["batch_name"] == "bob_martinez"
        assert rows_data[5]["criteria_name"] == "_composite"
        assert rows_data[5]["weighted_score"] == 5.5
        assert rows_data[5]["rank"] == 2
        assert rows_data[5]["percentile"] == 0
        assert rows_data[5]["percent_rank"] == 0

    def test_pivot_returns_none_when_no_normalized_criteria(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        builder = WorkbookParser(temp_workbook_with_data)

        results = [
            {
                "batch_name": "alice",
                "scores": {"skills_match": 8.0},
            }
        ]

        scoring_criteria = [
            {
                "criteria_name": "skills_match",
                "description": "Skills",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 1.0,
                "source_prompt": "eval",
                "score_type": "",
            },
        ]

        result = builder.write_scores_pivot(results, scoring_criteria)
        assert result is None

    def test_pivot_returns_none_when_no_eligible_scores(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        builder = WorkbookParser(temp_workbook_with_data)

        results = [
            {
                "batch_name": "alice",
                "scores": {},
            }
        ]

        scoring_criteria = [
            {
                "criteria_name": "skills_match",
                "description": "Skills",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 1.0,
                "source_prompt": "eval",
                "score_type": "normalized_score",
            },
        ]

        result = builder.write_scores_pivot(results, scoring_criteria)
        assert result is None

    def test_pivot_excludes_non_normalized_criteria(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        builder = WorkbookParser(temp_workbook_with_data)

        results = [
            {
                "batch_name": "alice",
                "scores": {"skills_match": 8.0, "raw_years": 12.0},
                "composite_score": 8.0,
            }
        ]

        scoring_criteria = [
            {
                "criteria_name": "skills_match",
                "description": "Skills",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 1.0,
                "source_prompt": "eval",
                "score_type": "normalized_score",
            },
            {
                "criteria_name": "raw_years",
                "description": "Years of experience",
                "scale_min": 0,
                "scale_max": 30,
                "weight": 0.5,
                "source_prompt": "eval",
                "score_type": "",
            },
        ]

        sheet_name = builder.write_scores_pivot(results, scoring_criteria)
        assert sheet_name is not None

        wb = load_workbook(temp_workbook_with_data)
        ws = wb[sheet_name]
        assert ws.max_row == 3
        assert ws.cell(row=2, column=2).value == "skills_match"
        assert ws.cell(row=2, column=10).value == 1
        assert ws.cell(row=2, column=11).value == 100
        assert ws.cell(row=3, column=2).value == "_composite"
        assert ws.cell(row=3, column=9).value == 8.0
        assert ws.cell(row=3, column=10).value == 1
        assert ws.cell(row=3, column=11).value == 100

    def test_pivot_per_criteria_ranking(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        builder = WorkbookParser(temp_workbook_with_data)

        results = [
            {
                "batch_id": 1,
                "batch_name": "alice",
                "scores": {"python": 9.0, "docker": 5.0},
                "composite_score": 7.0,
            },
            {
                "batch_id": 2,
                "batch_name": "bob",
                "scores": {"python": 6.0, "docker": 8.0},
                "composite_score": 7.0,
            },
            {
                "batch_id": 3,
                "batch_name": "carol",
                "scores": {"python": 6.0, "docker": 5.0},
                "composite_score": 5.5,
            },
        ]

        scoring_criteria = [
            {
                "criteria_name": "python",
                "description": "Python",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 1.0,
                "source_prompt": "eval",
                "score_type": "normalized_score",
            },
            {
                "criteria_name": "docker",
                "description": "Docker",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 1.0,
                "source_prompt": "eval",
                "score_type": "normalized_score",
            },
        ]

        sheet_name = builder.write_scores_pivot(results, scoring_criteria)
        assert sheet_name is not None

        wb = load_workbook(temp_workbook_with_data)
        ws = wb[sheet_name]

        rows_by_key = {}
        for row in range(2, ws.max_row + 1):
            bn = ws.cell(row=row, column=1).value
            cn = ws.cell(row=row, column=2).value
            rows_by_key[(bn, cn)] = {
                "normalized_score": ws.cell(row=row, column=6).value,
                "rank": ws.cell(row=row, column=10).value,
                "percentile": ws.cell(row=row, column=11).value,
                "percent_rank": ws.cell(row=row, column=12).value,
            }

        assert rows_by_key[("alice", "python")]["rank"] == 1
        assert rows_by_key[("alice", "python")]["percentile"] == 100
        assert rows_by_key[("alice", "python")]["percent_rank"] == 100
        assert rows_by_key[("bob", "python")]["rank"] == 2
        assert rows_by_key[("bob", "python")]["percentile"] == 50
        assert rows_by_key[("bob", "python")]["percent_rank"] == 0
        assert rows_by_key[("carol", "python")]["rank"] == 2
        assert rows_by_key[("carol", "python")]["percentile"] == 50
        assert rows_by_key[("carol", "python")]["percent_rank"] == 0

        assert rows_by_key[("bob", "docker")]["rank"] == 1
        assert rows_by_key[("bob", "docker")]["percentile"] == 100
        assert rows_by_key[("bob", "docker")]["percent_rank"] == 100
        assert rows_by_key[("alice", "docker")]["rank"] == 2
        assert rows_by_key[("alice", "docker")]["percentile"] == 50
        assert rows_by_key[("alice", "docker")]["percent_rank"] == 0
        assert rows_by_key[("carol", "docker")]["rank"] == 2
        assert rows_by_key[("carol", "docker")]["percentile"] == 50
        assert rows_by_key[("carol", "docker")]["percent_rank"] == 0

        assert rows_by_key[("alice", "_composite")]["rank"] == 1
        assert rows_by_key[("alice", "_composite")]["percentile"] == 100
        assert rows_by_key[("alice", "_composite")]["percent_rank"] == 50
        assert rows_by_key[("bob", "_composite")]["rank"] == 1
        assert rows_by_key[("bob", "_composite")]["percentile"] == 100
        assert rows_by_key[("bob", "_composite")]["percent_rank"] == 50
        assert rows_by_key[("carol", "_composite")]["rank"] == 2
        assert rows_by_key[("carol", "_composite")]["percentile"] == 50
        assert rows_by_key[("carol", "_composite")]["percent_rank"] == 0

    def test_pivot_deduplicates_per_batch(self, temp_workbook_with_data):
        from src.orchestrator.workbook_parser import WorkbookParser

        builder = WorkbookParser(temp_workbook_with_data)

        results = [
            {
                "batch_name": "alice",
                "scores": {"skills_match": 8.0},
                "composite_score": 8.0,
            },
            {
                "batch_name": "alice",
                "scores": {"skills_match": 8.0},
                "composite_score": 8.0,
            },
        ]

        scoring_criteria = [
            {
                "criteria_name": "skills_match",
                "description": "Skills",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 1.0,
                "source_prompt": "eval",
                "score_type": "normalized_score",
            },
        ]

        sheet_name = builder.write_scores_pivot(results, scoring_criteria)
        assert sheet_name is not None

        wb = load_workbook(temp_workbook_with_data)
        ws = wb[sheet_name]
        assert ws.max_row == 3
        assert ws.cell(row=2, column=2).value == "skills_match"
        assert ws.cell(row=2, column=10).value == 1
        assert ws.cell(row=2, column=11).value == 100
        assert ws.cell(row=2, column=12).value == 100
        assert ws.cell(row=3, column=2).value == "_composite"
        assert ws.cell(row=3, column=10).value == 1
        assert ws.cell(row=3, column=11).value == 100
        assert ws.cell(row=3, column=12).value == 100


class TestValidationWorkbookParsing:
    def test_prompts_headers_include_validation(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        assert "validation_prompt" in WorkbookParser.PROMPTS_HEADERS
        assert "max_validation_retries" in WorkbookParser.PROMPTS_HEADERS

    def test_results_headers_include_validation(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        assert "validation_passed" in WorkbookParser.RESULTS_HEADERS
        assert "validation_attempts" in WorkbookParser.RESULTS_HEADERS
        assert "validation_critique" in WorkbookParser.RESULTS_HEADERS

    def test_sample_prompt_spec_includes_validation_columns(self):
        import sys
        from pathlib import Path

        scripts_dir = str(Path(__file__).parent.parent / "scripts")
        if scripts_dir not in sys.path:
            sys.path.insert(0, scripts_dir)

        from scripts.sample_workbooks import DEFAULT_PROMPT_HEADERS, PromptSpec

        prompt = PromptSpec(
            sequence=1,
            name="validated",
            prompt="Use calculate to compute 9.",
            validation_prompt="The result must be 9.",
            max_validation_retries=2,
        )

        row = prompt.to_row()
        assert "validation_prompt" in DEFAULT_PROMPT_HEADERS
        assert "max_validation_retries" in DEFAULT_PROMPT_HEADERS
        assert row["validation_prompt"] == "The result must be 9."
        assert row["max_validation_retries"] == 2

    def test_sample_prompt_spec_includes_notes(self):
        import sys
        from pathlib import Path

        scripts_dir = str(Path(__file__).parent.parent / "scripts")
        if scripts_dir not in sys.path:
            sys.path.insert(0, scripts_dir)

        from scripts.sample_workbooks import DEFAULT_PROMPT_HEADERS, PromptSpec

        prompt = PromptSpec(
            sequence=1,
            name="annotated",
            prompt="Say hello",
            notes="Helpful note for workbook users.",
        )

        row = prompt.to_row()
        assert "notes" in DEFAULT_PROMPT_HEADERS
        assert row["notes"] == "Helpful note for workbook users."


class TestParseHistoryStringFallbacks:
    """Tests for parse_history_string fallback parsing paths."""

    def test_bracketed_invalid_json_extracts_quoted_items(self):
        from src.orchestrator.workbook_parser import parse_history_string

        result = parse_history_string('["a" "b"]')
        assert result == ["a", "b"]

    def test_bracketed_invalid_json_comma_split(self):
        from src.orchestrator.workbook_parser import parse_history_string

        result = parse_history_string("[item1, item2]")
        assert result == ["item1", "item2"]

    def test_bracketed_unquoted_items_comma_split(self):
        from src.orchestrator.workbook_parser import parse_history_string

        result = parse_history_string("[alpha, beta, gamma]")
        assert result == ["alpha", "beta", "gamma"]


class TestHasSheetCachingAndNoFile:
    """Tests for has_*_sheet cached returns and no-file returns."""

    def test_has_documents_sheet_false_when_no_file(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "nonexistent.xlsx"))
        assert parser.has_documents_sheet() is False

    def test_has_tools_sheet_false_when_no_file(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "nonexistent.xlsx"))
        assert parser.has_tools_sheet() is False

    def test_has_scoring_sheet_false_when_no_file(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "nonexistent.xlsx"))
        assert parser.has_scoring_sheet() is False

    def test_has_synthesis_sheet_false_when_no_file(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "nonexistent.xlsx"))
        assert parser.has_synthesis_sheet() is False

    def test_has_documents_sheet_cached_return(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        path = str(tmp_path / "cached.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        assert parser._has_documents_sheet is None
        result = parser.has_documents_sheet()
        assert result is False
        assert parser._has_documents_sheet is False
        cached = parser.has_documents_sheet()
        assert cached is False

    def test_has_tools_sheet_cached_return(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        path = str(tmp_path / "cached_tools.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        assert parser._has_tools_sheet is None
        parser.has_tools_sheet()
        assert parser._has_tools_sheet is False

    def test_has_scoring_sheet_cached_return(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        path = str(tmp_path / "cached_scoring.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        assert parser._has_scoring_sheet is None
        parser.has_scoring_sheet()
        assert parser._has_scoring_sheet is False

    def test_has_synthesis_sheet_cached_return(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        path = str(tmp_path / "cached_synth.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        assert parser._has_synthesis_sheet is None
        parser.has_synthesis_sheet()
        assert parser._has_synthesis_sheet is False


class TestValidateConfigErrorBranches:
    """Tests for validate_config error conditions."""

    def test_unknown_client_type(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"client_type": "nonexistent_client"})
        assert len(errors) == 1
        assert "Unknown client_type" in errors[0]
        assert "nonexistent_client" in errors[0]

    def test_temperature_out_of_range_high(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"temperature": 3.0})
        assert len(errors) == 1
        assert "temperature" in errors[0]
        assert "out of range" in errors[0]

    def test_temperature_out_of_range_negative(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"temperature": -0.5})
        assert len(errors) == 1
        assert "out of range" in errors[0]

    def test_temperature_not_a_number(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"temperature": "abc"})
        assert len(errors) == 1
        assert "not a valid number" in errors[0]

    def test_max_retries_out_of_range(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"max_retries": 11})
        assert len(errors) == 1
        assert "max_retries" in errors[0]
        assert "out of range" in errors[0]

    def test_max_retries_zero_out_of_range(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"max_retries": 0})
        assert len(errors) == 1
        assert "out of range" in errors[0]

    def test_max_retries_not_an_integer(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"max_retries": "abc"})
        assert len(errors) == 1
        assert "not a valid integer" in errors[0]

    def test_invalid_batch_mode(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"batch_mode": "invalid_mode"})
        assert len(errors) == 1
        assert "batch_mode" in errors[0]
        assert "not supported" in errors[0]

    def test_invalid_batch_output(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"batch_output": "invalid_output"})
        assert len(errors) == 1
        assert "batch_output" in errors[0]

    def test_invalid_on_batch_error(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config({"on_batch_error": "invalid_error"})
        assert len(errors) == 1
        assert "on_batch_error" in errors[0]

    def test_valid_config_returns_no_errors(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config(
            {
                "client_type": "mistral-small",
                "temperature": 0.8,
                "max_retries": 3,
                "batch_mode": "per_row",
                "batch_output": "combined",
                "on_batch_error": "continue",
            }
        )
        assert errors == []

    def test_multiple_errors_accumulate(self):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser("/tmp/test.xlsx")
        errors = parser.validate_config(
            {
                "temperature": 5.0,
                "max_retries": 20,
                "batch_mode": "bad",
            }
        )
        assert len(errors) == 3


class TestLoadPromptsSkipEmptyRow:
    """Tests for load_prompts skipping rows with empty first column."""

    def test_empty_sequence_row_is_skipped(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"

        ws_prompts.cell(2, 1, 1)
        ws_prompts.cell(2, 2, "first")
        ws_prompts.cell(2, 3, "Hello")
        ws_prompts.cell(2, 4, None)

        ws_prompts.cell(3, 1, None)
        ws_prompts.cell(3, 2, "empty_row")
        ws_prompts.cell(3, 3, "Should be skipped")
        ws_prompts.cell(3, 4, None)

        ws_prompts.cell(4, 1, 2)
        ws_prompts.cell(4, 2, "second")
        ws_prompts.cell(4, 3, "World")
        ws_prompts.cell(4, 4, None)

        path = str(tmp_path / "skip_empty.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        prompts = parser.load_prompts()
        assert len(prompts) == 2
        assert prompts[0]["prompt_name"] == "first"
        assert prompts[1]["prompt_name"] == "second"


class TestLoadScoringWithStringValues:
    """Tests for load_scoring coercing string scale/weight values."""

    def test_string_scale_and_weight_coerced(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"

        ws_scoring = wb.create_sheet(title="scoring")
        headers = [
            "criteria_name",
            "description",
            "scale_min",
            "scale_max",
            "weight",
            "source_prompt",
            "score_type",
            "label_1",
            "label_2",
            "label_3",
        ]
        for col, h in enumerate(headers, 1):
            ws_scoring.cell(1, col, h)
        ws_scoring.cell(2, 1, "skills")
        ws_scoring.cell(2, 2, "Skills evaluation")
        ws_scoring.cell(2, 3, "1")
        ws_scoring.cell(2, 4, "10")
        ws_scoring.cell(2, 5, "0.8")
        ws_scoring.cell(2, 6, "eval")
        ws_scoring.cell(2, 7, "normalized_score")

        path = str(tmp_path / "scoring_str.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        scoring = parser.load_scoring()
        assert len(scoring) == 1
        assert scoring[0]["scale_min"] == 1
        assert isinstance(scoring[0]["scale_min"], int)
        assert scoring[0]["scale_max"] == 10
        assert isinstance(scoring[0]["scale_max"], int)
        assert scoring[0]["weight"] == pytest.approx(0.8)
        assert isinstance(scoring[0]["weight"], float)


class TestLoadToolsWithStringEnabled:
    """Tests for load_tools parsing string enabled values."""

    def test_string_false_disables_tool(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"

        ws_tools = wb.create_sheet(title="tools")
        headers = ["name", "description", "parameters", "implementation", "enabled"]
        for col, h in enumerate(headers, 1):
            ws_tools.cell(1, col, h)
        ws_tools.cell(2, 1, "calculator")
        ws_tools.cell(2, 2, "A calculator")
        ws_tools.cell(2, 3, '{"type": "object"}')
        ws_tools.cell(2, 4, "builtin:calculator")
        ws_tools.cell(2, 5, "false")

        path = str(tmp_path / "tools_str.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        tools = parser.load_tools()
        assert len(tools) == 1
        assert tools[0]["enabled"] is False

    def test_string_zero_disables_tool(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"

        ws_tools = wb.create_sheet(title="tools")
        headers = ["name", "description", "parameters", "implementation", "enabled"]
        for col, h in enumerate(headers, 1):
            ws_tools.cell(1, col, h)
        ws_tools.cell(2, 1, "search")
        ws_tools.cell(2, 2, "A search tool")
        ws_tools.cell(2, 3, '{"type": "object"}')
        ws_tools.cell(2, 4, "builtin:search")
        ws_tools.cell(2, 5, "0")

        path = str(tmp_path / "tools_zero.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        tools = parser.load_tools()
        assert len(tools) == 1
        assert tools[0]["enabled"] is False


class TestLoadSynthesisEdgeCases:
    """Tests for load_synthesis with string sequence and bool include_scores."""

    def test_string_sequence_parsed_correctly(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"

        ws_synth = wb.create_sheet(title="synthesis")
        headers = [
            "sequence",
            "prompt_name",
            "prompt",
            "source_scope",
            "source_prompts",
            "include_scores",
            "history",
            "condition",
        ]
        for col, h in enumerate(headers, 1):
            ws_synth.cell(1, col, h)
        ws_synth.cell(2, 1, "2")
        ws_synth.cell(2, 2, "rank")
        ws_synth.cell(2, 3, "Rank candidates")
        ws_synth.cell(2, 4, "all")
        ws_synth.cell(2, 5, '["eval", "screen"]')
        ws_synth.cell(2, 6, True)
        ws_synth.cell(2, 7, None)
        ws_synth.cell(2, 8, None)

        path = str(tmp_path / "synth_str.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        synth = parser.load_synthesis()
        assert len(synth) == 1
        assert synth[0]["sequence"] == 2
        assert synth[0]["source_prompts"] == ["eval", "screen"]
        assert synth[0]["include_scores"] is True

    def test_invalid_string_sequence_skipped(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"

        ws_synth = wb.create_sheet(title="synthesis")
        headers = [
            "sequence",
            "prompt_name",
            "prompt",
            "source_scope",
            "source_prompts",
            "include_scores",
            "history",
            "condition",
        ]
        for col, h in enumerate(headers, 1):
            ws_synth.cell(1, col, h)
        ws_synth.cell(2, 1, "abc")
        ws_synth.cell(2, 2, "bad_row")
        ws_synth.cell(2, 3, "Should be skipped")
        ws_synth.cell(2, 4, "all")

        path = str(tmp_path / "synth_invalid.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        synth = parser.load_synthesis()
        assert synth == []

    def test_bool_include_scores_false(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"

        ws_synth = wb.create_sheet(title="synthesis")
        headers = [
            "sequence",
            "prompt_name",
            "prompt",
            "source_scope",
            "source_prompts",
            "include_scores",
            "history",
            "condition",
        ]
        for col, h in enumerate(headers, 1):
            ws_synth.cell(1, col, h)
        ws_synth.cell(2, 1, 1)
        ws_synth.cell(2, 2, "summary")
        ws_synth.cell(2, 3, "Summarize")
        ws_synth.cell(2, 4, "all")
        ws_synth.cell(2, 5, None)
        ws_synth.cell(2, 6, False)
        ws_synth.cell(2, 7, None)
        ws_synth.cell(2, 8, None)

        path = str(tmp_path / "synth_bool.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        synth = parser.load_synthesis()
        assert len(synth) == 1
        assert synth[0]["include_scores"] is False


class TestInferChunkingStrategy:
    """Tests for _infer_chunking_strategy by file extension."""

    def test_markdown_extension(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "test.xlsx"))
        assert parser._infer_chunking_strategy("guide.md") == "markdown"

    def test_python_extension(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "test.xlsx"))
        assert parser._infer_chunking_strategy("script.py") == "code"

    def test_javascript_extension(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "test.xlsx"))
        assert parser._infer_chunking_strategy("app.js") == "code"

    def test_typescript_extension(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "test.xlsx"))
        assert parser._infer_chunking_strategy("module.ts") == "code"

    def test_unknown_extension_defaults_recursive(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "test.xlsx"))
        assert parser._infer_chunking_strategy("document.txt") == "recursive"

    def test_case_insensitive_extension(self, tmp_path):
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(str(tmp_path / "test.xlsx"))
        assert parser._infer_chunking_strategy("README.MD") == "markdown"


class TestWriteBatchResultsCollision:
    """Tests for write_batch_results sheet name collision handling."""

    def test_colliding_sheet_name_gets_counter_suffix(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        wb.create_sheet("results_alice")

        path = str(tmp_path / "collision.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "hello",
                "response": "world",
                "status": "success",
                "attempts": 1,
                "history": None,
                "batch_id": 1,
                "batch_name": "alice",
            }
        ]
        sheet_name = parser.write_batch_results(results, "alice")
        assert sheet_name == "results_alice_1"

        wb2 = load_workbook(path)
        assert "results_alice_1" in wb2.sheetnames
        assert "results_alice" in wb2.sheetnames


class TestWriteScoresPivotCollision:
    """Tests for write_scores_pivot sheet name collision handling."""

    def test_colliding_pivot_sheet_gets_timestamp_suffix(self, tmp_path):
        from openpyxl import Workbook

        from src.orchestrator.workbook_parser import WorkbookParser

        wb = Workbook()
        wb.active.title = "config"
        wb["config"]["A1"] = "field"
        wb["config"]["B1"] = "value"
        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        wb.create_sheet("scores_pivot")

        path = str(tmp_path / "pivot_collision.xlsx")
        wb.save(path)

        parser = WorkbookParser(path)
        results = [
            {
                "batch_name": "alice",
                "scores": {"skills": 8.0},
                "composite_score": 8.0,
            }
        ]
        criteria = [
            {
                "criteria_name": "skills",
                "description": "Skills",
                "scale_min": 1,
                "scale_max": 10,
                "weight": 1.0,
                "source_prompt": "eval",
                "score_type": "normalized_score",
            },
        ]
        sheet_name = parser.write_scores_pivot(results, criteria)
        assert sheet_name.startswith("scores_pivot_")
        assert sheet_name != "scores_pivot"

        wb2 = load_workbook(path)
        assert sheet_name in wb2.sheetnames
        assert "scores_pivot" in wb2.sheetnames
