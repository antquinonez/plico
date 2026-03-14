# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Tests for execution strategies."""

from src.orchestrator.execution import (
    BatchParallelStrategy,
    BatchSequentialStrategy,
    ParallelStrategy,
    SequentialStrategy,
    get_execution_strategy,
)


class TestGetExecutionStrategy:
    """Tests for strategy factory."""

    def test_returns_sequential_for_concurrency_1(self):
        """Test that concurrency=1 returns SequentialStrategy."""
        strategy = get_execution_strategy(is_batch_mode=False, concurrency=1)
        assert isinstance(strategy, SequentialStrategy)

    def test_returns_parallel_for_concurrency_gt_1(self):
        """Test that concurrency>1 returns ParallelStrategy."""
        strategy = get_execution_strategy(is_batch_mode=False, concurrency=4)
        assert isinstance(strategy, ParallelStrategy)

    def test_returns_batch_sequential_for_batch_mode_concurrency_1(self):
        """Test that batch mode with concurrency=1 returns BatchSequentialStrategy."""
        strategy = get_execution_strategy(is_batch_mode=True, concurrency=1)
        assert isinstance(strategy, BatchSequentialStrategy)

    def test_returns_batch_parallel_for_batch_mode_concurrency_gt_1(self):
        """Test that batch mode with concurrency>1 returns BatchParallelStrategy."""
        strategy = get_execution_strategy(is_batch_mode=True, concurrency=4)
        assert isinstance(strategy, BatchParallelStrategy)


class TestStrategyNames:
    """Tests for strategy name property."""

    def test_sequential_name(self):
        """Test SequentialStrategy name."""
        assert SequentialStrategy().name == "sequential"

    def test_parallel_name(self):
        """Test ParallelStrategy name."""
        assert ParallelStrategy().name == "parallel"

    def test_batch_sequential_name(self):
        """Test BatchSequentialStrategy name."""
        assert BatchSequentialStrategy().name == "batch_sequential"

    def test_batch_parallel_name(self):
        """Test BatchParallelStrategy name."""
        assert BatchParallelStrategy().name == "batch_parallel"


class TestSequentialStrategy:
    """Tests for SequentialStrategy execution."""

    def test_execute_returns_results(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that execute returns list of results."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        strategy = SequentialStrategy()
        results = strategy.execute(orchestrator)

        assert len(results) == 3
        assert all(r["status"] == "success" for r in results)

    def test_execute_respects_dependencies(self, temp_workbook, mock_ffmistralsmall):
        """Test that dependencies are respected."""
        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "test-model"
        ws["A3"] = "max_retries"
        ws["B3"] = 1

        ws = wb.create_sheet("prompts")
        ws["A1"] = "sequence"
        ws["B1"] = "prompt_name"
        ws["C1"] = "prompt"
        ws["D1"] = "history"

        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "Hello"
        ws["D2"] = ""

        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "World"
        ws["D3"] = '["first"]'

        wb.save(temp_workbook)

        orchestrator = ExcelOrchestrator(temp_workbook, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        strategy = SequentialStrategy()
        results = strategy.execute(orchestrator)

        first_result = next(r for r in results if r["prompt_name"] == "first")
        second_result = next(r for r in results if r["prompt_name"] == "second")

        assert first_result["sequence"] == 1
        assert second_result["sequence"] == 2
        assert second_result["history"] == ["first"]


class TestParallelStrategy:
    """Tests for ParallelStrategy execution."""

    def test_execute_returns_results(self, temp_workbook_with_data, mock_ffmistralsmall):
        """Test that parallel execute returns list of results."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_data, mock_ffmistralsmall, concurrency=2
        )
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()

        strategy = ParallelStrategy()
        results = strategy.execute(orchestrator)

        assert len(results) == 3
        assert all(r["status"] == "success" for r in results)


class TestBatchSequentialStrategy:
    """Tests for BatchSequentialStrategy execution."""

    def test_execute_returns_batch_results(
        self, temp_workbook_with_batch_data, mock_ffmistralsmall
    ):
        """Test that batch sequential returns all batch results."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(temp_workbook_with_batch_data, mock_ffmistralsmall)
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()
        orchestrator.batch_data = orchestrator.builder.load_data()

        strategy = BatchSequentialStrategy()
        results = strategy.execute(orchestrator)

        assert len(results) == 9
        batch_ids = set(r["batch_id"] for r in results)
        assert batch_ids == {1, 2, 3}


class TestBatchParallelStrategy:
    """Tests for BatchParallelStrategy execution."""

    def test_execute_returns_batch_results(
        self, temp_workbook_with_batch_data, mock_ffmistralsmall
    ):
        """Test that batch parallel returns all batch results."""
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            temp_workbook_with_batch_data, mock_ffmistralsmall, concurrency=2
        )
        orchestrator._init_workbook()
        orchestrator._load_config()
        orchestrator.prompts = orchestrator.builder.load_prompts()
        orchestrator._init_client()
        orchestrator.batch_data = orchestrator.builder.load_data()

        strategy = BatchParallelStrategy()
        results = strategy.execute(orchestrator)

        assert len(results) == 9
        assert all(r["status"] == "success" for r in results)
