# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Tests for manifest-based orchestration."""

import os
import sys

import polars as pl
import yaml

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class TestWorkbookManifestExporter:
    """Tests for WorkbookManifestExporter class."""

    def test_export_creates_manifest_folder(self, temp_workbook_with_data, tmp_path):
        """Test that export creates a manifest folder."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        assert os.path.isdir(manifest_path)
        assert "manifest_" in manifest_path

    def test_export_creates_required_files(self, temp_workbook_with_data, tmp_path):
        """Test that export creates all required YAML files."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        assert os.path.exists(os.path.join(manifest_path, "manifest.yaml"))
        assert os.path.exists(os.path.join(manifest_path, "config.yaml"))
        assert os.path.exists(os.path.join(manifest_path, "prompts.yaml"))

    def test_export_manifest_yaml_content(self, temp_workbook_with_data, tmp_path):
        """Test manifest.yaml content."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "manifest.yaml"), encoding="utf-8") as f:
            manifest = yaml.safe_load(f)

        assert manifest["version"] == "1.0"
        assert "source_workbook" in manifest
        assert "exported_at" in manifest
        assert manifest["prompt_count"] == 3
        assert manifest["has_data"] is False
        assert manifest["has_clients"] is False
        assert manifest["has_documents"] is False

    def test_export_config_yaml_content(self, temp_workbook_with_data, tmp_path):
        """Test config.yaml content."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "config.yaml"), encoding="utf-8") as f:
            config = yaml.safe_load(f)

        assert config["model"] == "mistral-small-2503"
        assert config["max_retries"] == 3
        assert config["temperature"] == 0.8

    def test_export_prompts_yaml_content(self, temp_workbook_with_data, tmp_path):
        """Test prompts.yaml content."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "prompts.yaml"), encoding="utf-8") as f:
            prompts_data = yaml.safe_load(f)

        prompts = prompts_data["prompts"]
        assert len(prompts) == 3

        assert prompts[0]["sequence"] == 1
        assert prompts[0]["prompt_name"] == "greeting"
        assert prompts[0]["prompt"] == "Hello, how are you?"

        assert prompts[2]["sequence"] == 3
        assert prompts[2]["history"] == ["math", "greeting"]

    def test_export_with_batch_data(self, temp_workbook_with_batch_data, tmp_path):
        """Test export with batch data creates data.yaml."""
        from src.orchestrator.manifest import WorkbookManifestExporter

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_batch_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "manifest.yaml"), encoding="utf-8") as f:
            manifest = yaml.safe_load(f)

        assert manifest["has_data"] is True

        assert os.path.exists(os.path.join(manifest_path, "data.yaml"))

        with open(os.path.join(manifest_path, "data.yaml"), encoding="utf-8") as f:
            data = yaml.safe_load(f)

        assert "batches" in data
        assert len(data["batches"]) == 3


class TestManifestOrchestratorInit:
    """Tests for ManifestOrchestrator initialization."""

    def test_init_basic(self, tmp_path, mock_ffmistralsmall):
        """Test basic initialization."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        assert orchestrator.concurrency >= 1
        assert orchestrator.client == mock_ffmistralsmall

    def test_init_with_custom_concurrency(self, tmp_path, mock_ffmistralsmall):
        """Test initialization with custom concurrency."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
            concurrency=4,
        )

        assert orchestrator.concurrency == 4


class TestManifestOrchestratorLoad:
    """Tests for ManifestOrchestrator loading."""

    def test_load_manifest(self, temp_workbook_with_data, tmp_path, mock_ffmistralsmall):
        """Test loading manifest files."""
        from src.orchestrator.manifest import (
            ManifestOrchestrator,
            WorkbookManifestExporter,
        )

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_path,
            client=mock_ffmistralsmall,
        )
        orchestrator._load_manifest()

        assert len(orchestrator.prompts) == 3
        assert orchestrator.config["model"] == "mistral-small-2503"
        assert orchestrator.is_batch_mode is False

    def test_load_manifest_with_batch_data(
        self, temp_workbook_with_batch_data, tmp_path, mock_ffmistralsmall
    ):
        """Test loading manifest with batch data."""
        from src.orchestrator.manifest import (
            ManifestOrchestrator,
            WorkbookManifestExporter,
        )

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_batch_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_path,
            client=mock_ffmistralsmall,
        )
        orchestrator._load_manifest()

        assert orchestrator.is_batch_mode is True
        assert len(orchestrator.batch_data) == 3


class TestManifestOrchestratorOutput:
    """Tests for ManifestOrchestrator output generation."""

    def test_get_output_path_format(self, tmp_path, mock_ffmistralsmall):
        """Test output path format includes timestamp and manifest name."""
        import yaml

        from src.orchestrator.manifest import ManifestOrchestrator

        # Create manifest.yaml with name
        manifest_data = {"name": "My Prompts", "version": "1.0"}
        with open(tmp_path / "manifest.yaml", "w") as f:
            yaml.dump(manifest_data, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        output_path = orchestrator._get_output_path()

        # Name is sanitized: "My Prompts" -> "my_prompts"
        assert "my_prompts" in str(output_path)
        assert str(output_path).endswith(".parquet")

        filename = output_path.name
        # Timestamp format: YYYYMMDDHHMMSS (14 chars)
        assert len(filename) == 14 + len(".parquet")
        assert filename.startswith("20")

    def test_write_parquet_creates_file(self, tmp_path, mock_ffmistralsmall):
        """Test that _write_parquet creates a parquet file."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.source_workbook = "test_workbook.xlsx"

        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello",
                "history": None,
                "client": None,
                "condition": None,
                "condition_result": None,
                "condition_error": None,
                "response": "Response text",
                "status": "success",
                "attempts": 1,
                "error": None,
                "references": None,
                "semantic_query": None,
                "semantic_filter": None,
                "query_expansion": None,
                "rerank": None,
            }
        ]

        parquet_path = orchestrator._write_parquet(results)

        assert os.path.exists(parquet_path)

    def test_write_parquet_schema(self, tmp_path, mock_ffmistralsmall):
        """Test parquet file has expected schema."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.source_workbook = "test.xlsx"

        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello",
                "history": ["prev"],
                "client": "writer",
                "condition": "x > 5",
                "condition_result": "True",
                "condition_error": None,
                "response": "Response",
                "status": "success",
                "attempts": 1,
                "error": None,
                "references": ["doc1"],
                "semantic_query": "query",
                "semantic_filter": None,
                "query_expansion": "true",
                "rerank": None,
            }
        ]

        parquet_path = orchestrator._write_parquet(results)
        df = pl.read_parquet(parquet_path)

        expected_columns = [
            "sequence",
            "prompt_name",
            "prompt",
            "resolved_prompt",
            "history",
            "client",
            "condition",
            "condition_result",
            "condition_error",
            "response",
            "status",
            "attempts",
            "error",
            "references",
            "semantic_query",
            "semantic_filter",
            "query_expansion",
            "rerank",
        ]

        for col in expected_columns:
            assert col in df.columns

    def test_write_parquet_metadata(self, tmp_path, mock_ffmistralsmall):
        """Test parquet file includes manifest metadata."""
        import json

        import pyarrow.parquet as pq
        import yaml

        from src.orchestrator.manifest import ManifestOrchestrator

        # Create manifest.yaml with name and output_prompts
        manifest_data = {
            "name": "Test Manifest",
            "output_prompts": ["final_post", "summary"],
        }
        with open(tmp_path / "manifest.yaml", "w") as f:
            yaml.dump(manifest_data, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.source_workbook = "/path/to/test.xlsx"

        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello",
                "history": None,
                "client": None,
                "condition": None,
                "condition_result": None,
                "condition_error": None,
                "response": "Response",
                "status": "success",
                "attempts": 1,
                "error": None,
                "references": None,
                "semantic_query": None,
                "semantic_filter": None,
                "query_expansion": None,
                "rerank": None,
            }
        ]

        parquet_path = orchestrator._write_parquet(results)

        # Read metadata from parquet
        pf = pq.ParquetFile(parquet_path)
        metadata = pf.schema_arrow.metadata or {}

        assert b"manifest_name" in metadata
        assert metadata[b"manifest_name"] == b"test_manifest"
        assert b"output_prompts" in metadata
        assert json.loads(metadata[b"output_prompts"]) == ["final_post", "summary"]

    def test_write_parquet_with_batch_data(self, tmp_path, mock_ffmistralsmall):
        """Test parquet file includes batch columns when in batch mode."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.source_workbook = "batch_test.xlsx"

        results = [
            {
                "batch_id": 1,
                "batch_name": "batch_1",
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello",
                "history": None,
                "client": None,
                "condition": None,
                "condition_result": None,
                "condition_error": None,
                "response": "Response",
                "status": "success",
                "attempts": 1,
                "error": None,
                "references": None,
                "semantic_query": None,
                "semantic_filter": None,
                "query_expansion": None,
                "rerank": None,
            }
        ]

        parquet_path = orchestrator._write_parquet(results)
        df = pl.read_parquet(parquet_path)

        assert "batch_id" in df.columns
        assert "batch_name" in df.columns
        assert df["batch_id"][0] == 1
        assert df["batch_name"][0] == "batch_1"


class TestManifestOrchestratorSummary:
    """Tests for ManifestOrchestrator summary."""

    def test_get_summary_empty(self, tmp_path, mock_ffmistralsmall):
        """Test summary with no results."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        summary = orchestrator.get_summary()

        assert summary["status"] == "not_run"

    def test_get_summary_with_results(self, tmp_path, mock_ffmistralsmall):
        """Test summary with results."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.results = [
            {"status": "success", "attempts": 1},
            {"status": "success", "attempts": 1},
            {"status": "failed", "attempts": 3},
        ]

        summary = orchestrator.get_summary()

        assert summary["total_prompts"] == 3
        assert summary["successful"] == 2
        assert summary["failed"] == 1
        assert summary["total_attempts"] == 5


class TestManifestExporterAgentMode:
    """Tests for WorkbookManifestExporter with agent mode prompts."""

    def test_export_agent_mode_prompts(self, tmp_path):
        """Test export includes agent mode fields when present."""
        from openpyxl import Workbook

        from src.orchestrator.manifest import WorkbookManifestExporter

        wb = Workbook()
        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "model"
        ws_config["B2"] = "mistral-small-2503"

        ws_prompts = wb.create_sheet(title="prompts")
        headers = [
            "sequence",
            "prompt_name",
            "prompt",
            "history",
            "agent_mode",
            "tools",
            "max_tool_rounds",
            "validation_prompt",
            "max_validation_retries",
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws_prompts.cell(row=1, column=col_idx, value=h)

        ws_prompts.cell(row=2, column=1, value=1)
        ws_prompts.cell(row=2, column=2, value="research")
        ws_prompts.cell(row=2, column=3, value="Research topic")
        ws_prompts.cell(row=2, column=5, value=True)
        ws_prompts.cell(row=2, column=6, value='["calculate", "http_get"]')
        ws_prompts.cell(row=2, column=7, value=10)
        ws_prompts.cell(row=2, column=8, value="Check your answer")
        ws_prompts.cell(row=2, column=9, value=3)

        workbook_path = str(tmp_path / "agent_test.xlsx")
        wb.save(workbook_path)

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(workbook_path)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "prompts.yaml"), encoding="utf-8") as f:
            prompts_data = yaml.safe_load(f)

        prompt = prompts_data["prompts"][0]
        assert prompt["agent_mode"] is True
        assert prompt["tools"] == ["calculate", "http_get"]
        assert prompt["max_tool_rounds"] == 10
        assert prompt["validation_prompt"] == "Check your answer"
        assert prompt["max_validation_retries"] == 3

    def test_export_agent_mode_minimal_fields(self, tmp_path):
        """Test export with agent_mode true but no optional fields."""
        from openpyxl import Workbook

        from src.orchestrator.manifest import WorkbookManifestExporter

        wb = Workbook()
        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "model"
        ws_config["B2"] = "mistral-small-2503"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "agent_mode"
        ws_prompts["A2"] = 1
        ws_prompts["B2"] = "simple_agent"
        ws_prompts["C2"] = "Do something"
        ws_prompts["E2"] = True

        workbook_path = str(tmp_path / "agent_minimal.xlsx")
        wb.save(workbook_path)

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(workbook_path)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "prompts.yaml"), encoding="utf-8") as f:
            prompts_data = yaml.safe_load(f)

        prompt = prompts_data["prompts"][0]
        assert prompt["agent_mode"] is True
        assert prompt["tools"] == []
        assert "max_tool_rounds" not in prompt
        assert "validation_prompt" not in prompt
        assert "max_validation_retries" not in prompt


class TestManifestExporterWithTools:
    """Tests for WorkbookManifestExporter with tools sheet."""

    def test_export_with_tools_creates_tools_yaml(self, tmp_path):
        """Test export creates tools.yaml when tools sheet is present."""
        from openpyxl import Workbook

        from src.orchestrator.manifest import WorkbookManifestExporter

        wb = Workbook()
        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "model"
        ws_config["B2"] = "mistral-small-2503"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["A2"] = 1
        ws_prompts["B2"] = "test"
        ws_prompts["C2"] = "Hello"

        ws_tools = wb.create_sheet(title="tools")
        ws_tools["A1"] = "name"
        ws_tools["B1"] = "description"
        ws_tools["C1"] = "parameters"
        ws_tools["D1"] = "implementation"
        ws_tools["E1"] = "enabled"
        ws_tools["A2"] = "calculate"
        ws_tools["B2"] = "Evaluate math expressions"
        ws_tools["C2"] = "{}"
        ws_tools["D2"] = "builtin:calculate"
        ws_tools["E2"] = True

        workbook_path = str(tmp_path / "tools_test.xlsx")
        wb.save(workbook_path)

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(workbook_path)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "manifest.yaml"), encoding="utf-8") as f:
            manifest = yaml.safe_load(f)

        assert manifest["has_tools"] is True

        assert os.path.exists(os.path.join(manifest_path, "tools.yaml"))
        with open(os.path.join(manifest_path, "tools.yaml"), encoding="utf-8") as f:
            tools_data = yaml.safe_load(f)

        assert "tools" in tools_data
        assert len(tools_data["tools"]) == 1
        assert tools_data["tools"][0]["name"] == "calculate"


class TestManifestExporterDefaultManifestDir:
    """Tests for WorkbookManifestExporter using default manifest dir from config."""

    def test_export_uses_config_manifest_dir(self, tmp_path):
        """Test export uses config path when manifest_dir is None."""
        from unittest.mock import patch

        from src.orchestrator.manifest import WorkbookManifestExporter

        wb_path = str(tmp_path / "test.xlsx")
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"
        ws["A2"] = "model"
        ws["B2"] = "mistral-small-2503"
        ws_p = wb.create_sheet(title="prompts")
        ws_p["A1"] = "sequence"
        ws_p["B1"] = "prompt_name"
        ws_p["C1"] = "prompt"
        ws_p["D1"] = "history"
        ws_p["A2"] = 1
        ws_p["B2"] = "p1"
        ws_p["C2"] = "test"
        wb.save(wb_path)

        config_dir = str(tmp_path / "default_manifests")
        with patch("src.orchestrator.manifest.get_config") as mock_cfg:
            mock_cfg.return_value.paths.manifest_dir = config_dir
            exporter = WorkbookManifestExporter(wb_path)
            path = exporter.export()

        assert config_dir in path


class TestManifestOrchestratorProperties:
    """Tests for ManifestOrchestrator properties."""

    def test_source_path_property(self, tmp_path, mock_ffmistralsmall):
        """Test source_path returns manifest directory."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = str(tmp_path / "my_manifest")
        os.makedirs(manifest_dir)
        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_dir,
            client=mock_ffmistralsmall,
        )

        assert orchestrator.source_path == manifest_dir

    def test_manifest_meta_property(self, tmp_path, mock_ffmistralsmall):
        """Test manifest_meta getter and setter."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        assert orchestrator.manifest_meta == {}

        meta = {"name": "test", "version": "1.0"}
        orchestrator.manifest_meta = meta
        assert orchestrator.manifest_meta == meta

    def test_source_workbook_property(self, tmp_path, mock_ffmistralsmall):
        """Test source_workbook getter and setter."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        assert orchestrator.source_workbook == ""

        orchestrator.source_workbook = "/path/to/workbook.xlsx"
        assert orchestrator.source_workbook == "/path/to/workbook.xlsx"

    def test_get_cache_dir(self, tmp_path, mock_ffmistralsmall):
        """Test _get_cache_dir returns manifest_dir/doc_cache by default."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )

        cache_dir = orchestrator._get_cache_dir()
        assert str(tmp_path / "doc_cache") == cache_dir

    def test_get_cache_dir_from_config(self, tmp_path, mock_ffmistralsmall):
        """Test _get_cache_dir uses config override."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.config["document_cache_dir"] = "/custom/cache"

        cache_dir = orchestrator._get_cache_dir()
        assert cache_dir == "/custom/cache"


class TestManifestOrchestratorLoadWithExtras:
    """Tests for loading manifest with clients, documents, and tools."""

    def test_load_manifest_with_clients(self, tmp_path, mock_ffmistralsmall):
        """Test loading manifest loads client registry when has_clients is true."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest_with_clients"
        manifest_dir.mkdir()

        manifest_yaml = {"name": "test", "version": "1.0", "has_clients": True}
        with open(manifest_dir / "manifest.yaml", "w") as f:
            yaml.dump(manifest_yaml, f)

        config_yaml = {"model": "mistral-small-2503"}
        with open(manifest_dir / "config.yaml", "w") as f:
            yaml.dump(config_yaml, f)

        prompts_yaml = {"prompts": [{"sequence": 1, "prompt_name": "p1", "prompt": "test"}]}
        with open(manifest_dir / "prompts.yaml", "w") as f:
            yaml.dump(prompts_yaml, f)

        clients_yaml = {
            "clients": [
                {"name": "writer", "client_type": "mistral-small", "model": "mistral-small-2503"},
            ]
        }
        with open(manifest_dir / "clients.yaml", "w") as f:
            yaml.dump(clients_yaml, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )
        orchestrator._load_source()

        assert orchestrator.has_multi_client is True
        assert orchestrator.client_registry is not None

    def test_load_manifest_with_empty_clients(self, tmp_path, mock_ffmistralsmall):
        """Test loading manifest with has_clients but empty clients list."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest_empty_clients"
        manifest_dir.mkdir()

        manifest_yaml = {"name": "test", "version": "1.0", "has_clients": True}
        with open(manifest_dir / "manifest.yaml", "w") as f:
            yaml.dump(manifest_yaml, f)

        config_yaml = {"model": "mistral-small-2503"}
        with open(manifest_dir / "config.yaml", "w") as f:
            yaml.dump(config_yaml, f)

        prompts_yaml = {"prompts": [{"sequence": 1, "prompt_name": "p1", "prompt": "test"}]}
        with open(manifest_dir / "prompts.yaml", "w") as f:
            yaml.dump(prompts_yaml, f)

        clients_yaml = {"clients": []}
        with open(manifest_dir / "clients.yaml", "w") as f:
            yaml.dump(clients_yaml, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )
        orchestrator._load_source()

        assert orchestrator.has_multi_client is False

    def test_load_manifest_with_documents(self, tmp_path, mock_ffmistralsmall):
        """Test loading manifest with documents."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest_with_docs"
        manifest_dir.mkdir()

        docs_dir = tmp_path / "docs"
        docs_dir.mkdir()
        doc_file = docs_dir / "test.txt"
        doc_file.write_text("Hello world")

        manifest_yaml = {
            "name": "test",
            "version": "1.0",
            "has_documents": True,
            "source_workbook": str(tmp_path / "workbook.xlsx"),
        }
        with open(manifest_dir / "manifest.yaml", "w") as f:
            yaml.dump(manifest_yaml, f)

        config_yaml = {"model": "mistral-small-2503"}
        with open(manifest_dir / "config.yaml", "w") as f:
            yaml.dump(config_yaml, f)

        prompts_yaml = {"prompts": [{"sequence": 1, "prompt_name": "p1", "prompt": "test"}]}
        with open(manifest_dir / "prompts.yaml", "w") as f:
            yaml.dump(prompts_yaml, f)

        documents_yaml = {
            "documents": [
                {
                    "reference_name": "test_doc",
                    "common_name": "Test Doc",
                    "file_path": str(doc_file),
                }
            ]
        }
        with open(manifest_dir / "documents.yaml", "w") as f:
            yaml.dump(documents_yaml, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )
        orchestrator._load_source()

        assert orchestrator.has_documents is True
        assert orchestrator.document_registry is not None

    def test_load_manifest_with_tools(self, tmp_path, mock_ffmistralsmall):
        """Test loading manifest with tools."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest_with_tools"
        manifest_dir.mkdir()

        manifest_yaml = {"name": "test", "version": "1.0", "has_tools": True}
        with open(manifest_dir / "manifest.yaml", "w") as f:
            yaml.dump(manifest_yaml, f)

        config_yaml = {"model": "mistral-small-2503"}
        with open(manifest_dir / "config.yaml", "w") as f:
            yaml.dump(config_yaml, f)

        prompts_yaml = {"prompts": [{"sequence": 1, "prompt_name": "p1", "prompt": "test"}]}
        with open(manifest_dir / "prompts.yaml", "w") as f:
            yaml.dump(prompts_yaml, f)

        tools_yaml = {
            "tools": [
                {
                    "name": "calculate",
                    "description": "Math evaluator",
                    "parameters": {},
                    "implementation": "builtin:calculate",
                    "enabled": True,
                }
            ]
        }
        with open(manifest_dir / "tools.yaml", "w") as f:
            yaml.dump(tools_yaml, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )
        orchestrator._load_source()

        assert orchestrator.tool_registry is not None

    def test_init_documents_from_yaml(self, tmp_path, mock_ffmistralsmall):
        """Test _init_documents loads from YAML when data not provided."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest_docs_init"
        manifest_dir.mkdir()

        docs_dir = tmp_path / "docs2"
        docs_dir.mkdir()
        doc_file = docs_dir / "info.txt"
        doc_file.write_text("Some content")

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )
        orchestrator._manifest_meta = {"has_documents": True}

        documents_yaml = {
            "documents": [
                {
                    "reference_name": "info",
                    "common_name": "Info Doc",
                    "file_path": str(doc_file),
                }
            ]
        }
        with open(manifest_dir / "documents.yaml", "w") as f:
            yaml.dump(documents_yaml, f)

        orchestrator._init_documents()

        assert orchestrator.has_documents is True

    def test_init_documents_workbook_dir_from_source(self, tmp_path, mock_ffmistralsmall):
        """Test _init_documents uses source_workbook parent when no workbook_dir given."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest_wb_dir"
        manifest_dir.mkdir()

        docs_dir = tmp_path / "wb_parent" / "docs"
        docs_dir.mkdir(parents=True)
        doc_file = docs_dir / "doc.txt"
        doc_file.write_text("Content")

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )
        orchestrator._source_workbook = str(tmp_path / "wb_parent" / "workbook.xlsx")

        orchestrator._init_documents(
            documents_data=[
                {
                    "reference_name": "doc",
                    "common_name": "Doc",
                    "file_path": "docs/doc.txt",
                }
            ],
            workbook_dir=None,
        )

        assert orchestrator.has_documents is True

    def test_init_documents_explicit_workbook_dir(self, tmp_path, mock_ffmistralsmall):
        """Test _init_documents with explicit workbook_dir."""
        from src.orchestrator.manifest import ManifestOrchestrator

        docs_dir = tmp_path / "explicit_dir"
        docs_dir.mkdir()
        doc_file = docs_dir / "doc.txt"
        doc_file.write_text("Content")

        manifest_dir = tmp_path / "manifest_explicit"
        manifest_dir.mkdir()

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )
        orchestrator._source_workbook = "/some/other/path.xlsx"

        orchestrator._init_documents(
            documents_data=[
                {
                    "reference_name": "doc",
                    "common_name": "Doc",
                    "file_path": str(doc_file),
                }
            ],
            workbook_dir=str(docs_dir),
        )

        assert orchestrator.has_documents is True


class TestGetManifestName:
    """Tests for _get_manifest_name."""

    def test_get_manifest_name_from_manifest_yaml(self, tmp_path, mock_ffmistralsmall):
        """Test name is read and sanitized from manifest.yaml."""
        import yaml

        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest_my_prompts"
        manifest_dir.mkdir()
        manifest_yaml = {"name": "My-Custom Name", "version": "1.0"}
        with open(manifest_dir / "manifest.yaml", "w") as f:
            yaml.dump(manifest_yaml, f)

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )

        assert orchestrator._get_manifest_name() == "my_custom_name"

    def test_get_manifest_name_fallback_dir_no_prefix(self, tmp_path, mock_ffmistralsmall):
        """Test fallback to directory name without manifest_ prefix."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "custom_manifest_dir"
        manifest_dir.mkdir()

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )

        assert orchestrator._get_manifest_name() == "custom_manifest_dir"

    def test_get_manifest_name_fallback_dir_with_prefix(self, tmp_path, mock_ffmistralsmall):
        """Test fallback strips manifest_ prefix from directory name."""
        from src.orchestrator.manifest import ManifestOrchestrator

        manifest_dir = tmp_path / "manifest_my_workbook"
        manifest_dir.mkdir()

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(manifest_dir),
            client=mock_ffmistralsmall,
        )

        assert orchestrator._get_manifest_name() == "my_workbook"


class TestGetManifestMetadata:
    """Tests for get_manifest_metadata static method."""

    def test_get_manifest_metadata_extracts_fields(self, tmp_path, mock_ffmistralsmall):
        """Test static method extracts metadata from parquet file."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.source_workbook = "/path/to/src.xlsx"

        import yaml

        manifest_yaml = {"name": "Meta Test", "output_prompts": ["final_post"]}
        with open(tmp_path / "manifest.yaml", "w") as f:
            yaml.dump(manifest_yaml, f)

        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello",
                "history": None,
                "client": None,
                "condition": None,
                "condition_result": None,
                "condition_error": None,
                "response": "Response",
                "status": "success",
                "attempts": 1,
                "error": None,
                "references": None,
                "semantic_query": None,
                "semantic_filter": None,
                "query_expansion": None,
                "rerank": None,
            }
        ]

        parquet_path = orchestrator._write_parquet(results)
        metadata = ManifestOrchestrator.get_manifest_metadata(parquet_path)

        assert metadata["manifest_name"] == "meta_test"
        assert metadata["output_prompts"] == ["final_post"]
        assert metadata["source_workbook"] == "/path/to/src.xlsx"

    def test_get_manifest_metadata_empty_defaults(self, tmp_path, mock_ffmistralsmall):
        """Test static method returns defaults when metadata missing."""
        from src.orchestrator.manifest import ManifestOrchestrator

        orchestrator = ManifestOrchestrator(
            manifest_dir=str(tmp_path),
            client=mock_ffmistralsmall,
        )
        orchestrator.source_workbook = ""

        results = [
            {
                "sequence": 1,
                "prompt_name": "test",
                "prompt": "Hello",
                "history": None,
                "client": None,
                "condition": None,
                "condition_result": None,
                "condition_error": None,
                "response": "Response",
                "status": "success",
                "attempts": 1,
                "error": None,
                "references": None,
                "semantic_query": None,
                "semantic_filter": None,
                "query_expansion": None,
                "rerank": None,
            }
        ]

        parquet_path = orchestrator._write_parquet(results)
        metadata = ManifestOrchestrator.get_manifest_metadata(parquet_path)

        assert metadata["manifest_name"] != ""
        assert metadata["output_prompts"] == []
        assert metadata["source_workbook"] == ""


class TestManifestIntegration:
    """Integration tests for manifest workflow."""

    def test_full_workflow(self, temp_workbook_with_data, tmp_path, mock_ffmistralsmall):
        """Test full export -> run workflow without API calls."""
        from src.orchestrator.manifest import (
            ManifestOrchestrator,
            WorkbookManifestExporter,
        )

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(temp_workbook_with_data)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_path,
            client=mock_ffmistralsmall,
            concurrency=1,
        )
        orchestrator._load_manifest()
        orchestrator._validate()
        orchestrator._init_client()

        results = orchestrator.execute()

        assert len(results) == 3
        assert all(r["status"] == "success" for r in results)

    def test_export_notes_in_prompts_yaml(self, tmp_path):
        """Test that notes column is preserved in manifest export."""
        from openpyxl import Workbook

        from src.orchestrator.manifest import WorkbookManifestExporter

        wb = Workbook()
        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "model"
        ws_config["B2"] = "mistral-small-2503"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "notes"
        ws_prompts["A2"] = 1
        ws_prompts["B2"] = "hello"
        ws_prompts["C2"] = "Say hello"
        ws_prompts["E2"] = "This is a test note for the hello prompt"
        ws_prompts["A3"] = 2
        ws_prompts["B3"] = "goodbye"
        ws_prompts["C3"] = "Say goodbye"

        workbook_path = str(tmp_path / "notes_test.xlsx")
        wb.save(workbook_path)

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(workbook_path)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        with open(os.path.join(manifest_path, "prompts.yaml"), encoding="utf-8") as f:
            prompts_data = yaml.safe_load(f)

        prompts = prompts_data["prompts"]
        assert len(prompts) == 2
        assert prompts[0]["notes"] == "This is a test note for the hello prompt"
        assert prompts[1]["notes"] is None

    def test_export_notes_roundtrip_through_manifest_orchestrator(self, tmp_path):
        """Test that notes round-trips: workbook -> export -> load into ManifestOrchestrator."""
        from unittest.mock import MagicMock

        from openpyxl import Workbook

        from src.orchestrator.manifest import ManifestOrchestrator, WorkbookManifestExporter

        mock_client = MagicMock()

        wb = Workbook()
        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["A2"] = "model"
        ws_config["B2"] = "mistral-small-2503"

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "notes"
        ws_prompts["A2"] = 1
        ws_prompts["B2"] = "test_prompt"
        ws_prompts["C2"] = "Test"
        ws_prompts["E2"] = "Important note about this prompt"

        workbook_path = str(tmp_path / "notes_rt.xlsx")
        wb.save(workbook_path)

        manifest_dir = str(tmp_path / "manifests")
        exporter = WorkbookManifestExporter(workbook_path)
        manifest_path = exporter.export(manifest_dir=manifest_dir)

        orchestrator = ManifestOrchestrator(
            manifest_dir=manifest_path,
            client=mock_client,
        )
        orchestrator._load_source()

        assert len(orchestrator.prompts) == 1
        assert orchestrator.prompts[0]["notes"] == "Important note about this prompt"
