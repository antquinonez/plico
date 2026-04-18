# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Integration tests for multi-client workflows.

These tests verify that different clients can be used for different
prompts within the same workbook.
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from tests.integration.conftest import make_header_index


class TestMultiClientExecution:
    """Tests for multi-client execution."""

    def test_correct_client_per_prompt(self, integration_workbook_with_clients, spy_client):
        """
        Verify that prompt with client="fast" uses fast client config.
        """
        from unittest.mock import patch

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        def mock_create_client(registry_self, name):
            config = registry_self._client_configs[name]["config"]
            return spy_client.__class__(
                shared_calls=spy_client.calls,
                temperature=config.get("temperature", 0.3),
                max_tokens=config.get("max_tokens", 100),
            )

        with patch(
            "src.orchestrator.client_registry.ClientRegistry._create_client",
            mock_create_client,
        ):
            orchestrator = ExcelOrchestrator(
                integration_workbook_with_clients, spy_client, concurrency=1
            )

            results_sheet = orchestrator.run()

            assert len(spy_client.calls) == 3

            wb = load_workbook(integration_workbook_with_clients)
            ws = wb[results_sheet]
            h = make_header_index(ws)

            results = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[h["sequence"]] is not None:
                    results[row[h["prompt_name"]]] = {
                        "client": row[h["client"]],
                        "response": row[h["response"]],
                    }

            assert results["task1"]["client"] == "fast", (
                f"task1 should use 'fast' client, got {results['task1']['client']}"
            )
            assert results["task2"]["client"] is None, (
                f"task2 should use default (None), got {results['task2']['client']}"
            )
            assert results["task3"]["client"] == "detailed", (
                f"task3 should use 'detailed' client, got {results['task3']['client']}"
            )

    def test_fallback_to_default(self, integration_workbook_with_clients, spy_client):
        """
        Verify that prompt without 'client' uses default client.
        """
        from unittest.mock import patch

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        def mock_create_client(registry_self, name):
            config = registry_self._client_configs[name]["config"]
            return spy_client.__class__(
                shared_calls=spy_client.calls,
                temperature=config.get("temperature", 0.3),
                max_tokens=config.get("max_tokens", 100),
            )

        with patch(
            "src.orchestrator.client_registry.ClientRegistry._create_client",
            mock_create_client,
        ):
            orchestrator = ExcelOrchestrator(
                integration_workbook_with_clients, spy_client, concurrency=1
            )

            orchestrator.run()

            assert len(spy_client.calls) == 3

    def test_client_registry_clones_isolated(self, integration_workbook_with_clients):
        """
        Verify that named client clones don't share history.
        """
        from src.Clients.FFMistralSmall import FFMistralSmall
        from src.orchestrator.client_registry import ClientRegistry

        api_key = os.getenv("MISTRALSMALL_KEY")
        if not api_key:
            pytest.skip("MISTRALSMALL_KEY not set")

        default_client = FFMistralSmall(
            api_key=api_key, model="mistral-small-2503", temperature=0.7, max_tokens=100
        )

        registry = ClientRegistry(default_client)
        registry.register("test_client", "mistral-small", {"temperature": 0.3, "max_tokens": 50})

        clone1 = registry.clone("test_client")
        clone2 = registry.clone("test_client")

        clone1.generate_response("Hello")
        clone1.generate_response("World")

        assert len(clone1.conversation_history) == 4, (
            "clone1 should have 4 messages (2 user + 2 assistant)"
        )
        assert len(clone2.conversation_history) == 0, "clone2 should have empty history (isolated)"


class TestMultiClientRealAPI:
    """Tests using real API calls."""

    def test_real_api_multiclient_execution(
        self, integration_workbook_with_clients, real_mistral_client
    ):
        """
        Full multi-client execution test with real API.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_clients, real_mistral_client, concurrency=2
        )

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 3, f"Expected 3 successful, got {summary['successful']}"
        assert summary["failed"] == 0, f"Expected 0 failed, got {summary['failed']}"

        wb = load_workbook(integration_workbook_with_clients)
        ws = wb[results_sheet]
        h = make_header_index(ws)

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[h["sequence"]] is not None:
                results[row[h["prompt_name"]]] = {
                    "client": row[h["client"]],
                    "response": row[h["response"]],
                    "status": row[h["status"]],
                }

        for task_name, data in results.items():
            assert data["status"] == "success", f"{task_name} should succeed, got {data['status']}"
            assert data["response"] is not None, f"{task_name} should have response"

    def test_real_api_different_temperatures(self, tmp_path, real_mistral_client):
        """
        Test that different client configs produce expected behavior.

        Note: This test verifies that client configs are respected,
        but doesn't strictly verify temperature differences (LLM outputs vary).
        """

        from openpyxl import Workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        workbook_path = str(tmp_path / "temp_test.xlsx")

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        config_items = [
            ("model", "mistral-small-2503"),
            ("api_key_env", "MISTRALSMALL_KEY"),
            ("max_retries", 2),
            ("temperature", 0.7),
            ("max_tokens", 50),
            ("system_instructions", "You are a helpful assistant."),
        ]
        for idx, (field, value) in enumerate(config_items, start=2):
            ws_config[f"A{idx}"] = field
            ws_config[f"B{idx}"] = value

        ws_clients = wb.create_sheet(title="clients")
        ws_clients["A1"] = "name"
        ws_clients["B1"] = "client_type"
        ws_clients["C1"] = "temperature"
        ws_clients["D1"] = "max_tokens"
        ws_clients["A2"] = "low_temp"
        ws_clients["B2"] = "mistral-small"
        ws_clients["C2"] = 0.1
        ws_clients["D2"] = 20

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "client"

        ws_prompts["A2"] = 1
        ws_prompts["B2"] = "low"
        ws_prompts["C2"] = "What is 1+1? Just the number."
        ws_prompts["E2"] = "low_temp"

        wb.save(workbook_path)

        orchestrator = ExcelOrchestrator(workbook_path, real_mistral_client, concurrency=1)

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 1
        assert summary["failed"] == 0
