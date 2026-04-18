# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Integration tests for observability features.

Tests for execution plan preview, condition trace, scoring extraction trace,
and prompt preview using the ExcelOrchestrator with a spy client.
"""

import json
import os
import sys

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


@pytest.fixture
def explain_workbook(tmp_path):
    """Create a workbook with conditions for observability testing."""
    workbook_path = str(tmp_path / "explain_test.xlsx")

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", 2),
        ("temperature", 0.3),
        ("max_tokens", 100),
        (
            "system_instructions",
            "You are a helpful assistant. Give very brief answers.",
        ),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"
    ws_prompts["E1"] = "condition"

    wb.save(workbook_path)
    return workbook_path


@pytest.fixture
def explain_workbook_with_scoring(tmp_path):
    """Create a workbook with scoring sheet for extraction trace testing."""
    workbook_path = str(tmp_path / "explain_scoring_test.xlsx")

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", 2),
        ("temperature", 0.3),
        ("max_tokens", 100),
        (
            "system_instructions",
            "You are a helpful assistant. Respond with JSON.",
        ),
        ("batch_mode", "per_row"),
        ("batch_output", "combined"),
        ("on_batch_error", "continue"),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_scoring = wb.create_sheet(title="scoring")
    scoring_headers = [
        "criteria_name",
        "description",
        "scale_min",
        "scale_max",
        "weight",
        "source_prompt",
    ]
    for col_idx, header in enumerate(scoring_headers, start=1):
        ws_scoring.cell(row=1, column=col_idx, value=header)
    ws_scoring.cell(row=2, column=1, value="quality")
    ws_scoring.cell(row=2, column=2, value="Response quality")
    ws_scoring.cell(row=2, column=3, value=1)
    ws_scoring.cell(row=2, column=4, value=10)
    ws_scoring.cell(row=2, column=5, value=1.0)
    ws_scoring.cell(row=2, column=6, value="evaluate")

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    ws_prompts["A2"] = 10
    ws_prompts["B2"] = "evaluate"
    ws_prompts["C2"] = (
        'Rate the quality of this text on a scale of 1-10. Reply with JSON: {"quality": <score>}'
    )

    ws_data = wb.create_sheet(title="data")
    data_headers = ["id", "batch_name"]
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)
    ws_data.cell(row=2, column=1, value=1)
    ws_data.cell(row=2, column=2, value="batch_1")

    wb.save(workbook_path)
    return workbook_path


@pytest.fixture
def explain_workbook_with_batch(tmp_path):
    """Create a workbook with batch data for prompt preview testing."""
    workbook_path = str(tmp_path / "explain_batch_test.xlsx")

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", 2),
        ("temperature", 0.3),
        ("max_tokens", 50),
        (
            "system_instructions",
            "You are a helpful assistant. Give very brief answers.",
        ),
        ("batch_mode", "per_row"),
        ("batch_output", "combined"),
        ("on_batch_error", "continue"),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    ws_prompts["A2"] = 10
    ws_prompts["B2"] = "analyze"
    ws_prompts["C2"] = "Analyze {{region}} for {{product}}."

    ws_data = wb.create_sheet(title="data")
    data_headers = ["id", "batch_name", "region", "product"]
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)

    ws_data.cell(row=2, column=1, value=1)
    ws_data.cell(row=2, column=2, value="north_widget")
    ws_data.cell(row=2, column=3, value="north")
    ws_data.cell(row=2, column=4, value="widget_a")

    ws_data.cell(row=3, column=1, value=2)
    ws_data.cell(row=3, column=2, value="south_gadget")
    ws_data.cell(row=3, column=3, value="south")
    ws_data.cell(row=3, column=4, value="gadget_b")

    wb.save(workbook_path)
    return workbook_path


class StatusSpyClient:
    """Spy client that returns configurable responses by prompt name."""

    def __init__(self):
        self.calls = []
        self.conversation_history = []
        self.system_instructions = "Test"
        self.model = "test"
        self.temperature = 0.3
        self.max_tokens = 100
        self._responses = {
            "check": "success",
            "always_fails": "error: something went wrong",
        }

    def generate_response(self, prompt, **kwargs):
        self.calls.append({"prompt": prompt, "kwargs": kwargs})

        for key, resp in self._responses.items():
            if key in prompt.lower():
                return resp
        return "Default spy response"

    def clone(self):
        return self

    def clear_conversation(self):
        self.conversation_history = []

    def get_conversation_history(self):
        return self.conversation_history

    def set_conversation_history(self, history):
        self.conversation_history = history

    def test_connection(self):
        return True


class TestExecutionPlanPreview:
    """Tests for --explain execution plan preview."""

    def test_explain_no_api_calls(self, explain_workbook):
        """Verify --explain builds a plan without making any API calls."""
        from src.orchestrator.explain import build_explain_plan, format_explain
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(explain_workbook)
        prompts = parser.load_prompts()

        plan = build_explain_plan(prompts)
        output = format_explain(plan, title="explain_test.xlsx")

        assert "explain_test.xlsx" in output
        assert "No API calls made" in output or "Execution DAG" in output

    def test_explain_with_conditions(self, explain_workbook):
        """Verify --explain shows condition edges."""
        from openpyxl import load_workbook

        from src.orchestrator.explain import build_explain_plan, format_explain
        from src.orchestrator.workbook_parser import WorkbookParser

        wb = load_workbook(explain_workbook)
        ws = wb["prompts"]
        ws["A2"] = 10
        ws["B2"] = "check"
        ws["C2"] = "Check prompt"
        ws["A3"] = 20
        ws["B3"] = "proceed"
        ws["C3"] = "Proceed prompt"
        ws["E3"] = '{{check.status}} == "success"'
        wb.save(explain_workbook)

        parser = WorkbookParser(explain_workbook)
        prompts = parser.load_prompts()

        plan = build_explain_plan(prompts)
        output = format_explain(plan)

        assert "check" in output
        assert "proceed" in output
        assert "[condition]" in output
        assert "⚠" in output

    def test_explain_with_history(self, explain_workbook):
        """Verify --explain shows history edges."""
        from openpyxl import load_workbook

        from src.orchestrator.explain import build_explain_plan, format_explain
        from src.orchestrator.workbook_parser import WorkbookParser

        wb = load_workbook(explain_workbook)
        ws = wb["prompts"]
        ws["A2"] = 10
        ws["B2"] = "analyze"
        ws["C2"] = "Analyze the data"
        ws["A3"] = 20
        ws["B3"] = "summarize"
        ws["C3"] = "Summarize the analysis"
        ws["D3"] = json.dumps(["analyze"])
        wb.save(explain_workbook)

        parser = WorkbookParser(explain_workbook)
        prompts = parser.load_prompts()

        plan = build_explain_plan(prompts)
        output = format_explain(plan)

        assert "hist: analyze" in output
        assert "[history]" in output

    def test_explain_with_batch(self, explain_workbook_with_batch):
        """Verify --explain shows batch info."""
        from src.orchestrator.explain import build_explain_plan, format_explain
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(explain_workbook_with_batch)
        prompts = parser.load_prompts()
        batch_data = parser.load_data()

        plan = build_explain_plan(prompts, batch_data=batch_data)
        output = format_explain(plan)

        assert "Batch rows: 2" in output
        assert "2 LLM calls" in output


class TestConditionTraceIntegration:
    """Integration tests for condition trace in results."""

    def test_condition_trace_recorded_on_success(self, explain_workbook):
        """Verify condition_trace is populated when condition is true."""
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        client = StatusSpyClient()

        wb = load_workbook(explain_workbook)
        ws = wb["prompts"]
        ws["A2"] = 10
        ws["B2"] = "check"
        ws["C2"] = "Please check status"
        ws["A3"] = 20
        ws["B3"] = "proceed"
        ws["C3"] = "Proceeding"
        ws["E3"] = '{{check.status}} == "success"'
        wb.save(explain_workbook)

        orchestrator = ExcelOrchestrator(explain_workbook, client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(explain_workbook)
        ws = wb[results_sheet]
        headers = [cell.value for cell in ws[1]]

        condition_trace_idx = headers.index("condition_trace")
        prompt_name_idx = headers.index("prompt_name")

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[prompt_name_idx] is not None:
                results[row[prompt_name_idx]] = {
                    "condition_trace": row[condition_trace_idx],
                }

        assert results["proceed"]["condition_trace"] is not None, (
            "condition_trace should be populated for conditional prompts"
        )
        assert "success" in results["proceed"]["condition_trace"], (
            "condition_trace should contain the resolved value"
        )

    def test_condition_trace_recorded_on_skip(self, explain_workbook):
        """Verify condition_trace is populated when condition is false."""
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        client = StatusSpyClient()

        wb = load_workbook(explain_workbook)
        ws = wb["prompts"]
        ws["A2"] = 10
        ws["B2"] = "check"
        ws["C2"] = "Please check status"
        ws["A3"] = 20
        ws["B3"] = "never_runs"
        ws["C3"] = "Should not execute"
        ws["E3"] = '{{check.status}} == "failed"'
        wb.save(explain_workbook)

        orchestrator = ExcelOrchestrator(explain_workbook, client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(explain_workbook)
        ws = wb[results_sheet]
        headers = [cell.value for cell in ws[1]]

        condition_trace_idx = headers.index("condition_trace")
        prompt_name_idx = headers.index("prompt_name")
        status_idx = headers.index("status")

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[prompt_name_idx] is not None:
                results[row[prompt_name_idx]] = {
                    "condition_trace": row[condition_trace_idx],
                    "status": row[status_idx],
                }

        assert results["never_runs"]["status"] == "skipped"
        assert results["never_runs"]["condition_trace"] is not None, (
            "condition_trace should be populated even when skipped"
        )

    def test_condition_trace_none_without_condition(self, explain_workbook):
        """Verify condition_trace is None for prompts without conditions."""
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        client = StatusSpyClient()

        wb = load_workbook(explain_workbook)
        ws = wb["prompts"]
        ws["A2"] = 10
        ws["B2"] = "simple"
        ws["C2"] = "Just a prompt"
        wb.save(explain_workbook)

        orchestrator = ExcelOrchestrator(explain_workbook, client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(explain_workbook)
        ws = wb[results_sheet]
        headers = [cell.value for cell in ws[1]]

        condition_trace_idx = headers.index("condition_trace")
        prompt_name_idx = headers.index("prompt_name")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[prompt_name_idx] == "simple":
                assert row[condition_trace_idx] is None, (
                    "condition_trace should be None for prompts without conditions"
                )


class TestScoringExtractionTraceIntegration:
    """Integration tests for scoring extraction trace in results."""

    def test_extraction_trace_with_json_response(self, explain_workbook_with_scoring):
        """Verify extraction_trace is populated when scores are extracted."""
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        class JSONSpyClient:
            def __init__(self):
                self.calls = []
                self.conversation_history = []
                self.system_instructions = "Test"
                self.model = "test"
                self.temperature = 0.3
                self.max_tokens = 100

            def generate_response(self, prompt, **kwargs):
                self.calls.append({"prompt": prompt})
                return json.dumps({"quality": 8})

            def clone(self):
                return self

            def clear_conversation(self):
                pass

        client = JSONSpyClient()

        orchestrator = ExcelOrchestrator(explain_workbook_with_scoring, client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(explain_workbook_with_scoring)
        ws = wb[results_sheet]
        headers = [cell.value for cell in ws[1]]

        extraction_trace_idx = headers.index("extraction_trace")
        prompt_name_idx = headers.index("prompt_name")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[prompt_name_idx] == "evaluate":
                trace = row[extraction_trace_idx]
                assert trace is not None, (
                    "extraction_trace should be populated when scores are extracted"
                )
                assert "quality" in str(trace), (
                    "extraction_trace should reference the criteria name"
                )


class TestPromptPreviewIntegration:
    """Integration tests for prompt preview with batch variables."""

    def test_prompt_preview_shows_resolved_variables(self, explain_workbook_with_batch):
        """Verify prompt preview resolves batch variables."""
        from src.orchestrator.explain import format_prompt_preview
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(explain_workbook_with_batch)
        prompts = parser.load_prompts()

        evaluate_prompt = next(p for p in prompts if p["prompt_name"] == "analyze")
        batch_data = parser.load_data()

        output = format_prompt_preview(
            evaluate_prompt,
            batch_row=batch_data[0],
        )

        assert "Template Variables" in output
        assert "{{region}}  →  north" in output
        assert "{{product}}  →  widget_a" in output
        assert "Analyze north for widget_a." in output

    def test_prompt_preview_no_api_call(self, explain_workbook_with_batch):
        """Verify prompt preview explicitly states no API call was made."""
        from src.orchestrator.explain import format_prompt_preview
        from src.orchestrator.workbook_parser import WorkbookParser

        parser = WorkbookParser(explain_workbook_with_batch)
        prompts = parser.load_prompts()

        evaluate_prompt = next(p for p in prompts if p["prompt_name"] == "analyze")

        output = format_prompt_preview(evaluate_prompt)

        assert "No API call made" in output

    def test_prompt_preview_with_upstream_results(self, explain_workbook):
        """Verify prompt preview shows upstream response references."""
        from openpyxl import load_workbook

        from src.orchestrator.explain import format_prompt_preview
        from src.orchestrator.workbook_parser import WorkbookParser

        wb = load_workbook(explain_workbook)
        ws = wb["prompts"]
        ws["A2"] = 10
        ws["B2"] = "fetch"
        ws["C2"] = "Fetch data"
        ws["A3"] = 20
        ws["B3"] = "process"
        ws["C3"] = "Process {{fetch.response}}"
        ws["D3"] = json.dumps(["fetch"])
        wb.save(explain_workbook)

        parser = WorkbookParser(explain_workbook)
        prompts = parser.load_prompts()

        process_prompt = next(p for p in prompts if p["prompt_name"] == "process")

        upstream = {
            "fetch": {
                "response": "Data loaded successfully with 42 records",
                "status": "success",
            }
        }

        output = format_prompt_preview(
            process_prompt,
            upstream_results=upstream,
        )

        assert "Upstream References" in output
        assert "Data loaded successfully" in output
        assert "History Context" in output
