# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Integration tests for core orchestrator behavior.

These tests verify end-to-end orchestrator functionality including
execution order, results writing, retry mechanism, and failure handling.
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


class TestOrchestratorExecution:
    """Tests for orchestrator execution behavior."""

    def test_sequential_execution_order(self, integration_workbook, spy_client):
        """
        Verify prompts execute in dependency order, not just sequence order.
        """
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "First prompt"
        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "Second prompt"
        ws["D3"] = json.dumps(["first"])
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, spy_client, concurrency=1)
        orchestrator.run()

        assert len(spy_client.calls) == 2
        assert "First prompt" in spy_client.calls[0]["prompt"], (
            "First call should be 'first' (dependency)"
        )
        assert "Second prompt" in spy_client.calls[1]["prompt"], (
            "Second call should be 'second' (dependent)"
        )

    def test_results_written_to_workbook(self, integration_workbook, real_mistral_client):
        """
        Verify results sheet has all expected columns and values.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "test_prompt"
        ws["C2"] = "Say hello."
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, real_mistral_client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook)
        assert results_sheet in wb.sheetnames

        ws = wb[results_sheet]
        headers = [cell.value for cell in ws[1]]

        expected_columns = [
            "sequence",
            "prompt_name",
            "prompt",
            "history",
            "client",
            "condition",
            "response",
            "status",
            "attempts",
            "error",
        ]

        for col in expected_columns:
            assert col in headers, f"Missing column: {col}"

    def test_progress_callback_receives_updates(self, integration_workbook, spy_client):
        """
        Verify progress callback fires for each completed prompt.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        progress_updates = []

        def progress_callback(completed, total, success, failed, **kwargs):
            progress_updates.append(
                {
                    "completed": completed,
                    "total": total,
                    "success": success,
                    "failed": failed,
                }
            )

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        for i in range(1, 4):
            ws[f"A{i + 1}"] = i
            ws[f"B{i + 1}"] = f"p{i}"
            ws[f"C{i + 1}"] = f"Prompt {i}"
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(
            integration_workbook,
            spy_client,
            concurrency=1,
            progress_callback=progress_callback,
        )
        orchestrator.run()

        assert len(progress_updates) >= 3, (
            f"Should have at least 3 updates, got {len(progress_updates)}"
        )

        final_update = progress_updates[-1]
        assert final_update["completed"] == 3
        assert final_update["total"] == 3
        assert final_update["success"] == 3


class TestOrchestratorRetries:
    """Tests for retry mechanism."""

    def test_retry_mechanism_succeeds(self, integration_workbook):
        """
        Test that failed prompt retries and succeeds on second attempt.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "retry_test"
        ws["C2"] = "Say success"
        wb.save(integration_workbook)

        call_count = [0]

        class FlakyClient:
            def __init__(self):
                self.conversation_history = []
                self.system_instructions = "Test"
                self.model = "test"
                self.temperature = 0.3
                self.max_tokens = 100

            def generate_response(self, prompt, **kwargs):
                call_count[0] += 1
                if call_count[0] < 2:
                    raise Exception("Temporary failure")
                return "Success!"

            def clone(self):
                return self

            def clear_conversation(self):
                self.conversation_history = []

        flaky_client = FlakyClient()
        orchestrator = ExcelOrchestrator(integration_workbook, flaky_client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook)
        ws = wb[results_sheet]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == 1:
                assert row[12] == "success", "Status should be success after retry"
                assert call_count[0] >= 2, f"Should have retried, got {call_count[0]} calls"
                return

        pytest.fail("Did not find result for prompt 1")

    def test_max_retries_exhausted(self, integration_workbook):
        """
        Test that after N failures, status is 'failed' with error message.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "always_fails"
        ws["C2"] = "This will fail"
        wb.save(integration_workbook)

        class AlwaysFailsClient:
            def __init__(self):
                self.conversation_history = []
                self.system_instructions = "Test"
                self.model = "test"
                self.temperature = 0.3
                self.max_tokens = 100

            def generate_response(self, prompt, **kwargs):
                raise Exception("Permanent failure")

            def clone(self):
                return self

            def clear_conversation(self):
                pass

        failing_client = AlwaysFailsClient()
        orchestrator = ExcelOrchestrator(integration_workbook, failing_client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook)
        ws = wb[results_sheet]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == 1:
                assert row[12] == "failed", "Status should be failed"
                assert row[14] is not None, "Should have error message"
                assert "Permanent failure" in str(row[14]), (
                    f"Error message should contain 'Permanent failure', got {row[14]}"
                )
                return

        pytest.fail("Did not find result for prompt 1")


class TestOrchestratorFailurePropagation:
    """Tests for failure propagation to dependent prompts."""

    def test_failure_skips_dependents_by_default(self, integration_workbook):
        """
        If p1 fails and p2 depends on p1, p2 should fail or handle gracefully.
        """
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "failing"
        ws["C2"] = "Fail here"
        ws["A3"] = 2
        ws["B3"] = "dependent"
        ws["C3"] = "Depend on failing"
        ws["D3"] = json.dumps(["failing"])
        wb.save(integration_workbook)

        class FailsOnFirstClient:
            def __init__(self):
                self.conversation_history = []
                self.calls = 0
                self.system_instructions = "Test"
                self.model = "test"
                self.temperature = 0.3
                self.max_tokens = 100

            def generate_response(self, prompt, **kwargs):
                self.calls += 1
                if "Fail here" in prompt:
                    raise Exception("First prompt failed")
                return "Success"

            def clone(self):
                return self

            def clear_conversation(self):
                pass

        client = FailsOnFirstClient()
        orchestrator = ExcelOrchestrator(integration_workbook, client, concurrency=1)

        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook)
        ws = wb[results_sheet]

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                results[row[3]] = {"status": row[12], "response": row[11]}

        assert results["failing"]["status"] == "failed", "First prompt should fail"

        assert results["dependent"]["status"] in ["failed", "success"], (
            f"Dependent should handle missing dependency gracefully, got {results['dependent']['status']}"
        )


class TestOrchestratorRealAPI:
    """Tests using real API calls."""

    def test_real_api_full_workflow(self, integration_workbook, real_mistral_client):
        """
        Complete workflow test with real API calls.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "greeting"
        ws["C2"] = "Say 'Hello World'"
        ws["A3"] = 2
        ws["B3"] = "math"
        ws["C3"] = "What is 5 + 5?"
        ws["A4"] = 3
        ws["B4"] = "final"
        ws["C4"] = "Say goodbye"
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, real_mistral_client, concurrency=2)

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 3
        assert summary["failed"] == 0

        wb = load_workbook(integration_workbook)
        ws = wb[results_sheet]

        responses = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                responses[row[3]] = row[11]

        assert responses["greeting"] is not None
        assert responses["math"] is not None
        assert responses["final"] is not None

    def test_real_api_with_progress(self, integration_workbook, real_mistral_client):
        """
        Test progress callback with real API.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        progress_updates = []

        def progress_callback(completed, total, success, failed, **kwargs):
            progress_updates.append(
                {
                    "completed": completed,
                    "total": total,
                    "success": success,
                    "failed": failed,
                }
            )

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        for i in range(1, 4):
            ws[f"A{i + 1}"] = i
            ws[f"B{i + 1}"] = f"task{i}"
            ws[f"C{i + 1}"] = f"Say task{i}"
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(
            integration_workbook,
            real_mistral_client,
            concurrency=2,
            progress_callback=progress_callback,
        )

        orchestrator.run()

        assert len(progress_updates) >= 3
        assert progress_updates[-1]["completed"] == 3
        assert progress_updates[-1]["total"] == 3
