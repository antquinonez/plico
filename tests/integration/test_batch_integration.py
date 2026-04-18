# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Integration tests for batch mode.

These tests verify batch execution including variable resolution,
batch isolation, and combined results output.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from tests.integration.conftest import make_header_index


class TestBatchVariableResolution:
    """Tests for variable resolution in batch mode."""

    def test_variable_resolution_basic(self, integration_workbook_with_batch_data, spy_client):
        """
        Verify that {{region}} is replaced with actual value from data sheet.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, spy_client, concurrency=1
        )

        orchestrator.run()

        assert len(spy_client.calls) == 3

        prompts = [call["prompt"] for call in spy_client.calls]

        assert any("north" in p for p in prompts), "Should have 'north' region"
        assert any("south" in p for p in prompts), "Should have 'south' region"
        assert any("east" in p for p in prompts), "Should have 'east' region"

        for prompt in prompts:
            assert "{{region}}" not in prompt, f"Variable should be resolved, but got: {prompt}"

    def test_batch_name_template(self, integration_workbook_with_batch_data, spy_client):
        """
        Verify that {{region}}_{{product}} generates correct batch names.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, spy_client, concurrency=1
        )

        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook_with_batch_data)
        ws = wb[results_sheet]
        h = make_header_index(ws)

        batch_names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[h["batch_id"]] is not None:
                batch_names.add(row[h["batch_name"]])

        assert "north_widget_a" in batch_names, f"Should have 'north_widget_a', got {batch_names}"
        assert "south_widget_b" in batch_names, f"Should have 'south_widget_b', got {batch_names}"
        assert "east_widget_c" in batch_names, f"Should have 'east_widget_c', got {batch_names}"

    def test_all_batches_execute(self, integration_workbook_with_batch_data, spy_client):
        """
        Verify that 5 data rows = 5 batch executions.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook_with_batch_data)
        ws = wb["data"]
        ws.cell(row=5, column=1, value=4)
        ws.cell(row=5, column=2, value="{{region}}_{{product}}")
        ws.cell(row=5, column=3, value="west")
        ws.cell(row=5, column=4, value="widget_d")
        wb.save(integration_workbook_with_batch_data)

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, spy_client, concurrency=1
        )

        orchestrator.run()

        assert len(spy_client.calls) == 4, f"Should have 4 batch calls, got {len(spy_client.calls)}"


class TestBatchIsolation:
    """Tests for batch isolation."""

    def test_batch_isolation(self, integration_workbook_with_batch_data, spy_client):
        """
        Verify that batch 2's variables don't leak into batch 3.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, spy_client, concurrency=1
        )

        orchestrator.run()

        prompts = [call["prompt"] for call in spy_client.calls]

        north_idx = next(i for i, p in enumerate(prompts) if "north" in p)
        south_idx = next(i for i, p in enumerate(prompts) if "south" in p)
        east_idx = next(i for i, p in enumerate(prompts) if "east" in p)

        assert "south" not in prompts[north_idx], "North batch should not contain 'south'"
        assert "east" not in prompts[south_idx], "South batch should not contain 'east'"
        assert "north" not in prompts[east_idx], "East batch should not contain 'north'"

    def test_client_history_cleared_between_batches(
        self, integration_workbook_with_batch_data, spy_client
    ):
        """
        Verify client history is cleared between batch iterations.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, spy_client, concurrency=1
        )

        orchestrator.run()

        for i, call in enumerate(spy_client.calls):
            assert call["history_length"] == 0, (
                f"Batch call {i + 1} should have empty history, got {call['history_length']}"
            )


class TestBatchResults:
    """Tests for batch results output."""

    def test_combined_results_sheet(self, integration_workbook_with_batch_data, spy_client):
        """
        Verify all batch results in single sheet with batch_name column.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, spy_client, concurrency=1
        )

        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook_with_batch_data)
        ws = wb[results_sheet]
        h = make_header_index(ws)

        assert "batch_name" in h, "Should have batch_name column"

        results_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[h["batch_id"]] is not None:
                results_count += 1
                assert row[h["batch_name"]] is not None, "batch_name should be populated"

        assert results_count == 3, f"Should have 3 results, got {results_count}"

    def test_on_batch_error_continue(self, integration_workbook_with_batch_data):
        """
        Verify that one batch failure doesn't stop others.
        """

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        class FailsOnSouthClient:
            def __init__(self):
                self.conversation_history = []
                self.system_instructions = "Test"
                self.model = "test"
                self.temperature = 0.3
                self.max_tokens = 100

            def generate_response(self, prompt, **kwargs):
                if "south" in prompt:
                    raise Exception("South batch failed")
                return "Success"

            def clone(self):
                return self

            def clear_conversation(self):
                pass

        client = FailsOnSouthClient()
        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, client, concurrency=1
        )

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] >= 2, (
            f"Should have at least 2 successful (north, east), got {summary['successful']}"
        )
        assert summary["failed"] >= 1, (
            f"Should have at least 1 failed (south), got {summary['failed']}"
        )


class TestBatchRealAPI:
    """Tests using real API calls."""

    def test_real_api_batch_execution(
        self, integration_workbook_with_batch_data, real_mistral_client
    ):
        """
        Full batch execution test with real API.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, real_mistral_client, concurrency=2
        )

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 3, f"Should have 3 successful, got {summary['successful']}"
        assert summary["failed"] == 0, f"Should have 0 failed, got {summary['failed']}"

        wb = load_workbook(integration_workbook_with_batch_data)
        ws = wb[results_sheet]
        h = make_header_index(ws)

        batch_results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[h["batch_id"]] is not None:
                batch_name = row[h["batch_name"]]
                response = row[h["response"]]
                batch_results[batch_name] = response

        assert "north_widget_a" in batch_results
        assert "south_widget_b" in batch_results
        assert "east_widget_c" in batch_results

        assert "n" in str(batch_results["north_widget_a"]).lower(), (
            f"North batch should answer with 'n', got {batch_results['north_widget_a']}"
        )
        assert "s" in str(batch_results["south_widget_b"]).lower(), (
            f"South batch should answer with 's', got {batch_results['south_widget_b']}"
        )
        assert "e" in str(batch_results["east_widget_c"]).lower(), (
            f"East batch should answer with 'e', got {batch_results['east_widget_c']}"
        )
