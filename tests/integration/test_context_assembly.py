# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Integration tests for declarative context assembly.

These tests verify that FFAI correctly assembles context from named
prompt references and injects it into the prompt string.
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


class TestContextAssembly:
    """Tests for declarative context assembly."""

    def test_single_history_reference(self, integration_workbook, spy_client):
        """
        Verify that history=["p1"] injects p1's Q&A into the prompt.
        """
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "What is 2 + 2?"
        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "What was my previous question about?"
        ws["D3"] = json.dumps(["first"])
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, spy_client, concurrency=1)
        orchestrator.run()

        assert len(spy_client.calls) == 2

        second_prompt = spy_client.calls[1]["prompt"]

        assert "first" in second_prompt or "2 + 2" in second_prompt, (
            "Second prompt should reference first prompt"
        )
        assert "<conversation_history>" in second_prompt, "Should use conversation_history format"
        assert "<interaction" in second_prompt, "Should use interaction format"

    def test_multiple_history_references(self, integration_workbook, spy_client):
        """
        Verify that history=["p1", "p2"] injects both in order.
        """
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "alpha"
        ws["C2"] = "Say the word ALPHA."
        ws["A3"] = 2
        ws["B3"] = "beta"
        ws["C3"] = "Say the word BETA."
        ws["A4"] = 3
        ws["B4"] = "combined"
        ws["C4"] = "What words did I ask you to say?"
        ws["D4"] = json.dumps(["alpha", "beta"])
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, spy_client, concurrency=1)
        orchestrator.run()

        assert len(spy_client.calls) == 3

        third_prompt = spy_client.calls[2]["prompt"]

        assert "alpha" in third_prompt, "Should include alpha context"
        assert "beta" in third_prompt, "Should include beta context"
        assert third_prompt.find("alpha") < third_prompt.find("beta"), (
            "alpha should appear before beta (order preserved)"
        )

    def test_nested_dependencies(self, integration_workbook, tmp_path, spy_client):
        """
        Test that p3 depends on p2 which depends on p1 - all context available.
        """
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        workbook_path = str(tmp_path / "nested_deps.xlsx")

        wb = load_workbook(integration_workbook)
        wb["config"]["B7"] = "You are a helpful assistant."

        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "level0"
        ws["C2"] = "Remember the number 5."
        ws["A3"] = 2
        ws["B3"] = "level1"
        ws["C3"] = "What number should I remember? Just give the number."
        ws["D3"] = json.dumps(["level0"])
        ws["A4"] = 3
        ws["B4"] = "level2"
        ws["C4"] = "Add 10 to the number from the conversation. Just give the result."
        ws["D4"] = json.dumps(["level1"])
        wb.save(workbook_path)

        orchestrator = ExcelOrchestrator(workbook_path, spy_client, concurrency=1)
        orchestrator.run()

        assert len(spy_client.calls) == 3

        level2_prompt = spy_client.calls[2]["prompt"]

        assert "level1" in level2_prompt or "5" in level2_prompt, (
            "level2 should have context from level1 (which has context from level0)"
        )

    def test_history_only_includes_named_prompts(self, integration_workbook, spy_client):
        """
        Verify that unreferenced prompts are not included in context.
        """
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "included"
        ws["C2"] = "Say INCLUDE."
        ws["A3"] = 2
        ws["B3"] = "excluded"
        ws["C3"] = "Say EXCLUDE."
        ws["A4"] = 3
        ws["B4"] = "target"
        ws["C4"] = "What did I ask?"
        ws["D4"] = json.dumps(["included"])
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, spy_client, concurrency=1)
        orchestrator.run()

        target_prompt = spy_client.calls[2]["prompt"]

        assert "included" in target_prompt or "INCLUDE" in target_prompt, (
            "Should include 'included' prompt"
        )
        assert "excluded" not in target_prompt and "EXCLUDE" not in target_prompt, (
            "Should NOT include 'excluded' prompt"
        )

    def test_context_format_correct(self, integration_workbook, spy_client):
        """
        Verify output matches <conversation_history>...</conversation_history> format.
        """
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "q1"
        ws["C2"] = "Hello"
        ws["A3"] = 2
        ws["B3"] = "q2"
        ws["C3"] = "Reply"
        ws["D3"] = json.dumps(["q1"])
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, spy_client, concurrency=1)
        orchestrator.run()

        second_prompt = spy_client.calls[1]["prompt"]

        assert "<conversation_history>" in second_prompt, "Should start with <conversation_history>"
        assert "</conversation_history>" in second_prompt, "Should end with </conversation_history>"
        assert "<interaction" in second_prompt, "Should contain <interaction"
        assert "</interaction>" in second_prompt, "Should contain </interaction>"
        assert "USER:" in second_prompt, "Should contain USER:"
        assert "SYSTEM:" in second_prompt, "Should contain SYSTEM:"
        assert "===" in second_prompt, "Should contain separator ==="
        assert "Based on the conversation history above" in second_prompt, (
            "Should contain instruction text"
        )

    def test_missing_history_reference_fails(self, integration_workbook, spy_client):
        """
        Verify that referencing non-existent prompt_name raises error.
        """
        import json

        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "q1"
        ws["C2"] = "Hello"
        ws["A3"] = 2
        ws["B3"] = "q2"
        ws["C3"] = "Reply"
        ws["D3"] = json.dumps(["nonexistent"])
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, spy_client, concurrency=1)

        with pytest.raises(ValueError, match="nonexistent"):
            orchestrator.run()


class TestContextAssemblyRealAPI:
    """Tests using real API calls to verify context assembly."""

    def test_real_api_context_carryover(
        self, integration_workbook_with_dependencies, real_mistral_client
    ):
        """
        Test with real API that context is properly carried over.

        step3 asks to add results from step1 and step2.
        The AI should be able to answer correctly because the context
        from step1 (4) and step2 (6) is injected into the prompt.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_dependencies, real_mistral_client, concurrency=1
        )

        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook_with_dependencies)
        ws = wb[results_sheet]

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                results[row[3]] = row[11]

        assert (
            "4" in str(results.get("step1", "")) or "four" in str(results.get("step1", "")).lower()
        ), f"step1 should answer 4, got {results.get('step1')}"

        assert (
            "6" in str(results.get("step2", "")) or "six" in str(results.get("step2", "")).lower()
        ), f"step2 should answer 6, got {results.get('step2')}"

        step3_response = str(results.get("step3", ""))
        assert "10" in step3_response or "ten" in step3_response.lower(), (
            f"step3 should answer 10 (4+6), got {step3_response}"
        )

    def test_real_api_no_context_without_history(self, integration_workbook, real_mistral_client):
        """
        Verify that prompts without history references don't receive context.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "secret"
        ws["C2"] = "The secret word is BANANA."
        ws["A3"] = 2
        ws["B3"] = "unrelated"
        ws["C3"] = "What is 1 + 1? Just give the number."
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, real_mistral_client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook)
        ws = wb[results_sheet]

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                results[row[3]] = row[11]

        unrelated_response = str(results.get("unrelated", ""))
        assert "banana" not in unrelated_response.lower(), (
            f"Unrelated prompt should not know about banana, got {unrelated_response}"
        )
        assert "2" in unrelated_response or "two" in unrelated_response.lower(), (
            f"Unrelated prompt should answer 2, got {unrelated_response}"
        )
