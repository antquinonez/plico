# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Integration tests for client history isolation.

These tests verify that client conversation history is properly isolated
between prompt executions, ensuring no double-context or memory leaks.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


class TestClientHistoryIsolation:
    """Tests for client history isolation between prompts."""

    def test_client_history_cleared_between_prompts(self, integration_workbook, spy_client):
        """
        Verify that client conversation_history is empty before each prompt.

        The orchestrator should clone the client for each prompt execution,
        ensuring history from previous prompts doesn't accumulate.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "What is 1 + 1?"
        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "What is 2 + 2?"
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, spy_client, concurrency=1)
        orchestrator.run()

        assert len(spy_client.calls) == 2, "Should have 2 calls"

        assert spy_client.calls[0]["history_length"] == 0, (
            f"First call should have empty history, got {spy_client.calls[0]['history_length']}"
        )

        assert spy_client.calls[1]["history_length"] == 0, (
            f"Second call should have empty history, got {spy_client.calls[1]['history_length']}"
        )

    def test_no_double_context(self, integration_workbook_with_dependencies, spy_client):
        """
        Verify that API receives only declarative context, not accumulated history.

        When using history=["step1"], the context should be injected into the
        prompt STRING via FFAI's _build_prompt(), NOT accumulated in the client's
        conversation_history.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_dependencies, spy_client, concurrency=1
        )
        orchestrator.run()

        assert len(spy_client.calls) == 3, "Should have 3 calls"

        for i, call in enumerate(spy_client.calls):
            assert call["history_length"] == 0, (
                f"Call {i + 1} should have empty client history, got {call['history_length']}"
            )

        third_call = spy_client.calls[2]
        third_prompt = third_call["prompt"]

        assert "step1" in third_prompt or "2 + 2" in third_prompt or "4" in third_prompt, (
            "Third prompt should contain declarative context from step1"
        )
        assert "step2" in third_prompt or "3 + 3" in third_prompt or "6" in third_prompt, (
            "Third prompt should contain declarative context from step2"
        )
        assert "<conversation_history>" in third_prompt, (
            "Third prompt should use conversation_history format"
        )

    def test_parallel_clients_dont_share_state(self, integration_workbook, spy_client):
        """
        Verify that concurrent prompts don't interfere via shared client.

        With concurrency > 1, each prompt should get its own cloned client
        with empty history.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        for i, name in enumerate(["p1", "p2", "p3", "p4"], start=2):
            ws[f"A{i}"] = i - 1
            ws[f"B{i}"] = name
            ws[f"C{i}"] = f"What is {i - 1} + {i - 1}?"
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, spy_client, concurrency=3)
        orchestrator.run()

        assert len(spy_client.calls) == 4, "Should have 4 calls"

        for i, call in enumerate(spy_client.calls):
            assert call["history_length"] == 0, (
                f"Parallel call {i + 1} should have empty history, got {call['history_length']}"
            )

    def test_batch_clients_isolated(self, integration_workbook_with_batch_data, spy_client):
        """
        Verify that each batch iteration gets a fresh client.

        In batch mode, each row in the data sheet should execute with
        isolated client state.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_batch_data, spy_client, concurrency=1
        )
        orchestrator.run()

        assert len(spy_client.calls) == 3, "Should have 3 batch calls (3 data rows)"

        for i, call in enumerate(spy_client.calls):
            assert call["history_length"] == 0, (
                f"Batch call {i + 1} should have empty history, got {call['history_length']}"
            )

    def test_clone_preserves_config_but_clears_history(self, real_mistral_client):
        """
        Verify that cloned client has same model/temp/instructions but empty history.

        The clone() method should:
        - Preserve: api_key, model, temperature, max_tokens, system_instructions
        - Reset: conversation_history (should be empty)
        """
        real_mistral_client.conversation_history = [
            {"role": "user", "content": "Previous message"},
            {"role": "assistant", "content": "Previous response"},
        ]

        cloned = real_mistral_client.clone()

        assert cloned.api_key == real_mistral_client.api_key, "Clone should preserve api_key"
        assert cloned.model == real_mistral_client.model, "Clone should preserve model"
        assert cloned.temperature == real_mistral_client.temperature, (
            "Clone should preserve temperature"
        )
        assert cloned.max_tokens == real_mistral_client.max_tokens, (
            "Clone should preserve max_tokens"
        )
        assert cloned.system_instructions == real_mistral_client.system_instructions, (
            "Clone should preserve system_instructions"
        )

        assert len(cloned.conversation_history) == 0, (
            f"Clone should have empty history, got {len(cloned.conversation_history)}"
        )


class TestClientIsolationRealAPI:
    """Tests using real API calls to verify isolation."""

    def test_real_api_no_history_accumulation(
        self, integration_workbook_with_dependencies, real_mistral_client
    ):
        """
        Integration test with real API to verify no history accumulation.

        This test makes real API calls and verifies that:
        1. All prompts execute successfully
        2. Results are written correctly
        3. No errors occur from context overflow
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_dependencies, real_mistral_client, concurrency=1
        )

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 3, f"Expected 3 successful, got {summary['successful']}"
        assert summary["failed"] == 0, f"Expected 0 failed, got {summary['failed']}"

        wb = load_workbook(integration_workbook_with_dependencies)
        assert results_sheet in wb.sheetnames

        ws = wb[results_sheet]
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                results.append(
                    {
                        "sequence": row[2],
                        "prompt_name": row[3],
                        "response": row[11],
                        "status": row[12],
                    }
                )

        assert len(results) == 3

        for r in results:
            assert r["status"] == "success", f"Prompt {r['sequence']} should succeed"
            assert r["response"] is not None, f"Prompt {r['sequence']} should have response"

    def test_real_api_parallel_execution_isolated(self, integration_workbook, real_mistral_client):
        """
        Test parallel execution with real API to verify isolation.
        """
        from openpyxl import load_workbook

        from src.orchestrator.excel_orchestrator import ExcelOrchestrator

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        for i, name in enumerate(["a", "b", "c"], start=2):
            ws[f"A{i}"] = i - 1
            ws[f"B{i}"] = name
            ws[f"C{i}"] = f"Say the letter {name}. Just the letter."
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, real_mistral_client, concurrency=3)

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 3, f"Expected 3 successful, got {summary['successful']}"
        assert summary["failed"] == 0, f"Expected 0 failed, got {summary['failed']}"
