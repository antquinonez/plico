"""
Integration tests for conditional execution.

These tests verify that conditional expressions control prompt execution
correctly, including status checks, content checks, and boolean logic.
"""

import pytest
import sys
import os

sys.path.insert(
    0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
)


class TestConditionalExecution:
    """Tests for conditional execution."""

    def test_condition_skip_on_status(
        self, integration_workbook_with_conditions, spy_client
    ):
        """
        Verify that {{p1.status}} == "success" works correctly.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator
        from openpyxl import load_workbook

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_conditions, spy_client, concurrency=1
        )

        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook_with_conditions)
        ws = wb[results_sheet]

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                results[row[3]] = {
                    "status": row[11],
                    "condition": row[7],
                    "condition_result": row[8],
                }

        assert results["check1"]["status"] == "success", (
            "check1 should succeed (no condition)"
        )

        assert results["proceed"]["status"] == "success", (
            "proceed should run (condition true)"
        )

        assert results["combined"]["status"] == "success", (
            "combined should run (both conditions true)"
        )

        assert results["never_runs"]["status"] == "skipped", (
            "never_runs should be skipped (condition false)"
        )

    def test_condition_boolean_combination(
        self, integration_workbook_with_conditions, spy_client
    ):
        """
        Verify that and/or combinations work.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator
        from openpyxl import load_workbook

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_conditions, spy_client, concurrency=1
        )

        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook_with_conditions)
        ws = wb[results_sheet]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] == "combined":
                assert row[6] == True, (
                    f"combined condition_result should be True, got {row[6]}"
                )
                assert row[8] == "success", f"combined should run, got status {row[8]}"

    def test_condition_implies_dependency(self, integration_workbook, spy_client):
        """
        Verify that condition reference auto-adds to DAG.

        If p2 has condition {{p1.status}} == "success", then p2
        should wait for p1 to complete, even without explicit history.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator
        from openpyxl import load_workbook

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "first"
        ws["C2"] = "First prompt"
        ws["A3"] = 2
        ws["B3"] = "second"
        ws["C3"] = "Second prompt"
        ws["E3"] = '{{first.status}} == "success"'
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(
            integration_workbook, spy_client, concurrency=1
        )
        orchestrator.run()

        first_prompt = spy_client.calls[0]["prompt"]
        second_prompt = spy_client.calls[1]["prompt"]

        assert "First prompt" in first_prompt, "first should execute first"
        assert "Second prompt" in second_prompt, "second should execute second"

    def test_condition_with_response_content(self, integration_workbook, spy_client):
        """
        Verify that "text" in {{p.response}} works.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator
        from openpyxl import load_workbook

        class ResponseSpyClient:
            def __init__(self):
                self.calls = []
                self.conversation_history = []
                self.system_instructions = "Test"
                self.model = "test"
                self.temperature = 0.3
                self.max_tokens = 100
                self.call_count = 0

            def generate_response(self, prompt, **kwargs):
                self.call_count += 1
                self.calls.append(
                    {"prompt": prompt, "history_length": len(self.conversation_history)}
                )

                if "say_yes" in prompt.lower():
                    return "yes, I understand"
                elif "say_error" in prompt.lower():
                    return "error: something went wrong"
                else:
                    return "normal response"

            def clone(self):
                return self

            def clear_conversation(self):
                pass

        client = ResponseSpyClient()

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "has_yes"
        ws["C2"] = "Please say_yes"
        ws["A3"] = 2
        ws["B3"] = "has_error"
        ws["C3"] = "Please say_error"
        ws["A4"] = 3
        ws["B4"] = "check_yes"
        ws["C4"] = "Found yes"
        ws["E4"] = '"yes" in {{has_yes.response}}'
        ws["A5"] = 4
        ws["B5"] = "check_error"
        ws["C5"] = "Found error"
        ws["E5"] = '"error" in {{has_error.response}}'
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook)
        ws = wb[results_sheet]

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                results[row[3]] = {"status": row[11]}

        assert results["check_yes"]["status"] == "success", (
            "check_yes should run (response contains 'yes')"
        )
        assert results["check_error"]["status"] == "success", (
            "check_error should run (response contains 'error')"
        )

    def test_condition_with_functions(self, integration_workbook, spy_client):
        """
        Verify that len({{p.response}}) > 10 works.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator
        from openpyxl import load_workbook

        class LengthSpyClient:
            def __init__(self):
                self.calls = []
                self.conversation_history = []
                self.system_instructions = "Test"
                self.model = "test"
                self.temperature = 0.3
                self.max_tokens = 100

            def generate_response(self, prompt, **kwargs):
                self.calls.append({"prompt": prompt})
                return "This is a very long response that is definitely more than 10 characters"

            def clone(self):
                return self

            def clear_conversation(self):
                pass

        client = LengthSpyClient()

        wb = load_workbook(integration_workbook)
        ws = wb["prompts"]
        ws["A2"] = 1
        ws["B2"] = "long_response"
        ws["C2"] = "Give me a long response"
        ws["A3"] = 2
        ws["B3"] = "check_length"
        ws["C3"] = "Response was long enough"
        ws["E3"] = "len({{long_response.response}}) > 10"
        wb.save(integration_workbook)

        orchestrator = ExcelOrchestrator(integration_workbook, client, concurrency=1)
        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook)
        ws = wb[results_sheet]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None and row[3] == "check_length":
                assert row[11] == "success", (
                    f"check_length should run (len > 10), got {row[11]}"
                )
                return

        pytest.fail("check_length result not found")

    def test_skipped_status_recorded(
        self, integration_workbook_with_conditions, spy_client
    ):
        """
        Verify that skipped prompts have status="skipped".
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator
        from openpyxl import load_workbook

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_conditions, spy_client, concurrency=1
        )

        results_sheet = orchestrator.run()

        wb = load_workbook(integration_workbook_with_conditions)
        ws = wb[results_sheet]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None and row[3] == "never_runs":
                assert row[11] == "skipped", (
                    f"never_runs should have status 'skipped', got {row[11]}"
                )
                assert row[8] == False, (
                    f"never_runs condition_result should be False, got {row[8]}"
                )
                return

        pytest.fail("never_runs result not found")


class TestConditionalRealAPI:
    """Tests using real API calls."""

    def test_real_api_conditional_execution(
        self, integration_workbook_with_conditions, real_mistral_client
    ):
        """
        Full conditional execution test with real API.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator
        from openpyxl import load_workbook

        orchestrator = ExcelOrchestrator(
            integration_workbook_with_conditions, real_mistral_client, concurrency=2
        )

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 4, (
            f"Should have 4 successful (check1, check2, proceed, combined), got {summary['successful']}"
        )
        assert summary["skipped"] == 1, (
            f"Should have 1 skipped (never_runs), got {summary['skipped']}"
        )
        assert summary["failed"] == 0, f"Should have 0 failed, got {summary['failed']}"

        wb = load_workbook(integration_workbook_with_conditions)
        ws = wb[results_sheet]

        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:
                results[row[3]] = {
                    "status": row[11],
                    "response": row[10],
                    "condition_result": row[8],
                }

        assert results["check1"]["status"] == "success"
        assert results["check2"]["status"] == "success"
        assert results["proceed"]["status"] == "success"
        assert results["proceed"]["condition_result"] == True
        assert results["combined"]["status"] == "success"
        assert results["combined"]["condition_result"] == True
        assert results["never_runs"]["status"] == "skipped"
        assert results["never_runs"]["condition_result"] == False

    def test_real_api_response_based_condition(self, tmp_path, real_mistral_client):
        """
        Test condition based on actual response content with real API.
        """
        from src.orchestrator.excel_orchestrator import ExcelOrchestrator
        from openpyxl import load_workbook, Workbook

        workbook_path = str(tmp_path / "response_cond.xlsx")

        wb = Workbook()
        ws_config = wb.active
        ws_config.title = "config"
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        config_items = [
            ("model", "mistral-small-2503"),
            ("api_key_env", "MISTRALSMALL_KEY"),
            ("max_retries", 2),
            ("temperature", 0.3),
            ("max_tokens", 50),
            ("system_instructions", "Follow instructions exactly."),
        ]
        for idx, (field, value) in enumerate(config_items, start=2):
            ws_config[f"A{idx}"] = field
            ws_config[f"B{idx}"] = value

        ws_prompts = wb.create_sheet(title="prompts")
        ws_prompts["A1"] = "sequence"
        ws_prompts["B1"] = "prompt_name"
        ws_prompts["C1"] = "prompt"
        ws_prompts["D1"] = "history"
        ws_prompts["E1"] = "condition"

        ws_prompts["A2"] = 1
        ws_prompts["B2"] = "generate"
        ws_prompts["C2"] = "Say the word POSITIVE. Just that word."

        ws_prompts["A3"] = 2
        ws_prompts["B3"] = "conditional"
        ws_prompts["C3"] = "Good, you said positive."
        ws_prompts["E3"] = '"POSITIVE" in upper({{generate.response}})'

        wb.save(workbook_path)

        orchestrator = ExcelOrchestrator(
            workbook_path, real_mistral_client, concurrency=1
        )

        results_sheet = orchestrator.run()
        summary = orchestrator.get_summary()

        assert summary["successful"] == 2, (
            f"Should have 2 successful, got {summary['successful']}"
        )

        wb = load_workbook(workbook_path)
        ws = wb[results_sheet]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None and row[3] == "generate":
                assert "positive" in str(row[10]).lower(), (
                    f"generate response should contain 'positive', got {row[10]}"
                )
                return

        pytest.fail("generate result not found")
