import json
import os
import sys

import pytest
from dotenv import load_dotenv
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
load_dotenv()


@pytest.fixture
def integration_workbook(tmp_path):
    """Create a real workbook file with config and prompts sheets."""
    workbook_path = str(tmp_path / "integration_test.xlsx")

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", 2),
        ("temperature", 0.3),
        ("max_tokens", 100),
        (
            "system_instructions",
            "You are a helpful assistant. Give very brief answers.",
        ),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    wb.save(workbook_path)
    return workbook_path


@pytest.fixture
def integration_workbook_with_dependencies(tmp_path):
    """Create a workbook with prompt dependencies for context testing."""
    workbook_path = str(tmp_path / "integration_deps_test.xlsx")

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", 2),
        ("temperature", 0.3),
        ("max_tokens", 100),
        (
            "system_instructions",
            "You are a helpful assistant. Give very brief answers - just the answer, no explanation.",
        ),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    prompts = [
        (1, "step1", "What is 2 + 2? Just give the number.", None),
        (2, "step2", "What is 3 + 3? Just give the number.", None),
        (
            3,
            "step3",
            "Add the results from step1 and step2. Just give the number.",
            json.dumps(["step1", "step2"]),
        ),
    ]

    for idx, (seq, name, prompt, history) in enumerate(prompts, start=2):
        ws_prompts[f"A{idx}"] = seq
        ws_prompts[f"B{idx}"] = name
        ws_prompts[f"C{idx}"] = prompt
        if history:
            ws_prompts[f"D{idx}"] = history

    wb.save(workbook_path)
    return workbook_path


@pytest.fixture
def integration_workbook_with_batch_data(tmp_path):
    """Create a workbook with batch data sheet."""
    workbook_path = str(tmp_path / "integration_batch_test.xlsx")

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", 2),
        ("temperature", 0.3),
        ("max_tokens", 50),
        (
            "system_instructions",
            "You are a helpful assistant. Give very brief answers.",
        ),
        ("batch_mode", "per_row"),
        ("batch_output", "combined"),
        ("on_batch_error", "continue"),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    prompts = [
        (
            1,
            "analyze",
            "The region is {{region}} and product is {{product}}. What is the first letter of the region? Just give the letter.",
            None,
        ),
    ]

    for idx, (seq, name, prompt, history) in enumerate(prompts, start=2):
        ws_prompts[f"A{idx}"] = seq
        ws_prompts[f"B{idx}"] = name
        ws_prompts[f"C{idx}"] = prompt
        if history:
            ws_prompts[f"D{idx}"] = history

    ws_data = wb.create_sheet(title="data")
    data_headers = ["id", "batch_name", "region", "product"]
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)

    data_rows = [
        (1, "{{region}}_{{product}}", "north", "widget_a"),
        (2, "{{region}}_{{product}}", "south", "widget_b"),
        (3, "{{region}}_{{product}}", "east", "widget_c"),
    ]

    for row_idx, (row_id, batch_name, region, product) in enumerate(data_rows, start=2):
        ws_data.cell(row=row_idx, column=1, value=row_id)
        ws_data.cell(row=row_idx, column=2, value=batch_name)
        ws_data.cell(row=row_idx, column=3, value=region)
        ws_data.cell(row=row_idx, column=4, value=product)

    wb.save(workbook_path)
    return workbook_path


@pytest.fixture
def integration_workbook_with_clients(tmp_path):
    """Create a workbook with clients sheet for multi-client testing."""
    workbook_path = str(tmp_path / "integration_multiclient_test.xlsx")

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", 2),
        ("temperature", 0.7),
        ("max_tokens", 50),
        (
            "system_instructions",
            "You are a helpful assistant. Give very brief answers.",
        ),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_clients = wb.create_sheet(title="clients")
    ws_clients["A1"] = "name"
    ws_clients["B1"] = "client_type"
    ws_clients["C1"] = "temperature"
    ws_clients["D1"] = "max_tokens"

    ws_clients["A2"] = "fast"
    ws_clients["B2"] = "mistral-small"
    ws_clients["C2"] = 0.1
    ws_clients["D2"] = 30

    ws_clients["A3"] = "detailed"
    ws_clients["B3"] = "mistral-small"
    ws_clients["C3"] = 0.5
    ws_clients["D3"] = 100

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"
    ws_prompts["E1"] = "client"

    prompts = [
        (1, "task1", "What is 1 + 1? Just give the number.", None, "fast"),
        (2, "task2", "What is 2 + 2? Just give the number.", None, None),
        (3, "task3", "What is 3 + 3? Just give the number.", None, "detailed"),
    ]

    for idx, (seq, name, prompt, history, client) in enumerate(prompts, start=2):
        ws_prompts[f"A{idx}"] = seq
        ws_prompts[f"B{idx}"] = name
        ws_prompts[f"C{idx}"] = prompt
        if history:
            ws_prompts[f"D{idx}"] = history
        if client:
            ws_prompts[f"E{idx}"] = client

    wb.save(workbook_path)
    return workbook_path


@pytest.fixture
def integration_workbook_with_conditions(tmp_path):
    """Create a workbook with conditional execution."""
    workbook_path = str(tmp_path / "integration_conditional_test.xlsx")

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", 2),
        ("temperature", 0.3),
        ("max_tokens", 50),
        (
            "system_instructions",
            "You are a helpful assistant. Give very brief answers.",
        ),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"
    ws_prompts["E1"] = "condition"

    prompts = [
        (1, "check1", "Say 'success' if you understand. Just say success.", None, None),
        (2, "check2", "Say 'yes' if you agree. Just say yes.", None, None),
        (
            3,
            "proceed",
            "What is 5 + 5? Just give the number.",
            None,
            '{{check1.status}} == "success"',
        ),
        (
            4,
            "combined",
            "What is 10 + 10? Just give the number.",
            None,
            '{{check1.status}} == "success" and {{check2.status}} == "success"',
        ),
        (
            5,
            "never_runs",
            "This should never run.",
            None,
            '{{check1.status}} == "failed"',
        ),
    ]

    for idx, (seq, name, prompt, history, condition) in enumerate(prompts, start=2):
        ws_prompts[f"A{idx}"] = seq
        ws_prompts[f"B{idx}"] = name
        ws_prompts[f"C{idx}"] = prompt
        if history:
            ws_prompts[f"D{idx}"] = history
        if condition:
            ws_prompts[f"E{idx}"] = condition

    wb.save(workbook_path)
    return workbook_path


@pytest.fixture
def real_mistral_client():
    """Create a real FFMistralSmall client with API key from environment."""
    from src.Clients.FFMistralSmall import FFMistralSmall

    api_key = os.getenv("MISTRALSMALL_KEY")
    if not api_key:
        pytest.skip("MISTRALSMALL_KEY not set in environment")

    client = FFMistralSmall(
        api_key=api_key,
        model="mistral-small-2503",
        temperature=0.3,
        max_tokens=100,
        system_instructions="You are a helpful assistant. Give very brief answers.",
    )
    return client


@pytest.fixture
def spy_client():
    """
    Create a spy client that records all calls across all clones.
    Used to inspect what's being sent to the API without making real calls.
    """

    class SpyClient:
        def __init__(self, shared_calls=None, temperature=0.3, max_tokens=100):
            if shared_calls is None:
                self._calls = []
            else:
                self._calls = shared_calls
            self.conversation_history = []
            self.system_instructions = "You are a helpful assistant."
            self.model = "mistral-small-2503"
            self.temperature = temperature
            self.max_tokens = max_tokens

        @property
        def calls(self):
            return self._calls

        def generate_response(self, prompt, **kwargs):
            self._calls.append(
                {
                    "prompt": prompt,
                    "history_snapshot": list(self.conversation_history),
                    "history_length": len(self.conversation_history),
                    "kwargs": kwargs,
                }
            )
            self.conversation_history.append({"role": "user", "content": prompt})
            self.conversation_history.append({"role": "assistant", "content": "Spy response"})
            return "Spy response"

        def clone(self):
            cloned = SpyClient(
                shared_calls=self._calls,
                temperature=self.temperature,
                max_tokens=self.max_tokens,
            )
            cloned.system_instructions = self.system_instructions
            cloned.model = self.model
            return cloned

        def clear_conversation(self):
            self.conversation_history = []

        def get_conversation_history(self):
            return self.conversation_history

        def set_conversation_history(self, history):
            self.conversation_history = history

        def test_connection(self):
            return True

    return SpyClient()
