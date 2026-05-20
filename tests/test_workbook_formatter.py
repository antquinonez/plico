# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Tests for WorkbookFormatter and format_workbook."""

from __future__ import annotations

from unittest.mock import MagicMock

from openpyxl import Workbook

from src.orchestrator.workbook_formatter import WorkbookFormatter, format_workbook


class TestEstimateTextLinesEmptyParagraph:
    """Tests for _estimate_text_lines with empty paragraphs (line 151)."""

    def test_empty_paragraph_counts_as_line(self):
        """Empty paragraphs (from \\n\\n) count as 1 line each."""
        formatter = WorkbookFormatter.__new__(WorkbookFormatter)
        result = formatter._estimate_text_lines("hello\n\nworld", 50.0)
        assert result == 3

    def test_trailing_newline_adds_line(self):
        """Trailing newline adds an extra empty-paragraph line."""
        formatter = WorkbookFormatter.__new__(WorkbookFormatter)
        result = formatter._estimate_text_lines("hello\n", 50.0)
        assert result == 2


class TestFindColumnByHeaderNoHeaderRow:
    """Tests for _find_column_by_header when no header row exists (line 160)."""

    def test_empty_worksheet_returns_none(self):
        """Empty worksheet with no rows returns None."""
        wb = Workbook()
        ws = wb.active
        ws.delete_rows(1)
        formatter = WorkbookFormatter.__new__(WorkbookFormatter)
        result = formatter._find_column_by_header(ws, "missing")
        assert result is None


class TestDetectTextColumns:
    """Tests for _detect_text_columns (lines 185, 197)."""

    def test_no_header_columns_skipped(self):
        """Columns with no header are skipped (line 185)."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=None)
        ws.cell(row=2, column=1, value="short")
        formatter = WorkbookFormatter.__new__(WorkbookFormatter)
        result = formatter._detect_text_columns(ws)
        assert result == []

    def test_long_text_column_detected(self):
        """Columns with average content >30 chars are detected as text columns (line 197)."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="description")
        for row in range(2, 12):
            ws.cell(row=row, column=1, value="x" * 50)
        formatter = WorkbookFormatter.__new__(WorkbookFormatter)
        result = formatter._detect_text_columns(ws)
        assert "description" in result

    def test_short_text_column_not_detected(self):
        """Columns with average content <=30 chars are not flagged."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="name")
        for row in range(2, 12):
            ws.cell(row=row, column=1, value="short")
        formatter = WorkbookFormatter.__new__(WorkbookFormatter)
        result = formatter._detect_text_columns(ws)
        assert "name" not in result


class TestFormatWorkbookFunction:
    """Tests for format_workbook() module-level function (lines 210-213)."""

    def test_format_workbook_formats_all_sheets(self):
        """format_workbook iterates all sheets in the workbook."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "results"
        ws1.cell(row=1, column=1, value="header")
        ws1.cell(row=2, column=1, value="data")

        ws2 = wb.create_sheet("summary")
        ws2.cell(row=1, column=1, value="metric")
        ws2.cell(row=2, column=1, value="score")

        config_mock = MagicMock()
        config_mock.workbook.formatting.column_widths = {}
        config_mock.workbook.formatting.word_wrap = {"enabled": False}
        config_mock.workbook.formatting.features.freeze_panes = {"enabled": False}
        config_mock.workbook.formatting.features.auto_filter = {"enabled": False}
        config_mock.workbook.formatting.rows = {
            "auto_fit_height": False,
            "wrap_text_height_multiplier": 15,
        }

        format_workbook(wb, config_mock)

        assert ws1.freeze_panes is None
        assert ws2.freeze_panes is None
