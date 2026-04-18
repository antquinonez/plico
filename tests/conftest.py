# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

import os
import sys
from unittest.mock import MagicMock, patch

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

collect_ignore_glob = [
    "test_ffanthropic*.py",
    "test_ffazure_*.py",
    "test_ffnvidia_*.py",
    "test_ffopenai_*.py",
]


@pytest.fixture
def mock_mistral_response():
    """Mock response from Mistral API."""
    response = MagicMock()
    response.choices = [MagicMock()]
    response.choices[0].message.content = "This is a test response."
    response.choices[0].message.tool_calls = None
    response.usage = None
    return response


@pytest.fixture
def mock_mistral_client(mock_mistral_response):
    """Mock Mistral client."""
    client = MagicMock()
    client.chat.complete.return_value = mock_mistral_response
    return client


@pytest.fixture
def mock_anthropic_response():
    """Mock response from Anthropic API."""
    response = MagicMock()
    response.content = [MagicMock()]
    response.content[0].text = "This is a test response."
    return response


@pytest.fixture
def mock_anthropic_client(mock_anthropic_response):
    """Mock Anthropic client."""
    client = MagicMock()
    client.messages.create.return_value = mock_anthropic_response
    return client


@pytest.fixture
def mock_openai_response():
    """Mock response from OpenAI-compatible API."""
    response = MagicMock()
    response.choices = [MagicMock()]
    response.choices[0].message.content = "This is a test response."
    response.choices[0].message.tool_calls = None
    response.usage = None
    return response


@pytest.fixture
def mock_openai_client(mock_openai_response):
    """Mock OpenAI client."""
    client = MagicMock()
    client.chat.completions.create.return_value = mock_openai_response
    return client


@pytest.fixture
def mock_azure_response():
    """Mock response from Azure AI Inference API."""
    response = MagicMock()
    response.choices = [MagicMock()]
    response.choices[0].message.content = "This is a test response."
    response.choices[0].message.tool_calls = None
    response.usage = None
    return response


@pytest.fixture
def mock_azure_client(mock_azure_response):
    """Mock Azure AI Inference client."""
    client = MagicMock()
    client.complete.return_value = mock_azure_response
    return client


@pytest.fixture
def mock_gemini_client():
    """Mock Gemini/AsyncOpenAI client."""
    client = MagicMock()
    client.chat = MagicMock()
    client.chat.completions = MagicMock()
    return client


@pytest.fixture
def mock_ffmistral(mock_mistral_client):
    """Mock FFMistral instance."""
    with patch("src.Clients.FFMistral.Mistral") as MockMistral:
        MockMistral.return_value = mock_mistral_client
        from src.Clients.FFMistral import FFMistral

        client = FFMistral(
            api_key="test-api-key",
            model="mistral-large-latest",
            temperature=0.8,
            max_tokens=4096,
        )
        yield client


@pytest.fixture
def mock_ffmistralsmall(mock_mistral_client):
    """Mock FFMistralSmall instance."""
    with patch("src.Clients.FFMistralSmall.Mistral") as MockMistral:
        MockMistral.return_value = mock_mistral_client
        from src.Clients.FFMistralSmall import FFMistralSmall

        client = FFMistralSmall(
            api_key="test-api-key",
            model="mistral-small-2503",
            temperature=0.8,
            max_tokens=128000,
        )
        yield client


@pytest.fixture
def sample_prompts():
    """Sample prompts for testing."""
    return [
        {
            "sequence": 1,
            "prompt_name": "greeting",
            "prompt": "Hello, how are you?",
            "history": None,
        },
        {
            "sequence": 2,
            "prompt_name": "math",
            "prompt": "What is 2 + 2?",
            "history": None,
        },
        {
            "sequence": 3,
            "prompt_name": "followup",
            "prompt": "What was the answer to my math question?",
            "history": ["math", "greeting"],
        },
    ]


@pytest.fixture
def sample_config():
    """Sample configuration for testing."""
    return {
        "model": "mistral-small-2503",
        "api_key_env": "MISTRALSMALL_KEY",
        "max_retries": 3,
        "temperature": 0.8,
        "max_tokens": 4096,
        "system_instructions": "You are a helpful assistant.",
    }


@pytest.fixture
def temp_workbook(tmp_path):
    """Create a temporary workbook path."""
    return str(tmp_path / "sample_workbook.xlsx")


@pytest.fixture
def temp_workbook_with_data(temp_workbook, sample_prompts, sample_config):
    """Create a workbook with sample data."""
    from openpyxl import Workbook

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", sample_config["model"]),
        ("api_key_env", sample_config["api_key_env"]),
        ("max_retries", sample_config["max_retries"]),
        ("temperature", sample_config["temperature"]),
        ("max_tokens", sample_config["max_tokens"]),
        ("system_instructions", sample_config["system_instructions"]),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    for idx, p in enumerate(sample_prompts, start=2):
        ws_prompts[f"A{idx}"] = p["sequence"]
        ws_prompts[f"B{idx}"] = p["prompt_name"]
        ws_prompts[f"C{idx}"] = p["prompt"]
        if p["history"]:
            import json

            ws_prompts[f"D{idx}"] = json.dumps(p["history"])

    wb.save(temp_workbook)
    return temp_workbook


@pytest.fixture
def sample_batch_data():
    """Sample batch data for testing."""
    return [
        {
            "id": 1,
            "batch_name": "{{region}}_{{product}}",
            "region": "north",
            "product": "widget_a",
            "price": 10,
            "quantity": 100,
        },
        {
            "id": 2,
            "batch_name": "{{region}}_{{product}}",
            "region": "south",
            "product": "widget_b",
            "price": 15,
            "quantity": 75,
        },
        {
            "id": 3,
            "batch_name": "{{region}}_{{product}}",
            "region": "east",
            "product": "widget_c",
            "price": 20,
            "quantity": 50,
        },
    ]


@pytest.fixture
def temp_workbook_with_batch_data(temp_workbook, sample_prompts, sample_config, sample_batch_data):
    """Create a workbook with sample data including batch data sheet."""
    from openpyxl import Workbook

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_items = [
        ("model", sample_config["model"]),
        ("api_key_env", sample_config["api_key_env"]),
        ("max_retries", sample_config["max_retries"]),
        ("temperature", sample_config["temperature"]),
        ("max_tokens", sample_config["max_tokens"]),
        ("system_instructions", sample_config["system_instructions"]),
        ("batch_mode", "per_row"),
        ("batch_output", "combined"),
        ("on_batch_error", "continue"),
    ]

    for idx, (field, value) in enumerate(config_items, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_prompts = wb.create_sheet(title="prompts")
    ws_prompts["A1"] = "sequence"
    ws_prompts["B1"] = "prompt_name"
    ws_prompts["C1"] = "prompt"
    ws_prompts["D1"] = "history"

    batch_prompts = [
        {
            "sequence": 1,
            "prompt_name": "intro",
            "prompt": "Analyze {{region}} region, product {{product}}.",
            "history": None,
        },
        {
            "sequence": 2,
            "prompt_name": "calc",
            "prompt": "Price is {{price}}, quantity is {{quantity}}.",
            "history": None,
        },
        {
            "sequence": 3,
            "prompt_name": "summary",
            "prompt": "Summarize the analysis.",
            "history": ["intro", "calc"],
        },
    ]

    for idx, p in enumerate(batch_prompts, start=2):
        ws_prompts[f"A{idx}"] = p["sequence"]
        ws_prompts[f"B{idx}"] = p["prompt_name"]
        ws_prompts[f"C{idx}"] = p["prompt"]
        if p["history"]:
            import json

            ws_prompts[f"D{idx}"] = json.dumps(p["history"])

    ws_data = wb.create_sheet(title="data")
    data_headers = ["id", "batch_name", "region", "product", "price", "quantity"]
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(sample_batch_data, start=2):
        for col_idx, header in enumerate(data_headers, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=row_data.get(header))

    wb.save(temp_workbook)
    return temp_workbook
