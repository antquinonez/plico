#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Generate test workbook for document reference and RAG semantic search testing.

Creates a workbook with:
    - config sheet
    - prompts sheet with references and semantic_query columns
    - documents sheet
    - data sheet (optional batch data)

Demonstrates:
    - Full document injection via references column
    - Semantic search via semantic_query column (RAG)

Uses FFLiteLLMClient with LiteLLM routing for Mistral Small.

Usage:
    python scripts/create_test_workbook_documents.py [output_path]
"""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook

from src.config import get_config


def create_test_workbook(output_path: str):
    config = get_config()
    test_config = config.test

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_data = [
        ("model", test_config.default_model),
        ("max_retries", str(test_config.default_retries)),
        ("temperature", str(test_config.default_temperature)),
        ("max_tokens", "1000"),
        (
            "system_instructions",
            "You are a helpful assistant. Analyze documents and answer questions based on their content.",
        ),
        ("created_at", datetime.now().isoformat()),
    ]

    for idx, (field, value) in enumerate(config_data, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_config.column_dimensions["A"].width = 20
    ws_config.column_dimensions["B"].width = 70

    ws_documents = wb.create_sheet(title="documents")
    doc_headers = ["reference_name", "common_name", "file_path", "notes"]
    for col_idx, header in enumerate(doc_headers, start=1):
        ws_documents.cell(row=1, column=col_idx, value=header)

    app_config = get_config()

    documents = [
        (
            "product_spec",
            "Product Specification",
            f"{app_config.paths.library}/product_spec.md",
            "Main product documentation",
        ),
        (
            "api_ref",
            "API Reference",
            f"{app_config.paths.library}/api_reference.txt",
            "API documentation",
        ),
        (
            "config",
            "Configuration File",
            f"{app_config.paths.library}/config.json",
            "App configuration",
        ),
        (
            "troubleshoot",
            "Troubleshooting Guide",
            f"{app_config.paths.library}/troubleshooting.txt",
            "Common issues and solutions",
        ),
    ]

    for row_idx, (ref_name, common_name, file_path, notes) in enumerate(documents, start=2):
        ws_documents.cell(row=row_idx, column=1, value=ref_name)
        ws_documents.cell(row=row_idx, column=2, value=common_name)
        ws_documents.cell(row=row_idx, column=3, value=file_path)
        ws_documents.cell(row=row_idx, column=4, value=notes)

    ws_documents.column_dimensions["A"].width = 18
    ws_documents.column_dimensions["B"].width = 25
    ws_documents.column_dimensions["C"].width = 50
    ws_documents.column_dimensions["D"].width = 30

    ws_prompts = wb.create_sheet(title="prompts")
    prompt_headers = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "references",
        "semantic_query",
    ]
    for col_idx, header in enumerate(prompt_headers, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    prompts = [
        (
            1,
            "spec_summary",
            "Summarize the key features of the product specification.",
            None,
            None,
            None,
            '["product_spec"]',
            None,
        ),
        (
            2,
            "api_overview",
            "List the available client types and their purposes.",
            None,
            None,
            None,
            '["api_ref"]',
            None,
        ),
        (
            3,
            "config_analysis",
            "What features are enabled in the configuration?",
            None,
            None,
            None,
            '["config"]',
            None,
        ),
        (
            4,
            "combined_analysis",
            "Based on the product spec and API reference, describe the system architecture.",
            None,
            None,
            None,
            '["product_spec", "api_ref"]',
            None,
        ),
        (
            5,
            "troubleshoot_summary",
            "Summarize the common issues and their solutions.",
            None,
            None,
            None,
            '["troubleshoot"]',
            None,
        ),
        (
            6,
            "full_context",
            "Using all documents, provide a comprehensive overview of the FFClients system.",
            None,
            None,
            None,
            '["product_spec", "api_ref", "config", "troubleshoot"]',
            None,
        ),
        (
            7,
            "no_ref_prompt",
            "What is 2 + 2? Just give the number.",
            None,
            None,
            None,
            "",
            None,
        ),
        (
            8,
            "rag_authentication",
            "How do I authenticate with the API?",
            None,
            None,
            None,
            None,
            "authentication API key token",
        ),
        (
            9,
            "rag_errors",
            "What are common errors and how do I fix them?",
            None,
            None,
            None,
            None,
            "troubleshooting error import dependency",
        ),
        (
            10,
            "rag_performance",
            "What are the performance tips for this system?",
            None,
            None,
            None,
            None,
            "performance batch concurrency tokens",
        ),
    ]

    for row_idx, (seq, name, prompt, history, client, condition, refs, semantic) in enumerate(
        prompts, start=2
    ):
        ws_prompts.cell(row=row_idx, column=1, value=seq)
        ws_prompts.cell(row=row_idx, column=2, value=name)
        ws_prompts.cell(row=row_idx, column=3, value=prompt)
        ws_prompts.cell(row=row_idx, column=4, value=history if history else "")
        ws_prompts.cell(row=row_idx, column=5, value=client if client else "")
        ws_prompts.cell(row=row_idx, column=6, value=condition if condition else "")
        ws_prompts.cell(row=row_idx, column=7, value=refs if refs else "")
        ws_prompts.cell(row=row_idx, column=8, value=semantic if semantic else "")

    ws_prompts.column_dimensions["A"].width = 10
    ws_prompts.column_dimensions["B"].width = 20
    ws_prompts.column_dimensions["C"].width = 60
    ws_prompts.column_dimensions["D"].width = 15
    ws_prompts.column_dimensions["E"].width = 12
    ws_prompts.column_dimensions["F"].width = 15
    ws_prompts.column_dimensions["G"].width = 35
    ws_prompts.column_dimensions["H"].width = 30

    wb.save(output_path)

    print(f"\n{'=' * 60}")
    print(f"Created document reference test workbook: {output_path}")
    print(f"{'=' * 60}")
    print("\nUsing: FFLiteLLMClient with LiteLLM routing")
    print(f"\nDocuments defined: {len(documents)}")
    for ref_name, common_name, _, _ in documents:
        print(f"  - {ref_name}: {common_name}")
    print(f"\nPrompts defined: {len(prompts)}")
    print("  - 6 prompts with document references (full injection)")
    print("  - 3 prompts with semantic_query (RAG search)")
    print("  - 1 prompt without references or semantic search")
    print("\nColumns:")
    print("  - references: Full document injection (existing behavior)")
    print("  - semantic_query: RAG semantic search (new RAG feature)")
    print(f"\n{'=' * 60}")
    print(f"Run with: python scripts/run_orchestrator.py {output_path}")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    config = get_config()
    output = sys.argv[1] if len(sys.argv) > 1 else config.test.workbooks.documents
    create_test_workbook(output)
