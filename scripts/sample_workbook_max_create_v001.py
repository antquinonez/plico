#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Generate comprehensive sample workbook combining batch, conditional, multi-client, and RAG features.

This workbook demonstrates the full power of FFClients orchestrator:
- BATCH: Multiple data rows processed through the same prompt chain
- CONDITIONAL: Prompts execute/skip based on runtime conditions
- MULTI-CLIENT: Different model configurations for different task types
- RAG: Semantic search with query expansion, filtering, and reranking

Uses FFLiteLLMClient with LiteLLM routing for Mistral Small.

Creates 27 prompts across 6 sections with 5 batch data rows and 13 documents for RAG.

Paired with: sample_workbook_max_validate_v001.py

Usage:
    python scripts/sample_workbook_max_create_v001.py [output_path]

Version: 001
"""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook

from src.config import get_config


def create_max_sample_workbook(output_path: str):
    config = get_config()
    test_config = config.sample

    wb = Workbook()

    # ============================================
    # CONFIG SHEET
    # ============================================
    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_data = [
        ("model", test_config.default_model),
        ("max_retries", str(test_config.default_retries)),
        ("temperature", str(test_config.default_temperature)),
        ("max_tokens", str(test_config.default_max_tokens)),
        (
            "system_instructions",
            "You are a helpful assistant. Give brief, concise answers. "
            "For classification, respond with just the category. "
            "For ratings, respond with just the number. "
            "For yes/no, respond with just 'yes' or 'no'.",
        ),
        ("created_at", datetime.now().isoformat()),
    ]

    for idx, (field, value) in enumerate(config_data, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_config.column_dimensions["A"].width = 20
    ws_config.column_dimensions["B"].width = 70

    # ============================================
    # CLIENTS SHEET
    # ============================================
    ws_clients = wb.create_sheet(title="clients")

    clients_headers = [
        "name",
        "client_type",
        "api_key_env",
        "model",
        "temperature",
        "max_tokens",
    ]
    for col_idx, header in enumerate(clients_headers, start=1):
        ws_clients.cell(row=1, column=col_idx, value=header)

    test_clients = test_config.sample_clients
    clients_data = []
    for name in ["default", "fast", "creative", "analytical"]:
        if name in test_clients:
            cfg = test_clients[name]
            clients_data.append(
                (
                    name,
                    cfg.client_type,
                    cfg.api_key_env,
                    cfg.model,
                    cfg.temperature,
                    cfg.max_tokens,
                )
            )

    for row_idx, client_row in enumerate(clients_data, start=2):
        for col_idx, value in enumerate(client_row, start=1):
            ws_clients.cell(row=row_idx, column=col_idx, value=value)

    ws_clients.column_dimensions["A"].width = 12
    ws_clients.column_dimensions["B"].width = 15
    ws_clients.column_dimensions["C"].width = 18
    ws_clients.column_dimensions["D"].width = 20
    ws_clients.column_dimensions["E"].width = 12
    ws_clients.column_dimensions["F"].width = 12

    # ============================================
    # DATA SHEET (Batch data)
    # ============================================
    ws_data = wb.create_sheet(title="data")

    data_headers = ["batch_name", "product_name", "review_text", "priority"]
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)

    batch_data = [
        (
            "product_a",
            "Wireless Headphones",
            "The sound quality is amazing but the battery life could be better.",
            "high",
        ),
        (
            "product_b",
            "Coffee Maker",
            "Works perfectly every morning. Love the auto-brew feature!",
            "medium",
        ),
        (
            "product_c",
            "Running Shoes",
            "Uncomfortable after long runs. Disappointed with the fit.",
            "high",
        ),
        (
            "product_d",
            "Desk Lamp",
            "Good brightness levels. The adjustable arm is very useful.",
            "low",
        ),
        (
            "product_e",
            "Water Bottle",
            "Leaks when placed sideways. Cannot recommend.",
            "high",
        ),
    ]

    for row_idx, data_row in enumerate(batch_data, start=2):
        for col_idx, value in enumerate(data_row, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    ws_data.column_dimensions["A"].width = 12
    ws_data.column_dimensions["B"].width = 18
    ws_data.column_dimensions["C"].width = 60
    ws_data.column_dimensions["D"].width = 10

    # ============================================
    # DOCUMENTS SHEET (for RAG)
    # ============================================
    ws_documents = wb.create_sheet(title="documents")
    doc_headers = ["reference_name", "common_name", "file_path", "notes"]
    for col_idx, header in enumerate(doc_headers, start=1):
        ws_documents.cell(row=1, column=col_idx, value=header)

    app_config = get_config()

    documents = [
        (
            "product_spec",
            "Product Specification",
            f"{app_config.paths.library}/product_spec.md",
            "Main product documentation",
        ),
        (
            "api_ref",
            "API Reference",
            f"{app_config.paths.library}/api_reference.txt",
            "API documentation",
        ),
        (
            "config",
            "Configuration File",
            f"{app_config.paths.library}/config.json",
            "App configuration",
        ),
        (
            "troubleshoot",
            "Troubleshooting Guide",
            f"{app_config.paths.library}/troubleshooting.txt",
            "Common issues and solutions",
        ),
        (
            "architecture",
            "System Architecture",
            f"{app_config.paths.library}/ARCHITECTURE.md",
            "Overall system architecture documentation",
        ),
        (
            "client_api_guide",
            "Client API User Guide",
            f"{app_config.paths.library}/CLIENT API USER GUIDE.md",
            "User guide for client API usage",
        ),
        (
            "clients_arch",
            "Clients Architecture",
            f"{app_config.paths.library}/CLIENTS_ARCHITECTURE.md",
            "Architecture for AI client implementations",
        ),
        (
            "conditional_guide",
            "Conditional Expressions Guide",
            f"{app_config.paths.library}/CONDITIONAL EXPRESSIONS USER GUIDE.md",
            "Guide for conditional expressions in prompts",
        ),
        (
            "configuration_doc",
            "Configuration Documentation",
            f"{app_config.paths.library}/CONFIGURATION.md",
            "Configuration management documentation",
        ),
        (
            "orchestrator_arch",
            "Orchestrator Architecture",
            f"{app_config.paths.library}/ORCHESTRATOR_ARCHITECTURE.md",
            "Excel orchestrator architecture documentation",
        ),
        (
            "orchestrator_readme",
            "Orchestrator README",
            f"{app_config.paths.library}/ORCHESTRATOR README.md",
            "Orchestrator usage and examples",
        ),
        (
            "rag_architecture",
            "RAG Architecture",
            f"{app_config.paths.library}/RAG_ARCHITECTURE.md",
            "Retrieval-Augmented Generation architecture",
        ),
        (
            "shared_history",
            "Shared History Design",
            f"{app_config.paths.library}/SHARED_HISTORY_DESIGN.md",
            "Conversation history sharing design",
        ),
    ]

    for row_idx, (ref_name, common_name, file_path, notes) in enumerate(documents, start=2):
        ws_documents.cell(row=row_idx, column=1, value=ref_name)
        ws_documents.cell(row=row_idx, column=2, value=common_name)
        ws_documents.cell(row=row_idx, column=3, value=file_path)
        ws_documents.cell(row=row_idx, column=4, value=notes)

    ws_documents.column_dimensions["A"].width = 18
    ws_documents.column_dimensions["B"].width = 25
    ws_documents.column_dimensions["C"].width = 50
    ws_documents.column_dimensions["D"].width = 30

    # ============================================
    # PROMPTS SHEET
    # ============================================
    ws_prompts = wb.create_sheet(title="prompts")
    headers = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "references",
        "semantic_query",
        "semantic_filter",
        "query_expansion",
        "rerank",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    prompts = []
    row = 2

    # ============================================
    # SECTION 1: Input Classification (1-3)
    # ============================================

    prompts.append(
        (
            1,
            "classify_sentiment",
            "Classify this product review as positive, negative, or neutral. "
            "Just respond with the category name.\n\nProduct: {{product_name}}\nReview: {{review_text}}",
            None,
            "fast",
            None,
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            2,
            "rate_urgency",
            "Rate the urgency of this review on a scale of 1-5 (1=low, 5=critical). "
            "Consider the priority level: {{priority}}. Just respond with the number.\n\nReview: {{review_text}}",
            None,
            "fast",
            None,
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            3,
            "detect_issues",
            "Analyze this review for any specific issues mentioned. "
            "If issues found, list them briefly. If no issues, say 'none'.\n\nReview: {{review_text}}",
            None,
            "analytical",
            None,
            None,
            None,
            None,
            None,
            None,
        )
    )

    # ============================================
    # SECTION 2: Conditional Branching (4-8)
    # ============================================

    prompts.append(
        (
            4,
            "positive_response",
            "Write a warm thank-you response for this positive review. One sentence.\n\nProduct: {{product_name}}",
            '["classify_sentiment"]',
            "creative",
            '"positive" in lower({{classify_sentiment.response}})',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            5,
            "negative_response",
            "Write a professional apology and resolution offer for this negative review. One sentence.\n\nProduct: {{product_name}}",
            '["classify_sentiment"]',
            "default",
            '"negative" in lower({{classify_sentiment.response}})',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            6,
            "neutral_response",
            "Write a polite acknowledgment for this neutral review. One sentence.\n\nProduct: {{product_name}}",
            '["classify_sentiment"]',
            "default",
            '"neutral" in lower({{classify_sentiment.response}})',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            7,
            "escalate_high",
            "Flag this as HIGH PRIORITY for immediate attention. Confirm with 'ESCALATED'.",
            '["rate_urgency"]',
            "analytical",
            '{{rate_urgency.response}} == "5" or {{rate_urgency.response}} == "4"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            8,
            "normal_priority",
            "Note: Standard response time applies. Confirm with 'QUEUED'.",
            '["rate_urgency"]',
            "default",
            '{{rate_urgency.response}} != "5" and {{rate_urgency.response}} != "4"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    # ============================================
    # SECTION 3: Issue Resolution (9-12)
    # ============================================

    prompts.append(
        (
            9,
            "generate_solution",
            "Based on the issues detected, suggest a brief resolution or workaround. "
            "Keep it practical and concise.\n\nIssues: {{detect_issues.response}}",
            '["detect_issues"]',
            "creative",
            'lower({{detect_issues.response}}) != "none" and len({{detect_issues.response}}) > 4',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            10,
            "no_issues_note",
            "No action needed - customer had no specific issues. Say 'No issues to address.'",
            '["detect_issues"]',
            "fast",
            'lower({{detect_issues.response}}) == "none" or len({{detect_issues.response}}) <= 4',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            11,
            "detailed_analysis",
            "Provide a detailed analysis of why this review needs attention. "
            "Consider sentiment, urgency, and issues. Be thorough but concise.",
            '["classify_sentiment", "rate_urgency", "detect_issues"]',
            "analytical",
            '{{rate_urgency.response}} == "5" or {{rate_urgency.response}} == "4"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            12,
            "brief_summary",
            "Summarize the customer feedback in one sentence.",
            '["classify_sentiment", "detect_issues"]',
            "default",
            '{{rate_urgency.response}} != "5" and {{rate_urgency.response}} != "4"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    # ============================================
    # SECTION 4: Response Assembly (13-16)
    # ============================================

    prompts.append(
        (
            13,
            "assemble_response",
            "Create the final customer response email. "
            "Include the appropriate sentiment response and any solutions if applicable.\n\n"
            "Use the sentiment response from previous steps.",
            '["positive_response", "negative_response", "neutral_response", "generate_solution"]',
            "creative",
            '{{positive_response.status}} == "success" or {{negative_response.status}} == "success" or {{neutral_response.status}} == "success"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            14,
            "priority_note",
            "Add the priority handling status to the response context.",
            '["escalate_high", "normal_priority"]',
            "default",
            '{{escalate_high.status}} == "success" or {{normal_priority.status}} == "success"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            15,
            "internal_notes",
            "Create internal notes for the support team about this review. "
            "Include analysis if available.",
            '["detailed_analysis", "brief_summary"]',
            "analytical",
            '{{detailed_analysis.status}} == "success" or {{brief_summary.status}} == "success"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            16,
            "skip_reason",
            "Note: Some steps were skipped due to conditional logic. Say 'Partial processing completed.'",
            '["assemble_response"]',
            "fast",
            '{{assemble_response.status}} != "success"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    # ============================================
    # SECTION 5: Final Reporting (17-20)
    # ============================================

    prompts.append(
        (
            17,
            "metrics",
            "Based on this review processing, provide: sentiment category, urgency level (1-5), "
            "and whether issues were detected. Format: 'Sentiment: X, Urgency: Y, Issues: Z'",
            '["classify_sentiment", "rate_urgency", "detect_issues"]',
            "analytical",
            None,
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            18,
            "quality_check",
            "Quality check: Was a complete response generated? Answer yes or no.",
            '["assemble_response"]',
            "fast",
            '{{assemble_response.status}} == "success"',
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            19,
            "batch_item_summary",
            "Create a one-line summary for this batch item: {{batch_name}} - {{product_name}}.",
            '["metrics", "quality_check"]',
            "creative",
            None,
            None,
            None,
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            20,
            "final_confirmation",
            "Processing complete for {{product_name}}. "
            "Confirm with 'DONE' and include the sentiment classification result.",
            '["batch_item_summary"]',
            "default",
            None,
            None,
            None,
            None,
            None,
            None,
        )
    )

    # ============================================
    # SECTION 6: RAG-Enhanced Analysis (21-27)
    # ============================================

    prompts.append(
        (
            21,
            "rag_product_context",
            "Search documentation for best practices related to: {{product_name}}. "
            "Summarize any relevant guidance for handling this product category.",
            None,
            "analytical",
            None,
            None,
            "{{product_name}} product handling guidance",
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            22,
            "rag_sentiment_guidance",
            "Find guidance for handling negative customer feedback. Product: {{product_name}}",
            '["classify_sentiment"]',
            "default",
            '"negative" in lower({{classify_sentiment.response}})',
            None,
            "negative feedback customer complaint handling",
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            23,
            "rag_troubleshooting_search",
            "Search troubleshooting documentation for solutions to: {{detect_issues.response}}",
            '["detect_issues"]',
            "default",
            'lower({{detect_issues.response}}) != "none"',
            None,
            "troubleshooting error solution fix",
            None,
            None,
            None,
        )
    )

    prompts.append(
        (
            24,
            "rag_filtered_search",
            "Find escalation procedures from API documentation. "
            "Product: {{product_name}}, Urgency: {{rate_urgency.response}}",
            '["rate_urgency"]',
            "fast",
            '{{rate_urgency.response}} == "5" or {{rate_urgency.response}} == "4"',
            None,
            "escalation priority urgent handling",
            '{"reference_name": "api_ref"}',
            None,
            None,
        )
    )

    prompts.append(
        (
            25,
            "rag_expanded_search",
            "Research comprehensive approaches for: {{product_name}} customer satisfaction. "
            "Review context: {{review_text}}",
            None,
            "creative",
            None,
            None,
            "customer satisfaction improvement quality",
            None,
            "true",
            None,
        )
    )

    prompts.append(
        (
            26,
            "rag_reranked_search",
            "Find the most relevant solutions for reported issues: {{detect_issues.response}}",
            '["classify_sentiment", "detect_issues"]',
            "analytical",
            '"negative" in lower({{classify_sentiment.response}}) and lower({{detect_issues.response}}) != "none"',
            None,
            "error fix solution troubleshooting",
            None,
            None,
            "true",
        )
    )

    prompts.append(
        (
            27,
            "rag_comprehensive",
            "Based on the full API reference and semantic search results, "
            "provide a final recommendation for {{product_name}}.",
            '["rag_reranked_search"]',
            "creative",
            None,
            '["api_ref"]',
            "{{product_name}} recommendation guidance",
            None,
            None,
            None,
        )
    )

    for (
        seq,
        name,
        prompt,
        history,
        client,
        condition,
        refs,
        semantic,
        sem_filter,
        query_exp,
        rerank,
    ) in prompts:
        ws_prompts.cell(row=row, column=1, value=seq)
        ws_prompts.cell(row=row, column=2, value=name)
        ws_prompts.cell(row=row, column=3, value=prompt)
        ws_prompts.cell(row=row, column=4, value=history if history else "")
        ws_prompts.cell(row=row, column=5, value=client if client else "")
        ws_prompts.cell(row=row, column=6, value=condition if condition else "")
        ws_prompts.cell(row=row, column=7, value=refs if refs else "")
        ws_prompts.cell(row=row, column=8, value=semantic if semantic else "")
        ws_prompts.cell(row=row, column=9, value=sem_filter if sem_filter else "")
        ws_prompts.cell(row=row, column=10, value=query_exp if query_exp else "")
        ws_prompts.cell(row=row, column=11, value=rerank if rerank else "")
        row += 1

    ws_prompts.column_dimensions["A"].width = 10
    ws_prompts.column_dimensions["B"].width = 22
    ws_prompts.column_dimensions["C"].width = 80
    ws_prompts.column_dimensions["D"].width = 50
    ws_prompts.column_dimensions["E"].width = 12
    ws_prompts.column_dimensions["F"].width = 60
    ws_prompts.column_dimensions["G"].width = 15
    ws_prompts.column_dimensions["H"].width = 35
    ws_prompts.column_dimensions["I"].width = 30
    ws_prompts.column_dimensions["J"].width = 15
    ws_prompts.column_dimensions["K"].width = 10

    wb.save(output_path)

    rag_prompts = [p for p in prompts if p[7]]
    filtered_rag = [p for p in prompts if p[8]]
    expanded_rag = [p for p in prompts if p[9]]
    reranked_rag = [p for p in prompts if p[10]]

    print(f"\n{'=' * 70}")
    print(f"Created MAX sample workbook: {output_path}")
    print(f"{'=' * 70}")
    print("\nUsing: FFLiteLLMClient with LiteLLM routing")

    print(f"\n{'=' * 70}")
    print("FEATURES COMBINED:")
    print(f"{'=' * 70}")

    print("\n1. BATCH MODE:")
    print(f"   - {len(batch_data)} data rows")
    print(f"   - Each row processes through all {len(prompts)} prompts")
    print(f"   - Total executions: {len(batch_data) * len(prompts)} prompt calls")

    print("\n2. MULTI-CLIENT:")
    for name, _, _, _model, temp, tokens in clients_data:
        count = sum(1 for p in prompts if p[4] == name)
        print(f"   - {name}: temp={temp}, tokens={tokens}, {count} prompts")

    print("\n3. CONDITIONAL EXECUTION:")
    cond_count = sum(1 for p in prompts if p[5])
    print(f"   - {cond_count} prompts with conditions")
    print("   - Conditions include: sentiment branching, urgency levels, issue detection")

    print("\n4. RAG SEMANTIC SEARCH:")
    print(f"   - {len(documents)} documents defined for semantic search")
    print(f"   - {len(rag_prompts)} prompts with semantic_query (RAG search)")
    print(f"   - {len(filtered_rag)} prompts with semantic_filter (targeted RAG)")
    print(f"   - {len(expanded_rag)} prompts with query_expansion (multi-query)")
    print(f"   - {len(reranked_rag)} prompts with rerank (cross-encoder reranking)")

    print(f"\n{'=' * 70}")
    print("DOCUMENTS DEFINED:")
    print(f"{'=' * 70}")
    for ref_name, common_name, _, _ in documents[:5]:
        print(f"  - {ref_name}: {common_name}")
    print(f"  ... and {len(documents) - 5} more documents")

    print(f"\n{'=' * 70}")
    print("BATCH DATA:")
    print(f"{'=' * 70}")
    for name, product, review, priority in batch_data:
        print(f"  {name}: {product} (priority={priority})")
        print(f"    Review: {review[:50]}...")

    print(f"\n{'=' * 70}")
    print("PROMPT STRUCTURE:")
    print(f"{'=' * 70}")
    print("  Section 1 (Seq 1-3):   Input Classification")
    print("  Section 2 (Seq 4-8):   Conditional Branching (sentiment + urgency)")
    print("  Section 3 (Seq 9-12):  Issue Resolution")
    print("  Section 4 (Seq 13-16): Response Assembly")
    print("  Section 5 (Seq 17-20): Final Reporting")
    print("  Section 6 (Seq 21-27): RAG-Enhanced Analysis (NEW)")

    print(f"\n{'=' * 70}")
    print(f"Run with: python scripts/run_orchestrator.py {output_path} -c 3")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    config = get_config()
    output = sys.argv[1] if len(sys.argv) > 1 else config.sample.workbooks.max
    create_max_sample_workbook(output)
