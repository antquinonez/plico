#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Validate conditional execution workbook results.

Validates workbooks created by sample_workbook_conditional_create_v001.py
by checking condition evaluation results and execution status across all sections.

Features:
    - Section-by-section validation with detailed output
    - Condition evaluation tracking
    - JSON output for programmatic use
    - Exit code 0 for pass, 1 for failures

Usage:
    python scripts/sample_workbook_conditional_validate_v001.py <workbook_path>
    python scripts/sample_workbook_conditional_validate_v001.py <workbook_path> --json
    python scripts/sample_workbook_conditional_validate_v001.py <workbook_path> --results-sheet results_20250228_123456

Version: 001
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

VERSION = "001"

SECTION_DEFINITIONS = {
    "Section 1 - String Methods": {
        "range": (1, 10),
        "features": ["startswith", "endswith", "lower", "strip", "count"],
    },
    "Section 2 - JSON Simple": {
        "range": (11, 18),
        "features": ["json_get", "json_has", "json_type"],
    },
    "Section 3 - JSON Nested": {
        "range": (19, 26),
        "features": ["nested paths", "json_get_default"],
    },
    "Section 4 - JSON Array": {
        "range": (27, 34),
        "features": ["array indexing", "json_keys", "in operator"],
    },
    "Section 5 - JSON Complex": {"range": (35, 38), "features": ["deep nesting", "mixed access"]},
    "Section 6 - Math Functions": {"range": (39, 44), "features": ["abs", "min", "max"]},
    "Section 7 - Type Checking": {"range": (45, 47), "features": ["is_empty"]},
    "Section 8 - Combined": {
        "range": (48, 50),
        "features": ["chained conditions", "boolean logic"],
    },
}


def find_latest_results_sheet(workbook) -> str | None:
    """Find the most recent results sheet by timestamp in name."""
    results_sheets = [name for name in workbook.sheetnames if name.startswith("results_")]
    if not results_sheets:
        return None
    sorted_sheets = sorted(results_sheets, reverse=True)
    return sorted_sheets[0]


def parse_sequence_to_row(ws) -> dict[int, int]:
    """Build mapping from sequence number to row number."""
    seq_to_row = {}
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=3).value
        if seq is not None:
            seq_to_row[int(seq)] = row
    return seq_to_row


def validate_section(
    ws,
    seq_to_row: dict[int, int],
    section_name: str,
    section_def: dict,
) -> dict:
    """Validate a single section and return results."""
    start, end = section_def["range"]
    prompts = []
    passed = 0
    skipped = 0
    failed = 0
    conditions_true = 0
    conditions_false = 0

    for seq in range(start, end + 1):
        if seq not in seq_to_row:
            continue

        row = seq_to_row[seq]
        prompt_name = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=12).value
        cond_result = ws.cell(row=row, column=9).value
        cond_error = ws.cell(row=row, column=10).value

        prompt_info = {
            "sequence": seq,
            "name": prompt_name,
            "status": status,
            "condition_result": cond_result,
            "condition_error": cond_error,
        }
        prompts.append(prompt_info)

        if status == "success":
            passed += 1
            if cond_result is True:
                conditions_true += 1
        elif status == "skipped":
            skipped += 1
            if cond_result is False:
                conditions_false += 1
        elif status == "failed":
            failed += 1

    all_passed = failed == 0

    return {
        "name": section_name,
        "features": section_def["features"],
        "range": [start, end],
        "passed": passed,
        "skipped": skipped,
        "failed": failed,
        "conditions_true": conditions_true,
        "conditions_false": conditions_false,
        "all_passed": all_passed,
        "prompts": prompts,
    }


def validate_workbook(path: Path, results_sheet: str | None = None) -> dict:
    """Validate workbook and return comprehensive results."""
    workbook = openpyxl.load_workbook(path, data_only=True)

    if results_sheet:
        if results_sheet not in workbook.sheetnames:
            return {
                "valid": False,
                "error": f"Results sheet '{results_sheet}' not found",
                "available_sheets": workbook.sheetnames,
            }
        sheet_name = results_sheet
    else:
        sheet_name = find_latest_results_sheet(workbook)
        if not sheet_name:
            return {
                "valid": False,
                "error": "No results sheet found in workbook",
                "available_sheets": workbook.sheetnames,
            }

    ws = workbook[sheet_name]
    seq_to_row = parse_sequence_to_row(ws)

    sections = []
    total_passed = 0
    total_skipped = 0
    total_failed = 0
    total_conditions_true = 0
    total_conditions_false = 0

    for section_name, section_def in SECTION_DEFINITIONS.items():
        section_result = validate_section(ws, seq_to_row, section_name, section_def)
        sections.append(section_result)

        total_passed += section_result["passed"]
        total_skipped += section_result["skipped"]
        total_failed += section_result["failed"]
        total_conditions_true += section_result["conditions_true"]
        total_conditions_false += section_result["conditions_false"]

    all_passed = total_failed == 0

    skipped_prompts = []
    for section in sections:
        for prompt in section["prompts"]:
            if prompt["status"] == "skipped":
                skipped_prompts.append(prompt["name"])

    return {
        "valid": True,
        "workbook": str(path),
        "results_sheet": sheet_name,
        "validated_at": datetime.now().isoformat(),
        "validator_version": VERSION,
        "summary": {
            "total_prompts": sum(s["passed"] + s["skipped"] + s["failed"] for s in sections),
            "passed": total_passed,
            "skipped": total_skipped,
            "failed": total_failed,
            "conditions_true": total_conditions_true,
            "conditions_false": total_conditions_false,
        },
        "sections": sections,
        "skipped_prompts": skipped_prompts,
        "all_passed": all_passed,
    }


def print_report(results: dict) -> None:
    """Print human-readable validation report."""
    print("=" * 80)
    print("CONDITIONAL WORKBOOK VALIDATION RESULTS")
    print("=" * 80)
    print(f"Workbook: {results['workbook']}")
    print(f"Results Sheet: {results['results_sheet']}")
    print(f"Validator Version: v{results['validator_version']}")
    print(f"Validated: {results['validated_at']}")
    print()

    if not results["valid"]:
        print(f"ERROR: {results['error']}")
        return

    summary = results["summary"]
    print(f"Total Prompts: {summary['total_prompts']}")
    print(f"Passed: {summary['passed']}")
    print(f"Skipped: {summary['skipped']}")
    print(f"Failed: {summary['failed']}")
    print(f"Conditions True: {summary['conditions_true']}")
    print(f"Conditions False: {summary['conditions_false']}")
    print()

    for section in results["sections"]:
        status_icon = "✓" if section["all_passed"] else "✗"
        print(f"\n{status_icon} {section['name']}:")
        print(f"    Features: {', '.join(section['features'])}")
        print(f"    Range: sequences {section['range'][0]}-{section['range'][1]}")
        print(
            f"    Results: {section['passed']} passed, {section['skipped']} skipped, {section['failed']} failed"
        )

        for prompt in section["prompts"]:
            if prompt["status"] == "success":
                if prompt["condition_result"] is True:
                    print(f"      ✓ {prompt['name']}: condition=True")
                else:
                    print(f"      ✓ {prompt['name']}")
            elif prompt["status"] == "skipped":
                print(f"      ⊘ {prompt['name']}: SKIPPED (condition={prompt['condition_result']})")
            elif prompt["status"] == "failed":
                print(f"      ✗ {prompt['name']}: FAILED")

    if results["skipped_prompts"]:
        print()
        print("Skipped prompts (condition evaluated to False):")
        for name in results["skipped_prompts"]:
            print(f"  - {name}")

    print()
    print("=" * 80)
    if results["all_passed"]:
        print("✅ ALL SECTIONS PASSED!")
    else:
        print("❌ SOME SECTIONS HAD FAILURES")
    print("=" * 80)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Validate conditional execution workbook results",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    %(prog)s ./sample_workbook_conditional.xlsx
    %(prog)s ./sample_workbook_conditional.xlsx --json
    %(prog)s ./sample_workbook_conditional.xlsx --results-sheet results_20250228_123456
        """,
    )
    parser.add_argument("workbook", type=Path, help="Path to the workbook to validate")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output results as JSON",
    )
    parser.add_argument(
        "--results-sheet",
        type=str,
        help="Specific results sheet to validate (default: latest)",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s v{VERSION}",
    )

    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Error: Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    results = validate_workbook(args.workbook, args.results_sheet)

    if args.json:
        print(json.dumps(results, indent=2, default=str))
    else:
        print_report(results)

    return 0 if results.get("all_passed", False) else 1


if __name__ == "__main__":
    sys.exit(main())
