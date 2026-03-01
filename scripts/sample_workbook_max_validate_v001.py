#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Validate max workbook results.

Validates workbooks created by sample_workbook_max_create_v001.py
by checking batch execution, conditional execution, multi-client usage, and RAG features.

Features:
    - Batch data execution verification
    - Conditional branching validation
    - Multi-client usage tracking
    - RAG semantic search validation
    - Section-by-section validation with detailed output
    - JSON output for programmatic use
    - Exit code 0 for pass, 1 for failures

Usage:
    python scripts/sample_workbook_max_validate_v001.py <workbook_path>
    python scripts/sample_workbook_max_validate_v001.py <workbook_path> --json
    python scripts/sample_workbook_max_validate_v001.py <workbook_path> --results-sheet results_20250228_123456

Version: 001
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

VERSION = "001"

SECTION_DEFINITIONS = {
    "Section 1 - Input Classification": {
        "range": (1, 3),
        "features": ["batch", "multi-client"],
    },
    "Section 2 - Conditional Branching": {
        "range": (4, 8),
        "features": ["batch", "conditional", "multi-client"],
    },
    "Section 3 - Issue Resolution": {
        "range": (9, 12),
        "features": ["batch", "conditional", "multi-client"],
    },
    "Section 4 - Response Assembly": {
        "range": (13, 16),
        "features": ["batch", "conditional", "multi-client"],
    },
    "Section 5 - Final Reporting": {
        "range": (17, 20),
        "features": ["batch", "multi-client"],
    },
    "Section 6 - RAG Enhanced Analysis": {
        "range": (21, 27),
        "features": ["batch", "conditional", "multi-client", "rag"],
    },
}

RAG_FEATURE_RANGES = {
    "semantic_query": (21, 27),
    "semantic_filter": (24, 24),
    "query_expansion": (25, 25),
    "rerank": (26, 26),
}


def find_latest_results_sheet(workbook) -> str | None:
    """Find the most recent results sheet by timestamp in name."""
    results_sheets = [name for name in workbook.sheetnames if name.startswith("results_")]
    if not results_sheets:
        return None
    sorted_sheets = sorted(results_sheets, reverse=True)
    return sorted_sheets[0]


def parse_sequence_to_row(ws) -> dict[int, int]:
    """Build mapping from sequence number to row number."""
    seq_to_row = {}
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=3).value
        if seq is not None:
            seq_to_row[int(seq)] = row
    return seq_to_row


def validate_section(
    ws,
    seq_to_row: dict[int, int],
    section_name: str,
    section_def: dict,
) -> dict:
    """Validate a single section and return results."""
    start, end = section_def["range"]
    prompts = []
    passed = 0
    skipped = 0
    failed = 0
    conditions_evaluated = 0
    rag_queries = 0

    for seq in range(start, end + 1):
        if seq not in seq_to_row:
            continue

        row = seq_to_row[seq]
        prompt_name = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=12).value
        cond_result = ws.cell(row=row, column=9).value
        semantic_query = ws.cell(row=row, column=16).value

        prompt_info = {
            "sequence": seq,
            "name": prompt_name,
            "status": status,
            "condition_result": cond_result,
            "has_rag": bool(semantic_query),
        }
        prompts.append(prompt_info)

        if status == "success":
            passed += 1
        elif status == "skipped":
            skipped += 1
        elif status == "failed":
            failed += 1

        if cond_result is not None:
            conditions_evaluated += 1

        if semantic_query:
            rag_queries += 1

    all_passed = failed == 0

    return {
        "name": section_name,
        "features": section_def["features"],
        "range": [start, end],
        "passed": passed,
        "skipped": skipped,
        "failed": failed,
        "conditions_evaluated": conditions_evaluated,
        "rag_queries": rag_queries,
        "all_passed": all_passed,
        "prompts": prompts,
    }


def validate_rag_features(ws, seq_to_row: dict[int, int]) -> dict:
    """Validate RAG feature usage across all prompts."""
    rag_stats = {
        "semantic_query": {"count": 0, "passed": 0, "failed": 0},
        "semantic_filter": {"count": 0, "passed": 0, "failed": 0},
        "query_expansion": {"count": 0, "passed": 0, "failed": 0},
        "rerank": {"count": 0, "passed": 0, "failed": 0},
    }

    col_mapping = {
        "semantic_query": 16,
        "semantic_filter": 17,
        "query_expansion": 18,
        "rerank": 19,
    }

    for feature, start_end in RAG_FEATURE_RANGES.items():
        start, end = start_end
        for seq in range(start, end + 1):
            if seq not in seq_to_row:
                continue

            row = seq_to_row[seq]
            value = ws.cell(row=row, column=col_mapping[feature]).value
            status = ws.cell(row=row, column=12).value

            if value:
                rag_stats[feature]["count"] += 1
                if status == "success":
                    rag_stats[feature]["passed"] += 1
                elif status == "failed":
                    rag_stats[feature]["failed"] += 1

    return rag_stats


def validate_workbook(path: Path, results_sheet: str | None = None) -> dict:
    """Validate workbook and return comprehensive results."""
    workbook = openpyxl.load_workbook(path, data_only=True)

    if results_sheet:
        if results_sheet not in workbook.sheetnames:
            return {
                "valid": False,
                "error": f"Results sheet '{results_sheet}' not found",
                "available_sheets": workbook.sheetnames,
            }
        sheet_name = results_sheet
    else:
        sheet_name = find_latest_results_sheet(workbook)
        if not sheet_name:
            return {
                "valid": False,
                "error": "No results sheet found in workbook",
                "available_sheets": workbook.sheetnames,
            }

    ws = workbook[sheet_name]
    seq_to_row = parse_sequence_to_row(ws)

    sections = []
    total_passed = 0
    total_skipped = 0
    total_failed = 0
    total_conditions = 0
    total_rag_queries = 0

    for section_name, section_def in SECTION_DEFINITIONS.items():
        section_result = validate_section(ws, seq_to_row, section_name, section_def)
        sections.append(section_result)

        total_passed += section_result["passed"]
        total_skipped += section_result["skipped"]
        total_failed += section_result["failed"]
        total_conditions += section_result["conditions_evaluated"]
        total_rag_queries += section_result["rag_queries"]

    rag_stats = validate_rag_features(ws, seq_to_row)

    all_passed = total_failed == 0

    return {
        "valid": True,
        "workbook": str(path),
        "results_sheet": sheet_name,
        "validated_at": datetime.now().isoformat(),
        "validator_version": VERSION,
        "summary": {
            "total_prompts": sum(s["passed"] + s["skipped"] + s["failed"] for s in sections),
            "passed": total_passed,
            "skipped": total_skipped,
            "failed": total_failed,
            "conditions_evaluated": total_conditions,
            "rag_queries": total_rag_queries,
        },
        "rag_stats": rag_stats,
        "sections": sections,
        "all_passed": all_passed,
    }


def print_report(results: dict) -> None:
    """Print human-readable validation report."""
    print("=" * 80)
    print("MAX WORKBOOK VALIDATION RESULTS")
    print("=" * 80)

    if not results.get("valid", False):
        print(f"ERROR: {results.get('error', 'Unknown error')}")
        print(f"Available sheets: {results.get('available_sheets', [])}")
        return

    print(f"Workbook: {results['workbook']}")
    print(f"Results Sheet: {results['results_sheet']}")
    print(f"Validator Version: v{results['validator_version']}")
    print(f"Validated: {results['validated_at']}")
    print()

    summary = results["summary"]
    print(f"Total Prompts: {summary['total_prompts']}")
    print(f"Passed: {summary['passed']}")
    print(f"Skipped: {summary['skipped']}")
    print(f"Failed: {summary['failed']}")
    print(f"Conditions Evaluated: {summary['conditions_evaluated']}")
    print(f"RAG Queries: {summary['rag_queries']}")
    print()

    rag_stats = results.get("rag_stats", {})
    if rag_stats:
        print("=" * 80)
        print("RAG FEATURE BREAKDOWN:")
        print("=" * 80)
        for feature, stats in rag_stats.items():
            if stats["count"] > 0:
                status = "✓" if stats["failed"] == 0 else "✗"
                print(f"  {status} {feature}: {stats['passed']}/{stats['count']} passed")
        print()

    for section in results["sections"]:
        status_icon = "✓" if section["all_passed"] else "✗"
        print(f"\n{status_icon} {section['name']}:")
        print(f"    Features: {', '.join(section['features'])}")
        print(f"    Range: sequences {section['range'][0]}-{section['range'][1]}")
        print(
            f"    Results: {section['passed']} passed, {section['skipped']} skipped, {section['failed']} failed"
        )
        if section["rag_queries"] > 0:
            print(f"    RAG Queries: {section['rag_queries']}")

        for prompt in section["prompts"]:
            rag_indicator = " [RAG]" if prompt["has_rag"] else ""
            if prompt["status"] == "success":
                print(f"      ✓ {prompt['name']}{rag_indicator}")
            elif prompt["status"] == "skipped":
                print(f"      ⊘ {prompt['name']}: SKIPPED{rag_indicator}")
            elif prompt["status"] == "failed":
                print(f"      ✗ {prompt['name']}: FAILED{rag_indicator}")

    print()
    print("=" * 80)
    if results["all_passed"]:
        print("✅ ALL SECTIONS PASSED!")
    else:
        print("❌ SOME SECTIONS HAD FAILURES")
    print("=" * 80)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Validate max workbook results",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    %(prog)s ./sample_workbook_max.xlsx
    %(prog)s ./sample_workbook_max.xlsx --json
    %(prog)s ./sample_workbook_max.xlsx --results-sheet results_20250228_123456
        """,
    )
    parser.add_argument("workbook", type=Path, help="Path to the workbook to validate")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output results as JSON",
    )
    parser.add_argument(
        "--results-sheet",
        type=str,
        help="Specific results sheet to validate (default: latest)",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s v{VERSION}",
    )

    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Error: Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    results = validate_workbook(args.workbook, args.results_sheet)

    if args.json:
        print(json.dumps(results, indent=2, default=str))
    else:
        print_report(results)

    return 0 if results.get("all_passed", False) else 1


if __name__ == "__main__":
    sys.exit(main())
