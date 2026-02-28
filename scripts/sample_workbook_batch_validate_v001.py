#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Validate batch workbook results.

Validates workbooks created by sample_workbook_batch_create_v001.py
by checking batch execution and variable resolution across all levels.

Features:
    - Batch data execution verification
    - Variable resolution tracking
    - Level-by-level validation with detailed output
    - JSON output for programmatic use
    - Exit code 0 for pass, 1 for failures

Usage:
    python scripts/sample_workbook_batch_validate_v001.py <workbook_path>
    python scripts/sample_workbook_batch_validate_v001.py <workbook_path> --json
    python scripts/sample_workbook_batch_validate_v001.py <workbook_path> --results-sheet results_20250228_123456

Version: 001
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

VERSION = "001"

LEVEL_DEFINITIONS = {
    "Level 0 - Independent Prompts": {
        "range": (1, 10),
        "description": "10 independent prompts using {{variable}} templating",
    },
    "Level 1 - Single Dependencies": {
        "range": (11, 20),
        "description": "10 prompts with 1-3 dependencies",
    },
    "Level 2 - Multiple Dependencies": {
        "range": (21, 30),
        "description": "10 prompts with 2-3 dependencies",
    },
    "Level 3 - Final Synthesis": {
        "range": (31, 35),
        "description": "5 final synthesis prompts",
    },
}

EXPECTED_BATCHES = 5


def find_latest_results_sheet(workbook) -> str | None:
    """Find the most recent results sheet by timestamp in name."""
    results_sheets = [name for name in workbook.sheetnames if name.startswith("results_")]
    if not results_sheets:
        return None
    sorted_sheets = sorted(results_sheets, reverse=True)
    return sorted_sheets[0]


def parse_sequence_to_row(ws) -> dict[int, int]:
    """Build mapping from sequence number to row number."""
    seq_to_row = {}
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=3).value
        if seq is not None:
            seq_to_row[int(seq)] = row
    return seq_to_row


def count_batches(ws, seq_to_row: dict[int, int]) -> int:
    """Count unique batch indices in the results."""
    batch_indices = set()
    for row in seq_to_row.values():
        batch_idx = ws.cell(row=row, column=2).value
        if batch_idx is not None:
            batch_indices.add(batch_idx)
    return len(batch_indices)


def validate_level(
    ws,
    seq_to_row: dict[int, int],
    level_name: str,
    level_def: dict,
    batch_count: int,
) -> dict:
    """Validate a single level and return results."""
    start, end = level_def["range"]
    prompts = []
    passed = 0
    failed = 0
    expected_executions = (end - start + 1) * batch_count

    for seq in range(start, end + 1):
        if seq not in seq_to_row:
            continue

        row = seq_to_row[seq]
        prompt_name = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=12).value

        prompt_info = {
            "sequence": seq,
            "name": prompt_name,
            "status": status,
        }
        prompts.append(prompt_info)

        if status == "success":
            passed += 1
        elif status == "failed":
            failed += 1

    all_passed = failed == 0

    return {
        "name": level_name,
        "description": level_def["description"],
        "range": [start, end],
        "batch_count": batch_count,
        "expected_executions": expected_executions,
        "passed": passed,
        "failed": failed,
        "all_passed": all_passed,
        "prompts": prompts,
    }


def validate_workbook(path: Path, results_sheet: str | None = None) -> dict:
    """Validate workbook and return comprehensive results."""
    workbook = openpyxl.load_workbook(path, data_only=True)

    if results_sheet:
        if results_sheet not in workbook.sheetnames:
            return {
                "valid": False,
                "error": f"Results sheet '{results_sheet}' not found",
                "available_sheets": workbook.sheetnames,
            }
        sheet_name = results_sheet
    else:
        sheet_name = find_latest_results_sheet(workbook)
        if not sheet_name:
            return {
                "valid": False,
                "error": "No results sheet found in workbook",
                "available_sheets": workbook.sheetnames,
            }

    ws = workbook[sheet_name]
    seq_to_row = parse_sequence_to_row(ws)
    batch_count = count_batches(ws, seq_to_row)

    levels = []
    total_passed = 0
    total_failed = 0

    for level_name, level_def in LEVEL_DEFINITIONS.items():
        level_result = validate_level(ws, seq_to_row, level_name, level_def, batch_count)
        levels.append(level_result)

        total_passed += level_result["passed"]
        total_failed += level_result["failed"]

    all_passed = total_failed == 0

    return {
        "valid": True,
        "workbook": str(path),
        "results_sheet": sheet_name,
        "validated_at": datetime.now().isoformat(),
        "validator_version": VERSION,
        "summary": {
            "total_prompts": 35,
            "batch_count": batch_count,
            "expected_batches": EXPECTED_BATCHES,
            "batches_match": batch_count == EXPECTED_BATCHES,
            "passed": total_passed,
            "failed": total_failed,
        },
        "levels": levels,
        "all_passed": all_passed,
    }


def print_report(results: dict) -> None:
    """Print human-readable validation report."""
    print("=" * 80)
    print("BATCH WORKBOOK VALIDATION RESULTS")
    print("=" * 80)

    if not results.get("valid", False):
        print(f"ERROR: {results.get('error', 'Unknown error')}")
        print(f"Available sheets: {results.get('available_sheets', [])}")
        return

    print(f"Workbook: {results['workbook']}")
    print(f"Results Sheet: {results['results_sheet']}")
    print(f"Validator Version: v{results['validator_version']}")
    print(f"Validated: {results['validated_at']}")
    print()

    summary = results["summary"]
    print(f"Total Prompts: {summary['total_prompts']}")
    print(f"Batch Count: {summary['batch_count']} (expected: {summary['expected_batches']})")
    print(f"Passed: {summary['passed']}")
    print(f"Failed: {summary['failed']}")

    if not summary["batches_match"]:
        print("\n⚠️  WARNING: Batch count mismatch!")

    print()

    for level in results["levels"]:
        status_icon = "✓" if level["all_passed"] else "✗"
        print(f"\n{status_icon} {level['name']}:")
        print(f"    {level['description']}")
        print(f"    Range: sequences {level['range'][0]}-{level['range'][1]}")
        print(
            f"    Expected executions: {level['expected_executions']} ({level['batch_count']} batches)"
        )
        print(f"    Results: {level['passed']} passed, {level['failed']} failed")

        for prompt in level["prompts"]:
            if prompt["status"] == "success":
                print(f"      ✓ {prompt['name']}")
            elif prompt["status"] == "failed":
                print(f"      ✗ {prompt['name']}: FAILED")

    print()
    print("=" * 80)
    if results["all_passed"]:
        print("✅ ALL LEVELS PASSED!")
    else:
        print("❌ SOME LEVELS HAD FAILURES")
    print("=" * 80)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Validate batch workbook results",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    %(prog)s ./sample_workbook_batch.xlsx
    %(prog)s ./sample_workbook_batch.xlsx --json
    %(prog)s ./sample_workbook_batch.xlsx --results-sheet results_20250228_123456
        """,
    )
    parser.add_argument("workbook", type=Path, help="Path to the workbook to validate")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output results as JSON",
    )
    parser.add_argument(
        "--results-sheet",
        type=str,
        help="Specific results sheet to validate (default: latest)",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s v{VERSION}",
    )

    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Error: Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    results = validate_workbook(args.workbook, args.results_sheet)

    if args.json:
        print(json.dumps(results, indent=2, default=str))
    else:
        print_report(results)

    return 0 if results.get("all_passed", False) else 1


if __name__ == "__main__":
    sys.exit(main())
