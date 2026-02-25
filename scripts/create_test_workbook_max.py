#!/usr/bin/env python
"""
Generate comprehensive test workbook combining batch, conditional, and multi-client features.

This workbook demonstrates the full power of FFClients orchestrator:
- BATCH: Multiple data rows processed through the same prompt chain
- CONDITIONAL: Prompts execute/skip based on runtime conditions
- MULTI-CLIENT: Different model configurations for different task types

Creates 20 prompts across 5 sections with 5 batch data rows.

Usage:
    python scripts/create_test_workbook_max.py [output_path]
"""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook


def create_max_test_workbook(output_path: str):
    wb = Workbook()

    # ============================================
    # CONFIG SHEET
    # ============================================
    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_data = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", "2"),
        ("temperature", "0.7"),
        ("max_tokens", "300"),
        (
            "system_instructions",
            "You are a helpful assistant. Give brief, concise answers. "
            "For classification, respond with just the category. "
            "For ratings, respond with just the number. "
            "For yes/no, respond with just 'yes' or 'no'.",
        ),
        ("created_at", datetime.now().isoformat()),
    ]

    for idx, (field, value) in enumerate(config_data, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_config.column_dimensions["A"].width = 20
    ws_config.column_dimensions["B"].width = 70

    # ============================================
    # CLIENTS SHEET
    # ============================================
    ws_clients = wb.create_sheet(title="clients")

    clients_headers = [
        "name",
        "client_type",
        "api_key_env",
        "model",
        "temperature",
        "max_tokens",
    ]
    for col_idx, header in enumerate(clients_headers, start=1):
        ws_clients.cell(row=1, column=col_idx, value=header)

    clients_data = [
        (
            "default",
            "mistral-small",
            "MISTRALSMALL_KEY",
            "mistral-small-2503",
            0.7,
            300,
        ),
        ("fast", "mistral-small", "MISTRALSMALL_KEY", "mistral-small-2503", 0.3, 150),
        (
            "creative",
            "mistral-small",
            "MISTRALSMALL_KEY",
            "mistral-small-2503",
            0.9,
            500,
        ),
        (
            "analytical",
            "mistral-small",
            "MISTRALSMALL_KEY",
            "mistral-small-2503",
            0.2,
            400,
        ),
    ]

    for row_idx, client_row in enumerate(clients_data, start=2):
        for col_idx, value in enumerate(client_row, start=1):
            ws_clients.cell(row=row_idx, column=col_idx, value=value)

    ws_clients.column_dimensions["A"].width = 12
    ws_clients.column_dimensions["B"].width = 15
    ws_clients.column_dimensions["C"].width = 18
    ws_clients.column_dimensions["D"].width = 20
    ws_clients.column_dimensions["E"].width = 12
    ws_clients.column_dimensions["F"].width = 12

    # ============================================
    # DATA SHEET (Batch data)
    # ============================================
    ws_data = wb.create_sheet(title="data")

    data_headers = ["batch_name", "product_name", "review_text", "priority"]
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)

    batch_data = [
        (
            "product_a",
            "Wireless Headphones",
            "The sound quality is amazing but the battery life could be better.",
            "high",
        ),
        (
            "product_b",
            "Coffee Maker",
            "Works perfectly every morning. Love the auto-brew feature!",
            "medium",
        ),
        (
            "product_c",
            "Running Shoes",
            "Uncomfortable after long runs. Disappointed with the fit.",
            "high",
        ),
        (
            "product_d",
            "Desk Lamp",
            "Good brightness levels. The adjustable arm is very useful.",
            "low",
        ),
        (
            "product_e",
            "Water Bottle",
            "Leaks when placed sideways. Cannot recommend.",
            "high",
        ),
    ]

    for row_idx, data_row in enumerate(batch_data, start=2):
        for col_idx, value in enumerate(data_row, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    ws_data.column_dimensions["A"].width = 12
    ws_data.column_dimensions["B"].width = 18
    ws_data.column_dimensions["C"].width = 60
    ws_data.column_dimensions["D"].width = 10

    # ============================================
    # PROMPTS SHEET
    # ============================================
    ws_prompts = wb.create_sheet(title="prompts")
    headers = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "references",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    prompts = []
    row = 2

    # ============================================
    # SECTION 1: Input Classification (1-3)
    # ============================================

    # Classify sentiment - fast for simple classification
    prompts.append(
        (
            1,
            "classify_sentiment",
            "Classify this product review as positive, negative, or neutral. "
            "Just respond with the category name.\n\nProduct: {{product_name}}\nReview: {{review_text}}",
            None,
            "fast",
            None,
        )
    )

    # Rate urgency - fast for rating
    prompts.append(
        (
            2,
            "rate_urgency",
            "Rate the urgency of this review on a scale of 1-5 (1=low, 5=critical). "
            "Consider the priority level: {{priority}}. Just respond with the number.\n\nReview: {{review_text}}",
            None,
            "fast",
            None,
        )
    )

    # Detect issues - analytical for thorough analysis
    prompts.append(
        (
            3,
            "detect_issues",
            "Analyze this review for any specific issues mentioned. "
            "If issues found, list them briefly. If no issues, say 'none'.\n\nReview: {{review_text}}",
            None,
            "analytical",
            None,
        )
    )

    # ============================================
    # SECTION 2: Conditional Branching (4-8)
    # ============================================

    # Positive response - creative for warm response
    prompts.append(
        (
            4,
            "positive_response",
            "Write a warm thank-you response for this positive review. One sentence.\n\nProduct: {{product_name}}",
            '["classify_sentiment"]',
            "creative",
            '"positive" in lower({{classify_sentiment.response}})',
        )
    )

    # Negative response - default for professional response
    prompts.append(
        (
            5,
            "negative_response",
            "Write a professional apology and resolution offer for this negative review. One sentence.\n\nProduct: {{product_name}}",
            '["classify_sentiment"]',
            "default",
            '"negative" in lower({{classify_sentiment.response}})',
        )
    )

    # Neutral response - default
    prompts.append(
        (
            6,
            "neutral_response",
            "Write a polite acknowledgment for this neutral review. One sentence.\n\nProduct: {{product_name}}",
            '["classify_sentiment"]',
            "default",
            '"neutral" in lower({{classify_sentiment.response}})',
        )
    )

    # High urgency escalation - analytical
    prompts.append(
        (
            7,
            "escalate_high",
            "Flag this as HIGH PRIORITY for immediate attention. Confirm with 'ESCALATED'.",
            '["rate_urgency"]',
            "analytical",
            '{{rate_urgency.response}} == "5" or {{rate_urgency.response}} == "4"',
        )
    )

    # Normal priority handling - default
    prompts.append(
        (
            8,
            "normal_priority",
            "Note: Standard response time applies. Confirm with 'QUEUED'.",
            '["rate_urgency"]',
            "default",
            '{{rate_urgency.response}} != "5" and {{rate_urgency.response}} != "4"',
        )
    )

    # ============================================
    # SECTION 3: Issue Resolution (9-12)
    # ============================================

    # Generate solution if issues detected - creative
    prompts.append(
        (
            9,
            "generate_solution",
            "Based on the issues detected, suggest a brief resolution or workaround. "
            "Keep it practical and concise.\n\nIssues: {{detect_issues.response}}",
            '["detect_issues"]',
            "creative",
            'lower({{detect_issues.response}}) != "none" and len({{detect_issues.response}}) > 4',
        )
    )

    # Skip if no issues
    prompts.append(
        (
            10,
            "no_issues_note",
            "No action needed - customer had no specific issues. Say 'No issues to address.'",
            '["detect_issues"]',
            "fast",
            'lower({{detect_issues.response}}) == "none" or len({{detect_issues.response}}) <= 4',
        )
    )

    # Detailed analysis for high priority items - analytical
    prompts.append(
        (
            11,
            "detailed_analysis",
            "Provide a detailed analysis of why this review needs attention. "
            "Consider sentiment, urgency, and issues. Be thorough but concise.",
            '["classify_sentiment", "rate_urgency", "detect_issues"]',
            "analytical",
            '{{rate_urgency.response}} == "5" or {{rate_urgency.response}} == "4"',
        )
    )

    # Brief summary for normal items - default
    prompts.append(
        (
            12,
            "brief_summary",
            "Summarize the customer feedback in one sentence.",
            '["classify_sentiment", "detect_issues"]',
            "default",
            '{{rate_urgency.response}} != "5" and {{rate_urgency.response}} != "4"',
        )
    )

    # ============================================
    # SECTION 4: Response Assembly (13-16)
    # ============================================

    # Assemble final response - creative for polished output
    prompts.append(
        (
            13,
            "assemble_response",
            "Create the final customer response email. "
            "Include the appropriate sentiment response and any solutions if applicable.\n\n"
            "Use the sentiment response from previous steps.",
            '["positive_response", "negative_response", "neutral_response", "generate_solution"]',
            "creative",
            '{{positive_response.status}} == "success" or {{negative_response.status}} == "success" or {{neutral_response.status}} == "success"',
        )
    )

    # Add priority handling note
    prompts.append(
        (
            14,
            "priority_note",
            "Add the priority handling status to the response context.",
            '["escalate_high", "normal_priority"]',
            "default",
            '{{escalate_high.status}} == "success" or {{normal_priority.status}} == "success"',
        )
    )

    # Internal notes - analytical
    prompts.append(
        (
            15,
            "internal_notes",
            "Create internal notes for the support team about this review. "
            "Include analysis if available.",
            '["detailed_analysis", "brief_summary"]',
            "analytical",
            '{{detailed_analysis.status}} == "success" or {{brief_summary.status}} == "success"',
        )
    )

    # Skip reason if something went wrong
    prompts.append(
        (
            16,
            "skip_reason",
            "Note: Some steps were skipped due to conditional logic. Say 'Partial processing completed.'",
            '["assemble_response"]',
            "fast",
            '{{assemble_response.status}} != "success"',
        )
    )

    # ============================================
    # SECTION 5: Final Reporting (17-20)
    # ============================================

    # Calculate metrics - analytical
    prompts.append(
        (
            17,
            "metrics",
            "Based on this review processing, provide: sentiment category, urgency level (1-5), "
            "and whether issues were detected. Format: 'Sentiment: X, Urgency: Y, Issues: Z'",
            '["classify_sentiment", "rate_urgency", "detect_issues"]',
            "analytical",
            None,
        )
    )

    # Quality check - fast
    prompts.append(
        (
            18,
            "quality_check",
            "Quality check: Was a complete response generated? Answer yes or no.",
            '["assemble_response"]',
            "fast",
            '{{assemble_response.status}} == "success"',
        )
    )

    # Batch summary - creative
    prompts.append(
        (
            19,
            "batch_item_summary",
            "Create a one-line summary for this batch item: {{batch_name}} - {{product_name}}.",
            '["metrics", "quality_check"]',
            "creative",
            None,
        )
    )

    # Final confirmation - default
    prompts.append(
        (
            20,
            "final_confirmation",
            "Processing complete for {{product_name}}. "
            "Confirm with 'DONE' and include the sentiment classification result.",
            '["batch_item_summary"]',
            "default",
            None,
        )
    )

    # Write all prompts to sheet
    for seq, name, prompt, history, client, condition in prompts:
        ws_prompts.cell(row=row, column=1, value=seq)
        ws_prompts.cell(row=row, column=2, value=name)
        ws_prompts.cell(row=row, column=3, value=prompt)
        ws_prompts.cell(row=row, column=4, value=history if history else "")
        ws_prompts.cell(row=row, column=5, value=client if client else "")
        ws_prompts.cell(row=row, column=6, value=condition if condition else "")
        ws_prompts.cell(row=row, column=7, value="")
        row += 1

    ws_prompts.column_dimensions["A"].width = 10
    ws_prompts.column_dimensions["B"].width = 22
    ws_prompts.column_dimensions["C"].width = 80
    ws_prompts.column_dimensions["D"].width = 50
    ws_prompts.column_dimensions["E"].width = 12
    ws_prompts.column_dimensions["F"].width = 60
    ws_prompts.column_dimensions["G"].width = 15

    wb.save(output_path)

    print(f"\n{'=' * 70}")
    print(f"Created MAX test workbook: {output_path}")
    print(f"{'=' * 70}")

    print(f"\n{'=' * 70}")
    print("FEATURES COMBINED:")
    print(f"{'=' * 70}")

    print(f"\n1. BATCH MODE:")
    print(f"   - {len(batch_data)} data rows")
    print(f"   - Each row processes through all 20 prompts")
    print(f"   - Total executions: {len(batch_data) * len(prompts)} prompt calls")

    print(f"\n2. MULTI-CLIENT:")
    for name, _, _, model, temp, tokens in clients_data:
        count = sum(1 for p in prompts if p[4] == name)
        print(f"   - {name}: temp={temp}, tokens={tokens}, {count} prompts")

    print(f"\n3. CONDITIONAL EXECUTION:")
    cond_count = sum(1 for p in prompts if p[5])
    print(f"   - {cond_count} prompts with conditions")
    print(
        f"   - Conditions include: sentiment branching, urgency levels, issue detection"
    )

    print(f"\n{'=' * 70}")
    print("BATCH DATA:")
    print(f"{'=' * 70}")
    for name, product, review, priority in batch_data:
        print(f"  {name}: {product} (priority={priority})")
        print(f"    Review: {review[:50]}...")

    print(f"\n{'=' * 70}")
    print("PROMPT STRUCTURE:")
    print(f"{'=' * 70}")
    print(f"  Section 1 (Seq 1-3):  Input Classification")
    print(f"  Section 2 (Seq 4-8):  Conditional Branching (sentiment + urgency)")
    print(f"  Section 3 (Seq 9-12): Issue Resolution")
    print(f"  Section 4 (Seq 13-16): Response Assembly")
    print(f"  Section 5 (Seq 17-20): Final Reporting")

    print(f"\n{'=' * 70}")
    print(f"Run with: python scripts/run_orchestrator.py {output_path} -c 3")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "test_workbook_max.xlsx"
    create_max_test_workbook(output)
