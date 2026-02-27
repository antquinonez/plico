#!/usr/bin/env python
"""
Generate test workbook for parallel execution testing.

Creates 31 prompts with various dependency patterns:
- Level 0: 12 independent prompts (fully parallel)
- Level 1: 10 prompts with 1-2 dependencies each
- Level 2: 5 prompts with 2-4 dependencies each
- Level 3: 4 final synthesis prompts

Uses FFLiteLLMClient with LiteLLM routing for Mistral Small.

Usage:
    python scripts/create_test_workbook.py [output_path]
"""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from src.config import get_config


def create_test_workbook(output_path: str):
    config = get_config()
    test_config = config.test

    wb = Workbook()

    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_data = [
        ("model", test_config.default_model),
        ("max_retries", str(test_config.default_retries)),
        ("temperature", str(test_config.default_temperature)),
        ("max_tokens", str(test_config.default_max_tokens)),
        (
            "system_instructions",
            test_config.default_system_instructions,
        ),
        ("created_at", datetime.now().isoformat()),
    ]

    for idx, (field, value) in enumerate(config_data, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_config.column_dimensions["A"].width = 20
    ws_config.column_dimensions["B"].width = 70

    ws_prompts = wb.create_sheet(title="prompts")
    headers = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "references",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    prompts = []
    row = 2

    # ============================================
    # LEVEL 0: 12 Independent Prompts (Parallel)
    # ============================================

    # Math facts (6 prompts)
    for i in range(1, 7):
        prompts.append((i, f"math_{i}", f"What is {i} + {i}? Just give the number.", None))

    # Word facts (6 prompts)
    words = ["apple", "banana", "cherry", "dog", "elephant", "flower"]
    for i, word in enumerate(words, 7):
        prompts.append(
            (
                i,
                f"word_{i}",
                f"How many letters are in the word '{word}'? Just give the number.",
                None,
            )
        )

    # ============================================
    # LEVEL 1: 10 Prompts with Dependencies
    # ============================================

    # Double the math results (6 prompts)
    for i in range(1, 7):
        prompts.append(
            (
                12 + i,
                f"double_{i}",
                f"Take the result from math_{i} and multiply it by 2. Just give the number.",
                f'["math_{i}"]',
            )
        )

    # Compare adjacent word lengths (4 prompts)
    for i in range(7, 11):
        prompts.append(
            (
                12 + i,
                f"compare_{i}",
                f"Is the word length from word_{i} greater than word_{i + 1}? Answer yes or no.",
                f'["word_{i}", "word_{i + 1}"]',
            )
        )

    # ============================================
    # LEVEL 2: 5 Prompts with Multiple Dependencies
    # ============================================

    # Sum doubles (3 prompts)
    prompts.append(
        (
            23,
            "sum_1",
            "Add the results from double_1 and double_2. Just give the number.",
            '["double_1", "double_2"]',
        )
    )
    prompts.append(
        (
            24,
            "sum_2",
            "Add the results from double_3 and double_4. Just give the number.",
            '["double_3", "double_4"]',
        )
    )
    prompts.append(
        (
            25,
            "sum_3",
            "Add the results from double_5 and double_6. Just give the number.",
            '["double_5", "double_6"]',
        )
    )

    # Triple check (1 prompt)
    prompts.append(
        (
            26,
            "triple_check",
            "Add results from sum_1, sum_2, and sum_3. Just give the number.",
            '["sum_1", "sum_2", "sum_3"]',
        )
    )

    # Word analysis (1 prompt)
    prompts.append(
        (
            27,
            "word_analysis",
            "Based on compare_7, compare_8, compare_9, and compare_10, are most answers yes or no?",
            '["compare_7", "compare_8", "compare_9", "compare_10"]',
        )
    )

    # ============================================
    # LEVEL 3: 4 Final Prompts
    # ============================================

    prompts.append(
        (
            28,
            "final_math",
            "What is the final number from triple_check divided by 3? Just give the number.",
            '["triple_check"]',
        )
    )

    prompts.append(
        (
            29,
            "final_words",
            "Summarize: Are longer words typically before or after shorter words alphabetically? Based on word_analysis.",
            '["word_analysis"]',
        )
    )

    # Independent bonus prompts (parallel with entire chain)
    prompts.append((30, "bonus_math", "What is 100 divided by 4? Just give the number.", None))

    prompts.append((31, "bonus_fact", "Name one interesting fact about the number 42.", None))

    # Write all prompts to sheet
    for seq, name, prompt, history in prompts:
        ws_prompts.cell(row=row, column=1, value=seq)
        ws_prompts.cell(row=row, column=2, value=name)
        ws_prompts.cell(row=row, column=3, value=prompt)
        ws_prompts.cell(row=row, column=4, value=history if history else "")
        ws_prompts.cell(row=row, column=5, value="")
        ws_prompts.cell(row=row, column=6, value="")
        ws_prompts.cell(row=row, column=7, value="")
        row += 1

    ws_prompts.column_dimensions["A"].width = 10
    ws_prompts.column_dimensions["B"].width = 18
    ws_prompts.column_dimensions["C"].width = 75
    ws_prompts.column_dimensions["D"].width = 45
    ws_prompts.column_dimensions["E"].width = 12
    ws_prompts.column_dimensions["F"].width = 25
    ws_prompts.column_dimensions["G"].width = 25

    wb.save(output_path)

    print(f"\n{'=' * 60}")
    print(f"Created test workbook: {output_path}")
    print(f"{'=' * 60}")
    print(f"\nUsing: FFLiteLLMClient with LiteLLM routing")
    print(f"Total prompts: {len(prompts)}")
    print("\nDependency Structure:")
    print("  Level 0: 12 independent prompts (sequences 1-12)")
    print("    - 6 math prompts: math_1 to math_6")
    print("    - 6 word prompts: word_7 to word_12")
    print("\n  Level 1: 10 prompts with 1-2 dependencies (sequences 13-22)")
    print("    - 6 double prompts: double_1 to double_6")
    print("    - 4 compare prompts: compare_7 to compare_10")
    print("\n  Level 2: 5 prompts with 2-4 dependencies (sequences 23-27)")
    print("    - 3 sum prompts: sum_1 to sum_3")
    print("    - 1 triple_check prompt")
    print("    - 1 word_analysis prompt")
    print("\n  Level 3: 4 final prompts (sequences 28-31)")
    print("    - 2 synthesis prompts: final_math, final_words")
    print("    - 2 independent bonus prompts: bonus_math, bonus_fact")
    print(f"\n{'=' * 60}")
    print(f"Run with: python scripts/run_orchestrator.py {output_path} -c 3")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    config = get_config()
    output = sys.argv[1] if len(sys.argv) > 1 else config.test.workbooks.basic
    create_test_workbook(output)
