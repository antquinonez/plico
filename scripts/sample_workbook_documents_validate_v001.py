#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Validate documents workbook results.

Validates workbooks created by sample_workbook_documents_create_v001.py
by checking document references and RAG semantic search execution.

Features:
    - Document reference injection verification
    - RAG semantic query validation
    - Prompt type breakdown (references vs semantic_query)
    - JSON output for programmatic use
    - Exit code 0 for pass, 1 for failures

Usage:
    python scripts/sample_workbook_documents_validate_v001.py <workbook_path>
    python scripts/sample_workbook_documents_validate_v001.py <workbook_path> --json
    python scripts/sample_workbook_documents_validate_v001.py <workbook_path> --results-sheet results_20250228_123456

Version: 001
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

VERSION = "001"

PROMPT_TYPES = {
    "Full Document References": {
        "range": (1, 6),
        "description": "Prompts with full document injection via references column",
    },
    "No Reference Control": {
        "range": (7, 7),
        "description": "Control prompt without references or semantic search",
    },
    "RAG Semantic Search": {
        "range": (8, 20),
        "description": "Prompts using RAG semantic search via semantic_query column",
    },
}


def find_latest_results_sheet(workbook) -> str | None:
    """Find the most recent results sheet by timestamp in name."""
    results_sheets = [name for name in workbook.sheetnames if name.startswith("results_")]
    if not results_sheets:
        return None
    sorted_sheets = sorted(results_sheets, reverse=True)
    return sorted_sheets[0]


def parse_sequence_to_row(ws) -> dict[int, int]:
    """Build mapping from sequence number to row number."""
    seq_to_row = {}
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=3).value
        if seq is not None:
            seq_to_row[int(seq)] = row
    return seq_to_row


def validate_prompt_type(
    ws,
    seq_to_row: dict[int, int],
    type_name: str,
    type_def: dict,
) -> dict:
    """Validate a single prompt type and return results."""
    start, end = type_def["range"]
    prompts = []
    passed = 0
    failed = 0

    for seq in range(start, end + 1):
        if seq not in seq_to_row:
            continue

        row = seq_to_row[seq]
        prompt_name = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=12).value

        prompt_info = {
            "sequence": seq,
            "name": prompt_name,
            "status": status,
        }
        prompts.append(prompt_info)

        if status == "success":
            passed += 1
        elif status == "failed":
            failed += 1

    all_passed = failed == 0

    return {
        "name": type_name,
        "description": type_def["description"],
        "range": [start, end],
        "passed": passed,
        "failed": failed,
        "all_passed": all_passed,
        "prompts": prompts,
    }


def validate_workbook(path: Path, results_sheet: str | None = None) -> dict:
    """Validate workbook and return comprehensive results."""
    workbook = openpyxl.load_workbook(path, data_only=True)

    if results_sheet:
        if results_sheet not in workbook.sheetnames:
            return {
                "valid": False,
                "error": f"Results sheet '{results_sheet}' not found",
                "available_sheets": workbook.sheetnames,
            }
        sheet_name = results_sheet
    else:
        sheet_name = find_latest_results_sheet(workbook)
        if not sheet_name:
            return {
                "valid": False,
                "error": "No results sheet found in workbook",
                "available_sheets": workbook.sheetnames,
            }

    ws = workbook[sheet_name]
    seq_to_row = parse_sequence_to_row(ws)

    prompt_types = []
    total_passed = 0
    total_failed = 0

    for type_name, type_def in PROMPT_TYPES.items():
        type_result = validate_prompt_type(ws, seq_to_row, type_name, type_def)
        prompt_types.append(type_result)

        total_passed += type_result["passed"]
        total_failed += type_result["failed"]

    all_passed = total_failed == 0

    return {
        "valid": True,
        "workbook": str(path),
        "results_sheet": sheet_name,
        "validated_at": datetime.now().isoformat(),
        "validator_version": VERSION,
        "summary": {
            "total_prompts": sum(t["passed"] + t["failed"] for t in prompt_types),
            "passed": total_passed,
            "failed": total_failed,
        },
        "prompt_types": prompt_types,
        "all_passed": all_passed,
    }


def print_report(results: dict) -> None:
    """Print human-readable validation report."""
    print("=" * 80)
    print("DOCUMENTS WORKBOOK VALIDATION RESULTS")
    print("=" * 80)

    if not results.get("valid", False):
        print(f"ERROR: {results.get('error', 'Unknown error')}")
        print(f"Available sheets: {results.get('available_sheets', [])}")
        return

    print(f"Workbook: {results['workbook']}")
    print(f"Results Sheet: {results['results_sheet']}")
    print(f"Validator Version: v{results['validator_version']}")
    print(f"Validated: {results['validated_at']}")
    print()

    summary = results["summary"]
    print(f"Total Prompts: {summary['total_prompts']}")
    print(f"Passed: {summary['passed']}")
    print(f"Failed: {summary['failed']}")
    print()

    for prompt_type in results["prompt_types"]:
        status_icon = "✓" if prompt_type["all_passed"] else "✗"
        print(f"\n{status_icon} {prompt_type['name']}:")
        print(f"    {prompt_type['description']}")
        print(f"    Range: sequences {prompt_type['range'][0]}-{prompt_type['range'][1]}")
        print(f"    Results: {prompt_type['passed']} passed, {prompt_type['failed']} failed")

        for prompt in prompt_type["prompts"]:
            if prompt["status"] == "success":
                print(f"      ✓ {prompt['name']}")
            elif prompt["status"] == "failed":
                print(f"      ✗ {prompt['name']}: FAILED")

    print()
    print("=" * 80)
    if results["all_passed"]:
        print("✅ ALL PROMPTS PASSED!")
    else:
        print("❌ SOME PROMPTS FAILED")
    print("=" * 80)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Validate documents workbook results",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    %(prog)s ./sample_workbook_documents.xlsx
    %(prog)s ./sample_workbook_documents.xlsx --json
    %(prog)s ./sample_workbook_documents.xlsx --results-sheet results_20250228_123456
        """,
    )
    parser.add_argument("workbook", type=Path, help="Path to the workbook to validate")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output results as JSON",
    )
    parser.add_argument(
        "--results-sheet",
        type=str,
        help="Specific results sheet to validate (default: latest)",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s v{VERSION}",
    )

    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Error: Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    results = validate_workbook(args.workbook, args.results_sheet)

    if args.json:
        print(json.dumps(results, indent=2, default=str))
    else:
        print_report(results)

    return 0 if results.get("all_passed", False) else 1


if __name__ == "__main__":
    sys.exit(main())
