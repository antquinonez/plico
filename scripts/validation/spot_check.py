#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Spot check responses from sample workbooks."""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from openpyxl import load_workbook

from src.config import get_config


def spot_check_responses(path: str, prompts_to_check: list[str], verbose: bool = False) -> None:
    """Spot check specific prompts in a workbook."""
    wb = load_workbook(path)
    results_sheets = [s for s in wb.sheetnames if s.startswith("results_")]

    if not results_sheets:
        print(f"No results sheet found in {path}")
        return

    ws = wb[results_sheets[-1]]
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        prompt_name = row[col["prompt_name"]]
        if prompt_name in prompts_to_check:
            response = row[col["response"]]
            status = row[col["status"]]
            print(f"\n{prompt_name} ({status}):")
            if response:
                preview = str(response)[:200] + "..." if len(str(response)) > 200 else response
                print(f"  {preview}")
            else:
                print("  (no response)")


def main():
    parser = argparse.ArgumentParser(description="Spot check responses from sample workbooks")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose output")
    args = parser.parse_args()

    config = get_config()
    workbooks = config.test.workbooks

    print("=" * 80)
    print("SPOT CHECK: Basic Workbook")
    print("=" * 80)
    spot_check_responses(workbooks.basic, ["math_1", "double_1", "final_math"], args.verbose)

    print("\n" + "=" * 80)
    print("SPOT CHECK: Conditional Workbook")
    print("=" * 80)
    spot_check_responses(
        workbooks.conditional,
        ["fetch_data", "process_success", "classify_sentiment", "positive_response"],
        args.verbose,
    )

    print("\n" + "=" * 80)
    print("SPOT CHECK: Max Workbook")
    print("=" * 80)
    spot_check_responses(
        workbooks.max,
        ["classify_sentiment", "rate_urgency", "final_confirmation"],
        args.verbose,
    )


if __name__ == "__main__":
    main()
