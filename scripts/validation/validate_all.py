#!/usr/bin/env python3
"""Validate all test workbooks after orchestration."""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from openpyxl import load_workbook

from src.config import get_config


def validate_workbook(
    path: str, name: str, verbose: bool = False
) -> tuple[int, int, int, list[str]]:
    """Validate a single workbook.

    Returns:
        Tuple of (success, failed, skipped, issues)
    """
    wb = load_workbook(path)
    results_sheets = [s for s in wb.sheetnames if s.startswith("results_")]

    if not results_sheets:
        return 0, 0, 0, [f"No results sheet found in {path}"]

    ws = wb[results_sheets[-1]]
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    success = 0
    failed = 0
    skipped = 0
    issues = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        status = row[col["status"]]
        prompt_name = row[col["prompt_name"]]
        cond_error = row[col["condition_error"]]

        if status == "success":
            success += 1
        elif status == "failed":
            failed += 1
            error = row[col["error"]]
            issues.append(f"FAILED: {prompt_name} - {error}")
        elif status == "skipped":
            skipped += 1
            if cond_error:
                issues.append(f"CONDITION ERROR: {prompt_name} - {cond_error}")

    if verbose and success > 0:
        print(f"  ✓ {name}: {success} success, {failed} failed, {skipped} skipped")

    return success, failed, skipped, issues


def validate_skipped_conditions(path: str, verbose: bool = False) -> list[dict]:
    """Validate that skipped prompts have correct condition results.

    Returns:
        List of dicts with skipped prompt info
    """
    wb = load_workbook(path)
    results_sheets = [s for s in wb.sheetnames if s.startswith("results_")]

    if not results_sheets:
        return []

    ws = wb[results_sheets[-1]]
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    skipped_info = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col["status"]] == "skipped":
            info = {
                "name": row[col["prompt_name"]],
                "condition": row[col["condition"]],
                "cond_result": row[col["condition_result"]],
                "cond_error": row[col["condition_error"]],
            }
            skipped_info.append(info)

    return skipped_info


def check_variable_substitution(path: str, verbose: bool = False) -> list[str]:
    """Check that batch variables are properly substituted.

    Returns:
        List of issues (empty if all OK)
    """
    wb = load_workbook(path)
    results_sheets = [s for s in wb.sheetnames if s.startswith("results_")]

    if not results_sheets:
        return []

    ws = wb[results_sheets[-1]]
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    issues = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        prompt = row[col["prompt"]]
        if prompt and "{{" in prompt and "}}" in prompt:
            issues.append(f"Unsubstituted variable in: {row[col['prompt_name']]}")

    if verbose and not issues:
        print("  ✓ All variables correctly substituted")

    return issues


def main():
    parser = argparse.ArgumentParser(description="Validate all test workbooks")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose output")
    args = parser.parse_args()

    config = get_config()
    workbooks = config.test.workbooks

    workbook_configs = [
        (workbooks.basic, "Basic", 31),
        (workbooks.multiclient, "Multi-client", 13),
        (workbooks.conditional, "Conditional", 30),
        (workbooks.documents, "Documents", 7),
        (workbooks.batch, "Batch", 175),
        (workbooks.max, "Max", 100),
    ]

    all_issues = []

    print("\n" + "=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80 + "\n")

    print(f"{'Workbook':<25} {'Success':>8} {'Failed':>8} {'Skipped':>8} {'Status':<10}")
    print("-" * 70)

    total_success = 0
    total_failed = 0
    total_skipped = 0

    for path, name, expected in workbook_configs:
        if not Path(path).exists():
            print(f"{name:<25} {'MISSING':>8} {'':>8} {'':>8} {'✗ FAIL':<10}")
            all_issues.append(f"{name}: workbook not found at {path}")
            continue

        success, failed, skipped, issues = validate_workbook(path, name, args.verbose)
        total_success += success
        total_failed += failed
        total_skipped += skipped
        all_issues.extend(issues)

        status = "✓ PASS" if failed == 0 and not issues else "✗ FAIL"
        print(f"{name:<25} {success:>8} {failed:>8} {skipped:>8} {status:<10}")

    print("-" * 70)
    print(f"{'TOTAL':<25} {total_success:>8} {total_failed:>8} {total_skipped:>8}")

    if all_issues:
        print(f"\n{'=' * 80}")
        print(f"ISSUES FOUND ({len(all_issues)}):")
        print("=" * 80)
        for issue in all_issues:
            print(f"  - {issue}")
        sys.exit(1)
    else:
        print("\n✓ All validations passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
