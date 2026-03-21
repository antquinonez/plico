#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Workbook builder for creating sample Excel workbooks.

This module provides the WorkbookBuilder class for creating test workbooks
with config, prompts, clients, documents, and data sheets.

Usage:
    from sample_workbooks import WorkbookBuilder, PromptSpec

    builder = WorkbookBuilder("output.xlsx")
    builder.add_config_sheet()
    builder.add_prompts_sheet([PromptSpec(...), ...])
    builder.save()
"""

from __future__ import annotations

import os
import sys
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from src.config import get_config

from .base import (
    DEFAULT_CLIENTS_COLUMN_WIDTHS,
    DEFAULT_CLIENTS_HEADERS,
    DEFAULT_CONFIG_COLUMN_WIDTHS,
    DEFAULT_CONFIG_FIELDS,
    DEFAULT_DOCUMENTS_COLUMN_WIDTHS,
    DEFAULT_DOCUMENTS_HEADERS,
    DEFAULT_PROMPT_COLUMN_WIDTHS,
    DEFAULT_PROMPT_HEADERS,
    PromptSpec,
)


class WorkbookBuilder:
    """Builds Excel workbooks for orchestrator testing.

    This class provides a fluent interface for creating workbooks with
    standard sheets (config, prompts, clients, documents, data).

    Example:
        builder = WorkbookBuilder("test.xlsx")
        builder.add_config_sheet(
            overrides={"system_instructions": "Custom instructions"}
        ).add_prompts_sheet(prompts).add_clients_sheet().save()

    """

    def __init__(self, output_path: str, config: Any = None):
        """Initialize the workbook builder.

        Args:
            output_path: Path where the workbook will be saved
            config: Optional config object (defaults to get_config())

        """
        self.output_path = output_path
        self.config = config or get_config()
        self.sample_config = self.config.sample
        self.wb = Workbook()
        self._prompts_count = 0

    def add_config_sheet(
        self,
        overrides: dict[str, Any] | None = None,
        extra_fields: list[tuple[str, str]] | None = None,
    ) -> WorkbookBuilder:
        """Add the config sheet to the workbook.

        Args:
            overrides: Dict of field -> value to override defaults
            extra_fields: Additional (field, value) tuples to append

        Returns:
            self for method chaining

        """
        ws = self.wb.active
        ws.title = "config"
        ws["A1"] = "field"
        ws["B1"] = "value"

        config_data = []
        overrides = overrides or {}
        for field_name, config_attr in DEFAULT_CONFIG_FIELDS:
            if config_attr is None:
                value = datetime.now().isoformat() if field_name == "created_at" else ""
            else:
                value = overrides.get(field_name, getattr(self.sample_config, config_attr, ""))
                if not isinstance(value, str):
                    value = str(value)
            config_data.append((field_name, value))

        if extra_fields:
            config_data.extend(extra_fields)

        for idx, (field, value) in enumerate(config_data, start=2):
            ws[f"A{idx}"] = field
            ws[f"B{idx}"] = value

        for col, width in DEFAULT_CONFIG_COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

        return self

    def add_prompts_sheet(
        self,
        prompts: list[PromptSpec],
        headers: list[str] | None = None,
        column_widths: dict[str, int] | None = None,
        include_extra_columns: bool = True,
    ) -> WorkbookBuilder:
        """Add the prompts sheet to the workbook.

        Args:
            prompts: List of PromptSpec objects
            headers: Custom headers (defaults to DEFAULT_PROMPT_HEADERS)
            column_widths: Custom column widths (defaults to DEFAULT_PROMPT_COLUMN_WIDTHS)
            include_extra_columns: Whether to include semantic_query, etc.

        Returns:
            self for method chaining

        """
        ws = self.wb.create_sheet(title="prompts")

        if headers is None:
            if include_extra_columns:
                headers = DEFAULT_PROMPT_HEADERS
            else:
                headers = DEFAULT_PROMPT_HEADERS[:7]

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, prompt in enumerate(prompts, start=2):
            row_data = prompt.to_row()
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        widths = column_widths or DEFAULT_PROMPT_COLUMN_WIDTHS
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        self._prompts_count = len(prompts)
        return self

    def add_clients_sheet(
        self,
        clients: list[dict] | None = None,
        client_names: list[str] | None = None,
        headers: list[str] | None = None,
        column_widths: dict[str, int] | None = None,
        sample_clients_overrides: dict[str, dict[str, Any]] | None = None,
    ) -> WorkbookBuilder:
        """Add the clients sheet to the workbook.

        Args:
            clients: List of client dicts with keys matching headers
            client_names: Named clients to include from config (e.g., ["default", "fast"])
            headers: Custom headers (defaults to DEFAULT_CLIENTS_HEADERS)
            column_widths: Custom column widths
            sample_clients_overrides: Override dict for sample_clients (from build_sample_clients_overrides)

        Returns:
            self for method chaining

        """
        ws = self.wb.create_sheet(title="clients")

        if headers is None:
            headers = DEFAULT_CLIENTS_HEADERS

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        if clients is None:
            clients = []
            names_to_use = client_names or ["default", "fast", "creative"]

            if sample_clients_overrides:
                sample_clients = sample_clients_overrides
            else:
                sample_clients = {}
                for name in names_to_use:
                    if name in self.sample_config.sample_clients:
                        cfg = self.sample_config.sample_clients[name]
                        sample_clients[name] = {
                            "client_type": cfg.client_type,
                            "api_key_env": cfg.api_key_env,
                            "model": cfg.model,
                            "temperature": cfg.temperature,
                            "max_tokens": cfg.max_tokens,
                        }

            for name in names_to_use:
                if name in sample_clients:
                    cfg = sample_clients[name]
                    clients.append(
                        {
                            "name": name,
                            "client_type": cfg["client_type"],
                            "api_key_env": cfg["api_key_env"],
                            "model": cfg["model"],
                            "temperature": str(cfg["temperature"]),
                            "max_tokens": str(cfg["max_tokens"]),
                        }
                    )

        for row_idx, client in enumerate(clients, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=client.get(header, ""))

        widths = column_widths or DEFAULT_CLIENTS_COLUMN_WIDTHS
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        return self

    def add_documents_sheet(
        self,
        documents: list[dict],
        headers: list[str] | None = None,
        column_widths: dict[str, int] | None = None,
    ) -> WorkbookBuilder:
        """Add the documents sheet to the workbook.

        Args:
            documents: List of dicts with keys: reference_name, common_name, file_path, notes
            headers: Custom headers (defaults to DEFAULT_DOCUMENTS_HEADERS)
            column_widths: Custom column widths

        Returns:
            self for method chaining

        """
        ws = self.wb.create_sheet(title="documents")

        if headers is None:
            headers = DEFAULT_DOCUMENTS_HEADERS

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, doc in enumerate(documents, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=doc.get(header, ""))

        widths = column_widths or DEFAULT_DOCUMENTS_COLUMN_WIDTHS
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        return self

    def add_data_sheet(
        self,
        data: list[dict],
        headers: list[str] | None = None,
    ) -> WorkbookBuilder:
        """Add the data sheet for batch execution.

        Args:
            data: List of dicts representing batch data rows
            headers: Custom headers (defaults to keys of first dict)

        Returns:
            self for method chaining

        """
        ws = self.wb.create_sheet(title="data")

        if not data:
            return self

        if headers is None:
            headers = list(data[0].keys())

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        return self

    def get_worksheet(self, name: str) -> Worksheet | None:
        """Get a worksheet by name.

        Args:
            name: Sheet name

        Returns:
            Worksheet or None if not found

        """
        return self.wb.get(name)

    def save(self) -> None:
        """Save the workbook to the output path."""
        self.wb.save(self.output_path)

    def print_summary(
        self,
        workbook_type: str,
        details: dict[str, Any] | None = None,
        run_command: str | None = None,
    ) -> None:
        """Print a summary of the created workbook.

        Args:
            workbook_type: Type of workbook (e.g., "basic", "conditional")
            details: Additional details to print
            run_command: Custom run command (defaults to standard orchestrator command)

        """
        separator = "=" * 60
        print(f"\n{separator}")
        print(f"Created {workbook_type} sample workbook: {self.output_path}")
        print(separator)

        if details:
            for key, value in details.items():
                if isinstance(value, list):
                    print(f"\n{key}:")
                    for item in value:
                        print(f"  - {item}")
                elif isinstance(value, dict):
                    print(f"\n{key}:")
                    for k, v in value.items():
                        print(f"  {k}: {v}")
                else:
                    print(f"\n{key}: {value}")

        cmd = run_command or f"python scripts/run_orchestrator.py {self.output_path}"
        print(f"\n{separator}")
        print(f"Run with: {cmd}")
        print(f"{separator}\n")
