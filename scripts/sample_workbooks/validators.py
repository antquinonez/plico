#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Workbook validator for validating orchestrator results.

This module provides the WorkbookValidator class for validating test workbooks
after orchestrator execution.

Usage:
    from sample_workbooks import WorkbookValidator, SectionDefinition

    sections = {
        "Level 0": SectionDefinition((1, 12), "Independent prompts"),
        "Level 1": SectionDefinition((13, 22), "With dependencies"),
    }

    validator = WorkbookValidator(sections, version="001")
    results = validator.validate_workbook(path)
    validator.print_report(results, "Basic Workbook")
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime

# Fix Windows console encoding for Unicode characters
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
from pathlib import Path
from typing import Any

import openpyxl

from .base import (
    BATCH_INDEX_COLUMN,
    CONDITION_ERROR_COLUMN,
    CONDITION_RESULT_COLUMN,
    PROMPT_NAME_COLUMN,
    SEQUENCE_COLUMN,
    STATUS_COLUMN,
    SectionDefinition,
)


class WorkbookValidator:
    """Validates orchestrator results in Excel workbooks.

    This class provides utilities for validating workbook execution results
    organized by sections/levels.

    Example:
        sections = {
            "Level 0": SectionDefinition((1, 12), "Independent prompts"),
        }
        validator = WorkbookValidator(sections, version="001")
        results = validator.validate_workbook(Path("workbook.xlsx"))
        validator.print_report(results, "My Workbook")

    """

    def __init__(
        self,
        sections: dict[str, SectionDefinition],
        version: str = "001",
        title: str = "WORKBOOK",
        status_column: int = STATUS_COLUMN,
        track_conditions: bool = False,
        track_batches: bool = False,
    ):
        """Initialize the validator.

        Args:
            sections: Dict mapping section names to SectionDefinitions
            version: Validator version string
            title: Title for reports (e.g., "BASIC WORKBOOK")
            status_column: Column number for status field
            track_conditions: Whether to track condition results
            track_batches: Whether to track batch indices

        """
        self.sections = sections
        self.version = version
        self.title = title
        self.status_column = status_column
        self.track_conditions = track_conditions
        self.track_batches = track_batches

    @staticmethod
    def find_latest_results_sheet(workbook) -> str | None:
        """Find the most recent results sheet by timestamp in name."""
        results_sheets = [name for name in workbook.sheetnames if name.startswith("results_")]
        if not results_sheets:
            return None
        return sorted(results_sheets, reverse=True)[0]

    @staticmethod
    def parse_sequence_to_row(ws, sequence_column: int = SEQUENCE_COLUMN) -> dict[int, int]:
        """Build mapping from sequence number to row number."""
        seq_to_row = {}
        for row in range(2, ws.max_row + 1):
            seq = ws.cell(row=row, column=sequence_column).value
            if seq is not None:
                seq_to_row[int(seq)] = row
        return seq_to_row

    @staticmethod
    def count_batches(
        ws, seq_to_row: dict[int, int], batch_column: int = BATCH_INDEX_COLUMN
    ) -> int:
        """Count unique batch indices in the results."""
        batch_indices = set()
        for row in seq_to_row.values():
            batch_idx = ws.cell(row=row, column=batch_column).value
            if batch_idx is not None:
                batch_indices.add(batch_idx)
        return len(batch_indices)

    def validate_section(
        self,
        ws,
        seq_to_row: dict[int, int],
        section_name: str,
        section: SectionDefinition,
        batch_count: int = 1,
    ) -> dict[str, Any]:
        """Validate a single section and return results."""
        start, end = section.range
        prompts = []
        passed = 0
        failed = 0
        skipped = 0
        conditions_true = 0
        conditions_false = 0
        field_check_failures: list[dict[str, Any]] = []

        for seq in range(start, end + 1):
            if seq not in seq_to_row:
                continue

            row = seq_to_row[seq]
            prompt_name = ws.cell(row=row, column=PROMPT_NAME_COLUMN).value
            status = ws.cell(row=row, column=self.status_column).value

            prompt_info = {
                "sequence": seq,
                "name": prompt_name,
                "status": status,
            }

            if self.track_conditions:
                cond_result = ws.cell(row=row, column=CONDITION_RESULT_COLUMN).value
                cond_error = ws.cell(row=row, column=CONDITION_ERROR_COLUMN).value
                prompt_info["condition_result"] = cond_result
                prompt_info["condition_error"] = cond_error

            if seq in section.field_checks:
                check_results = self._check_fields(
                    ws, row, seq, prompt_name, section.field_checks[seq]
                )
                prompt_info["field_checks"] = check_results
                field_check_failures.extend(check_results["failures"])

            prompts.append(prompt_info)

            if status == "success":
                passed += 1
                if self.track_conditions and prompt_info.get("condition_result") is True:
                    conditions_true += 1
            elif status == "skipped":
                skipped += 1
                if self.track_conditions and prompt_info.get("condition_result") is False:
                    conditions_false += 1
            elif status == "failed":
                failed += 1

        all_passed = failed == 0 and len(field_check_failures) == 0
        expected_executions = (end - start + 1) * batch_count

        result = {
            "name": section_name,
            "description": section.description,
            "features": section.features,
            "range": [start, end],
            "passed": passed,
            "skipped": skipped,
            "failed": failed,
            "all_passed": all_passed,
            "prompts": prompts,
        }

        if self.track_batches:
            result["batch_count"] = batch_count
            result["expected_executions"] = expected_executions

        if self.track_conditions:
            result["conditions_true"] = conditions_true
            result["conditions_false"] = conditions_false

        if field_check_failures:
            result["field_check_failures"] = field_check_failures

        return result

    @staticmethod
    def _check_fields(
        ws, row: int, seq: int, prompt_name: str, checks: dict[str, Any]
    ) -> dict[str, Any]:
        """Check field values against validation rules.

        Args:
            ws: openpyxl worksheet
            row: Row number
            seq: Sequence number
            prompt_name: Prompt name
            checks: Dict of check rules (column, expected, expected_one_of, etc.)

        Returns:
            Dict with 'failures' list of failure descriptions.

        """
        failures: list[dict[str, Any]] = []

        for check_name, rule in checks.items():
            column = rule.get("column")
            if column is None:
                continue

            actual = ws.cell(row=row, column=column).value

            if "expected" in rule:
                expected = rule["expected"]
                if actual != expected:
                    failures.append(
                        {
                            "sequence": seq,
                            "prompt_name": prompt_name,
                            "check": check_name,
                            "column": column,
                            "expected": expected,
                            "actual": actual,
                        }
                    )

            elif "expected_one_of" in rule:
                allowed = rule["expected_one_of"]
                if actual not in allowed:
                    failures.append(
                        {
                            "sequence": seq,
                            "prompt_name": prompt_name,
                            "check": check_name,
                            "column": column,
                            "expected_one_of": allowed,
                            "actual": actual,
                        }
                    )

            if rule.get("not_empty") and (actual is None or actual == ""):
                failures.append(
                    {
                        "sequence": seq,
                        "prompt_name": prompt_name,
                        "check": check_name,
                        "column": column,
                        "expected": "not empty",
                        "actual": actual,
                    }
                )

            if "greater_than" in rule and isinstance(actual, (int, float)):
                threshold = rule["greater_than"]
                if actual <= threshold:
                    failures.append(
                        {
                            "sequence": seq,
                            "prompt_name": prompt_name,
                            "check": check_name,
                            "column": column,
                            "expected": f"> {threshold}",
                            "actual": actual,
                        }
                    )

        return {"failures": failures}

    def validate_workbook(self, path: Path, results_sheet: str | None = None) -> dict[str, Any]:
        """Validate workbook and return comprehensive results."""
        workbook = openpyxl.load_workbook(path, data_only=True)

        if results_sheet:
            if results_sheet not in workbook.sheetnames:
                return {
                    "valid": False,
                    "error": f"Results sheet '{results_sheet}' not found",
                    "available_sheets": workbook.sheetnames,
                }
            sheet_name = results_sheet
        else:
            sheet_name = self.find_latest_results_sheet(workbook)
            if not sheet_name:
                return {
                    "valid": False,
                    "error": "No results sheet found in workbook",
                    "available_sheets": workbook.sheetnames,
                }

        ws = workbook[sheet_name]
        seq_to_row = self.parse_sequence_to_row(ws)

        batch_count = 1
        if self.track_batches:
            batch_count = self.count_batches(ws, seq_to_row)

        section_results = []
        total_passed = 0
        total_skipped = 0
        total_failed = 0
        total_conditions_true = 0
        total_conditions_false = 0

        for section_name, section in self.sections.items():
            result = self.validate_section(ws, seq_to_row, section_name, section, batch_count)
            section_results.append(result)

            total_passed += result["passed"]
            total_skipped += result.get("skipped", 0)
            total_failed += result["failed"]

            if self.track_conditions:
                total_conditions_true += result.get("conditions_true", 0)
                total_conditions_false += result.get("conditions_false", 0)

        all_passed = total_failed == 0 and all(s["all_passed"] for s in section_results)

        summary = {
            "total_prompts": sum(
                s["passed"] + s.get("skipped", 0) + s["failed"] for s in section_results
            ),
            "passed": total_passed,
            "skipped": total_skipped,
            "failed": total_failed,
        }

        if self.track_batches:
            summary["batch_count"] = batch_count

        if self.track_conditions:
            summary["conditions_true"] = total_conditions_true
            summary["conditions_false"] = total_conditions_false

        skipped_prompts = []
        if self.track_conditions:
            for section in section_results:
                for prompt in section["prompts"]:
                    if prompt["status"] == "skipped":
                        skipped_prompts.append(prompt["name"])

        return {
            "valid": True,
            "workbook": str(path),
            "results_sheet": sheet_name,
            "validated_at": datetime.now().isoformat(),
            "validator_version": self.version,
            "summary": summary,
            "sections": section_results,
            "skipped_prompts": skipped_prompts if skipped_prompts else None,
            "all_passed": all_passed,
        }

    def print_report(self, results: dict[str, Any], title: str | None = None) -> None:
        """Print human-readable validation report."""
        report_title = title or self.title
        print("=" * 80)
        print(f"{report_title} VALIDATION RESULTS")
        print("=" * 80)

        if not results.get("valid", False):
            print(f"ERROR: {results.get('error', 'Unknown error')}")
            print(f"Available sheets: {results.get('available_sheets', [])}")
            return

        print(f"Workbook: {results['workbook']}")
        print(f"Results Sheet: {results['results_sheet']}")
        print(f"Validator Version: v{results['validator_version']}")
        print(f"Validated: {results['validated_at']}")
        print()

        summary = results["summary"]
        print(f"Total Prompts: {summary['total_prompts']}")
        print(f"Passed: {summary['passed']}")

        if summary.get("skipped", 0) > 0:
            print(f"Skipped: {summary['skipped']}")

        print(f"Failed: {summary['failed']}")

        if self.track_batches and "batch_count" in summary:
            print(f"Batch Count: {summary['batch_count']}")

        if self.track_conditions:
            print(f"Conditions True: {summary.get('conditions_true', 0)}")
            print(f"Conditions False: {summary.get('conditions_false', 0)}")

        print()

        for section in results["sections"]:
            status_icon = "✓" if section["all_passed"] else "✗"
            print(f"\n{status_icon} {section['name']}:")
            print(f"    {section['description']}")

            if section.get("features"):
                print(f"    Features: {', '.join(section['features'])}")

            print(f"    Range: sequences {section['range'][0]}-{section['range'][1]}")

            if self.track_batches and "expected_executions" in section:
                print(
                    f"    Expected executions: {section['expected_executions']} ({section.get('batch_count', 1)} batches)"
                )

            status_parts = [f"{section['passed']} passed"]
            if section.get("skipped", 0) > 0:
                status_parts.append(f"{section['skipped']} skipped")
            status_parts.append(f"{section['failed']} failed")
            print(f"    Results: {', '.join(status_parts)}")

            for prompt in section["prompts"]:
                if prompt["status"] == "success":
                    if self.track_conditions and prompt.get("condition_result") is True:
                        print(f"      ✓ {prompt['name']}: condition=True")
                    else:
                        print(f"      ✓ {prompt['name']}")
                elif prompt["status"] == "skipped":
                    cond_str = (
                        f" (condition={prompt.get('condition_result')})"
                        if self.track_conditions
                        else ""
                    )
                    print(f"      ⊘ {prompt['name']}: SKIPPED{cond_str}")
                elif prompt["status"] == "failed":
                    print(f"      ✗ {prompt['name']}: FAILED")

            if section.get("field_check_failures"):
                for fc in section["field_check_failures"]:
                    print(
                        f"      ✗ {fc['prompt_name']} [{fc['check']}]: "
                        f"expected {fc.get('expected', fc.get('expected_one_of'))}, "
                        f"got {fc['actual']!r}"
                    )

        if results.get("skipped_prompts"):
            print()
            print("Skipped prompts (condition evaluated to False):")
            for name in results["skipped_prompts"]:
                print(f"  - {name}")

        print()
        print("=" * 80)
        if results["all_passed"]:
            print("✅ ALL SECTIONS PASSED!")
        else:
            print("❌ SOME SECTIONS HAD FAILURES")
        print("=" * 80)

    def run_cli(self, args: list[str] | None = None) -> int:
        """Run the validator as a CLI application.

        Args:
            args: Command line arguments (defaults to sys.argv[1:])

        Returns:
            Exit code (0 for pass, 1 for failures)

        """
        parser = argparse.ArgumentParser(
            description=f"Validate {self.title.lower()} workbook results",
            formatter_class=argparse.RawDescriptionHelpFormatter,
        )
        parser.add_argument("workbook", type=Path, help="Path to the workbook to validate")
        parser.add_argument(
            "--json",
            action="store_true",
            help="Output results as JSON",
        )
        parser.add_argument(
            "--results-sheet",
            type=str,
            help="Specific results sheet to validate (default: latest)",
        )
        parser.add_argument(
            "--version",
            action="version",
            version=f"%(prog)s v{self.version}",
        )

        parsed_args = parser.parse_args(args)

        if not parsed_args.workbook.exists():
            print(f"Error: Workbook not found: {parsed_args.workbook}", file=sys.stderr)
            return 1

        results = self.validate_workbook(parsed_args.workbook, parsed_args.results_sheet)

        if parsed_args.json:
            print(json.dumps(results, indent=2, default=str))
        else:
            self.print_report(results)

        return 0 if results.get("all_passed", False) else 1


def create_validator(
    sections: dict[str, SectionDefinition],
    version: str = "001",
    title: str = "WORKBOOK",
    **kwargs,
) -> WorkbookValidator:
    """Factory function to create a validator with common settings.

    Args:
        sections: Dict mapping section names to SectionDefinitions
        version: Validator version string
        title: Title for reports
        **kwargs: Additional arguments passed to WorkbookValidator

    Returns:
        Configured WorkbookValidator instance

    """
    return WorkbookValidator(sections, version=version, title=title, **kwargs)
