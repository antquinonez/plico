#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Generate test workbook for multi-client execution testing.

Creates a workbook with:
- config sheet
- prompts sheet with client column
- clients sheet with named client configurations using FFLiteLLMClient

Usage:
    python scripts/create_test_workbook_multiclient.py [output_path]
"""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from src.config import get_config


def create_multiclient_test_workbook(output_path: str):
    config = get_config()
    test_config = config.test

    wb = Workbook()

    # ==========================================
    # CONFIG SHEET
    # ==========================================
    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_data = [
        ("model", test_config.default_model),
        ("max_retries", str(test_config.default_retries)),
        ("temperature", str(test_config.default_temperature)),
        ("max_tokens", str(test_config.default_max_tokens)),
        (
            "system_instructions",
            "You are a helpful assistant. Give brief, concise answers.",
        ),
        ("created_at", datetime.now().isoformat()),
    ]

    for idx, (field, value) in enumerate(config_data, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_config.column_dimensions["A"].width = 20
    ws_config.column_dimensions["B"].width = 70

    # ==========================================
    # CLIENTS SHEET
    # ==========================================
    ws_clients = wb.create_sheet(title="clients")

    clients_headers = [
        "name",
        "client_type",
        "api_key_env",
        "model",
        "temperature",
        "max_tokens",
    ]
    for col_idx, header in enumerate(clients_headers, start=1):
        ws_clients.cell(row=1, column=col_idx, value=header)

    # Define multiple clients from config
    test_clients = test_config.test_clients
    clients_data = []
    for name in ["default", "fast", "creative"]:
        if name in test_clients:
            cfg = test_clients[name]
            clients_data.append(
                (
                    name,
                    cfg["client_type"],
                    cfg["api_key_env"],
                    cfg["model"],
                    cfg["temperature"],
                    cfg["max_tokens"],
                )
            )

    for row_idx, client_row in enumerate(clients_data, start=2):
        for col_idx, value in enumerate(client_row, start=1):
            ws_clients.cell(row=row_idx, column=col_idx, value=value)

    ws_clients.column_dimensions["A"].width = 15
    ws_clients.column_dimensions["B"].width = 18
    ws_clients.column_dimensions["C"].width = 20
    ws_clients.column_dimensions["D"].width = 22
    ws_clients.column_dimensions["E"].width = 14
    ws_clients.column_dimensions["F"].width = 12

    # ==========================================
    # PROMPTS SHEET (with client column)
    # ==========================================
    ws_prompts = wb.create_sheet(title="prompts")
    headers = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "references",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    prompts = []

    # Level 0: Independent prompts using different clients
    prompts.append(
        (
            1,
            "classify",
            "Classify this sentiment: 'I love this product!'. Answer: positive, negative, or neutral.",
            None,
            "fast",
        )
    )
    prompts.append(
        (
            2,
            "analyze",
            "Analyze the sentence 'The weather is nice today' for linguistic features.",
            None,
            None,
        )
    )
    prompts.append((3, "creative", "Write a one-line poem about clouds.", None, "creative"))
    prompts.append(
        (
            4,
            "summarize",
            "Summarize: 'AI is transforming many industries including healthcare and finance.'",
            None,
            "fast",
        )
    )
    prompts.append((5, "expand", "Expand on this topic: 'Machine learning basics'", None, None))

    # Level 1: Prompts with dependencies
    prompts.append(
        (
            6,
            "classify_context",
            "Based on the classification, explain why it was classified that way.",
            '["classify"]',
            None,
        )
    )
    prompts.append(
        (
            7,
            "analyze_deep",
            "Provide a deeper linguistic analysis based on the initial analysis.",
            '["analyze"]',
            "creative",
        )
    )
    prompts.append((8, "poem_explain", "Explain the imagery in the poem.", '["creative"]', None))
    prompts.append(
        (
            9,
            "summary_validate",
            "Is this summary accurate and complete? Answer yes or no with reason.",
            '["summarize"]',
            "fast",
        )
    )
    prompts.append(
        (
            10,
            "expand_outline",
            "Create an outline based on the expansion.",
            '["expand"]',
            None,
        )
    )

    # Level 2: Synthesis prompts
    prompts.append(
        (
            11,
            "compare",
            "Compare the classification and analysis approaches used.",
            '["classify", "analyze"]',
            None,
        )
    )
    prompts.append(
        (
            12,
            "creative_summary",
            "Summarize both the poem and its explanation creatively.",
            '["creative", "poem_explain"]',
            "creative",
        )
    )
    prompts.append(
        (
            13,
            "final_report",
            "Create a brief report summarizing all findings.",
            '["classify_context", "analyze_deep", "summary_validate"]',
            None,
        )
    )

    # Write all prompts to sheet
    for row_idx, (seq, name, prompt, history, client) in enumerate(prompts, start=2):
        ws_prompts.cell(row=row_idx, column=1, value=seq)
        ws_prompts.cell(row=row_idx, column=2, value=name)
        ws_prompts.cell(row=row_idx, column=3, value=prompt)
        ws_prompts.cell(row=row_idx, column=4, value=history if history else "")
        ws_prompts.cell(row=row_idx, column=5, value=client if client else "")
        ws_prompts.cell(row=row_idx, column=6, value="")
        ws_prompts.cell(row=row_idx, column=7, value="")

    ws_prompts.column_dimensions["A"].width = 10
    ws_prompts.column_dimensions["B"].width = 20
    ws_prompts.column_dimensions["C"].width = 70
    ws_prompts.column_dimensions["D"].width = 40
    ws_prompts.column_dimensions["E"].width = 12
    ws_prompts.column_dimensions["F"].width = 15
    ws_prompts.column_dimensions["G"].width = 15

    wb.save(output_path)

    print(f"\n{'=' * 70}")
    print(f"Created MULTI-CLIENT test workbook: {output_path}")
    print(f"{'=' * 70}")
    print("\nUsing: FFLiteLLMClient with LiteLLM routing")
    print("\nClients defined:")
    for name, client_type, _, model, temp, tokens in clients_data:
        print(f"  - {name}: {client_type} (model={model}, temp={temp}, tokens={tokens})")

    print(f"\nTotal prompts: {len(prompts)}")
    print("\nPrompt client assignments:")
    for seq, name, _, _, client in prompts:
        client_str = client if client else "(default)"
        print(f"  Seq {seq:2d} ({name:18s}): {client_str}")

    print("\nPrompt Structure:")
    print("  Level 0: 5 independent prompts (sequences 1-5)")
    print("  Level 1: 5 prompts with 1 dependency (sequences 6-10)")
    print("  Level 2: 3 synthesis prompts (sequences 11-13)")

    print(f"\n{'=' * 70}")
    print(f"Run with: python scripts/run_orchestrator.py {output_path} -c 2")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    config = get_config()
    output = sys.argv[1] if len(sys.argv) > 1 else config.test.workbooks.multiclient
    create_multiclient_test_workbook(output)
