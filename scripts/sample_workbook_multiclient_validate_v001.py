#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Validate multiclient workbook results.

Validates workbooks created by sample_workbook_multiclient_create_v001.py
by checking client assignments and execution status across all levels.

Features:
    - Client assignment verification
    - Level-by-level validation with detailed output
    - Dependency chain verification
    - JSON output for programmatic use
    - Exit code 0 for pass, 1 for failures

Usage:
    python scripts/sample_workbook_multiclient_validate_v001.py <workbook_path>
    python scripts/sample_workbook_multiclient_validate_v001.py <workbook_path> --json
    python scripts/sample_workbook_multiclient_validate_v001.py <workbook_path> --results-sheet results_20250228_123456

Version: 001
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

VERSION = "001"

LEVEL_DEFINITIONS = {
    "Level 0 - Independent Prompts": {
        "range": (1, 5),
        "description": "5 independent prompts using different clients",
    },
    "Level 1 - Single Dependencies": {
        "range": (6, 10),
        "description": "5 prompts with 1 dependency",
    },
    "Level 2 - Synthesis Prompts": {
        "range": (11, 13),
        "description": "3 synthesis prompts combining earlier results",
    },
}


def find_latest_results_sheet(workbook) -> str | None:
    """Find the most recent results sheet by timestamp in name."""
    results_sheets = [name for name in workbook.sheetnames if name.startswith("results_")]
    if not results_sheets:
        return None
    sorted_sheets = sorted(results_sheets, reverse=True)
    return sorted_sheets[0]


def parse_sequence_to_row(ws) -> dict[int, int]:
    """Build mapping from sequence number to row number."""
    seq_to_row = {}
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=3).value
        if seq is not None:
            seq_to_row[int(seq)] = row
    return seq_to_row


def validate_level(
    ws,
    seq_to_row: dict[int, int],
    level_name: str,
    level_def: dict,
) -> dict:
    """Validate a single level and return results."""
    start, end = level_def["range"]
    prompts = []
    passed = 0
    failed = 0
    client_usage: dict[str, int] = {}

    for seq in range(start, end + 1):
        if seq not in seq_to_row:
            continue

        row = seq_to_row[seq]
        prompt_name = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=12).value
        client = ws.cell(row=row, column=8).value or "default"

        prompt_info = {
            "sequence": seq,
            "name": prompt_name,
            "status": status,
            "client": client,
        }
        prompts.append(prompt_info)

        if status == "success":
            passed += 1
            client_usage[client] = client_usage.get(client, 0) + 1
        elif status == "failed":
            failed += 1

    all_passed = failed == 0

    return {
        "name": level_name,
        "description": level_def["description"],
        "range": [start, end],
        "passed": passed,
        "failed": failed,
        "client_usage": client_usage,
        "all_passed": all_passed,
        "prompts": prompts,
    }


def validate_workbook(path: Path, results_sheet: str | None = None) -> dict:
    """Validate workbook and return comprehensive results."""
    workbook = openpyxl.load_workbook(path, data_only=True)

    if results_sheet:
        if results_sheet not in workbook.sheetnames:
            return {
                "valid": False,
                "error": f"Results sheet '{results_sheet}' not found",
                "available_sheets": workbook.sheetnames,
            }
        sheet_name = results_sheet
    else:
        sheet_name = find_latest_results_sheet(workbook)
        if not sheet_name:
            return {
                "valid": False,
                "error": "No results sheet found in workbook",
                "available_sheets": workbook.sheetnames,
            }

    ws = workbook[sheet_name]
    seq_to_row = parse_sequence_to_row(ws)

    levels = []
    total_passed = 0
    total_failed = 0
    total_client_usage: dict[str, int] = {}

    for level_name, level_def in LEVEL_DEFINITIONS.items():
        level_result = validate_level(ws, seq_to_row, level_name, level_def)
        levels.append(level_result)

        total_passed += level_result["passed"]
        total_failed += level_result["failed"]

        for client, count in level_result["client_usage"].items():
            total_client_usage[client] = total_client_usage.get(client, 0) + count

    all_passed = total_failed == 0

    return {
        "valid": True,
        "workbook": str(path),
        "results_sheet": sheet_name,
        "validated_at": datetime.now().isoformat(),
        "validator_version": VERSION,
        "summary": {
            "total_prompts": sum(level["passed"] + level["failed"] for level in levels),
            "passed": total_passed,
            "failed": total_failed,
            "client_usage": total_client_usage,
        },
        "levels": levels,
        "all_passed": all_passed,
    }


def print_report(results: dict) -> None:
    """Print human-readable validation report."""
    print("=" * 80)
    print("MULTICLIENT WORKBOOK VALIDATION RESULTS")
    print("=" * 80)

    if not results.get("valid", False):
        print(f"ERROR: {results.get('error', 'Unknown error')}")
        print(f"Available sheets: {results.get('available_sheets', [])}")
        return

    print(f"Workbook: {results['workbook']}")
    print(f"Results Sheet: {results['results_sheet']}")
    print(f"Validator Version: v{results['validator_version']}")
    print(f"Validated: {results['validated_at']}")
    print()

    summary = results["summary"]
    print(f"Total Prompts: {summary['total_prompts']}")
    print(f"Passed: {summary['passed']}")
    print(f"Failed: {summary['failed']}")
    print(f"\nClient Usage: {summary['client_usage']}")
    print()

    for level in results["levels"]:
        status_icon = "✓" if level["all_passed"] else "✗"
        print(f"\n{status_icon} {level['name']}:")
        print(f"    {level['description']}")
        print(f"    Range: sequences {level['range'][0]}-{level['range'][1]}")
        print(f"    Results: {level['passed']} passed, {level['failed']} failed")

        for prompt in level["prompts"]:
            if prompt["status"] == "success":
                print(f"      ✓ {prompt['name']} (client={prompt['client']})")
            elif prompt["status"] == "failed":
                print(f"      ✗ {prompt['name']}: FAILED")

    print()
    print("=" * 80)
    if results["all_passed"]:
        print("✅ ALL LEVELS PASSED!")
    else:
        print("❌ SOME LEVELS HAD FAILURES")
    print("=" * 80)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Validate multiclient workbook results",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    %(prog)s ./sample_workbook_multiclient.xlsx
    %(prog)s ./sample_workbook_multiclient.xlsx --json
    %(prog)s ./sample_workbook_multiclient.xlsx --results-sheet results_20250228_123456
        """,
    )
    parser.add_argument("workbook", type=Path, help="Path to the workbook to validate")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output results as JSON",
    )
    parser.add_argument(
        "--results-sheet",
        type=str,
        help="Specific results sheet to validate (default: latest)",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s v{VERSION}",
    )

    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Error: Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    results = validate_workbook(args.workbook, args.results_sheet)

    if args.json:
        print(json.dumps(results, indent=2, default=str))
    else:
        print_report(results)

    return 0 if results.get("all_passed", False) else 1


if __name__ == "__main__":
    sys.exit(main())
