#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Generate sample workbook for conditional execution testing with multi-client support.

Creates 30 prompts with various conditional patterns using multiple client configurations:
- Basic status checks (success/failed/skipped)
- Boolean operators (and, or, not)
- Response content checks (contains, not contains)
- Numeric comparisons (attempts, len)
- Error recovery patterns
- Content-based branching

Different clients are used for different task types:
- fast: Low temperature for classification/yes-no tasks
- default: Standard settings for general tasks
- creative: High temperature for generative tasks

Uses FFLiteLLMClient with LiteLLM routing for Mistral Small.

Usage:
    python scripts/create_sample_workbook_conditional.py [output_path]
"""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from src.config import get_config


def create_conditional_sample_workbook(output_path: str):
    config = get_config()
    test_config = config.test

    wb = Workbook()

    # ============================================
    # CONFIG SHEET
    # ============================================
    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_data = [
        ("model", test_config.default_model),
        ("max_retries", str(test_config.default_retries)),
        ("temperature", str(test_config.default_temperature)),
        ("max_tokens", str(test_config.default_max_tokens)),
        (
            "system_instructions",
            "You are a helpful assistant. Give brief, concise answers. "
            "For classification questions, respond with just the category name. "
            "For yes/no questions, respond with just 'yes' or 'no'.",
        ),
        ("created_at", datetime.now().isoformat()),
    ]

    for idx, (field, value) in enumerate(config_data, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_config.column_dimensions["A"].width = 20
    ws_config.column_dimensions["B"].width = 70

    # ============================================
    # CLIENTS SHEET - Multiple client configurations
    # ============================================
    ws_clients = wb.create_sheet(title="clients")

    clients_headers = [
        "name",
        "client_type",
        "api_key_env",
        "model",
        "temperature",
        "max_tokens",
    ]
    for col_idx, header in enumerate(clients_headers, start=1):
        ws_clients.cell(row=1, column=col_idx, value=header)

    # Define multiple clients from config
    test_clients = test_config.test_clients
    clients_data = []
    for name in ["default", "fast", "creative"]:
        if name in test_clients:
            cfg = test_clients[name]
            clients_data.append(
                (
                    name,
                    cfg["client_type"],
                    cfg["api_key_env"],
                    cfg["model"],
                    cfg["temperature"],
                    cfg["max_tokens"],
                )
            )

    for row_idx, client_row in enumerate(clients_data, start=2):
        for col_idx, value in enumerate(client_row, start=1):
            ws_clients.cell(row=row_idx, column=col_idx, value=value)

    ws_clients.column_dimensions["A"].width = 12
    ws_clients.column_dimensions["B"].width = 15
    ws_clients.column_dimensions["C"].width = 18
    ws_clients.column_dimensions["D"].width = 20
    ws_clients.column_dimensions["E"].width = 12
    ws_clients.column_dimensions["F"].width = 12

    # ============================================
    # PROMPTS SHEET
    # ============================================
    ws_prompts = wb.create_sheet(title="prompts")
    headers = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "references",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    prompts = []
    row = 2

    # ============================================
    # SECTION 1: Error Recovery Pattern (1-5)
    # ============================================

    # Initial fetch - use fast client for simple response
    prompts.append(
        (
            1,
            "fetch_data",
            "Simulate fetching data. Just respond with: 'Data fetched successfully' or 'Fetch failed' (choose randomly).",
            None,
            "fast",
            None,
        )
    )

    # Branch A: Success path - creative for processing
    prompts.append(
        (
            2,
            "process_success",
            "The data was fetched successfully. Process it by saying 'Data processed'.",
            '["fetch_data"]',
            "default",
            '{{fetch_data.status}} == "success"',
        )
    )

    # Branch B: Failure path - creative for fallback message
    prompts.append(
        (
            3,
            "process_failure",
            "The data fetch failed. Provide a fallback message saying 'Using cached data'.",
            None,
            "creative",
            '{{fetch_data.status}} == "failed"',
        )
    )

    # Continue from either branch - fast for simple validation
    prompts.append(
        (
            4,
            "validate_result",
            "Based on the previous processing, say 'Validation passed' or describe any issues.",
            '["process_success", "process_failure"]',
            "fast",
            '{{process_success.status}} == "success" or {{process_failure.status}} == "success"',
        )
    )

    # Summary - creative for nice summary
    prompts.append(
        (
            5,
            "finalize_section1",
            "Summarize the data fetch and processing outcome in one sentence.",
            '["fetch_data", "validate_result"]',
            "creative",
            None,
        )
    )

    # ============================================
    # SECTION 2: Classification Branching (6-11)
    # ============================================

    # Classification - fast client for simple classification
    prompts.append(
        (
            6,
            "classify_sentiment",
            "Classify this text as positive, negative, or neutral: 'I love this amazing product!'",
            None,
            "fast",
            None,
        )
    )

    # Positive branch - creative for celebratory response
    prompts.append(
        (
            7,
            "positive_response",
            "Write a celebratory response for positive feedback. Just one sentence.",
            '["classify_sentiment"]',
            "creative",
            '"positive" in lower({{classify_sentiment.response}})',
        )
    )

    # Negative branch - default for empathetic
    prompts.append(
        (
            8,
            "negative_response",
            "Write an empathetic response for negative feedback. Just one sentence.",
            '["classify_sentiment"]',
            "default",
            '"negative" in lower({{classify_sentiment.response}})',
        )
    )

    # Neutral branch - default for balanced
    prompts.append(
        (
            9,
            "neutral_response",
            "Write a balanced response for neutral feedback. Just one sentence.",
            '["classify_sentiment"]',
            "default",
            '"neutral" in lower({{classify_sentiment.response}})',
        )
    )

    # Format response - creative for nice formatting
    prompts.append(
        (
            10,
            "format_response",
            "Based on the classification, format the appropriate response for the customer.",
            '["positive_response", "negative_response", "neutral_response"]',
            "creative",
            '{{positive_response.status}} == "success" or {{negative_response.status}} == "success" or {{neutral_response.status}} == "success"',
        )
    )

    # Log - fast for simple log
    prompts.append(
        (
            11,
            "log_interaction",
            "Log this customer interaction with the sentiment classification. One line summary.",
            '["classify_sentiment", "format_response"]',
            "fast",
            None,
        )
    )

    # ============================================
    # SECTION 3: Response Length Branching (12-16)
    # ============================================

    # Generate summary - default
    prompts.append(
        (
            12,
            "generate_summary",
            "Write a brief summary of machine learning. Keep it under 50 words.",
            None,
            "default",
            None,
        )
    )

    # If too short, expand - creative for expansion
    prompts.append(
        (
            13,
            "expand_summary",
            "The previous summary was too short. Add more detail about supervised learning. Keep under 100 words.",
            '["generate_summary"]',
            "creative",
            "len({{generate_summary.response}}) < 100",
        )
    )

    # If too long, condense - fast for condensation
    prompts.append(
        (
            14,
            "condense_summary",
            "The summary was too long. Condense it to under 30 words.",
            '["generate_summary"]',
            "fast",
            "len({{generate_summary.response}}) > 200",
        )
    )

    # If just right, approve - fast for simple approval
    prompts.append(
        (
            15,
            "approve_summary",
            "The summary length is acceptable. Confirm by saying 'Summary approved'.",
            '["generate_summary"]',
            "fast",
            "len({{generate_summary.response}}) >= 100 and len({{generate_summary.response}}) <= 200",
        )
    )

    # Final summary - creative
    prompts.append(
        (
            16,
            "final_summary",
            "Provide the final approved summary on machine learning.",
            '["expand_summary", "condense_summary", "approve_summary"]',
            "creative",
            '{{expand_summary.status}} == "success" or {{condense_summary.status}} == "success" or {{approve_summary.status}} == "success"',
        )
    )

    # ============================================
    # SECTION 4: Retry Pattern (17-21)
    # ============================================

    # First attempt - fast for simple number generation
    prompts.append(
        (
            17,
            "attempt_1",
            "Generate a random 4-digit code. Just give the numbers.",
            None,
            "fast",
            None,
        )
    )

    # If first attempt was short, try again - fast
    prompts.append(
        (
            18,
            "attempt_2",
            "Generate a different random 4-digit code. Just give the numbers.",
            None,
            "fast",
            "len({{attempt_1.response}}) < 4",
        )
    )

    # If still short, try once more - fast
    prompts.append(
        (
            19,
            "attempt_3",
            "Generate a final random 4-digit code. Just give the numbers.",
            None,
            "fast",
            '{{attempt_1.status}} == "failed" and {{attempt_2.status}} == "failed"',
        )
    )

    # Validate code - fast
    prompts.append(
        (
            20,
            "validate_code",
            "Verify the generated code is valid (4 digits). Say 'Valid' or 'Invalid'.",
            '["attempt_1", "attempt_2", "attempt_3"]',
            "fast",
            '{{attempt_1.status}} == "success" or {{attempt_2.status}} == "success" or {{attempt_3.status}} == "success"',
        )
    )

    # Store code - default
    prompts.append(
        (
            21,
            "store_code",
            "Confirm the final code has been stored. Just say 'Code stored: [the code]'.",
            '["validate_code"]',
            "default",
            None,
        )
    )

    # ============================================
    # SECTION 5: Complex Boolean Logic (22-26)
    # ============================================

    # Check A - fast for yes/no
    prompts.append(
        (
            22,
            "check_a",
            "Say either 'pass' or 'fail' (choose randomly).",
            None,
            "fast",
            None,
        )
    )

    # Check B - fast
    prompts.append(
        (
            23,
            "check_b",
            "Say either 'pass' or 'fail' (choose randomly).",
            None,
            "fast",
            None,
        )
    )

    # Check C - fast
    prompts.append(
        (
            24,
            "check_c",
            "Say either 'pass' or 'fail' (choose randomly).",
            None,
            "fast",
            None,
        )
    )

    # Both A and B passed - default
    prompts.append(
        (
            25,
            "both_passed",
            "Both checks A and B passed. Say 'Dual validation complete'.",
            '["check_a", "check_b"]',
            "default",
            '"pass" in lower({{check_a.response}}) and "pass" in lower({{check_b.response}})',
        )
    )

    # Any failed - creative for warning
    prompts.append(
        (
            26,
            "any_failed",
            "At least one check failed. Say 'Warning: Some checks failed'.",
            '["check_a", "check_b", "check_c"]',
            "creative",
            '"fail" in lower({{check_a.response}}) or "fail" in lower({{check_b.response}}) or "fail" in lower({{check_c.response}})',
        )
    )

    # ============================================
    # SECTION 6: Has Response Check (27-30)
    # ============================================

    # Optional generation - creative
    prompts.append(
        (
            27,
            "generate_optional",
            "Optionally generate a fun fact about the moon. You may choose to respond or not.",
            None,
            "creative",
            None,
        )
    )

    # Only process if response exists - creative
    prompts.append(
        (
            28,
            "process_fact",
            "You provided a moon fact. Add one more interesting detail.",
            '["generate_optional"]',
            "creative",
            "{{generate_optional.has_response}} == True",
        )
    )

    # Fallback if no response - default
    prompts.append(
        (
            29,
            "default_fact",
            "No fact was provided. Here is a default fact: The moon has no atmosphere.",
            None,
            "default",
            "{{generate_optional.has_response}} == False",
        )
    )

    # Final report - creative for nice report
    prompts.append(
        (
            30,
            "final_report",
            "Create a brief summary report of all the conditional execution tests that ran.",
            '["finalize_section1", "log_interaction", "final_summary", "store_code", "both_passed", "any_failed", "process_fact", "default_fact"]',
            "creative",
            '{{finalize_section1.status}} == "success"',
        )
    )

    # Write all prompts to sheet
    for seq, name, prompt, history, client, condition in prompts:
        ws_prompts.cell(row=row, column=1, value=seq)
        ws_prompts.cell(row=row, column=2, value=name)
        ws_prompts.cell(row=row, column=3, value=prompt)
        ws_prompts.cell(row=row, column=4, value=history if history else "")
        ws_prompts.cell(row=row, column=5, value=client if client else "")
        ws_prompts.cell(row=row, column=6, value=condition if condition else "")
        ws_prompts.cell(row=row, column=7, value="")
        row += 1

    ws_prompts.column_dimensions["A"].width = 10
    ws_prompts.column_dimensions["B"].width = 22
    ws_prompts.column_dimensions["C"].width = 80
    ws_prompts.column_dimensions["D"].width = 45
    ws_prompts.column_dimensions["E"].width = 10
    ws_prompts.column_dimensions["F"].width = 60
    ws_prompts.column_dimensions["G"].width = 15

    wb.save(output_path)

    print(f"\n{'=' * 70}")
    print(f"Created conditional execution sample workbook: {output_path}")
    print(f"{'=' * 70}")

    print("\nUsing: FFLiteLLMClient with LiteLLM routing")
    print("\nClients defined:")
    for name, client_type, _, model, temp, tokens in clients_data:
        print(f"  - {name}: {client_type} (model={model}, temp={temp}, tokens={tokens})")

    print(f"\nTotal prompts: {len(prompts)}")

    # Count client usage
    client_counts = {}
    for p in prompts:
        c = p[4] or "default"
        client_counts[c] = client_counts.get(c, 0) + 1
    print("\nClient usage:")
    for c, count in sorted(client_counts.items()):
        print(f"  - {c}: {count} prompts")

    print("\nConditional Patterns Tested:")
    print("\n  SECTION 1 - Error Recovery Pattern (sequences 1-5):")
    print("    - Status checks: {fetch_data.status} == 'success'")
    print("    - OR branching: process_success OR process_failure")
    print("\n  SECTION 2 - Classification Branching (sequences 6-11):")
    print("    - Response content: 'positive' in lower({response})")
    print("    - Multi-way branching based on content")
    print("\n  SECTION 3 - Response Length Branching (sequences 12-16):")
    print("    - Length checks: len({response}) < 100")
    print("    - Range checks: len({response}) >= 100 and len({response}) <= 200")
    print("\n  SECTION 4 - Retry Pattern (sequences 17-21):")
    print("    - Chained conditions: {a.status} == 'failed' and {b.status} == 'failed'")
    print("    - Multiple OR conditions")
    print("\n  SECTION 5 - Complex Boolean Logic (sequences 22-26):")
    print("    - AND operator: check_a and check_b both pass")
    print("    - OR operator: any check fails")
    print("\n  SECTION 6 - Has Response Check (sequences 27-30):")
    print("    - has_response property: {prompt.has_response} == True")
    print("    - Fallback when no response")

    print(f"\n{'=' * 70}")
    print(f"Run with: python scripts/run_orchestrator.py {output_path}")
    print(f"Or with parallel execution: python scripts/run_orchestrator.py {output_path} -c 3")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    config = get_config()
    output = sys.argv[1] if len(sys.argv) > 1 else config.test.workbooks.conditional
    create_conditional_sample_workbook(output)
