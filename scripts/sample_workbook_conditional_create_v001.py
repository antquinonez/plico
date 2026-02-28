#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Generate sample workbook for conditional execution testing with extensive functionality.

Creates 50 prompts testing:
    - String methods: startswith, endswith, lower, upper, strip, replace, split, count, etc.
    - JSON functions: json_get, json_get_default, json_has, json_keys, json_parse, json_type
    - List/dict operations: membership, subscript access, methods
    - Math functions: abs, min, max, round
    - Type functions: is_null, is_empty, bool
    - Complex combined conditions

Paired with: sample_workbook_conditional_validate_v001.py

Usage:
    python scripts/sample_workbook_conditional_create_v001.py [output_path]

Version: 001
"""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook

from src.config import get_config


def create_conditional_sample_workbook(output_path: str):
    config = get_config()
    test_config = config.sample

    wb = Workbook()

    # ============================================
    # CONFIG SHEET
    # ============================================
    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_data = [
        ("model", test_config.default_model),
        ("max_retries", str(test_config.default_retries)),
        ("temperature", str(test_config.default_temperature)),
        ("max_tokens", str(test_config.default_max_tokens)),
        (
            "system_instructions",
            "You are a helpful assistant. For JSON questions, return valid JSON. "
            "For classification, respond with just the category. "
            "For yes/no questions, respond with just 'yes' or 'no'. "
            "Keep responses concise.",
        ),
        ("created_at", datetime.now().isoformat()),
    ]

    for idx, (field, value) in enumerate(config_data, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_config.column_dimensions["A"].width = 20
    ws_config.column_dimensions["B"].width = 70

    # ============================================
    # CLIENTS SHEET
    # ============================================
    ws_clients = wb.create_sheet(title="clients")

    clients_headers = [
        "name",
        "client_type",
        "api_key_env",
        "model",
        "temperature",
        "max_tokens",
    ]
    for col_idx, header in enumerate(clients_headers, start=1):
        ws_clients.cell(row=1, column=col_idx, value=header)

    test_clients = test_config.sample_clients
    clients_data = []
    for name in ["default", "fast", "creative"]:
        if name in test_clients:
            cfg = test_clients[name]
            clients_data.append(
                (
                    name,
                    cfg.client_type,
                    cfg.api_key_env,
                    cfg.model,
                    str(cfg.temperature),
                    str(cfg.max_tokens),
                )
            )

    for row_idx, client_row in enumerate(clients_data, start=2):
        for col_idx, value in enumerate(client_row, start=1):
            ws_clients.cell(row=row_idx, column=col_idx, value=value)

    ws_clients.column_dimensions["A"].width = 12
    ws_clients.column_dimensions["B"].width = 15
    ws_clients.column_dimensions["C"].width = 18
    ws_clients.column_dimensions["D"].width = 20
    ws_clients.column_dimensions["E"].width = 12
    ws_clients.column_dimensions["F"].width = 12

    # ============================================
    # PROMPTS SHEET
    # ============================================
    ws_prompts = wb.create_sheet(title="prompts")
    headers = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "references",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    prompts = []
    row = 2

    # ============================================
    # SECTION 1: String Method Tests (1-10)
    # ============================================

    prompts.append(
        (
            1,
            "gen_status",
            "Respond with exactly: 'SUCCESS: Operation completed'",
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            2,
            "check_prefix",
            "The previous response started with SUCCESS. Confirm by saying 'Prefix confirmed'.",
            '["gen_status"]',
            "fast",
            '{{gen_status.response}}.startswith("SUCCESS")',
        )
    )

    prompts.append(
        (
            3,
            "check_suffix",
            "The previous response ended with 'completed'. Confirm by saying 'Suffix confirmed'.",
            '["gen_status"]',
            "fast",
            '{{gen_status.response}}.endswith("completed")',
        )
    )

    prompts.append(
        (
            4,
            "gen_uppercase",
            "Respond with exactly: 'HELLO WORLD'",
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            5,
            "check_lower",
            "Convert the previous response to lowercase and confirm it says 'hello world'.",
            '["gen_uppercase"]',
            "fast",
            '{{gen_uppercase.response}}.lower() == "hello world"',
        )
    )

    prompts.append(
        (
            6,
            "gen_whitespace",
            "Respond with exactly: '  padded  ' (include the spaces)",
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            7,
            "check_strip",
            "The trimmed response should equal 'padded'. Confirm by saying 'Strip works'.",
            '["gen_whitespace"]',
            "fast",
            '{{gen_whitespace.response}}.strip() == "padded"',
        )
    )

    prompts.append(
        (
            8,
            "gen_countable",
            "Respond with exactly: 'error error error'",
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            9,
            "check_count",
            "The word 'error' appeared 3 times. Confirm by saying 'Count verified'.",
            '["gen_countable"]',
            "fast",
            '{{gen_countable.response}}.count("error") == 3',
        )
    )

    prompts.append(
        (
            10,
            "section1_summary",
            "All string method tests passed. Say 'String methods: OK'.",
            '["check_prefix", "check_suffix", "check_lower", "check_strip", "check_count"]',
            "default",
            '{{check_prefix.status}} == "success" and {{check_count.status}} == "success"',
        )
    )

    # ============================================
    # SECTION 2: JSON Simple Access (11-18)
    # ============================================

    prompts.append(
        (
            11,
            "gen_json_simple",
            'Return exactly this JSON: {"status": "active", "count": 42}',
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            12,
            "check_json_status",
            "The JSON status was 'active'. Confirm by saying 'Status is active'.",
            '["gen_json_simple"]',
            "fast",
            'json_get({{gen_json_simple.response}}, "status") == "active"',
        )
    )

    prompts.append(
        (
            13,
            "check_json_count",
            "The JSON count was 42. Confirm by saying 'Count is 42'.",
            '["gen_json_simple"]',
            "fast",
            'json_get({{gen_json_simple.response}}, "count") == 42',
        )
    )

    prompts.append(
        (
            14,
            "check_json_has_status",
            "The JSON has a 'status' key. Confirm by saying 'Key exists'.",
            '["gen_json_simple"]',
            "fast",
            'json_has({{gen_json_simple.response}}, "status") == True',
        )
    )

    prompts.append(
        (
            15,
            "check_json_missing_key",
            "The JSON does not have an 'error' key. Confirm by saying 'No error key'.",
            '["gen_json_simple"]',
            "fast",
            'json_has({{gen_json_simple.response}}, "error") == False',
        )
    )

    prompts.append(
        (
            16,
            "check_json_type_string",
            "The 'status' value is a string type. Confirm by saying 'Type is string'.",
            '["gen_json_simple"]',
            "fast",
            'json_type({{gen_json_simple.response}}, "status") == "string"',
        )
    )

    prompts.append(
        (
            17,
            "check_json_type_number",
            "The 'count' value is a number type. Confirm by saying 'Type is number'.",
            '["gen_json_simple"]',
            "fast",
            'json_type({{gen_json_simple.response}}, "count") == "number"',
        )
    )

    prompts.append(
        (
            18,
            "section2_summary",
            "JSON simple access tests completed. Say 'JSON simple: OK'.",
            '["check_json_status", "check_json_count", "check_json_has_status"]',
            "default",
            '{{check_json_status.status}} == "success"',
        )
    )

    # ============================================
    # SECTION 3: JSON Nested Access (19-26)
    # ============================================

    prompts.append(
        (
            19,
            "gen_json_nested",
            'Return exactly this JSON: {"data": {"user": {"name": "Alice", "age": 30}}}',
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            20,
            "check_nested_name",
            "The nested name was 'Alice'. Confirm by saying 'Name is Alice'.",
            '["gen_json_nested"]',
            "fast",
            'json_get({{gen_json_nested.response}}, "data.user.name") == "Alice"',
        )
    )

    prompts.append(
        (
            21,
            "check_nested_age",
            "The nested age was 30. Confirm by saying 'Age is 30'.",
            '["gen_json_nested"]',
            "fast",
            'json_get({{gen_json_nested.response}}, "data.user.age") == 30',
        )
    )

    prompts.append(
        (
            22,
            "check_deep_path_exists",
            "The deep path 'data.user.name' exists. Confirm by saying 'Path exists'.",
            '["gen_json_nested"]',
            "fast",
            'json_has({{gen_json_nested.response}}, "data.user.name") == True',
        )
    )

    prompts.append(
        (
            23,
            "check_deep_path_missing",
            "The path 'data.user.email' does not exist. Confirm by saying 'Email missing'.",
            '["gen_json_nested"]',
            "fast",
            'json_has({{gen_json_nested.response}}, "data.user.email") == False',
        )
    )

    prompts.append(
        (
            24,
            "check_json_default",
            "Using default value for missing key. Confirm 'Default works'.",
            '["gen_json_nested"]',
            "fast",
            'json_get_default({{gen_json_nested.response}}, "data.user.email", "none") == "none"',
        )
    )

    prompts.append(
        (
            25,
            "check_json_default_returns_value",
            "Existing key returns value not default. Confirm 'Value not default'.",
            '["gen_json_nested"]',
            "fast",
            'json_get_default({{gen_json_nested.response}}, "data.user.age", 0) == 30',
        )
    )

    prompts.append(
        (
            26,
            "section3_summary",
            "JSON nested access tests completed. Say 'JSON nested: OK'.",
            '["check_nested_name", "check_nested_age", "check_json_default"]',
            "default",
            '{{check_nested_name.status}} == "success"',
        )
    )

    # ============================================
    # SECTION 4: JSON Array Access (27-34)
    # ============================================

    prompts.append(
        (
            27,
            "gen_json_array",
            'Return exactly this JSON: {"items": ["apple", "banana", "cherry"]}',
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            28,
            "check_array_first",
            "The first item was 'apple'. Confirm by saying 'First is apple'.",
            '["gen_json_array"]',
            "fast",
            'json_get({{gen_json_array.response}}, "items[0]") == "apple"',
        )
    )

    prompts.append(
        (
            29,
            "check_array_second",
            "The second item was 'banana'. Confirm by saying 'Second is banana'.",
            '["gen_json_array"]',
            "fast",
            'json_get({{gen_json_array.response}}, "items[1]") == "banana"',
        )
    )

    prompts.append(
        (
            30,
            "check_array_third",
            "The third item was 'cherry'. Confirm by saying 'Third is cherry'.",
            '["gen_json_array"]',
            "fast",
            'json_get({{gen_json_array.response}}, "items[2]") == "cherry"',
        )
    )

    prompts.append(
        (
            31,
            "check_array_type",
            "The 'items' value is an array type. Confirm by saying 'Type is array'.",
            '["gen_json_array"]',
            "fast",
            'json_type({{gen_json_array.response}}, "items") == "array"',
        )
    )

    prompts.append(
        (
            32,
            "check_keys_list",
            "The JSON has an 'items' key in its keys. Confirm by saying 'Items in keys'.",
            '["gen_json_array"]',
            "fast",
            '"items" in json_keys({{gen_json_array.response}})',
        )
    )

    prompts.append(
        (
            33,
            "check_keys_count",
            "The JSON has exactly 1 key. Confirm by saying 'One key'.",
            '["gen_json_array"]',
            "fast",
            "len(json_keys({{gen_json_array.response}})) == 1",
        )
    )

    prompts.append(
        (
            34,
            "section4_summary",
            "JSON array access tests completed. Say 'JSON array: OK'.",
            '["check_array_first", "check_array_third", "check_keys_list"]',
            "default",
            '{{check_array_first.status}} == "success"',
        )
    )

    # ============================================
    # SECTION 5: JSON Complex Nested (35-38)
    # ============================================

    prompts.append(
        (
            35,
            "gen_json_complex",
            'Return exactly this JSON: {"data": {"items": [{"id": 1, "name": "first"}, {"id": 2, "name": "second"}]}}',
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            36,
            "check_complex_id",
            "The second item's id was 2. Confirm by saying 'ID is 2'.",
            '["gen_json_complex"]',
            "fast",
            'json_get({{gen_json_complex.response}}, "data.items[1].id") == 2',
        )
    )

    prompts.append(
        (
            37,
            "check_complex_name",
            "The first item's name was 'first'. Confirm by saying 'Name is first'.",
            '["gen_json_complex"]',
            "fast",
            'json_get({{gen_json_complex.response}}, "data.items[0].name") == "first"',
        )
    )

    prompts.append(
        (
            38,
            "section5_summary",
            "JSON complex nested tests completed. Say 'JSON complex: OK'.",
            '["check_complex_id", "check_complex_name"]',
            "default",
            '{{check_complex_id.status}} == "success" and {{check_complex_name.status}} == "success"',
        )
    )

    # ============================================
    # SECTION 6: Math Functions (39-44)
    # ============================================

    prompts.append(
        (
            39,
            "gen_negative",
            "Respond with exactly: '-42'",
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            40,
            "check_abs",
            "The absolute value is 42. Confirm by saying 'Abs is 42'.",
            '["gen_negative"]',
            "fast",
            "abs(int({{gen_negative.response}})) == 42",
        )
    )

    prompts.append(
        (
            41,
            "gen_number_5",
            "Respond with exactly: '5'",
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            42,
            "check_min",
            "The minimum of 5 and 10 is 5. Confirm by saying 'Min is 5'.",
            '["gen_number_5"]',
            "fast",
            "min(int({{gen_number_5.response}}), 10) == 5",
        )
    )

    prompts.append(
        (
            43,
            "check_max",
            "The maximum of 5 and 3 is 5. Confirm by saying 'Max is 5'.",
            '["gen_number_5"]',
            "fast",
            "max(int({{gen_number_5.response}}), 3) == 5",
        )
    )

    prompts.append(
        (
            44,
            "section6_summary",
            "Math function tests completed. Say 'Math functions: OK'.",
            '["check_abs", "check_min", "check_max"]',
            "default",
            '{{check_abs.status}} == "success"',
        )
    )

    # ============================================
    # SECTION 7: Type Checking Functions (45-47)
    # ============================================

    prompts.append(
        (
            45,
            "gen_empty_test",
            "Respond with exactly: '' (empty string, just quotes)",
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            46,
            "check_is_empty",
            "The response was empty. Confirm by saying 'Response is empty'.",
            '["gen_empty_test"]',
            "fast",
            "is_empty({{gen_empty_test.response}}) == True",
        )
    )

    prompts.append(
        (
            47,
            "section7_summary",
            "Type checking tests completed. Say 'Type checks: OK'.",
            '["check_is_empty"]',
            "default",
            '{{check_is_empty.status}} == "success"',
        )
    )

    # ============================================
    # SECTION 8: Combined Complex Conditions (48-50)
    # ============================================

    prompts.append(
        (
            48,
            "gen_combined_json",
            'Return exactly this JSON: {"status": "SUCCESS", "data": {"score": 85, "items": ["a", "b"]}}',
            None,
            "fast",
            None,
        )
    )

    prompts.append(
        (
            49,
            "check_combined_all",
            "Complex check: status starts with SUCCESS, score > 80, items array exists. Say 'All checks pass'.",
            '["gen_combined_json"]',
            "fast",
            'json_get({{gen_combined_json.response}}, "status").startswith("SUCCESS") and json_get({{gen_combined_json.response}}, "data.score") > 80 and json_has({{gen_combined_json.response}}, "data.items")',
        )
    )

    prompts.append(
        (
            50,
            "final_report",
            "Create a one-line summary of all 50 conditional execution tests using new features: string methods, JSON access, math functions, and combined conditions.",
            '["section1_summary", "section2_summary", "section3_summary", "section4_summary", "section5_summary", "section6_summary", "section7_summary", "check_combined_all"]',
            "creative",
            '{{section1_summary.status}} == "success" and {{section2_summary.status}} == "success" and {{check_combined_all.status}} == "success"',
        )
    )

    # Write all prompts to sheet
    for seq, name, prompt, history, client, condition in prompts:
        ws_prompts.cell(row=row, column=1, value=seq)
        ws_prompts.cell(row=row, column=2, value=name)
        ws_prompts.cell(row=row, column=3, value=prompt)
        ws_prompts.cell(row=row, column=4, value=history if history else "")
        ws_prompts.cell(row=row, column=5, value=client if client else "")
        ws_prompts.cell(row=row, column=6, value=condition if condition else "")
        ws_prompts.cell(row=row, column=7, value="")
        row += 1

    ws_prompts.column_dimensions["A"].width = 10
    ws_prompts.column_dimensions["B"].width = 25
    ws_prompts.column_dimensions["C"].width = 90
    ws_prompts.column_dimensions["D"].width = 50
    ws_prompts.column_dimensions["E"].width = 10
    ws_prompts.column_dimensions["F"].width = 80
    ws_prompts.column_dimensions["G"].width = 15

    wb.save(output_path)

    print(f"\n{'=' * 80}")
    print(f"Created conditional execution sample workbook: {output_path}")
    print(f"{'=' * 80}")

    print("\nTotal prompts: 50")

    print("\nConditional Patterns Tested:")
    print("\n  SECTION 1 - String Methods (1-10):")
    print("    .startswith(), .endswith(), .lower(), .strip(), .count()")
    print("\n  SECTION 2 - JSON Simple Access (11-18):")
    print("    json_get(), json_has(), json_type()")
    print("\n  SECTION 3 - JSON Nested Access (19-26):")
    print("    Nested path: json_get(response, 'data.user.name')")
    print("    json_get_default() for missing keys")
    print("\n  SECTION 4 - JSON Array Access (27-34):")
    print("    Array indexing: json_get(response, 'items[0]')")
    print("    json_keys(), list membership with 'in'")
    print("\n  SECTION 5 - JSON Complex Nested (35-38):")
    print("    Deep nesting: json_get(response, 'data.items[1].id')")
    print("\n  SECTION 6 - Math Functions (39-44):")
    print("    abs(), min(), max()")
    print("\n  SECTION 7 - Type Checking (45-47):")
    print("    is_empty()")
    print("\n  SECTION 8 - Combined Conditions (48-50):")
    print("    Chained: .startswith() AND json_get() > N AND json_has()")

    print(f"\n{'=' * 80}")
    print(f"Run with: python scripts/run_orchestrator.py {output_path}")
    print(f"{'=' * 80}\n")


if __name__ == "__main__":
    config = get_config()
    output = sys.argv[1] if len(sys.argv) > 1 else config.sample.workbooks.conditional
    create_conditional_sample_workbook(output)
