#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Export Parquet File to Excel Workbook

Usage:
    python scripts/parquet_to_excel.py <parquet_file> [options]

Examples:
    # Export to Excel (same directory as parquet, with .xlsx extension)
    python scripts/parquet_to_excel.py ./outputs/results.parquet

    # Export with custom output path
    python scripts/parquet_to_excel.py ./outputs/results.parquet -o ./exported.xlsx

    # Include only specific columns
    python scripts/parquet_to_excel.py ./outputs/results.parquet --columns sequence,prompt_name,response

    # Filter by status
    python scripts/parquet_to_excel.py ./outputs/results.parquet --status success

    # Filter by batch_id
    python scripts/parquet_to_excel.py ./outputs/results.parquet --batch-id 1 2 3
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

COLUMN_WIDTHS = {
    "batch_id": 10,
    "batch_name": 15,
    "sequence": 10,
    "prompt_name": 25,
    "prompt": 60,
    "resolved_prompt": 60,
    "history": 20,
    "client": 12,
    "condition": 20,
    "condition_result": 15,
    "condition_error": 20,
    "response": 80,
    "status": 10,
    "attempts": 10,
    "error": 30,
    "references": 20,
    "semantic_query": 20,
    "semantic_filter": 20,
    "query_expansion": 15,
    "rerank": 10,
}

WRAP_COLUMNS = {"prompt", "resolved_prompt", "response", "error", "condition", "condition_error"}


def export_to_excel(
    parquet_path: str,
    output_path: str | None = None,
    columns: list[str] | None = None,
    status_filter: list[str] | None = None,
    batch_ids: list[int] | None = None,
) -> str:
    """Export parquet file to Excel workbook.

    Args:
        parquet_path: Path to the parquet file.
        output_path: Optional output Excel file path.
        columns: Optional list of columns to include.
        status_filter: Optional list of statuses to include.
        batch_ids: Optional list of batch IDs to include.

    Returns:
        Path to the created Excel file.

    """
    df = pl.read_parquet(parquet_path)

    if status_filter:
        df = df.filter(pl.col("status").is_in(status_filter))

    if batch_ids:
        df = df.filter(pl.col("batch_id").is_in(batch_ids))

    if columns:
        available = [c for c in columns if c in df.columns]
        df = df.select(available)

    if output_path is None:
        parquet_file = Path(parquet_path)
        output_path = str(parquet_file.with_suffix(".xlsx"))

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    headers = list(df.columns)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_idx, row in enumerate(df.iter_rows(named=True), start=2):
        for col_idx, col_name in enumerate(headers, start=1):
            value = row.get(col_name)

            if value is None:
                cell_value = ""
            elif isinstance(value, list):
                cell_value = json.dumps(value)
            elif isinstance(value, dict):
                cell_value = json.dumps(value)
            else:
                cell_value = str(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

            if col_name in WRAP_COLUMNS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = COLUMN_WIDTHS.get(col_name, 15)
        ws.column_dimensions[col_letter].width = width

    if "sequence" in df.columns:
        pass

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"

    wb.save(output_path)

    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Export parquet file to Excel workbook",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("parquet", help="Path to parquet file")
    parser.add_argument(
        "--output",
        "-o",
        help="Output Excel file path (default: same as parquet with .xlsx extension)",
    )
    parser.add_argument(
        "--columns",
        "-c",
        help="Columns to include (comma-separated)",
    )
    parser.add_argument(
        "--status",
        choices=["success", "failed", "skipped"],
        nargs="+",
        help="Filter by status (can specify multiple)",
    )
    parser.add_argument(
        "--batch-id",
        type=int,
        nargs="+",
        help="Filter by batch IDs",
    )

    args = parser.parse_args()

    if not os.path.exists(args.parquet):
        print(f"Error: File not found: {args.parquet}")
        return 1

    columns = None
    if args.columns:
        columns = [c.strip() for c in args.columns.split(",")]

    try:
        output_path = export_to_excel(
            parquet_path=args.parquet,
            output_path=args.output,
            columns=columns,
            status_filter=args.status,
            batch_ids=args.batch_id,
        )

        df = pl.read_parquet(args.parquet)
        if args.status:
            df = df.filter(pl.col("status").is_in(args.status))
        if args.batch_id:
            df = df.filter(pl.col("batch_id").is_in(args.batch_id))

        print(f"\nExcel exported to: {output_path}")
        print(f"  Rows: {len(df)}")
        print(f"  Columns: {len(df.columns)}")
        print()

        return 0

    except Exception as e:
        print(f"Error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
