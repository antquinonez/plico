#!/usr/bin/env python
# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""
Generate sample workbook for batch execution testing.

Creates 35 prompts with {{variable}} templating:
- Uses 5 data rows (batches) with different regions/products
- Tests variable resolution across all prompt types
- Tests dependency chains within batches

Uses FFLiteLLMClient with LiteLLM routing for Mistral Small.

Usage:
    python scripts/create_sample_workbook_batch.py [output_path]
"""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from src.config import get_config


def create_batch_sample_workbook(output_path: str):
    config = get_config()
    test_config = config.test
    batch_config = config.workbook.batch

    wb = Workbook()

    # ==========================================
    # CONFIG SHEET
    # ==========================================
    ws_config = wb.active
    ws_config.title = "config"
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"

    config_data = [
        ("model", test_config.default_model),
        ("max_retries", str(test_config.default_retries)),
        ("temperature", str(test_config.default_temperature)),
        ("max_tokens", str(test_config.default_max_tokens)),
        (
            "system_instructions",
            test_config.default_system_instructions,
        ),
        ("batch_mode", batch_config.mode),
        ("batch_output", batch_config.output),
        ("on_batch_error", batch_config.on_error),
        ("created_at", datetime.now().isoformat()),
    ]

    for idx, (field, value) in enumerate(config_data, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value

    ws_config.column_dimensions["A"].width = 20
    ws_config.column_dimensions["B"].width = 70

    # ==========================================
    # DATA SHEET (5 batches)
    # ==========================================
    ws_data = wb.create_sheet(title="data")

    data_headers = ["id", "batch_name", "region", "product", "price", "quantity"]
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)

    data_rows = [
        (1, "{{region}}_{{product}}", "north", "widget_a", 10, 100),
        (2, "{{region}}_{{product}}", "south", "widget_b", 15, 75),
        (3, "{{region}}_{{product}}", "east", "widget_c", 20, 50),
        (4, "{{region}}_{{product}}", "west", "widget_a", 12, 80),
        (5, "{{region}}_{{product}}", "central", "widget_b", 18, 60),
    ]

    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    ws_data.column_dimensions["A"].width = 8
    ws_data.column_dimensions["B"].width = 25
    ws_data.column_dimensions["C"].width = 12
    ws_data.column_dimensions["D"].width = 15
    ws_data.column_dimensions["E"].width = 10
    ws_data.column_dimensions["F"].width = 12

    # ==========================================
    # PROMPTS SHEET (35 prompts)
    # ==========================================
    ws_prompts = wb.create_sheet(title="prompts")
    headers = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "references",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    prompts = []
    row = 2

    # ==========================================
    # LEVEL 0: 10 Independent Prompts (uses variables)
    # ==========================================

    # Basic info prompts (5)
    prompts.append(
        (
            1,
            "intro",
            "You are analyzing sales data for {{region}} region, product {{product}}. Confirm by saying 'Analyzing {{region}} {{product}}'.",
            None,
        )
    )
    prompts.append(
        (
            2,
            "price_check",
            "The price of {{product}} is {{price}}. Just confirm the price.",
            None,
        )
    )
    prompts.append(
        (
            3,
            "quantity_check",
            "The quantity sold of {{product}} in {{region}} is {{quantity}}. Just confirm the quantity.",
            None,
        )
    )
    prompts.append(
        (
            4,
            "region_fact",
            "Name one interesting fact about the {{region}} region. Keep it brief.",
            None,
        )
    )
    prompts.append(
        (
            5,
            "product_desc",
            "Describe what {{product}} might be in 10 words or less.",
            None,
        )
    )

    # Math prompts with variables (5)
    prompts.append(
        (
            6,
            "calc_revenue",
            "Calculate total revenue: {{price}} times {{quantity}}. Just give the number.",
            None,
        )
    )
    prompts.append((7, "calc_double", "What is {{quantity}} times 2? Just give the number.", None))
    prompts.append(
        (
            8,
            "calc_half",
            "What is {{quantity}} divided by 2? Just give the number.",
            None,
        )
    )
    prompts.append(
        (
            9,
            "calc_price_plus_tax",
            "If tax is 10%, what is the final price for {{price}}? Just give the number.",
            None,
        )
    )
    prompts.append(
        (
            10,
            "calc_quantity_10pct",
            "What is 10% of {{quantity}}? Just give the number.",
            None,
        )
    )

    # ==========================================
    # LEVEL 1: 10 Prompts with Dependencies
    # ==========================================

    prompts.append(
        (
            11,
            "revenue_check",
            "Confirm: the revenue from calc_revenue should be price times quantity. Is this correct for {{region}}?",
            '["calc_revenue", "price_check", "quantity_check"]',
        )
    )
    prompts.append(
        (
            12,
            "double_check",
            "If we doubled sales in {{region}}, what would the new quantity be? Answer based on calc_double.",
            '["calc_double"]',
        )
    )
    prompts.append(
        (
            13,
            "half_check",
            "If we sold half as much in {{region}}, what would the quantity be? Based on calc_half.",
            '["calc_half"]',
        )
    )
    prompts.append(
        (
            14,
            "tax_check",
            "Is the taxed price from calc_price_plus_tax higher than {{price}}? Answer yes or no.",
            '["calc_price_plus_tax", "price_check"]',
        )
    )
    prompts.append(
        (
            15,
            "pct_check",
            "10% of our {{quantity}} units is how many? Confirm with calc_quantity_10pct.",
            '["calc_quantity_10pct", "quantity_check"]',
        )
    )

    # Analysis prompts (5)
    prompts.append(
        (
            16,
            "market_size",
            "Based on {{quantity}} units in {{region}}, is this a large or small market? Answer briefly.",
            '["quantity_check"]',
        )
    )
    prompts.append(
        (
            17,
            "pricing_analysis",
            "At {{price}} per unit, is {{product}} premium or budget? Answer briefly.",
            '["price_check", "product_desc"]',
        )
    )
    prompts.append(
        (
            18,
            "region_potential",
            "Given {{region}} region facts, what's the growth potential? Answer in one sentence.",
            '["region_fact", "quantity_check"]',
        )
    )
    prompts.append(
        (
            19,
            "product_fit",
            "Does {{product}} fit well with {{region}} region? Answer yes/no with brief reason.",
            '["product_desc", "region_fact"]',
        )
    )
    prompts.append(
        (
            20,
            "sales_trend",
            "If we project {{quantity}} units growing 10% (calc_quantity_10pct), what's the outlook?",
            '["calc_quantity_10pct", "quantity_check"]',
        )
    )

    # ==========================================
    # LEVEL 2: 10 Prompts with Multiple Dependencies
    # ==========================================

    prompts.append(
        (
            21,
            "revenue_analysis",
            "Analyze the revenue situation for {{region}} {{product}} using revenue_check and pricing_analysis.",
            '["revenue_check", "pricing_analysis"]',
        )
    )
    prompts.append(
        (
            22,
            "volume_analysis",
            "Analyze sales volume for {{region}} using double_check, half_check, and market_size.",
            '["double_check", "half_check", "market_size"]',
        )
    )
    prompts.append(
        (
            23,
            "competitiveness",
            "Is {{product}} at {{price}} competitive? Use pricing_analysis and product_fit.",
            '["pricing_analysis", "product_fit"]',
        )
    )
    prompts.append(
        (
            24,
            "growth_analysis",
            "Analyze growth potential using pct_check, sales_trend, and region_potential.",
            '["pct_check", "sales_trend", "region_potential"]',
        )
    )
    prompts.append(
        (
            25,
            "market_position",
            "Summarize market position using market_size and competitiveness.",
            '["market_size", "competitiveness"]',
        )
    )

    # Cross-analysis (5)
    prompts.append(
        (
            26,
            "profit_estimate",
            "If cost is 60% of {{price}}, estimate profit for {{quantity}} units. Give just the number.",
            '["calc_revenue", "price_check", "quantity_check"]',
        )
    )
    prompts.append(
        (
            27,
            "break_even",
            "If fixed costs are 1000, how many {{product}} units at {{price}} to break even? Give number.",
            '["price_check"]',
        )
    )
    prompts.append(
        (
            28,
            "margin_check",
            "At price {{price}} with estimated profit, what's the profit margin percentage?",
            '["profit_estimate", "price_check"]',
        )
    )
    prompts.append(
        (
            29,
            "volume_impact",
            "If quantity doubles (calc_double), what's the new revenue? Use calc_revenue as base.",
            '["calc_revenue", "calc_double"]',
        )
    )
    prompts.append(
        (
            30,
            "scenario_compare",
            "Compare current revenue (calc_revenue) vs doubled quantity (volume_impact). Which is higher?",
            '["calc_revenue", "volume_impact"]',
        )
    )

    # ==========================================
    # LEVEL 3: 5 Final Synthesis Prompts
    # ==========================================

    prompts.append(
        (
            31,
            "executive_summary",
            "Provide a 2-sentence executive summary for {{region}} {{product}} using revenue_analysis and market_position.",
            '["revenue_analysis", "market_position"]',
        )
    )
    prompts.append(
        (
            32,
            "recommendations",
            "Based on growth_analysis and competitiveness, give 2 recommendations for {{region}}.",
            '["growth_analysis", "competitiveness"]',
        )
    )
    prompts.append(
        (
            33,
            "risk_assessment",
            "Assess risks for {{product}} in {{region}} using market_position and scenario_compare.",
            '["market_position", "scenario_compare"]',
        )
    )
    prompts.append(
        (
            34,
            "action_plan",
            "Create a 3-step action plan for {{region}} using all previous analyses.",
            '["executive_summary", "recommendations", "risk_assessment"]',
        )
    )
    prompts.append(
        (
            35,
            "final_score",
            "On a scale of 1-10, score the {{region}} {{product}} business potential. Just give the number.",
            '["executive_summary", "market_position", "growth_analysis"]',
        )
    )

    # Write all prompts to sheet
    for seq, name, prompt, history in prompts:
        ws_prompts.cell(row=row, column=1, value=seq)
        ws_prompts.cell(row=row, column=2, value=name)
        ws_prompts.cell(row=row, column=3, value=prompt)
        ws_prompts.cell(row=row, column=4, value=history if history else "")
        ws_prompts.cell(row=row, column=5, value="")
        ws_prompts.cell(row=row, column=6, value="")
        ws_prompts.cell(row=row, column=7, value="")
        row += 1

    ws_prompts.column_dimensions["A"].width = 10
    ws_prompts.column_dimensions["B"].width = 20
    ws_prompts.column_dimensions["C"].width = 90
    ws_prompts.column_dimensions["D"].width = 50
    ws_prompts.column_dimensions["E"].width = 12
    ws_prompts.column_dimensions["F"].width = 15
    ws_prompts.column_dimensions["G"].width = 15

    wb.save(output_path)

    print(f"\n{'=' * 70}")
    print(f"Created BATCH sample workbook: {output_path}")
    print(f"{'=' * 70}")
    print(f"\nUsing: FFLiteLLMClient with LiteLLM routing")
    print(f"Total prompts: {len(prompts)}")
    print("Total batches: 5")
    print("\nData Variables:")
    print("  - region: north, south, east, west, central")
    print("  - product: widget_a, widget_b, widget_c")
    print("  - price: 10, 12, 15, 18, 20")
    print("  - quantity: 50, 60, 75, 80, 100")
    print("\nPrompt Structure:")
    print("  Level 0: 10 independent prompts (sequences 1-10)")
    print("  Level 1: 10 prompts with 1-3 dependencies (sequences 11-20)")
    print("  Level 2: 10 prompts with 2-3 dependencies (sequences 21-30)")
    print("  Level 3: 5 synthesis prompts (sequences 31-35)")
    print(f"\nTotal executions: {len(prompts)} prompts x 5 batches = {len(prompts) * 5} total")
    print(f"\n{'=' * 70}")
    print(f"Run with: python scripts/run_orchestrator.py {output_path} -c 3")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    config = get_config()
    output = sys.argv[1] if len(sys.argv) > 1 else config.test.workbooks.batch
    create_batch_sample_workbook(output)
