
import os
import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

class WorkbookManager:
    """Manages Excel workbook creation and updates based on a source Parquet file."""

    def __init__(self, parquet_path: str):
        if not os.path.isfile(parquet_path):
            raise FileNotFoundError(f"Parquet file not found: {parquet_path}")

        self.parquet_path = parquet_path
        self.base_name = os.path.splitext(os.path.basename(parquet_path))[0]
        self.workbook_path = os.path.join(os.path.dirname(parquet_path), f"{self.base_name}.xlsx")

    def create_or_update_workbook(self):
        """Create a new workbook with 'input' sheet if not exists, or add a result sheet."""
        # Read parquet file
        df = pl.read_parquet(self.parquet_path)
        if df.is_empty():
            raise ValueError("Parquet file is empty")

        # Validate datetime column
        if "datetime" not in df.columns:
            raise ValueError("Parquet file does not contain 'datetime' column")

        # Get latest datetime for naming result sheet
        latest_dt = df["datetime"].max()
        if isinstance(latest_dt, datetime):
            timestamp_str = latest_dt.strftime("%d%m%y_%H%M%S")
        else:
            # Convert string to datetime if needed
            timestamp_str = datetime.fromisoformat(str(latest_dt)).strftime("%d%m%y_%H%M%S")

        sheet_name = f"result_{timestamp_str}"

        # Check if workbook exists
        if not os.path.exists(self.workbook_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "input"
            wb.save(self.workbook_path)
            print(f"Workbook created with 'input' sheet: {self.workbook_path}")

        # Load workbook and add result sheet
        wb = load_workbook(self.workbook_path)
        if sheet_name in wb.sheetnames:
            print(f"Sheet {sheet_name} already exists. Skipping creation.")
        else:
            ws = wb.create_sheet(title=sheet_name)
            # Write headers
            headers = df.columns
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Write data rows
            for row_idx, row in enumerate(df.rows(), start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, start=1):
                max_length = max(len(str(header)), *(len(str(row[col_idx-1])) for row in df.rows()))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            wb.save(self.workbook_path)
            print(f"Added sheet '{sheet_name}' to workbook: {self.workbook_path}")
