
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class PromptWorkbookManager:
    """Manages an Excel workbook for prompt engineering with columns:
    sequence_number, model, prompt_name, prompt, history."""

    HEADERS = ["sequence_number", "model", "prompt_name", "prompt", "history"]

    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path

    def create_or_update_workbook(self):
        """Create the workbook with an 'input' sheet and required headers if missing."""
        if not os.path.exists(self.workbook_path):
            # Create new workbook with headers
            wb = Workbook()
            ws = wb.active
            ws.title = "input"
            for col_idx, header in enumerate(self.HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            for col_idx, header in enumerate(self.HEADERS, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = len(header) + 2
            wb.save(self.workbook_path)
            print(f"Workbook created with 'input' sheet: {self.workbook_path}")
            return

        # If workbook exists, check if 'input' sheet has headers
        wb = load_workbook(self.workbook_path)
        if "input" not in wb.sheetnames:
            ws = wb.create_sheet(title="input")
            for col_idx, header in enumerate(self.HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            for col_idx, header in enumerate(self.HEADERS, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = len(header) + 2
            wb.save(self.workbook_path)
            print(f"Added 'input' sheet with headers to workbook: {self.workbook_path}")
            return

        ws = wb["input"]
        # Check if first row contains headers
        existing_headers = [ws.cell(row=1, column=i).value for i in range(1, len(self.HEADERS)+1)]
        if all(h is None for h in existing_headers):
            for col_idx, header in enumerate(self.HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            for col_idx, header in enumerate(self.HEADERS, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = len(header) + 2
            wb.save(self.workbook_path)
            print(f"Updated 'input' sheet with headers in workbook: {self.workbook_path}")
        else:
            print(f"Workbook already has headers in 'input' sheet: {self.workbook_path}")
