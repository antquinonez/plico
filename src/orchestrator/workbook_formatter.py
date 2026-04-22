# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Centralized workbook formatting utilities.

Provides consistent formatting across all workbook creation:
- Column widths from config
- Word wrap for text-heavy columns
- Frozen header rows
- Auto-filter
- Row height auto-fit
"""

from __future__ import annotations

import logging
import math
from typing import Any

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..config import get_config

logger = logging.getLogger(__name__)


class WorkbookFormatter:
    """Apply consistent formatting to workbook sheets."""

    def __init__(self, config: Any = None):
        """Initialize formatter with config."""
        self.config = config or get_config()
        self.fmt_config = self.config.workbook.formatting

    def apply_formatting(self, ws: Worksheet, sheet_name: str) -> None:
        """Apply all formatting to a worksheet.

        Applies:
        - Column widths
        - Word wrap
        - Freeze panes
        - Auto-filter
        - Row height auto-fit (if enabled)
        """
        self.set_column_widths(ws, sheet_name)
        self.apply_vertical_top(ws)

        if self.fmt_config.word_wrap.get("enabled", True):
            self.apply_word_wrap(ws, sheet_name)

        if self.fmt_config.features.freeze_panes.get("enabled", True):
            self.freeze_header_row(ws)

        if self.fmt_config.features.auto_filter.get("enabled", True):
            self.apply_auto_filter(ws)

        if self.fmt_config.rows.get("auto_fit_height", True):
            self.auto_fit_row_heights(ws, sheet_name)

    def set_column_widths(self, ws: Worksheet, sheet_name: str) -> None:
        """Set column widths from config."""
        widths = self.fmt_config.column_widths.get(sheet_name, {})

        if widths == "auto":
            self._auto_fit_columns(ws)
        elif isinstance(widths, dict):
            for col_name, width in widths.items():
                col_idx = self._find_column_by_header(ws, col_name)
                if col_idx:
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = width

    def apply_vertical_top(self, ws: Worksheet) -> None:
        """Apply vertical top alignment to all data cells."""
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row_cells:
                if cell.value is not None:
                    cell.alignment = Alignment(
                        wrap_text=cell.alignment.wrap_text,
                        vertical="top",
                        horizontal="left",
                    )

    def apply_word_wrap(self, ws: Worksheet, sheet_name: str) -> None:
        """Apply word wrap to specified columns."""
        wrap_columns = self.fmt_config.word_wrap.get("columns", {}).get(sheet_name, [])

        if wrap_columns == "auto":
            wrap_columns = self._detect_text_columns(ws)

        for col_name in wrap_columns:
            col_idx = self._find_column_by_header(ws, col_name)
            if col_idx:
                wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
                for row_cells in ws.iter_rows(
                    min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
                ):
                    row_cells[0].alignment = wrap_align

    def freeze_header_row(self, ws: Worksheet) -> None:
        """Freeze the first row (header row)."""
        ws.freeze_panes = "A2"

    def apply_auto_filter(self, ws: Worksheet) -> None:
        """Apply auto-filter to all columns."""
        if ws.max_row > 1 and ws.max_column > 0:
            end_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{end_col}{ws.max_row}"

    def auto_fit_row_heights(self, ws: Worksheet, sheet_name: str) -> None:
        """Auto-fit row heights for wrapped text."""
        wrap_columns = self.fmt_config.word_wrap.get("columns", {}).get(sheet_name, [])

        if wrap_columns == "auto":
            wrap_columns = self._detect_text_columns(ws)

        if not wrap_columns:
            return

        multiplier = self.fmt_config.rows.get("wrap_text_height_multiplier", 15)

        for row in range(2, ws.max_row + 1):
            max_lines = 1
            for col_name in wrap_columns:
                col_idx = self._find_column_by_header(ws, col_name)
                if col_idx:
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value:
                        col_letter = get_column_letter(col_idx)
                        col_width = ws.column_dimensions[col_letter].width or 50
                        estimated_lines = self._estimate_text_lines(str(cell.value), col_width)
                        max_lines = max(max_lines, estimated_lines)

            ws.row_dimensions[row].height = max_lines * multiplier

    def _estimate_text_lines(self, text: str, col_width: float) -> int:
        """Estimate the number of display lines for text in a given column width.

        Splits text by newlines into paragraphs, then estimates how many
        display lines each paragraph occupies based on word-wrap at the
        effective column width. Empty paragraphs (blank lines) count as 1 line.

        """
        chars_per_line = max(1, col_width * 0.9)
        lines = 0
        for paragraph in text.split("\n"):
            if not paragraph:
                lines += 1
            else:
                lines += math.ceil(len(paragraph) / chars_per_line)
        return max(1, lines)

    def _find_column_by_header(self, ws: Worksheet, header_name: str) -> int | None:
        """Find column index by header name."""
        header_row = list(ws.iter_rows(min_row=1, max_row=1))
        if not header_row:
            return None
        for cell in header_row[0]:
            if cell.value == header_name:
                return cell.column
        return None

    def _auto_fit_columns(self, ws: Worksheet) -> None:
        """Auto-fit all columns based on content."""
        for col in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col)

            for row in range(1, min(100, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(100, max(max_length + 2, 12))

    def _detect_text_columns(self, ws: Worksheet) -> list[str]:
        """Detect columns that likely contain long text."""
        text_columns = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if not header:
                continue

            avg_length = 0
            sample_count = min(10, ws.max_row - 1)
            for row in range(2, min(12, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    avg_length += len(str(cell.value))

            if sample_count > 0:
                avg_length = avg_length / sample_count
                if avg_length > 30:
                    text_columns.append(str(header))

        return text_columns


def format_workbook(wb: Any, config: Any = None) -> None:
    """Format all sheets in a workbook.

    Args:
        wb: openpyxl Workbook object
        config: Optional config object

    """
    formatter = WorkbookFormatter(config)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        formatter.apply_formatting(ws, sheet_name)
