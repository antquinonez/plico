# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Auto-discovery utility for document evaluation workbooks.

Scans folders for documents and generates workbook structures (documents
sheet, data sheet, skeleton scoring/synthesis sheets). Used to bootstrap
evaluation workbooks from a folder of files (resumes, contracts, etc.).
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from ..config import get_config
from .workbook_formatter import WorkbookFormatter
from .workbook_parser import WorkbookParser

logger = logging.getLogger(__name__)

DEFAULT_EXTENSIONS: set[str] = {".pdf", ".docx", ".doc", ".txt", ".md"}

_FILENAME_SANITIZE_RE = re.compile(r"[^a-z0-9_]+")


def _sanitize_for_reference(name: str) -> str:
    """Convert a name to a snake_case reference_name.

    Strips extension, lowercases, replaces non-alphanumeric chars with
    underscores, collapses consecutive underscores, strips leading/trailing.

    Args:
        name: Raw filename stem (e.g., "Alice Chen CV").

    Returns:
        Sanitized reference name (e.g., "alice_chen_cv").

    """
    stem = Path(name).stem
    return _FILENAME_SANITIZE_RE.sub("_", stem.lower()).strip("_")


def _derive_names(filepath: Path) -> tuple[str, str, str]:
    """Derive reference_name, common_name, and batch_name from a file path.

    Args:
        filepath: Path to the document file.

    Returns:
        Tuple of (reference_name, common_name, batch_name).

    """
    stem = filepath.stem
    return (
        _sanitize_for_reference(stem),
        stem,
        stem,
    )


def discover_documents(
    folder_path: str | Path,
    extensions: set[str] | None = None,
    tags: list[str] | None = None,
) -> list[dict[str, Any]]:
    """Scan a folder for documents and return definitions matching DOCUMENTS_HEADERS.

    Files are sorted alphabetically by filename stem. Unsupported file
    extensions are skipped with a warning.

    Args:
        folder_path: Path to the folder containing documents.
        extensions: Set of file extensions to include. Defaults to
            ``{".pdf", ".docx", ".doc", ".txt", ".md"}``.
        tags: Optional tags to assign to all discovered documents.

    Returns:
        List of dicts with keys matching ``DOCUMENTS_HEADERS``:
        ``reference_name``, ``common_name``, ``file_path``, ``tags``, ``notes``.

    """
    folder = Path(folder_path)
    if not folder.is_dir():
        raise ValueError(f"Folder does not exist: {folder_path}")

    exts = extensions or DEFAULT_EXTENSIONS
    tags_str = ", ".join(tags) if tags else ""

    document_defs: list[dict[str, Any]] = []
    for filepath in sorted(folder.iterdir()):
        if not filepath.is_file():
            continue

        if filepath.suffix.lower() not in exts:
            logger.warning(f"Skipping unsupported file: {filepath.name}")
            continue

        reference_name, common_name, _batch_name = _derive_names(filepath)

        if not reference_name:
            logger.warning(f"Skipping file with empty reference name: {filepath.name}")
            continue

        relative_path = filepath.relative_to(folder).as_posix()

        document_defs.append(
            {
                "reference_name": reference_name,
                "common_name": common_name,
                "file_path": relative_path,
                "tags": tags_str,
                "chunking_strategy": "",
                "notes": "",
            }
        )

    logger.info(
        f"Discovered {len(document_defs)} documents in {folder} "
        f"(extensions: {', '.join(sorted(exts))})"
    )
    return document_defs


def create_data_rows_from_documents(
    document_defs: list[dict[str, Any]],
    documents_column: str = "_documents",
) -> list[dict[str, Any]]:
    """Generate data rows for batch execution from document definitions.

    Each document becomes one data row with ``id``, ``batch_name``,
    ``candidate_name``, and ``_documents`` columns.

    Args:
        document_defs: List of dicts as returned by ``discover_documents()``
            or matching ``DOCUMENTS_HEADERS`` format.
        documents_column: Name of the column that binds documents to rows.

    Returns:
        List of dicts with keys: ``id``, ``batch_name``, ``candidate_name``,
        ``_documents``.

    """
    rows: list[dict[str, Any]] = []
    for idx, doc in enumerate(document_defs, start=1):
        ref_name = doc.get("reference_name", "")
        common_name = doc.get("common_name", ref_name)
        rows.append(
            {
                "id": idx,
                "batch_name": doc.get("reference_name", ""),
                "candidate_name": common_name,
                documents_column: f'["{ref_name}"]',
            }
        )

    logger.info(f"Created {len(rows)} data rows from documents")
    return rows


def create_evaluation_workbook(
    output_path: str | Path,
    documents_folder: str | Path,
    shared_documents: list[dict[str, Any]] | None = None,
    evaluation_strategy: str = "balanced",
    extensions: set[str] | None = None,
    tags: list[str] | None = None,
) -> str:
    """Create a complete evaluation workbook from a folder of documents.

    Creates an ``.xlsx`` workbook with:
    - **config** sheet: defaults plus ``evaluation_strategy`` and batch fields
    - **documents** sheet: discovered documents + shared documents
    - **data** sheet: one row per discovered document with ``_documents`` column
    - **scoring** sheet: headers only (user fills in criteria)
    - **synthesis** sheet: headers only (user fills in prompts)
    - **prompts** sheet: headers only (domain-specific, user-provided)

    Args:
        output_path: Path for the output ``.xlsx`` file.
        documents_folder: Folder to scan for documents.
        shared_documents: Optional list of additional document defs (e.g., a
            job description) matching ``WorkbookParser.DOCUMENTS_HEADERS``
            format. These are prepended to the documents sheet.
        evaluation_strategy: Strategy name to set in the config sheet.
        extensions: File extensions to include in discovery.
        tags: Tags to assign to discovered documents.

    Returns:
        Absolute path to the created workbook.

    """
    output = Path(output_path)
    config = get_config()
    sheet_names = config.workbook.sheet_names

    discovered = discover_documents(documents_folder, extensions, tags)
    if not discovered:
        raise ValueError(f"No documents found in {documents_folder}")

    all_documents = list(shared_documents or []) + discovered
    data_rows = create_data_rows_from_documents(discovered)

    wb = Workbook()

    # --- Config sheet ---
    ws_config = wb.active
    ws_config.title = sheet_names.config
    ws_config["A1"] = "field"
    ws_config["B1"] = "value"
    ws_config["C1"] = "notes"

    defaults = config.workbook.defaults
    batch = config.workbook.batch
    config_rows = [
        ("name", "", "Human-readable name for this process/workbook"),
        ("description", "", "Brief description of what this process does"),
        ("client_type", "", "AI client type from config/clients.yaml client_types"),
        ("model", defaults.model, "Model identifier (e.g., mistral-small-latest)"),
        ("api_key_env", defaults.api_key_env, "Environment variable name for API key"),
        ("max_retries", str(defaults.max_retries), "Maximum retry attempts (1-10)"),
        ("temperature", str(defaults.temperature), "Sampling temperature (0.0-2.0)"),
        ("max_tokens", str(defaults.max_tokens), "Maximum response tokens"),
        (
            "system_instructions",
            defaults.system_instructions,
            "System prompt for AI",
        ),
        ("created_at", datetime.now().isoformat(), "ISO timestamp when created"),
        ("evaluation_strategy", evaluation_strategy, "Evaluation strategy from config"),
        ("batch_mode", batch.mode, "Batch mode: per_row"),
        (
            "batch_output",
            batch.output,
            "Output format: combined or separate_sheets",
        ),
        ("on_batch_error", batch.on_error, "Error handling: continue or stop"),
    ]

    for idx, (field, value, notes) in enumerate(config_rows, start=2):
        ws_config[f"A{idx}"] = field
        ws_config[f"B{idx}"] = value
        ws_config[f"C{idx}"] = notes

    # --- Prompts sheet (headers only) ---
    ws_prompts = wb.create_sheet(title=sheet_names.prompts)
    for col_idx, header in enumerate(WorkbookParser.PROMPTS_HEADERS, start=1):
        ws_prompts.cell(row=1, column=col_idx, value=header)

    # --- Data sheet ---
    ws_data = wb.create_sheet(title=sheet_names.data)
    data_headers = list(data_rows[0].keys())
    for col_idx, header in enumerate(data_headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, header in enumerate(data_headers, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=row.get(header))

    # --- Documents sheet ---
    ws_docs = wb.create_sheet(title=sheet_names.documents)
    for col_idx, header in enumerate(WorkbookParser.DOCUMENTS_HEADERS, start=1):
        ws_docs.cell(row=1, column=col_idx, value=header)
    for row_idx, doc in enumerate(all_documents, start=2):
        for col_idx, header in enumerate(WorkbookParser.DOCUMENTS_HEADERS, start=1):
            ws_docs.cell(row=row_idx, column=col_idx, value=doc.get(header, ""))

    # --- Scoring sheet (headers only) ---
    ws_scoring = wb.create_sheet(title=sheet_names.scoring)
    for col_idx, header in enumerate(WorkbookParser.SCORING_HEADERS, start=1):
        ws_scoring.cell(row=1, column=col_idx, value=header)

    # --- Synthesis sheet (headers only) ---
    ws_synthesis = wb.create_sheet(title=sheet_names.synthesis)
    for col_idx, header in enumerate(WorkbookParser.SYNTHESIS_HEADERS, start=1):
        ws_synthesis.cell(row=1, column=col_idx, value=header)

    # --- Formatting ---
    formatter = WorkbookFormatter(config)
    formatter.apply_formatting(ws_config, "config")
    formatter.apply_formatting(ws_prompts, "prompts")
    formatter.apply_formatting(ws_data, "data")
    formatter.apply_formatting(ws_docs, "documents")
    formatter.apply_formatting(ws_scoring, "scoring")
    formatter.apply_formatting(ws_synthesis, "synthesis")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    logger.info(f"Evaluation workbook created: {output}")
    return str(output.resolve())
