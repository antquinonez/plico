import os
import json
import re
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class WorkbookBuilder:
    """Creates and validates Excel workbooks for prompt orchestration."""

    CONFIG_SHEET = "config"
    PROMPTS_SHEET = "prompts"
    DATA_SHEET = "data"
    CLIENTS_SHEET = "clients"

    CONFIG_FIELDS = [
        ("model", "mistral-small-2503"),
        ("api_key_env", "MISTRALSMALL_KEY"),
        ("max_retries", "3"),
        ("temperature", "0.8"),
        ("max_tokens", "4096"),
        (
            "system_instructions",
            "You are a helpful assistant. Respond accurately to user queries.",
        ),
        ("created_at", ""),
    ]

    BATCH_CONFIG_FIELDS = [
        ("batch_mode", "per_row"),
        ("batch_output", "combined"),
        ("on_batch_error", "continue"),
    ]

    PROMPTS_HEADERS = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
    ]
    REQUIRED_PROMPTS_HEADERS = ["sequence", "prompt_name", "prompt", "history"]
    CLIENTS_HEADERS = [
        "name",
        "client_type",
        "api_key_env",
        "model",
        "temperature",
        "max_tokens",
        "system_instructions",
    ]
    RESULTS_HEADERS = [
        "batch_id",
        "batch_name",
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "client",
        "condition",
        "condition_result",
        "condition_error",
        "response",
        "status",
        "attempts",
        "error",
    ]

    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self._has_data_sheet: Optional[bool] = None
        self._has_clients_sheet: Optional[bool] = None

    def has_data_sheet(self) -> bool:
        """Check if workbook has a data sheet for batch execution."""
        if self._has_data_sheet is not None:
            return self._has_data_sheet

        if not os.path.exists(self.workbook_path):
            self._has_data_sheet = False
            return False

        wb = load_workbook(self.workbook_path)
        self._has_data_sheet = self.DATA_SHEET in wb.sheetnames
        return self._has_data_sheet

    def has_clients_sheet(self) -> bool:
        """Check if workbook has a clients sheet for multi-client execution."""
        if self._has_clients_sheet is not None:
            return self._has_clients_sheet

        if not os.path.exists(self.workbook_path):
            self._has_clients_sheet = False
            return False

        wb = load_workbook(self.workbook_path)
        self._has_clients_sheet = self.CLIENTS_SHEET in wb.sheetnames
        return self._has_clients_sheet

    def create_template_workbook(
        self, with_data_sheet: bool = False, with_clients_sheet: bool = False
    ) -> str:
        """Create a new workbook with template structure."""
        logger.info(f"Creating template workbook: {self.workbook_path}")

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = self.CONFIG_SHEET
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"

        all_config = list(self.CONFIG_FIELDS)
        if with_data_sheet:
            all_config.extend(self.BATCH_CONFIG_FIELDS)

        for idx, (field, default) in enumerate(all_config, start=2):
            ws_config[f"A{idx}"] = field
            if field == "created_at":
                ws_config[f"B{idx}"] = datetime.now().isoformat()
            else:
                ws_config[f"B{idx}"] = default

        ws_config.column_dimensions["A"].width = 20
        ws_config.column_dimensions["B"].width = 60

        ws_prompts = wb.create_sheet(title=self.PROMPTS_SHEET)
        for col_idx, header in enumerate(self.PROMPTS_HEADERS, start=1):
            ws_prompts.cell(row=1, column=col_idx, value=header)
            ws_prompts.column_dimensions[get_column_letter(col_idx)].width = max(
                15, len(header) + 5
            )

        ws_prompts.column_dimensions["C"].width = 50
        ws_prompts.column_dimensions["D"].width = 30
        ws_prompts.column_dimensions["E"].width = 15
        ws_prompts.column_dimensions["F"].width = 40

        if with_data_sheet:
            ws_data = wb.create_sheet(title=self.DATA_SHEET)
            ws_data["A1"] = "id"
            ws_data["B1"] = "batch_name"
            ws_data.column_dimensions["A"].width = 10
            ws_data.column_dimensions["B"].width = 30

        if with_clients_sheet:
            ws_clients = wb.create_sheet(title=self.CLIENTS_SHEET)
            for col_idx, header in enumerate(self.CLIENTS_HEADERS, start=1):
                ws_clients.cell(row=1, column=col_idx, value=header)
                ws_clients.column_dimensions[get_column_letter(col_idx)].width = max(
                    15, len(header) + 5
                )

        wb.save(self.workbook_path)
        logger.info(f"Template workbook created: {self.workbook_path}")
        return self.workbook_path

    def validate_workbook(self) -> bool:
        """Validate workbook has required structure."""
        if not os.path.exists(self.workbook_path):
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        wb = load_workbook(self.workbook_path)

        if self.CONFIG_SHEET not in wb.sheetnames:
            raise ValueError(f"Missing required sheet: {self.CONFIG_SHEET}")

        if self.PROMPTS_SHEET not in wb.sheetnames:
            raise ValueError(f"Missing required sheet: {self.PROMPTS_SHEET}")

        ws_prompts = wb[self.PROMPTS_SHEET]

        actual_headers = []
        for col in range(1, ws_prompts.max_column + 1):
            header = ws_prompts.cell(row=1, column=col).value
            if header:
                actual_headers.append(str(header).strip().lower())

        normalized_required = [h.lower() for h in self.REQUIRED_PROMPTS_HEADERS]
        missing = set(normalized_required) - set(actual_headers)
        if missing:
            raise ValueError(f"Missing required columns in prompts sheet: {missing}")

        logger.info(f"Workbook validated: {self.workbook_path}")
        return True

    def load_config(self) -> Dict[str, Any]:
        """Load configuration from config sheet."""
        wb = load_workbook(self.workbook_path)
        ws = wb[self.CONFIG_SHEET]

        config = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                key = str(row[0]).strip()
                value = row[1]

                if key == "max_retries":
                    value = int(value)
                elif key == "temperature":
                    value = float(value)
                elif key == "max_tokens":
                    value = int(value)

                config[key] = value

        logger.debug(f"Loaded config: {config}")
        return config

    def load_prompts(self) -> List[Dict[str, Any]]:
        """Load prompts from prompts sheet."""
        wb = load_workbook(self.workbook_path)
        ws = wb[self.PROMPTS_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip().lower() if header else f"col_{col}")

        prompts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            row_dict = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_dict[header] = row[col_idx]

            prompt_data = {
                "sequence": int(row_dict.get("sequence"))
                if row_dict.get("sequence")
                else None,
                "prompt_name": str(row_dict.get("prompt_name", "")).strip()
                if row_dict.get("prompt_name")
                else None,
                "prompt": str(row_dict.get("prompt", "")).strip()
                if row_dict.get("prompt")
                else None,
                "history": self.parse_history_string(row_dict.get("history"))
                if row_dict.get("history")
                else None,
                "client": str(row_dict.get("client", "")).strip()
                if row_dict.get("client")
                else None,
                "condition": str(row_dict.get("condition", "")).strip()
                if row_dict.get("condition")
                else None,
            }

            if prompt_data["sequence"] and prompt_data["prompt"]:
                prompts.append(prompt_data)

        prompts.sort(key=lambda x: x["sequence"])
        logger.info(f"Loaded {len(prompts)} prompts")
        return prompts

    def load_data(self) -> List[Dict[str, Any]]:
        """Load batch data from data sheet."""
        if not self.has_data_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.DATA_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip() if header else f"col_{col}")

        data_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value is not None:
                    has_content = True

            if has_content:
                row_data["_row_idx"] = row_idx
                data_rows.append(row_data)

        logger.info(f"Loaded {len(data_rows)} data rows for batch execution")
        return data_rows

    def get_data_columns(self) -> List[str]:
        """Get column names from data sheet."""
        if not self.has_data_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.DATA_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header).strip())

        return headers

    def load_clients(self) -> List[Dict[str, Any]]:
        """
        Load client configurations from clients sheet.

        Returns:
            List of client config dictionaries with keys:
            - name: unique identifier for this client
            - client_type: type of client (e.g., "mistral-small", "anthropic")
            - api_key_env: environment variable for API key (optional)
            - model: model override (optional)
            - temperature: temperature override (optional)
            - max_tokens: max tokens override (optional)
            - system_instructions: system prompt override (optional)
        """
        if not self.has_clients_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.CLIENTS_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip() if header else f"col_{col}")

        clients = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value is not None:
                    has_content = True

            if has_content and row_data.get("name") and row_data.get("client_type"):
                clients.append(row_data)

        logger.info(f"Loaded {len(clients)} client configurations")
        return clients

    def parse_history_string(self, history_str: Any) -> Optional[List[str]]:
        """Parse history string like '["a", "b"]' into list."""
        if not history_str:
            return None

        if isinstance(history_str, list):
            return history_str

        s = str(history_str).strip()

        if s.startswith("[") and s.endswith("]"):
            try:
                parsed = json.loads(s)
                if isinstance(parsed, list):
                    return [str(item).strip().strip("\"'") for item in parsed]
            except json.JSONDecodeError:
                pass

            inner = s[1:-1]
            items = re.findall(r'["\']([^"\']+)["\']', inner)
            if items:
                return items

            items = [x.strip().strip("\"'") for x in inner.split(",") if x.strip()]
            if items:
                return items

        return [s.strip().strip("\"'")]

    def write_results(self, results: List[Dict[str, Any]], sheet_name: str) -> str:
        """Write execution results to a new sheet."""
        wb = load_workbook(self.workbook_path)

        if sheet_name in wb.sheetnames:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            sheet_name = f"results_{timestamp}"

        ws = wb.create_sheet(title=sheet_name)

        for col_idx, header in enumerate(self.RESULTS_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, result in enumerate(results, start=2):
            ws.cell(row=row_idx, column=1, value=result.get("batch_id"))
            ws.cell(row=row_idx, column=2, value=result.get("batch_name"))
            ws.cell(row=row_idx, column=3, value=result.get("sequence"))
            ws.cell(row=row_idx, column=4, value=result.get("prompt_name"))
            ws.cell(row=row_idx, column=5, value=result.get("prompt"))

            history = result.get("history")
            history_str = json.dumps(history) if history else ""
            ws.cell(row=row_idx, column=6, value=history_str)

            ws.cell(row=row_idx, column=7, value=result.get("client"))
            ws.cell(row=row_idx, column=8, value=result.get("condition"))
            ws.cell(row=row_idx, column=9, value=result.get("condition_result"))
            ws.cell(row=row_idx, column=10, value=result.get("condition_error"))
            ws.cell(row=row_idx, column=11, value=result.get("response"))
            ws.cell(row=row_idx, column=12, value=result.get("status"))
            ws.cell(row=row_idx, column=13, value=result.get("attempts"))
            ws.cell(row=row_idx, column=14, value=result.get("error"))

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 30
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 40
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 25
        ws.column_dimensions["K"].width = 60
        ws.column_dimensions["L"].width = 10
        ws.column_dimensions["M"].width = 10
        ws.column_dimensions["N"].width = 40

        wb.save(self.workbook_path)
        logger.info(f"Results written to sheet: {sheet_name}")
        return sheet_name

    def write_batch_results(
        self,
        results: List[Dict[str, Any]],
        batch_name: str,
        base_sheet_name: str = "results",
    ) -> str:
        """Write results for a single batch to a separate sheet."""
        wb = load_workbook(self.workbook_path)

        sheet_name = f"{base_sheet_name}_{batch_name}"
        original_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "sequence",
            "prompt_name",
            "prompt",
            "history",
            "condition",
            "condition_result",
            "response",
            "status",
            "attempts",
            "error",
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, result in enumerate(results, start=2):
            ws.cell(row=row_idx, column=1, value=result.get("sequence"))
            ws.cell(row=row_idx, column=2, value=result.get("prompt_name"))
            ws.cell(row=row_idx, column=3, value=result.get("prompt"))

            history = result.get("history")
            history_str = json.dumps(history) if history else ""
            ws.cell(row=row_idx, column=4, value=history_str)

            ws.cell(row=row_idx, column=5, value=result.get("condition"))
            ws.cell(row=row_idx, column=6, value=result.get("condition_result"))
            ws.cell(row=row_idx, column=7, value=result.get("response"))
            ws.cell(row=row_idx, column=8, value=result.get("status"))
            ws.cell(row=row_idx, column=9, value=result.get("attempts"))
            ws.cell(row=row_idx, column=10, value=result.get("error"))

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 60
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 40

        wb.save(self.workbook_path)
        logger.info(f"Batch results written to sheet: {sheet_name}")
        return sheet_name
