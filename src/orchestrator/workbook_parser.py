# Copyright (c) 2025 Antonio Quinonez / Far Finer LLC
# SPDX-License-Identifier: MIT
# Contact: antquinonez@farfiner.com

"""Excel workbook parser for prompt orchestration.

Provides utilities for validating, reading, and writing orchestrator workbooks
with config, prompts, data, clients, and documents sheets.
"""

from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from ..config import get_config
from .models import ConfigSpec, DocumentSpec, PromptSpec
from .results.frame import (
    PIVOT_COLUMNS,
    RESULTS_COLUMNS,
    _serialize_for_excel,
)
from .workbook_formatter import WorkbookFormatter

logger = logging.getLogger(__name__)


def parse_history_string(history_str: Any) -> list[str] | None:
    """Parse history string like '["a", "b"]' into list.

    Handles smart quotes (curly quotes) by normalizing them to ASCII quotes
    before parsing. This allows users to copy/paste from Word, Google Docs, etc.
    """
    if not history_str:
        return None

    if isinstance(history_str, list):
        return history_str

    s = str(history_str).strip()

    quote_map = {
        0x201C: 0x22,
        0x201D: 0x22,
        0x2018: 0x27,
        0x2019: 0x27,
        0x201E: 0x22,
        0x201F: 0x22,
    }
    s = s.translate(quote_map)

    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [str(item).strip().strip("\"'") for item in parsed]
        except json.JSONDecodeError:
            pass

        inner = s[1:-1]
        items = re.findall(r'["\']([^"\']+)["\']', inner)
        if items:
            return items

        items = [x.strip().strip("\"'") for x in inner.split(",") if x.strip()]
        if items:
            return items

    return [s.strip().strip("\"'")]


class WorkbookParser:
    """Parses and validates Excel workbooks for prompt orchestration."""

    def __init__(self, workbook_path: str) -> None:
        """Initialize the WorkbookParser.

        Args:
            workbook_path: Path to the Excel workbook file.

        """
        self.workbook_path = workbook_path
        self._has_data_sheet: bool | None = None
        self._has_clients_sheet: bool | None = None
        self._has_documents_sheet: bool | None = None
        self._has_tools_sheet: bool | None = None
        self._has_scoring_sheet: bool | None = None
        self._has_synthesis_sheet: bool | None = None
        self._config = get_config()
        self.formatter = WorkbookFormatter(self._config)

    @property
    def CONFIG_SHEET(self) -> str:
        return self._config.workbook.sheet_names.config

    @property
    def PROMPTS_SHEET(self) -> str:
        return self._config.workbook.sheet_names.prompts

    @property
    def DATA_SHEET(self) -> str:
        return self._config.workbook.sheet_names.data

    @property
    def CLIENTS_SHEET(self) -> str:
        return self._config.workbook.sheet_names.clients

    @property
    def DOCUMENTS_SHEET(self) -> str:
        return self._config.workbook.sheet_names.documents

    @property
    def TOOLS_SHEET(self) -> str:
        return self._config.workbook.sheet_names.tools

    @property
    def SCORING_SHEET(self) -> str:
        return self._config.workbook.sheet_names.scoring

    @property
    def SYNTHESIS_SHEET(self) -> str:
        return self._config.workbook.sheet_names.synthesis

    @property
    def CONFIG_FIELDS(self) -> list[tuple[str, str, str]]:
        defaults = self._config.workbook.defaults
        return [
            ("name", "", "Human-readable name for this process/workbook"),
            ("description", "", "Brief description of what this process does"),
            ("client_type", "", "AI client type from config/clients.yaml client_types"),
            ("model", defaults.model, "Model identifier (e.g., mistral-small-latest)"),
            ("api_key_env", defaults.api_key_env, "Environment variable name for API key"),
            ("max_retries", str(defaults.max_retries), "Maximum retry attempts (1-10)"),
            ("temperature", str(defaults.temperature), "Sampling temperature (0.0-2.0)"),
            ("max_tokens", str(defaults.max_tokens), "Maximum response tokens"),
            ("system_instructions", defaults.system_instructions, "System prompt for AI"),
            ("created_at", "", "ISO timestamp when created"),
            ("evaluation_strategy", "", "Evaluation strategy name from config/main.yaml"),
        ]

    @property
    def BATCH_CONFIG_FIELDS(self) -> list[tuple[str, str, str]]:
        batch = self._config.workbook.batch
        return [
            (
                "batch_mode",
                batch.mode,
                "Batch execution mode: 'per_row' (execute for each data row)",
            ),
            (
                "batch_output",
                batch.output,
                "Output format: 'combined' (single sheet) or 'separate_sheets'",
            ),
            (
                "on_batch_error",
                batch.on_error,
                "Error handling: 'continue' (skip failed) or 'stop' (halt on error)",
            ),
        ]

    PROMPTS_HEADERS = [
        "sequence",
        "prompt_name",
        "prompt",
        "history",
        "notes",
        "client",
        "condition",
        "abort_condition",
        "references",
        "semantic_query",
        "semantic_filter",
        "query_expansion",
        "rerank",
        "agent_mode",
        "tools",
        "max_tool_rounds",
        "validation_prompt",
        "max_validation_retries",
        "phase",
        "generator",
    ]
    REQUIRED_PROMPTS_HEADERS = ["sequence", "prompt_name", "prompt", "history"]
    DOCUMENTS_HEADERS = [
        "reference_name",
        "common_name",
        "file_path",
        "tags",
        "chunking_strategy",
        "notes",
    ]
    CLIENTS_HEADERS = [
        "name",
        "client_type",
        "api_key_env",
        "model",
        "temperature",
        "max_tokens",
        "system_instructions",
    ]
    RESULTS_HEADERS = [
        "batch_id",
        "batch_name",
        *[c for c in RESULTS_COLUMNS if c not in ("batch_id", "batch_name")],
    ]
    TOOLS_HEADERS = [
        "name",
        "description",
        "parameters",
        "implementation",
        "enabled",
    ]
    SCORING_HEADERS = [
        "criteria_name",
        "description",
        "scale_min",
        "scale_max",
        "weight",
        "source_prompt",
        "score_type",
        "label_1",
        "label_2",
        "label_3",
    ]

    SYNTHESIS_HEADERS = [
        "sequence",
        "prompt_name",
        "prompt",
        "source_scope",
        "source_prompts",
        "include_scores",
        "history",
        "condition",
    ]

    def has_data_sheet(self) -> bool:
        """Check if workbook has a data sheet for batch execution."""
        if self._has_data_sheet is not None:
            return self._has_data_sheet

        if not os.path.exists(self.workbook_path):
            self._has_data_sheet = False
            return False

        wb = load_workbook(self.workbook_path)
        self._has_data_sheet = self.DATA_SHEET in wb.sheetnames
        return self._has_data_sheet

    def has_clients_sheet(self) -> bool:
        """Check if workbook has a clients sheet for multi-client execution."""
        if self._has_clients_sheet is not None:
            return self._has_clients_sheet

        if not os.path.exists(self.workbook_path):
            self._has_clients_sheet = False
            return False

        wb = load_workbook(self.workbook_path)
        self._has_clients_sheet = self.CLIENTS_SHEET in wb.sheetnames
        return self._has_clients_sheet

    def has_documents_sheet(self) -> bool:
        """Check if workbook has a documents sheet for document references."""
        if self._has_documents_sheet is not None:
            return self._has_documents_sheet

        if not os.path.exists(self.workbook_path):
            self._has_documents_sheet = False
            return False

        wb = load_workbook(self.workbook_path)
        self._has_documents_sheet = self.DOCUMENTS_SHEET in wb.sheetnames
        return self._has_documents_sheet

    def has_tools_sheet(self) -> bool:
        """Check if workbook has a tools sheet for agentic execution."""
        if self._has_tools_sheet is not None:
            return self._has_tools_sheet

        if not os.path.exists(self.workbook_path):
            self._has_tools_sheet = False
            return False

        wb = load_workbook(self.workbook_path)
        self._has_tools_sheet = self.TOOLS_SHEET in wb.sheetnames
        return self._has_tools_sheet

    def has_scoring_sheet(self) -> bool:
        """Check if workbook has a scoring sheet for evaluation."""
        if self._has_scoring_sheet is not None:
            return self._has_scoring_sheet

        if not os.path.exists(self.workbook_path):
            self._has_scoring_sheet = False
            return False

        wb = load_workbook(self.workbook_path)
        self._has_scoring_sheet = self.SCORING_SHEET in wb.sheetnames
        return self._has_scoring_sheet

    def has_synthesis_sheet(self) -> bool:
        """Check if workbook has a synthesis sheet for cross-row comparison."""
        if self._has_synthesis_sheet is not None:
            return self._has_synthesis_sheet

        if not os.path.exists(self.workbook_path):
            self._has_synthesis_sheet = False
            return False

        wb = load_workbook(self.workbook_path)
        self._has_synthesis_sheet = self.SYNTHESIS_SHEET in wb.sheetnames
        return self._has_synthesis_sheet

    def create_template_workbook(
        self,
        with_data_sheet: bool = False,
        with_clients_sheet: bool = False,
        with_documents_sheet: bool = False,
        with_tools_sheet: bool = False,
        with_scoring_sheet: bool = False,
        with_synthesis_sheet: bool = False,
    ) -> str:
        """Create a new workbook with template structure."""
        logger.info(f"Creating template workbook: {self.workbook_path}")

        wb = Workbook()

        ws_config = wb.active
        ws_config.title = self.CONFIG_SHEET
        ws_config["A1"] = "field"
        ws_config["B1"] = "value"
        ws_config["C1"] = "notes"

        all_config = list(self.CONFIG_FIELDS)
        if with_data_sheet:
            all_config.extend(self.BATCH_CONFIG_FIELDS)

        for idx, (field, default, notes) in enumerate(all_config, start=2):
            ws_config[f"A{idx}"] = field
            if field == "created_at":
                ws_config[f"B{idx}"] = datetime.now().isoformat()
            else:
                ws_config[f"B{idx}"] = default
            ws_config[f"C{idx}"] = notes

        self.formatter.apply_formatting(ws_config, "config")

        ws_prompts = wb.create_sheet(title=self.PROMPTS_SHEET)
        for col_idx, header in enumerate(self.PROMPTS_HEADERS, start=1):
            ws_prompts.cell(row=1, column=col_idx, value=header)

        self.formatter.apply_formatting(ws_prompts, "prompts")

        if with_data_sheet:
            ws_data = wb.create_sheet(title=self.DATA_SHEET)
            ws_data["A1"] = "id"
            ws_data["B1"] = "batch_name"
            self.formatter.apply_formatting(ws_data, "data")

        if with_clients_sheet:
            ws_clients = wb.create_sheet(title=self.CLIENTS_SHEET)
            for col_idx, header in enumerate(self.CLIENTS_HEADERS, start=1):
                ws_clients.cell(row=1, column=col_idx, value=header)
            self.formatter.apply_formatting(ws_clients, "clients")

        if with_documents_sheet:
            ws_documents = wb.create_sheet(title=self.DOCUMENTS_SHEET)
            for col_idx, header in enumerate(self.DOCUMENTS_HEADERS, start=1):
                ws_documents.cell(row=1, column=col_idx, value=header)
            self.formatter.apply_formatting(ws_documents, "documents")

        if with_tools_sheet:
            ws_tools = wb.create_sheet(title=self.TOOLS_SHEET)
            for col_idx, header in enumerate(self.TOOLS_HEADERS, start=1):
                ws_tools.cell(row=1, column=col_idx, value=header)
            self.formatter.apply_formatting(ws_tools, "tools")

        if with_scoring_sheet:
            ws_scoring = wb.create_sheet(title=self.SCORING_SHEET)
            for col_idx, header in enumerate(self.SCORING_HEADERS, start=1):
                ws_scoring.cell(row=1, column=col_idx, value=header)
            self.formatter.apply_formatting(ws_scoring, "scoring")

        if with_synthesis_sheet:
            ws_synthesis = wb.create_sheet(title=self.SYNTHESIS_SHEET)
            for col_idx, header in enumerate(self.SYNTHESIS_HEADERS, start=1):
                ws_synthesis.cell(row=1, column=col_idx, value=header)
            self.formatter.apply_formatting(ws_synthesis, "synthesis")

        wb.save(self.workbook_path)
        logger.info(f"Template workbook created: {self.workbook_path}")
        return self.workbook_path

    def validate_workbook(self) -> bool:
        """Validate workbook has required structure."""
        if not os.path.exists(self.workbook_path):
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        wb = load_workbook(self.workbook_path)

        if self.CONFIG_SHEET not in wb.sheetnames:
            raise ValueError(f"Missing required sheet: {self.CONFIG_SHEET}")

        if self.PROMPTS_SHEET not in wb.sheetnames:
            raise ValueError(f"Missing required sheet: {self.PROMPTS_SHEET}")

        ws_prompts = wb[self.PROMPTS_SHEET]

        actual_headers = []
        for col in range(1, ws_prompts.max_column + 1):
            header = ws_prompts.cell(row=1, column=col).value
            if header:
                actual_headers.append(str(header).strip().lower())

        normalized_required = [h.lower() for h in self.REQUIRED_PROMPTS_HEADERS]
        missing = set(normalized_required) - set(actual_headers)
        if missing:
            raise ValueError(f"Missing required columns in prompts sheet: {missing}")

        logger.info(f"Workbook validated: {self.workbook_path}")
        return True

    def load_config(self) -> ConfigSpec:
        """Load configuration from config sheet.

        Handles both 2-column format (field, value) and 3-column format
        (field, value, notes). The notes column is ignored if present.

        Returns:
            Dict of config field -> value pairs.

        """
        wb = load_workbook(self.workbook_path)
        ws = wb[self.CONFIG_SHEET]

        config = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            key = str(row[0]).strip()
            value = row[1] if len(row) > 1 else None

            if value is None:
                continue

            if key == "max_retries" or key == "max_tokens":
                value = int(value)
            elif key == "temperature":
                value = float(value)

            config[key] = value

        logger.debug(f"Loaded config: {config}")
        return config

    def validate_config(self, config: ConfigSpec) -> list[str]:
        """Validate config values and return list of error messages.

        Args:
            config: Configuration dictionary from load_config().

        Returns:
            List of error message strings. Empty list if valid.

        """
        errors: list[str] = []

        client_type = config.get("client_type")
        if client_type:
            available = self._config.get_available_client_types()
            if client_type not in available:
                errors.append(
                    f"Unknown client_type '{client_type}'. Available: {', '.join(sorted(available))}"
                )

        temperature = config.get("temperature")
        if temperature is not None:
            try:
                temp_float = float(temperature)
                if not (0.0 <= temp_float <= 2.0):
                    errors.append(f"temperature {temperature} out of range [0.0, 2.0]")
            except (TypeError, ValueError):
                errors.append(f"temperature '{temperature}' is not a valid number")

        max_retries = config.get("max_retries")
        if max_retries is not None:
            try:
                retries_int = int(max_retries)
                if not (1 <= retries_int <= 10):
                    errors.append(f"max_retries {max_retries} out of range [1, 10]")
            except (TypeError, ValueError):
                errors.append(f"max_retries '{max_retries}' is not a valid integer")

        batch_mode = config.get("batch_mode")
        if batch_mode and batch_mode not in ["per_row"]:
            errors.append(f"batch_mode '{batch_mode}' not supported. Use: per_row")

        batch_output = config.get("batch_output")
        if batch_output and batch_output not in ["combined", "separate_sheets"]:
            errors.append("batch_output must be 'combined' or 'separate_sheets'")

        on_batch_error = config.get("on_batch_error")
        if on_batch_error and on_batch_error not in ["continue", "stop"]:
            errors.append("on_batch_error must be 'continue' or 'stop'")

        return errors

    def load_prompts(self) -> list[PromptSpec]:
        """Load prompts from prompts sheet."""
        wb = load_workbook(self.workbook_path)
        ws = wb[self.PROMPTS_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip().lower() if header else f"col_{col}")

        prompts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            row_dict = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_dict[header] = row[col_idx]

            prompt_data = {
                "sequence": int(row_dict.get("sequence")) if row_dict.get("sequence") else None,
                "prompt_name": str(row_dict.get("prompt_name", "")).strip()
                if row_dict.get("prompt_name")
                else None,
                "prompt": str(row_dict.get("prompt", "")).strip()
                if row_dict.get("prompt")
                else None,
                "history": self._parse_history_string(row_dict.get("history"))
                if row_dict.get("history")
                else None,
                "notes": str(row_dict.get("notes", "")).strip() if row_dict.get("notes") else None,
                "client": str(row_dict.get("client", "")).strip()
                if row_dict.get("client")
                else None,
                "condition": str(row_dict.get("condition", "")).strip()
                if row_dict.get("condition")
                else None,
                "abort_condition": str(row_dict.get("abort_condition", "")).strip()
                if row_dict.get("abort_condition")
                else None,
                "references": self._parse_history_string(row_dict.get("references"))
                if row_dict.get("references")
                else None,
                "semantic_query": str(row_dict.get("semantic_query", "")).strip()
                if row_dict.get("semantic_query")
                else None,
                "semantic_filter": str(row_dict.get("semantic_filter", "")).strip()
                if row_dict.get("semantic_filter")
                else None,
                "query_expansion": str(row_dict.get("query_expansion", "")).strip().lower()
                if row_dict.get("query_expansion")
                else None,
                "rerank": str(row_dict.get("rerank", "")).strip().lower()
                if row_dict.get("rerank")
                else None,
                "agent_mode": str(row_dict.get("agent_mode", "")).strip().lower()
                in ("true", "1", "yes")
                if row_dict.get("agent_mode")
                else False,
                "tools": self._parse_history_string(row_dict.get("tools"))
                if row_dict.get("tools")
                else None,
                "max_tool_rounds": int(row_dict["max_tool_rounds"])
                if row_dict.get("max_tool_rounds")
                else None,
                "validation_prompt": str(row_dict.get("validation_prompt", "")).strip()
                if row_dict.get("validation_prompt")
                else None,
                "max_validation_retries": int(row_dict["max_validation_retries"])
                if row_dict.get("max_validation_retries")
                else None,
                "phase": str(row_dict.get("phase", "")).strip().lower()
                if row_dict.get("phase")
                else "execution",
                "generator": str(row_dict.get("generator", "")).strip().lower()
                in ("true", "1", "yes")
                if row_dict.get("generator")
                else False,
            }

            if prompt_data["sequence"] and prompt_data["prompt"]:
                prompts.append(prompt_data)

        prompts.sort(key=lambda x: x["sequence"])
        logger.info(f"Loaded {len(prompts)} prompts")
        return prompts

    def load_data(self) -> list[dict[str, Any]]:
        """Load batch data from data sheet."""
        if not self.has_data_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.DATA_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip() if header else f"col_{col}")

        data_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value is not None:
                    has_content = True

            if has_content:
                row_data["_row_idx"] = row_idx
                data_rows.append(row_data)

        logger.info(f"Loaded {len(data_rows)} data rows for batch execution")
        return data_rows

    def get_data_columns(self) -> list[str]:
        """Get column names from data sheet."""
        if not self.has_data_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.DATA_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header).strip())

        return headers

    def load_clients(self) -> list[dict[str, Any]]:
        """Load client configurations from clients sheet.

        Returns:
            List of client config dictionaries with keys:
            - name: unique identifier for this client
            - client_type: type of client (e.g., "mistral-small", "anthropic")
            - api_key_env: environment variable for API key (optional)
            - model: model override (optional)
            - temperature: temperature override (optional)
            - max_tokens: max tokens override (optional)
            - system_instructions: system prompt override (optional)

        """
        if not self.has_clients_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.CLIENTS_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip() if header else f"col_{col}")

        clients = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value is not None:
                    has_content = True

            if has_content and row_data.get("name") and row_data.get("client_type"):
                clients.append(row_data)

        logger.info(f"Loaded {len(clients)} client configurations")
        return clients

    def load_documents(self) -> list[DocumentSpec]:
        """Load document definitions from documents sheet.

        Returns:
            List of document config dictionaries with keys:
            - reference_name: unique identifier for referencing in prompts
            - common_name: human-readable name
            - file_path: path to document (relative to workbook)
            - tags: comma-separated tags for filtering (optional)
            - notes: optional description
            - chunking_strategy: inferred from file extension

        """
        if not self.has_documents_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.DOCUMENTS_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip() if header else f"col_{col}")

        documents = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value is not None:
                    has_content = True

            if has_content and row_data.get("reference_name") and row_data.get("file_path"):
                file_path = str(row_data.get("file_path", "")).strip()
                tags_raw = row_data.get("tags")
                tags_list = None
                if tags_raw:
                    tags_str = str(tags_raw).strip()
                    tags_list = [t.strip() for t in tags_str.split(",") if t.strip()]

                documents.append(
                    {
                        "reference_name": str(row_data.get("reference_name", "")).strip(),
                        "common_name": str(row_data.get("common_name", "")).strip()
                        if row_data.get("common_name")
                        else row_data.get("reference_name", ""),
                        "file_path": file_path,
                        "tags": tags_list,
                        "notes": str(row_data.get("notes", "")).strip()
                        if row_data.get("notes")
                        else None,
                        "chunking_strategy": self._infer_chunking_strategy(file_path),
                    }
                )

        logger.info(f"Loaded {len(documents)} document configurations")
        return documents

    def load_tools(self) -> list[dict[str, Any]]:
        """Load tool definitions from tools sheet.

        Returns:
            List of tool config dictionaries with keys:
            - name: unique identifier for the tool
            - description: human-readable description sent to LLM
            - parameters: JSON Schema for tool parameters
            - implementation: implementation reference (builtin:<name> or python:<module.func>)
            - enabled: whether the tool is available (default True)

        """
        if not self.has_tools_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.TOOLS_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip() if header else f"col_{col}")

        tools = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value is not None:
                    has_content = True

            if has_content and row_data.get("name"):
                enabled = row_data.get("enabled", True)
                if isinstance(enabled, str):
                    enabled = enabled.strip().lower() not in ("false", "0", "no")
                tools.append(
                    {
                        "name": str(row_data["name"]).strip(),
                        "description": str(row_data.get("description", "")).strip(),
                        "parameters": row_data.get("parameters", {}),
                        "implementation": str(row_data.get("implementation", "")).strip(),
                        "enabled": bool(enabled),
                    }
                )

        logger.info(f"Loaded {len(tools)} tool definitions")
        return tools

    def load_scoring(self) -> list[dict[str, Any]]:
        """Load scoring criteria from scoring sheet.

        Returns:
            List of scoring criteria dicts with keys:
            - criteria_name: machine-readable key
            - description: human-readable description
            - scale_min: minimum score value
            - scale_max: maximum score value
            - weight: base weight for aggregation
            - source_prompt: prompt name containing this score

        """
        if not self.has_scoring_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.SCORING_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip() if header else f"col_{col}")

        scoring = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value is not None:
                    has_content = True

            if has_content and row_data.get("criteria_name"):
                scale_min = row_data.get("scale_min", 1)
                scale_max = row_data.get("scale_max", 10)
                weight = row_data.get("weight", 1.0)
                if isinstance(scale_min, str):
                    scale_min = int(scale_min)
                if isinstance(scale_max, str):
                    scale_max = int(scale_max)
                if isinstance(weight, str):
                    weight = float(weight)

                scoring.append(
                    {
                        "criteria_name": str(row_data["criteria_name"]).strip(),
                        "description": str(row_data.get("description", "")).strip(),
                        "scale_min": int(scale_min),
                        "scale_max": int(scale_max),
                        "weight": float(weight),
                        "source_prompt": str(row_data.get("source_prompt", "")).strip(),
                        "score_type": str(row_data.get("score_type", "")).strip(),
                        "label_1": str(row_data.get("label_1", "")).strip(),
                        "label_2": str(row_data.get("label_2", "")).strip(),
                        "label_3": str(row_data.get("label_3", "")).strip(),
                    }
                )

        logger.info(f"Loaded {len(scoring)} scoring criteria")
        return scoring

    def load_synthesis(self) -> list[dict[str, Any]]:
        """Load synthesis prompts from synthesis sheet.

        Returns:
            List of synthesis prompt dicts with keys from SYNTHESIS_HEADERS.

        """
        if not self.has_synthesis_sheet():
            return []

        wb = load_workbook(self.workbook_path)
        ws = wb[self.SYNTHESIS_SHEET]

        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(str(header).strip().lower() if header else f"col_{col}")

        synthesis = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value is not None:
                    has_content = True

            if not has_content or not row_data.get("sequence") or not row_data.get("prompt"):
                continue

            sequence = row_data["sequence"]
            if isinstance(sequence, str):
                sequence = int(sequence.strip()) if sequence.strip().isdigit() else None
            if sequence is None:
                continue

            include_scores = row_data.get("include_scores")
            if isinstance(include_scores, str):
                include_scores_val = include_scores.strip().lower() in ("true", "1", "yes")
            elif isinstance(include_scores, bool):
                include_scores_val = include_scores
            else:
                include_scores_val = True

            synthesis.append(
                {
                    "sequence": int(sequence),
                    "prompt_name": str(row_data.get("prompt_name", "")).strip()
                    if row_data.get("prompt_name")
                    else None,
                    "prompt": str(row_data["prompt"]).strip(),
                    "source_scope": str(row_data.get("source_scope", "all")).strip(),
                    "source_prompts": parse_history_string(row_data.get("source_prompts"))
                    if row_data.get("source_prompts")
                    else [],
                    "include_scores": include_scores_val,
                    "history": parse_history_string(row_data.get("history"))
                    if row_data.get("history")
                    else None,
                    "condition": str(row_data.get("condition", "")).strip()
                    if row_data.get("condition")
                    else None,
                    "client": str(row_data.get("client", "")).strip()
                    if row_data.get("client")
                    else None,
                }
            )

        synthesis.sort(key=lambda x: x["sequence"])
        logger.info(f"Loaded {len(synthesis)} synthesis prompts")
        return synthesis

    def _infer_chunking_strategy(self, file_path: str) -> str:
        """Infer chunking strategy from file extension.

        Args:
            file_path: Path to the document file.

        Returns:
            Chunking strategy name: 'markdown', 'code', or 'recursive' (default).

        """
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".md":
            return "markdown"
        elif ext in {
            ".py",
            ".js",
            ".ts",
            ".jsx",
            ".tsx",
            ".java",
            ".go",
            ".rs",
            ".c",
            ".cpp",
            ".h",
            ".rb",
            ".php",
            ".swift",
            ".kt",
        }:
            return "code"
        else:
            return "recursive"

    def _parse_history_string(self, history_str: Any) -> list[str] | None:
        return parse_history_string(history_str)

    def write_results(self, results: list[dict[str, Any]], sheet_name: str) -> str:
        """Write execution results to a new sheet."""
        wb = load_workbook(self.workbook_path)
        ordered_results = sorted(
            results, key=lambda r: ((r.get("batch_id") or 0), r.get("sequence", 0))
        )

        if sheet_name in wb.sheetnames:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            sheet_name = f"results_{timestamp}"

        ws = wb.create_sheet(title=sheet_name)

        ws.append(self.RESULTS_HEADERS)

        for result in ordered_results:
            row = [
                _serialize_for_excel(header, result.get(header)) for header in self.RESULTS_HEADERS
            ]
            ws.append(row)

        self.formatter.apply_formatting(ws, "results")

        wb.save(self.workbook_path)
        logger.info(f"Results written to sheet: {sheet_name}")
        return sheet_name

    def write_batch_results(
        self,
        results: list[dict[str, Any]],
        batch_name: str,
        base_sheet_name: str = "results",
    ) -> str:
        """Write results for a single batch to a separate sheet."""
        wb = load_workbook(self.workbook_path)
        ordered_results = sorted(results, key=lambda r: r.get("sequence", 0))

        sheet_name = f"{base_sheet_name}_{batch_name}"
        original_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=sheet_name)

        ws.append(self.RESULTS_HEADERS)

        for result in ordered_results:
            row = [
                _serialize_for_excel(header, result.get(header)) for header in self.RESULTS_HEADERS
            ]
            ws.append(row)

        self.formatter.apply_formatting(ws, "results")

        wb.save(self.workbook_path)
        logger.info(f"Batch results written to sheet: {sheet_name}")
        return sheet_name

    PIVOT_HEADERS = PIVOT_COLUMNS

    def write_scores_pivot(
        self,
        results: list[dict[str, Any]],
        scoring_criteria: list[dict[str, Any]],
        sheet_name: str = "scores_pivot",
    ) -> str | None:
        """Write a scores pivot sheet using ResultsFrame for computation.

        Only criteria with score_type='normalized_score' are included.
        Delegates pivot computation to ResultsFrame.scores_pivot() and
        writes the resulting DataFrame to an Excel worksheet.

        Args:
            results: List of result dictionaries (must include batch_name, scores).
            scoring_criteria: List of scoring criteria dicts (from load_scoring).
            sheet_name: Name for the pivot worksheet.

        Returns:
            Name of the pivot sheet created, or None if no pivot-eligible criteria exist.

        """
        from .results import ResultsFrame

        frame = ResultsFrame(results)
        pivot_df = frame.scores_pivot(scoring_criteria)

        if pivot_df.is_empty():
            return None

        wb = load_workbook(self.workbook_path)

        if sheet_name in wb.sheetnames:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            sheet_name = f"scores_pivot_{timestamp}"

        ws = wb.create_sheet(title=sheet_name)

        for col_idx, header in enumerate(self.PIVOT_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row_data in enumerate(pivot_df.iter_rows(named=True), start=2):
            for col_idx, header in enumerate(self.PIVOT_HEADERS, start=1):
                val = row_data.get(header)
                if val is None:
                    val = ""
                ws.cell(row=row_idx, column=col_idx, value=val)

        self.formatter.apply_formatting(ws, "scoring")

        wb.save(self.workbook_path)
        logger.info(f"Scores pivot written to sheet: {sheet_name}")
        return sheet_name
